from __future__ import annotations

import base64
import json
import logging
import sqlite3
import zlib
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.dashboard import dashboard_data
from scripts.reports import build_storebook_zr_report as report

REMOVED_ZR_VISIBLE_COLUMNS = {
    "Cleared/open items symbol",
    "Posting Key",
    "Clearing date",
    "Clearing Document",
    "Net due date symbol",
}


def _storebook_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append([""] * 22)
    ws.append(
        [
            "Payment Batches",
            "Payment Number",
            "Status",
            "Blocked Reason",
            "Ext.Reference",
            "Header Doc.",
            "Created Date",
            "Auth. Date",
            "Export Date",
            "Manually Auth. By",
            "Type",
            "Financial Net Price COGS",
            "Financial Net Price OPEX",
            "Manual Authorisation Reason",
            "Valued Logistics",
            "Site",
            "Site",
            "Supplier",
            "Supplier",
            "Main Storebook #",
            "Main Vendor Doc.",
        ]
    )
    ws.append(
        [
            "",
            "7100001",
            "Blocked",
            "",
            "",
            "7200001",
            "2026-01-20 14:38:34",
            "",
            "",
            "",
            "",
            733.74,
            0,
            "",
            551.05,
            "9101",
            "Synthetic Site Alpha",
            "700001",
            "Synthetic Supplier Delta",
            "7200001",
            "8300000001",
        ]
    )
    wb.save(path)


def _zr_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(
        [
            "",
            "Unique",
            "Document Number",
            "Cleared/open items symbol",
            "Vendor",
            "Company Code",
            "Vendor Name 1",
            "Document Type",
            "Posting Key",
            "Reference",
            "Document Date",
            "Amount in local currency",
            "Local Currency",
            "Net due date",
            "Tax code",
            "Posting Date",
            "Text",
            "Clearing date",
            "Clearing Document",
            "Net due date symbol",
            "Payment Block",
            "User name",
        ]
    )
    ws.append(
        [
            "#N/A",
            "DO NOT USE",
            "8000000001",
            "",
            "700002",
            "9002",
            "Synthetic Supplier Gamma",
            "RE",
            "31",
            "SYNTHETIC-REF-2026-07",
            "2026-07-01 00:00:00",
            -916.3,
            "EUR",
            "2026-06-26 00:00:00",
            "VL",
            "2026-06-26 00:00:00",
            "MIETE",
            "2026-06-28 00:00:00",
            "",
            "",
            "Z",
            "SYNTHETIC-USER",
        ]
    )
    wb.save(path)


def _vendor_matrix(path: Path) -> None:
    path.write_text(
        "Company Code,Unique ref,Vendor,Name 1,Category\n"
        "9001,9001 700001,700001,Synthetic Supplier Delta,Bakery\n"
        "9002,9002 700002,700002,Synthetic Supplier Gamma,Rent\n",
        encoding="utf-8",
    )


def _previous_output(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Storebook"
    ws.append(report.STOREBOOK_OUTPUT_COLUMNS)
    ws.append(
        [
            "Synthetic Owner 001",
            "Blocked",
            "2026-01-20",
            733.74,
            "9101",
            "Synthetic Site Alpha",
            "700001",
            "Synthetic Supplier Delta",
            "7200001",
            "8300000001",
            "9001 700001",
            "Bakery",
            "historic note",
            "2026-06-29",
            "Resolved",
        ]
    )
    ws_zr = wb.create_sheet("Z & R")
    ws_zr.append(report.ZR_OUTPUT_COLUMNS)
    wb.save(path)


def test_zr_key_ignores_unique_column() -> None:
    row = {
        "Unique": "WRONG",
        "Company Code": "9002",
        "Document Number": "8000000001",
        "Vendor": "700002",
        "Reference": "SYNTHETIC-REF-2026-07",
        "Posting Date": "2026-06-26",
        "Payment Block": "Z",
    }

    assert report.make_zr_key(row) == "Z & R|9002|8000000001|700002|SYNTHETIC-REF-2026-07|2026-06-26|Z"


def test_storebook_zr_config_defaults_to_local_only_db_sync() -> None:
    assert report.StorebookZRConfig(snapshot_date=date(2026, 6, 30)).sync_db is False


def test_storebook_status_dropdown_matches_operational_options() -> None:
    assert report.STOREBOOK_STATUS_OPTIONS == [
        "Awaiting Invoice",
        "Waiting for site to respond",
        "Resolved",
        "Awaiting SMC/AM to confirm prices.",
        "Waiting for CS to correct.",
        "GR Issue",
    ]


def test_build_report_with_missing_sources_writes_empty_feed(tmp_path: Path, caplog) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    output_dir = tmp_path / "out"
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"
    db_path.parent.mkdir()
    columns = report.DB_COLUMNS + ["loaded_at"]
    stale_row = {
        "source": report.STOREBOOK_SOURCE,
        "snapshot_date": "2026-06-30",
        "source_key": "Storebook|stale",
        "loaded_at": "2026-06-30T08:00:00",
    }

    with sqlite3.connect(db_path) as conn:
        report.init_db(conn)
        conn.execute(
            f"INSERT INTO storebook_zr_lines ({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
            [stale_row.get(column, "") for column in columns],
        )
        conn.commit()

    caplog.set_level(logging.INFO, logger=report.__name__)
    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 7, 1),
            output_dir=output_dir,
            archive_dir=output_dir / "archive",
            source_dir=source_dir,
            reference_workbook=None,
            vendor_matrix_path=tmp_path / "missing_vendor_matrix.csv",
            db_path=db_path,
        )
    )

    assert result.rows_by_source == {report.STOREBOOK_SOURCE: 0, report.ZR_SOURCE: 0}
    assert result.output_path.exists()
    assert "No Storebook/Z&R source files found" in caplog.text

    wb = load_workbook(result.output_path, read_only=True, data_only=True)
    try:
        assert wb.sheetnames == [report.STOREBOOK_SOURCE, report.ZR_SOURCE]
        assert wb[report.STOREBOOK_SOURCE].max_row == 1
        assert wb[report.ZR_SOURCE].max_row == 1
    finally:
        wb.close()

    with sqlite3.connect(db_path) as conn:
        count = conn.execute("SELECT COUNT(*) FROM storebook_zr_lines").fetchone()[0]
    assert count == 0


def test_resolve_sources_accepts_dated_storebook_and_zr_files(tmp_path: Path) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    storebook_path = source_dir / "Storebook 06.07.2026.xlsx"
    zr_path = source_dir / "ZR 06.07.2026.XLSX"
    _storebook_source(storebook_path)
    _zr_source(zr_path)

    storebook_source, zr_source = report.resolve_sources(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 7, 6),
            source_dir=source_dir,
            reference_workbook=None,
        )
    )

    assert storebook_source == storebook_path
    assert zr_source == zr_path


def test_build_report_keeps_zr_rows_when_storebook_source_is_missing(tmp_path: Path) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    zr_path = source_dir / "ZR 06.07.2026.XLSX"
    _zr_source(zr_path)
    matrix_path = source_dir / "Vendor Master Matrix.csv"
    _vendor_matrix(matrix_path)
    output_dir = tmp_path / "out"
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"

    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 7, 6),
            output_dir=output_dir,
            archive_dir=output_dir / "archive",
            source_dir=source_dir,
            reference_workbook=None,
            vendor_matrix_path=matrix_path,
            db_path=db_path,
        )
    )

    assert result.rows_by_source == {report.STOREBOOK_SOURCE: 0, report.ZR_SOURCE: 1}
    with sqlite3.connect(db_path) as conn:
        count = conn.execute("SELECT COUNT(*) FROM storebook_zr_lines").fetchone()[0]
        sources = conn.execute("SELECT source FROM storebook_zr_lines").fetchall()
    assert count == 1
    assert sources == [(report.ZR_SOURCE,)]


def test_build_report_writes_workbook_sqlite_and_preserves_tail(tmp_path: Path) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    storebook_path = source_dir / "PaymentBatch.xlsx"
    zr_path = source_dir / "30.06.2026.XLSX"
    _storebook_source(storebook_path)
    _zr_source(zr_path)
    matrix_path = source_dir / "Vendor Master Matrix.csv"
    _vendor_matrix(matrix_path)

    output_dir = tmp_path / "out"
    archive_dir = output_dir / "archive"
    output_dir.mkdir()
    _previous_output(output_dir / "Storebook_ZR_2026-06-29.xlsx")
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"

    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 6, 30),
            output_dir=output_dir,
            archive_dir=archive_dir,
            storebook_source=storebook_path,
            zr_source=zr_path,
            reference_workbook=None,
            vendor_matrix_path=matrix_path,
            db_path=db_path,
        )
    )

    assert result.output_path.name == "Storebook_ZR_2026-06-30.xlsx"
    assert result.rows_by_source == {"Storebook": 1, "Z & R": 1}
    assert (archive_dir / "Storebook_ZR_2026-06-29.xlsx").exists()
    assert result.output_path.exists()

    wb = load_workbook(result.output_path, read_only=True, data_only=True)
    try:
        assert wb.sheetnames == ["Storebook", "Z & R"]
        headers = [cell.value for cell in next(wb["Storebook"].iter_rows(min_row=1, max_row=1))]
        assert headers == report.STOREBOOK_OUTPUT_COLUMNS
        assert headers[0] == "Owner"
        zr_headers = [cell.value for cell in next(wb["Z & R"].iter_rows(min_row=1, max_row=1))]
        assert zr_headers == report.ZR_OUTPUT_COLUMNS
        assert "Total Amount" not in zr_headers
        assert REMOVED_ZR_VISIBLE_COLUMNS.isdisjoint(zr_headers)
        status_idx = headers.index("Status")
        comments_idx = headers.index("Comments")
        action_date_idx = headers.index("Action Date")
        unique_ref_idx = headers.index("Unique Ref")
        category_idx = headers.index("Category")
        row = [cell.value for cell in next(wb["Storebook"].iter_rows(min_row=2, max_row=2))]
        assert row[0] == "Synthetic Owner 001"
        assert row[unique_ref_idx] == "9001 700001"
        assert row[category_idx] == "Bakery"
        assert row[status_idx] == "Resolved"
        assert row[comments_idx] == "historic note"
        assert row[action_date_idx] == "2026-06-29"
    finally:
        wb.close()

    with sqlite3.connect(db_path) as conn:
        count = conn.execute("SELECT COUNT(*) FROM storebook_zr_lines").fetchone()[0]
        statuses = conn.execute(
            "SELECT status FROM storebook_zr_lines WHERE source = ?",
            ("Storebook",),
        ).fetchone()
    assert count == 2
    assert statuses == ("Resolved",)


def test_output_dropdown_values_and_visible_columns(tmp_path: Path) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    storebook_path = source_dir / "PaymentBatch.xlsx"
    zr_path = source_dir / "30.06.2026.XLSX"
    matrix_path = source_dir / "Vendor Master Matrix.csv"
    _storebook_source(storebook_path)
    _zr_source(zr_path)
    _vendor_matrix(matrix_path)

    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 6, 30),
            output_dir=tmp_path / "out",
            archive_dir=tmp_path / "out" / "archive",
            storebook_source=storebook_path,
            zr_source=zr_path,
            reference_workbook=None,
            vendor_matrix_path=matrix_path,
            db_path=tmp_path / "db" / "storebook_zr_daily.sqlite",
            sync_db=False,
        )
    )

    wb = load_workbook(result.output_path, data_only=True)
    try:
        assert [cell.value for cell in wb["Storebook"][1]] == report.STOREBOOK_OUTPUT_COLUMNS
        zr_headers = [cell.value for cell in wb["Z & R"][1]]
        assert zr_headers == report.ZR_OUTPUT_COLUMNS
        assert "Total Amount" not in zr_headers
        assert REMOVED_ZR_VISIBLE_COLUMNS.isdisjoint(zr_headers)
        validation_formulas = {
            ws_name: {dv.formula1 for dv in wb[ws_name].data_validations.dataValidation}
            for ws_name in ("Storebook", "Z & R")
        }
        assert f'"{",".join(report.STOREBOOK_STATUS_OPTIONS)}"' in validation_formulas["Storebook"]
        assert f'"{",".join(report.ZR_STATUS_OPTIONS)}"' in validation_formulas["Z & R"]
    finally:
        wb.close()


def test_source_parsing_row_counts_and_contract_fields(tmp_path: Path) -> None:
    storebook_path = tmp_path / "PaymentBatch.xlsx"
    zr_path = tmp_path / "30.06.2026.XLSX"
    _storebook_source(storebook_path)
    _zr_source(zr_path)
    matrix_path = tmp_path / "Vendor Master Matrix.csv"
    _vendor_matrix(matrix_path)
    category_lookup = report.load_vendor_category_lookup(matrix_path)

    storebook_rows = report._read_storebook_rows(storebook_path, date(2026, 6, 30), category_lookup)
    zr_rows = report._read_zr_rows(zr_path, date(2026, 6, 30), category_lookup)

    assert len(storebook_rows) == 1
    assert len(zr_rows) == 1
    assert set(report.COMMON_FIELDS).issubset(storebook_rows[0])
    assert set(report.COMMON_FIELDS).issubset(zr_rows[0])
    assert storebook_rows[0]["owner"] == "Synthetic Owner 001"
    assert zr_rows[0]["owner"] == "Synthetic Owner 001"
    assert report.STOREBOOK_OUTPUT_COLUMNS[0] == "Owner"
    assert storebook_rows[0]["Owner"] == "Synthetic Owner 001"
    assert storebook_rows[0]["Unique Ref"] == "9001 700001"
    assert storebook_rows[0]["unique_ref"] == "9001 700001"
    assert storebook_rows[0]["company_code"] == "9001"
    assert storebook_rows[0]["company_or_entity"] == "9001"
    assert storebook_rows[0]["status_system"] == "Blocked"
    assert storebook_rows[0]["Category"] == "Bakery"
    assert zr_rows[0]["unique_ref"] == "9002 700002"
    assert zr_rows[0]["company_code"] == "9002"
    assert zr_rows[0]["company_or_entity"] == "9002"
    assert zr_rows[0]["Category"] == "Rent"
    assert zr_rows[0]["resolved_date"] == ""
    assert REMOVED_ZR_VISIBLE_COLUMNS.isdisjoint(report.ZR_OUTPUT_COLUMNS)
    assert REMOVED_ZR_VISIBLE_COLUMNS.isdisjoint(zr_rows[0])


def test_zr_resolved_date_uses_action_date_for_removed_status_not_clearing_date() -> None:
    rows = report.normalize_zr_rows(
        report.pd.DataFrame(
            [
                {
                    "Document Number": "8000000001",
                    "Vendor": "700002",
                    "Company Code": "9002",
                    "Vendor Name 1": "Synthetic Supplier Gamma",
                    "Document Type": "RE",
                    "Reference": "SYNTHETIC-REF-2026-07",
                    "Document Date": "2026-07-01 00:00:00",
                    "Amount in local currency": -916.3,
                    "Local Currency": "EUR",
                    "Net due date": "2026-06-26 00:00:00",
                    "Tax code": "VL",
                    "Posting Date": "2026-06-26 00:00:00",
                    "Text": "MIETE",
                    "Clearing date": "2026-06-28 00:00:00",
                    "Payment Block": "Z",
                    "User name": "SYNTHETIC-USER",
                    "Status": "Removed",
                    "action date": "2026-06-30",
                }
            ]
        ),
        date(2026, 6, 30),
        {"unique_ref:9002 700002": "Rent"},
    )

    assert rows[0]["resolved_date"] == "2026-06-30"
    metrics = report.compute_window_metrics(rows, date(2026, 6, 30), date(2026, 6, 30))
    assert metrics["resolved_count"] == 1
    assert metrics["productivity_count"] == 1


def test_zr_normalization_skips_blank_source_rows() -> None:
    rows = report.normalize_zr_rows(
        report.pd.DataFrame(
            [
                {
                    "Document Number": "",
                    "Vendor": "",
                    "Company Code": "",
                    "Reference": "",
                    "Posting Date": "",
                    "Payment Block": "",
                },
                {
                    "Document Number": "8000000001",
                    "Vendor": "700002",
                    "Company Code": "9002",
                    "Reference": "SYNTHETIC-REF-2026-07",
                    "Posting Date": "2026-06-26",
                    "Payment Block": "Z",
                },
            ]
        ),
        date(2026, 6, 30),
        {"unique_ref:9002 700002": "Rent"},
    )

    assert len(rows) == 1
    assert rows[0]["source_key"] == "Z & R|9002|8000000001|700002|SYNTHETIC-REF-2026-07|2026-06-26|Z"


def test_tail_recomputes_resolved_date_from_manual_status_and_action_date() -> None:
    rows = [
        {
            "source": report.STOREBOOK_SOURCE,
            "source_key": "sb1",
            "Status": "",
            "Comments": "",
            "Action Date": "",
            "resolved_date": "",
        },
        {
            "source": report.ZR_SOURCE,
            "source_key": "zr1",
            "Status": "",
            "Comments": "",
            "action date": "",
            "resolved_date": "",
        },
    ]
    tail = {
        "sb1": {"Status": "Resolved", "Action Date": "2026-06-29"},
        "zr1": {"Status": "Removed", "Action Date": "2026-06-28"},
    }

    report.apply_tail(rows, tail)

    assert rows[0]["resolved_date"] == "2026-06-29"
    assert rows[1]["action date"] == "2026-06-28"
    assert rows[1]["resolved_date"] == "2026-06-28"


def test_tail_reads_legacy_storebook_blocked_invocies_sheet(tmp_path: Path) -> None:
    path = tmp_path / "Storebook_ZR_2026-07-06.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Storebook Blocked Invocies"
    ws.append(report.STOREBOOK_OUTPUT_COLUMNS)
    ws.append(
        [
            "Synthetic Owner 001",
            "Blocked",
            "2026-01-20",
            733.74,
            "9101",
            "Synthetic Site Alpha",
            "700001",
            "Synthetic Supplier Delta",
            "7200001",
            "8300000001",
            "9001 700001",
            "Bakery",
            "legacy comment",
            "2026-07-06",
            "Price Variance",
        ]
    )
    wb.save(path)

    tail = report._capture_tail_from_workbook(path)
    key = "Storebook|tail|7200001|8300000001|9001 700001"

    assert tail[key]["Comments"] == "legacy comment"
    assert tail[key]["Action Date"] == "2026-07-06"
    assert tail[key]["Status"] == "Price Variance"


def test_tail_candidates_include_current_output_first(tmp_path: Path) -> None:
    output_dir = tmp_path / "out"
    archive_dir = output_dir / "archive"
    archive_dir.mkdir(parents=True)
    current = output_dir / "Storebook_ZR_2026-07-07.xlsx"
    previous = archive_dir / "Storebook_ZR_2026-07-06.xlsx"
    current.touch()
    previous.touch()

    candidates = report._previous_output_candidates(output_dir, archive_dir, current)

    assert candidates[0] == current
    assert previous in candidates


def _previous_zr_row(source_key: str, status: str = "Awaiting Response") -> dict[str, object]:
    return {
        "source": report.ZR_SOURCE,
        "snapshot_date": "2026-06-30",
        "source_key": source_key,
        "unique_ref": "9002 700002",
        "owner": "Synthetic Owner 001",
        "supplier_id": "700002",
        "supplier_name": "Synthetic Supplier Gamma",
        "company_or_entity": "9002",
        "company_code": "9002",
        "category": "Rent",
        "status_system": "",
        "value": -916.3,
        "opened_date": "2026-06-01",
        "action_date": "",
        "resolved_date": "",
        "status": status,
        "comments": "",
        "document_number": "8000000001",
        "vendor_id": "700002",
        "reference": "SYNTHETIC-REF-2026-07",
        "posting_date": "2026-06-26",
        "payment_block": "Z",
    }


def test_zr_previous_open_row_missing_from_current_source_auto_resolves() -> None:
    rows: list[dict[str, object]] = []
    previous = [_previous_zr_row("Z & R|9002|8000000001|700002|SYNTHETIC-REF-2026-07|2026-06-26|Z")]

    auto_rows = report.apply_zr_missing_source_auto_resolutions(rows, previous, date(2026, 7, 1))

    assert rows == []
    assert len(auto_rows) == 1
    assert auto_rows[0]["snapshot_date"] == "2026-07-01"
    assert auto_rows[0]["resolved_date"] == "2026-07-01"
    assert auto_rows[0]["status"] == "Auto Resolved - Missing From Source"
    assert auto_rows[0]["resolution_source"] == "auto_missing_from_source"
    assert auto_rows[0]["comments"] == "Auto Resolved - Missing From Source"


def test_build_report_writes_auto_resolved_only_to_sqlite_not_workbook(
    tmp_path: Path,
    monkeypatch,
) -> None:
    current = [_previous_zr_row("Z & R|current")]
    current[0]["snapshot_date"] = "2026-07-01"
    missing = [_previous_zr_row("Z & R|missing")]
    captured: dict[str, list[dict[str, object]]] = {}

    monkeypatch.setattr(report, "load_rows", lambda config: current)
    monkeypatch.setattr(report, "capture_tail", lambda config, output_path: {})
    monkeypatch.setattr(report, "load_previous_zr_snapshot_rows", lambda db_path, snapshot_date: missing)
    monkeypatch.setattr(
        report,
        "write_workbook",
        lambda rows, output_path: captured.setdefault("workbook", list(rows)) or 0,
    )
    monkeypatch.setattr(
        report,
        "write_sqlite",
        lambda rows, db_path: captured.setdefault("sqlite", list(rows)),
    )
    monkeypatch.setattr(report, "archive_previous_day_output", lambda *args: [])

    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 7, 1),
            output_dir=tmp_path / "out",
            archive_dir=tmp_path / "archive",
            storebook_source=None,
            zr_source=None,
            reference_workbook=None,
            vendor_matrix_path=tmp_path / "matrix.csv",
            db_path=tmp_path / "storebook_zr.sqlite",
            sync_db=False,
        )
    )

    assert [row["source_key"] for row in captured["workbook"]] == ["Z & R|current"]
    assert [row["source_key"] for row in captured["sqlite"]] == ["Z & R|current", "Z & R|missing"]
    assert result.rows_by_source == {report.STOREBOOK_SOURCE: 0, report.ZR_SOURCE: 1}


def test_previous_zr_tail_status_prevents_false_auto_resolution(tmp_path: Path) -> None:
    source_dir = tmp_path / "sources"
    source_dir.mkdir()
    storebook_path = source_dir / "PaymentBatch.xlsx"
    zr_path = source_dir / "ZR 01.07.2026.XLSX"
    matrix_path = source_dir / "Vendor Master Matrix.csv"
    _storebook_source(storebook_path)
    _zr_source(zr_path)
    _vendor_matrix(matrix_path)

    output_dir = tmp_path / "out"
    output_dir.mkdir()
    archive_dir = output_dir / "archive"
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"
    db_path.parent.mkdir()
    previous_key = "Z & R|9002|8000000000|700002|OLD REF|2026-06-26|Z"
    previous_row = _previous_zr_row(previous_key, status="")
    previous_row.update(
        {
            "document_number": "8000000000",
            "reference": "OLD REF",
            "comments": "",
            "action_date": "",
            "resolved_date": "",
        }
    )

    with sqlite3.connect(db_path) as conn:
        report.init_db(conn)
        columns = report.DB_COLUMNS + ["loaded_at"]
        conn.execute(
            f"INSERT INTO storebook_zr_lines ({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
            [previous_row.get(column, "2026-06-30T08:00:00" if column == "loaded_at" else "") for column in columns],
        )
        conn.commit()

    previous_wb = Workbook()
    previous_wb.remove(previous_wb.active)
    previous_wb.create_sheet("Storebook").append(report.STOREBOOK_OUTPUT_COLUMNS)
    previous_ws = previous_wb.create_sheet("Z & R")
    previous_ws.append(report.ZR_OUTPUT_COLUMNS)
    previous_ws.append(
        [
            "Synthetic Owner 001",
            "9002 700002",
            "8000000000",
            "700002",
            "9002",
            "Synthetic Supplier Gamma",
            "",
            "OLD REF",
            "2026-06-01",
            -916.3,
            "",
            "",
            "",
            "2026-06-26",
            "",
            "Z",
            "",
            "Rent",
            "Removed",
            "2026-06-30",
            "Removed by operator",
        ]
    )
    previous_wb.save(output_dir / "Storebook_ZR_2026-06-30.xlsx")

    result = report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 7, 1),
            output_dir=output_dir,
            archive_dir=archive_dir,
            storebook_source=storebook_path,
            zr_source=zr_path,
            reference_workbook=None,
            vendor_matrix_path=matrix_path,
            db_path=db_path,
            sync_db=False,
        )
    )

    assert result.rows_by_source == {"Storebook": 1, "Z & R": 1}
    with sqlite3.connect(db_path) as conn:
        auto_rows = conn.execute(
            """
            SELECT COUNT(*)
            FROM storebook_zr_lines
            WHERE snapshot_date = ?
              AND source = ?
              AND status = ?
            """,
            ("2026-07-01", report.ZR_SOURCE, report.AUTO_MISSING_FROM_SOURCE_STATUS),
        ).fetchone()[0]
    assert auto_rows == 0


def test_current_zr_removed_blank_action_date_does_not_get_fake_resolved_date() -> None:
    rows = report.normalize_zr_rows(
        report.pd.DataFrame(
            [
                {
                    "Document Number": "8000000001",
                    "Vendor": "700002",
                    "Company Code": "9002",
                    "Reference": "SYNTHETIC-REF-2026-07",
                    "Posting Date": "2026-06-26",
                    "Payment Block": "Z",
                    "Status": "Removed",
                    "action date": "",
                }
            ]
        ),
        date(2026, 7, 1),
        {},
    )

    report.apply_zr_missing_source_auto_resolutions(rows, [], date(2026, 7, 1))

    assert rows[0]["resolved_date"] == ""
    assert rows[0].get("resolution_source", "") == ""


def test_reappeared_zr_key_stays_active_from_current_source() -> None:
    source_key = "Z & R|9002|8000000001|700002|SYNTHETIC-REF-2026-07|2026-06-26|Z"
    rows = [_previous_zr_row(source_key)]
    rows[0]["snapshot_date"] = "2026-07-01"
    previous = [_previous_zr_row(source_key)]

    report.apply_zr_missing_source_auto_resolutions(rows, previous, date(2026, 7, 1))

    assert len(rows) == 1
    assert rows[0]["resolved_date"] == ""
    assert rows[0]["status"] == "Awaiting Response"
    assert rows[0].get("resolution_source", "") == ""


def test_storebook_rows_are_not_auto_resolved_by_zr_missing_rule() -> None:
    rows: list[dict[str, object]] = []
    previous = [
        {
            "source": report.STOREBOOK_SOURCE,
            "snapshot_date": "2026-06-30",
            "source_key": "Storebook|old",
            "resolved_date": "",
            "status": "Price Variance",
        }
    ]

    report.apply_zr_missing_source_auto_resolutions(rows, previous, date(2026, 7, 1))

    assert rows == []


def test_duplicate_detection_reports_stable_keys(tmp_path: Path) -> None:
    storebook_path = tmp_path / "PaymentBatch.xlsx"
    _storebook_source(storebook_path)
    rows = report._read_storebook_rows(storebook_path, date(2026, 6, 30), {})
    duplicated = rows + [dict(rows[0])]

    duplicate_keys = report.detect_duplicate_keys(duplicated)
    deduped = report.dedupe_rows(duplicated)

    assert duplicate_keys == ["Storebook|7200001|8300000001|7100001"]
    assert len(deduped) == 1


def test_tail_carry_forward_only_status_comments_action_date() -> None:
    rows = [
        {
            "source_key": "k1",
            "Status": "",
            "Comments": "",
            "Action Date": "",
            "Category": "Keep Category",
            "resolved_date": "",
        }
    ]
    tail = {
        "k1": {
            "Status": "Resolved",
            "Comments": "historic note",
            "Action Date": "2026-06-29",
            "Category": "Do not carry",
            "resolved_date": "2026-06-29",
        }
    }

    report.apply_tail(rows, tail)

    assert rows[0]["Status"] == "Resolved"
    assert rows[0]["Comments"] == "historic note"
    assert rows[0]["Action Date"] == "2026-06-29"
    assert rows[0]["Category"] == "Keep Category"
    assert rows[0]["resolved_date"] == "2026-06-29"


def test_dashboard_payload_exports_stable_storebook_zr_ui_fields(tmp_path: Path, monkeypatch) -> None:
    storebook_path = tmp_path / "PaymentBatch.xlsx"
    zr_path = tmp_path / "30.06.2026.XLSX"
    matrix_path = tmp_path / "Vendor Master Matrix.csv"
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"
    _storebook_source(storebook_path)
    _zr_source(zr_path)
    _vendor_matrix(matrix_path)
    reference_path = tmp_path / "Storebook ZR KPI.xlsx"
    _previous_output(reference_path)

    report.build_report(
        report.StorebookZRConfig(
            snapshot_date=date(2026, 6, 30),
            output_dir=tmp_path / "out",
            archive_dir=tmp_path / "out" / "archive",
            storebook_source=storebook_path,
            zr_source=zr_path,
            reference_workbook=reference_path,
            vendor_matrix_path=matrix_path,
            db_path=db_path,
            sync_db=False,
        )
    )
    monkeypatch.setattr(dashboard_data, "STOREBOOK_ZR_DB", db_path)

    payload = dashboard_data.load_storebook_zr_from_sqlite()
    compressed = dashboard_data._compress_dashboard_payload(payload)
    decoded = json.loads(zlib.decompress(base64.b64decode(compressed)).decode("utf-8"))
    rows = decoded["rows"]
    storebook = next(row for row in rows if row["source"] == report.STOREBOOK_SOURCE)
    zr = next(row for row in rows if row["source"] == report.ZR_SOURCE)

    assert storebook["status"] == "Resolved"
    assert storebook["status_system"] == "Blocked"
    assert storebook["unique_ref"] == "9001 700001"
    assert storebook["company_code"] == "9001"
    assert storebook["company_or_entity"] == "9001"
    assert storebook["category"] == "Bakery"
    assert storebook["action_date"] == "2026-06-29"
    assert storebook["resolved_date"] == "2026-06-29"
    assert storebook["site_id"] == "9101"
    assert storebook["site_name"] == "Synthetic Site Alpha"
    assert zr["unique_ref"] == "9002 700002"
    assert zr["company_code"] == "9002"
    assert zr["company_or_entity"] == "9002"
    assert zr["category"] == "Rent"


def test_dashboard_payload_includes_storebook_history_for_csv_export(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"
    db_path.parent.mkdir()
    matrix_dir = tmp_path / "master"
    matrix_dir.mkdir()
    _vendor_matrix(matrix_dir / "Synthetic_Vendor_Master_Matrix.csv")

    columns = report.DB_COLUMNS + ["loaded_at"]
    rows = [
        {
            "source": report.STOREBOOK_SOURCE,
            "snapshot_date": "2026-06-30",
            "source_key": "Storebook|old",
            "unique_ref": None,
            "owner": "Synthetic Owner 001",
            "supplier_id": "700001",
            "supplier_name": "Synthetic Supplier Delta",
            "company_or_entity": "Synthetic Site Alpha",
            "company_code": "",
            "category": "",
            "status_system": "Blocked",
            "value": 733.74,
            "opened_date": "2026-01-20",
            "action_date": "2026-06-29",
            "resolved_date": "2026-06-29",
            "status": "Resolved",
            "comments": "historic manual status",
            "site_id": "9101",
            "site_name": "Synthetic Site Alpha",
            "loaded_at": "2026-06-30T10:00:00",
        },
        {
            "source": report.STOREBOOK_SOURCE,
            "snapshot_date": "2026-07-01",
            "source_key": "Storebook|new",
            "unique_ref": "9001 700001",
            "owner": "Synthetic Owner 001",
            "supplier_id": "700001",
            "supplier_name": "Synthetic Supplier Delta",
            "company_or_entity": "9001",
            "company_code": "9001",
            "category": "Bakery",
            "status_system": "Blocked",
            "value": 733.74,
            "opened_date": "2026-01-20",
            "action_date": "",
            "resolved_date": "",
            "status": "Price Variance",
            "comments": "",
            "site_id": "9101",
            "site_name": "Synthetic Site Alpha",
            "loaded_at": "2026-07-01T10:00:00",
        },
    ]

    with sqlite3.connect(db_path) as conn:
        report.init_db(conn)
        column_sql = ", ".join(columns)
        placeholders = ", ".join("?" for _ in columns)
        conn.executemany(
            f"INSERT INTO storebook_zr_lines ({column_sql}) VALUES ({placeholders})",
            [tuple(row.get(column, "") for column in columns) for row in rows],
        )
        conn.commit()

    monkeypatch.setattr(dashboard_data, "STOREBOOK_ZR_DB", db_path)
    monkeypatch.setattr(dashboard_data, "MASTER_DATA", matrix_dir, raising=False)

    payload = dashboard_data.load_storebook_zr_from_sqlite()
    storebook_rows = [row for row in payload["rows"] if row["source"] == report.STOREBOOK_SOURCE]

    assert len(storebook_rows) == 2
    assert {row["snapshot_date"] for row in storebook_rows} == {"2026-06-30", "2026-07-01"}
    assert all(row["unique_ref"] == "9001 700001" for row in storebook_rows)
    assert all(row["company_code"] == "9001" for row in storebook_rows)
    assert all(row["company_or_entity"] == "9001" for row in storebook_rows)
    assert all(row["category"] == "Bakery" for row in storebook_rows)
    latest = next(row for row in storebook_rows if row["snapshot_date"] == "2026-07-01")
    assert latest["status"] == "Price Variance"
    assert latest["status_system"] == "Blocked"


def test_dashboard_payload_derives_zr_human_action_date_with_provenance(
    tmp_path: Path,
    monkeypatch,
) -> None:
    db_path = tmp_path / "db" / "storebook_zr_daily.sqlite"
    db_path.parent.mkdir()
    matrix_dir = tmp_path / "master"
    matrix_dir.mkdir()
    columns = report.DB_COLUMNS + ["loaded_at"]
    base = {
        "source": report.ZR_SOURCE,
        "source_key": "Z & R|human",
        "unique_ref": "9002 700002",
        "owner": "Synthetic Owner 001",
        "supplier_id": "700002",
        "supplier_name": "Supplier",
        "company_or_entity": "9002",
        "company_code": "9002",
        "opened_date": "2026-06-30",
        "action_date": "",
        "resolved_date": "",
        "resolution_source": "",
    }
    rows = [
        {**base, "snapshot_date": "2026-06-30", "status": "", "comments": ""},
        {**base, "snapshot_date": "2026-07-01", "status": "Awaiting Response", "comments": ""},
        {**base, "snapshot_date": "2026-07-02", "status": "Awaiting Response", "comments": ""},
        {
            **base,
            "source_key": "Z & R|auto",
            "snapshot_date": "2026-07-02",
            "status": report.AUTO_MISSING_FROM_SOURCE_STATUS,
            "comments": report.AUTO_MISSING_FROM_SOURCE_STATUS,
            "resolved_date": "2026-07-02",
            "resolution_source": report.AUTO_MISSING_FROM_SOURCE_CODE,
        },
        {
            **base,
            "source_key": "Z & R|first-manual",
            "snapshot_date": "2026-07-02",
            "status": "Awaiting TL",
            "comments": "",
        },
        {
            **base,
            "source_key": "Z & R|explicit-manual",
            "snapshot_date": "2026-07-02",
            "action_date": "2026-06-29",
            "status": "Awaiting Response",
            "comments": "Operator supplied the date",
        },
    ]
    with sqlite3.connect(db_path) as conn:
        report.init_db(conn)
        conn.executemany(
            f"INSERT INTO storebook_zr_lines ({', '.join(columns)}) VALUES ({', '.join('?' for _ in columns)})",
            [
                tuple(
                    row.get(column, "2026-07-02T10:00:00" if column == "loaded_at" else "")
                    for column in columns
                )
                for row in rows
            ],
        )
        conn.commit()

    monkeypatch.setattr(dashboard_data, "STOREBOOK_ZR_DB", db_path)
    monkeypatch.setattr(dashboard_data, "MASTER_DATA", matrix_dir, raising=False)

    payload = dashboard_data.load_storebook_zr_from_sqlite()
    human_rows = sorted(
        (row for row in payload["rows"] if row["source_key"] == "Z & R|human"),
        key=lambda row: row["snapshot_date"],
    )
    auto_row = next(row for row in payload["rows"] if row["source_key"] == "Z & R|auto")
    first_manual = next(row for row in payload["rows"] if row["source_key"] == "Z & R|first-manual")
    explicit_manual = next(row for row in payload["rows"] if row["source_key"] == "Z & R|explicit-manual")

    assert human_rows[0]["action_date"] == ""
    assert human_rows[1]["action_date"] == "2026-07-01"
    assert human_rows[1]["action_date_source"] == "derived_manual_change"
    assert human_rows[2]["action_date"] == "2026-07-01"
    assert human_rows[2]["action_date_source"] == "carried_from_history"
    assert auto_row["action_date"] == ""
    assert auto_row["action_date_source"] == ""
    assert first_manual["action_date"] == "2026-07-02"
    assert first_manual["action_date_source"] == "derived_manual_change"
    assert explicit_manual["action_date"] == "2026-06-29"
    assert explicit_manual["action_date_source"] == "manual"


def test_metric_window_semantics() -> None:
    rows = [
        {"opened_date": "2026-06-01", "resolved_date": "", "action_date": "2026-06-30"},
        {"opened_date": "2026-06-01", "resolved_date": "2026-06-15", "action_date": "2026-06-15"},
        {"opened_date": "2026-05-01", "resolved_date": "2026-05-30", "action_date": "2026-05-30"},
    ]

    metrics = report.compute_window_metrics(rows, date(2026, 6, 1), date(2026, 6, 30))

    assert metrics["open_count"] == 1
    assert metrics["resolved_count"] == 1
    assert metrics["productivity_count"] == 2
    assert metrics["detail_active_count"] == 2


def test_runner_integration_commands_present() -> None:
    root = Path(__file__).resolve().parents[1]
    command = "scripts.reports.build_storebook_zr_report"

    assert command in (root / "automation" / "RUN_DAILY.bat").read_text(encoding="utf-8")
    assert command in (root / "automation" / "RUN_FULL.bat").read_text(encoding="utf-8")
    assert command in (root / "automation" / "RUN_REBUILD_ALL.bat").read_text(encoding="utf-8")
