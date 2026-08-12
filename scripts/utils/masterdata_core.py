# -*- coding: utf-8 -*-
"""MasterData Core  shared functions for ERP vendor master data processing."""
from __future__ import annotations

import csv
import logging
import shutil
import glob
import os
import re
from functools import lru_cache
from pathlib import Path
from datetime import date, datetime
from typing import Callable, Dict, Iterable, Optional, Tuple, List, Set

import pandas as pd


logger = logging.getLogger(__name__)


def norm_str_series(s):
    """Normalize a Series or scalar: fill NaN, replace NBSP, strip whitespace."""
    if isinstance(s, pd.Series):
        return s.fillna("").astype(str).str.replace("\u00A0", " ", regex=False).str.strip()
    return str(s).replace("\u00A0", " ").strip() if s is not None else ""


def append_df_to_csv(
    df: pd.DataFrame,
    out_csv: str,
    *,
    header: bool,
    encoding: str = "utf-8-sig",
    retries: int = 5,
    retry_sleep_s: float = 0.5,
) -> None:
    """
    Append a DataFrame to a CSV on disk.

    Mantem o comportamento original (append em CSV), mas trata um problema comum
    em Windows/Local Fixture Store: OSError: [Errno 22] Invalid argument.

    Estrategia:
      1) Tenta o caminho original (pandas to_csv com mode="a") com retry
      2) Se falhar com Errno 22, faz fallback "text-safe":
         - escreve o chunk em um arquivo temporario
         - le como texto (utf-8-sig) e anexa em modo texto ("a") com retry
      3) Se ainda falhar com Errno 22, faz fallback "replace-safe":
         - cria um arquivo temporario final copiando o existente (se houver)
         - anexa o chunk e faz os.replace(temp, out_csv)
    """
    out_dir = os.path.dirname(out_csv) or "."
    os.makedirs(out_dir, exist_ok=True)

    import time
    last_err: OSError | None = None

    # 1) Tentativa padrao (igual original)
    for _ in range(max(1, retries)):
        size_before = os.path.getsize(out_csv) if os.path.exists(out_csv) else 0
        try:
            df.to_csv(out_csv, mode="a", index=False, header=header, encoding=encoding)
            return
        except OSError as e:
            last_err = e
            if getattr(e, "errno", None) != 22:
                raise
            # If file grew, data was already written despite the error  do NOT retry
            size_after = os.path.getsize(out_csv) if os.path.exists(out_csv) else 0
            if size_after > size_before:
                return
            time.sleep(retry_sleep_s)

    # 2) Fallback text-safe
    import tempfile

    fd, tmp_chunk = tempfile.mkstemp(prefix="md_chunk_", suffix=".csv", dir=out_dir)
    os.close(fd)

    try:
        df.to_csv(tmp_chunk, index=False, header=header, encoding=encoding)

        for _ in range(max(1, retries)):
            try:
                with open(tmp_chunk, "rb") as fin:
                    data_bytes = fin.read()

                with open(out_csv, "ab") as fout:
                    fout.write(data_bytes)
                return
            except OSError as e:
                last_err = e
                if getattr(e, "errno", None) != 22:
                    raise
                time.sleep(retry_sleep_s)
    finally:
        try:
            os.remove(tmp_chunk)
        except OSError:
            pass

    # 3) Fallback replace-safe (evita append direto no arquivo do Local Fixture Store)
    fd, tmp_final = tempfile.mkstemp(prefix="md_out_", suffix=".csv", dir=out_dir)
    os.close(fd)

    try:
        # copia o existente (se houver) para o temp final
        if os.path.exists(out_csv):
            with open(out_csv, "rb") as src_f, open(tmp_final, "wb") as dst_f:
                shutil.copyfileobj(src_f, dst_f)

        # escreve o chunk como bytes e anexa no temp final
        fd2, tmp_chunk2 = tempfile.mkstemp(prefix="md_chunk_", suffix=".csv", dir=out_dir)
        os.close(fd2)
        try:
            df.to_csv(tmp_chunk2, index=False, header=header, encoding=encoding)
            with open(tmp_chunk2, "rb") as csrc, open(tmp_final, "ab") as cdst:
                shutil.copyfileobj(csrc, cdst)
        finally:
            try:
                os.remove(tmp_chunk2)
            except OSError:
                pass

        os.replace(tmp_final, out_csv)
        return
    except OSError as e:
        last_err = e
        raise
    finally:
        try:
            if os.path.exists(tmp_final):
                os.remove(tmp_final)
        except OSError:
            pass

    if last_err:
        raise last_err

# Tokens que identificam os arquivos CSV de entrada
TOKENS_DEFAULT = ["ERP1", "ERP2", "ERP3", "UK", "ERP4"]

# Colunas finais do Master Data
MASTER_COLS = [
    "Country",
    "Vendor category",
    "Sheet",  # Key / ROL
    "Owner",
    "PM Name",
    "Company Code",
    "Document currency",
    "Supplier",
    "Document Number",
    "Reference",
    "Posting Period",
    "Amount in doc. curr.",
    "Document Type",
    "Name 1",
    "Document Date",
    "Net due date",
    "Text",
    "User name",
    "Posting Date",
    "Payment Block",
    "Unique Ref",
    "Payment Issues",
]

KEY_TEAM_FALLBACK_OWNERS: Set[str] = {"Synthetic Owner 011", "Synthetic Owner 014", "Synthetic Owner 012", "Synthetic Owner 010", "Synthetic Owner 015"}
KEY_TEAM_EXCLUDED_OWNERS: Set[str] = {"fuel", "unassigned"}
ROL_JEBIN_DAISY_EFFECTIVE_DATE = date(2026, 6, 15)
ROL_LEGACY_OWNER = "Synthetic Owner 020"
ROL_TRANSFER_OWNER = "Synthetic Owner 021"
ROL_TRANSFER_CATEGORIES: Set[str] = {"bakery", "maintenance", "telecom", "telecoms"}
LEDGER_OWNER_MAP_FILENAME = "Owner_map_Ledger.xlsx"
LEDGER_OWNER_MAP_SHEET = "Owner Map"
LEDGER_OWNER_MAP_TABLE = "LedgerOwnerMap"
LEDGER_OWNER_IGNORED_RULE_TYPE = "ignored"
LEDGER_NON_TEAM_OWNERS: Set[str] = {"No Owner", "ROL Uncategorised"}
LEDGER_HISTORICAL_OWNER_FALLBACKS: Set[str] = {"Synthetic Owner 020", "Synthetic Owner 023", "Synthetic Owner 025"}


# -------------------------
# Payment issue mapping (ERP payment codes)
# -------------------------
ISSUE_CODE_MAP: Dict[int, str] = {
    # Principais
    651: "BANK_DETAILS_MISSING_OR_NOT_ALLOWED",
    608: "NO_PAYMENT_METHOD_PERMITTED",
    390: "PAYMENT_PROCEDURE_NOT_CONFIGURED",
    347: "PAYMENT_METHOD_NOT_USABLE_FOR_ITEMS",
    # Extras uteis (aparecem nos logs):
    368: "VENDOR_MARKED_FOR_DELETION_OR_POSTING_BLOCK",
    349: "CONTAINED_IN_ANOTHER_PROPOSAL",
    716: "ITEM_BLOCKED_WITH_BLOCK_KEY",
    717: "ACCOUNT_BLOCKED_WITH_BLOCK_KEY",
    605: "PAYMENT_METHOD_NOT_MAINTAINED_IN_VENDOR_OR_DOC",
}

# Codigos de company logs (arquivos *.txt) que queremos ler (um code por arquivo)
COMPANY_LOG_CODES: List[str] = [
    # BE
    "BE21", "BE28", "BE32",
    # NL
    "NL00", "NL11", "NL19",
    "NL77", "NL79",
    "NL81", "NL82", "NL83", "NL84", "NL85", "NL86",
    # FR
    "FR02", "FR05",
    # DE (normais)
    "SYN-CC-004", "SYN-CC-005", "SYN-CC-006", "SYN-CC-007",
        "SYN-CC-008", "SYN-CC-009",
    # GB
    "SYN-CC-010", "SYN-CC-011", "SYN-CC-012",
    # Outros
    "SYN-CC-003",
        "SYN-CC-001", "SYN-CC-002",
    # LU
    "LU30", "LU31",
]

# Aliases: quando 1 unico TXT representa 2 company codes (token no NOME do arquivo)
COMPANY_LOG_ALIASES: Dict[str, List[str]] = {
    "SYN-CC-008 SYN-CC-009": ["SYN-CC-008", "SYN-CC-009"],
    "SYN-CC-001 SYN-CC-002": ["SYN-CC-001", "SYN-CC-002"],
}

def _txt_matches_any_company_code(txt_path: str) -> bool:
    name = os.path.basename(txt_path).upper()
    if any(code.upper() in name for code in COMPANY_LOG_CODES):
        return True
    if any(alias.upper() in name for alias in COMPANY_LOG_ALIASES.keys()):
        return True
    return False



# -------------------------
# Text helpers
# -------------------------
def clean_text(s: str) -> str:
    """
    Normaliza texto (estilo VBA CleanText):
    - lower case
    - remove pontuacoes comuns
    - remove espacos duplicados
    """
    s = (s or "").strip().lower()
    s = re.sub(r"[.:/\\()\[\]\-]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def add_many(d: Dict[str, str], owner: str, keys: Iterable[str]) -> None:
    for k in keys:
        kk = clean_text(k)
        if kk:
            d[kk] = owner


def _coerce_as_of_date(as_of: Optional[date]) -> date:
    if as_of is None:
        return date.today()
    if isinstance(as_of, datetime):
        return as_of.date()
    return as_of


def active_rol_transfer_owner(as_of: Optional[date] = None) -> str:
    """Return the active owner for the ROL categories moving from Synthetic Owner 020 to Synthetic Owner 021."""
    effective_as_of = _coerce_as_of_date(as_of)
    if effective_as_of >= ROL_JEBIN_DAISY_EFFECTIVE_DATE:
        return ROL_TRANSFER_OWNER
    return ROL_LEGACY_OWNER


def normalize_rol_owner_for_date(owner: str, as_of: Optional[date] = None) -> str:
    """Preserve historical Synthetic Owner 020, but shift active ROL rows to Synthetic Owner 021 from 2026-06-15."""
    owner_clean = (owner or "").strip()
    if owner_clean.casefold() == ROL_LEGACY_OWNER.casefold():
        return active_rol_transfer_owner(as_of)
    return owner_clean


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def default_ledger_owner_map_path() -> Path:
    return _project_root() / "data" / "master" / LEDGER_OWNER_MAP_FILENAME


def _default_owner_map_rol(as_of: Optional[date] = None) -> Dict[str, str]:
    """Built-in ROL owner map used only when the external map file is missing."""
    d: Dict[str, str] = {}

    add_many(d, "Synthetic Owner 018", ["authorities", "authority"])
    add_many(d, active_rol_transfer_owner(as_of), ROL_TRANSFER_CATEGORIES)
    add_many(d, "Synthetic Owner 019", ["it", "professional services", "professional service"])
    add_many(d, "Synthetic Owner 022", ["contracts", "contract", "office overheads", "office overhead", "overheads"])
    add_many(d, "Synthetic Owner 026", [
        "employees", "employee", "emp", "payroll hr", "payroll",
        "payroll/hr", "payroll / hr", "transport",
    ])
    add_many(d, "Synthetic Owner 017", ["franchise fees", "franchise fee", "franchise", "ftg"])
    add_many(d, "Synthetic Owner 016", [
        "litigation", "rent", "rent ", "rentals", "Fuel AP",
        "press", "fixed assets", "fixed asset",
    ])
    add_many(d, "Synthetic Owner 024", ["cng", "utilities", "utility"])
    add_many(d, "No Owner", [
        "shop", "store", "stores", "# uncategorised (shop?) #",
        "# uncategorised(Shop?) #", "tobacco",
    ])
    d[clean_text("uncategorised")] = "ROL Uncategorised"
    d[clean_text("# uncategorised #")] = "ROL Uncategorised"
    d[clean_text("intercompany")] = ""
    d[clean_text("fuel")] = ""
    return d


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_text(value)
    if not text:
        return False
    if text in {"true", "yes", "y", "1", "x", "active"}:
        return True
    if text in {"false", "no", "n", "0", "inactive"}:
        return False
    raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} has invalid Active value: {value!r}")


def _ledger_owner_map_cache_key(path: Path) -> tuple[str, int, int]:
    stat = path.stat()
    return str(path), int(stat.st_mtime_ns), int(stat.st_size)


def _read_ledger_owner_map_table(path: Path) -> pd.DataFrame:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError as exc:
        raise ImportError(f"openpyxl is required to read {LEDGER_OWNER_MAP_FILENAME}") from exc

    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise ValueError(f"Unable to read {LEDGER_OWNER_MAP_FILENAME}: {path}") from exc

    try:
        if LEDGER_OWNER_MAP_SHEET not in workbook.sheetnames:
            raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} must contain an '{LEDGER_OWNER_MAP_SHEET}' sheet")
        worksheet = workbook[LEDGER_OWNER_MAP_SHEET]
        table = worksheet.tables.get(LEDGER_OWNER_MAP_TABLE)
        if table is None:
            raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} must contain table '{LEDGER_OWNER_MAP_TABLE}'")

        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        rows = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f"{LEDGER_OWNER_MAP_TABLE} table is empty")

    headers = [norm_str_series(value) for value in rows[0]]
    if any(not header for header in headers):
        raise ValueError(f"{LEDGER_OWNER_MAP_TABLE} table has blank header cells")
    return pd.DataFrame(rows[1:], columns=headers)


@lru_cache(maxsize=8)
def _load_ledger_owner_rows_cached(path_text: str, _mtime_ns: int, _size: int) -> tuple[tuple[str, str, str, bool, bool, str], ...]:
    path = Path(path_text)
    df = _read_ledger_owner_map_table(path)
    cols = {clean_text(c): c for c in df.columns}

    active_col = cols.get("active")
    rule_type_col = cols.get("rule type") or cols.get("ruletype")
    alias_col = cols.get("category alias") or cols.get("category") or cols.get("vendor category")
    owner_col = cols.get("owner")
    required_owner_col = cols.get("required owner") or cols.get("requiredowner")
    sheet_col = cols.get("sheet")

    missing = [
        name
        for name, col in {
            "Active": active_col,
            "Rule Type": rule_type_col,
            "Category Alias": alias_col,
            "Owner": owner_col,
        }.items()
        if not col
    ]
    if missing:
        raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} missing columns: {', '.join(missing)}")

    rows: list[tuple[str, str, str, bool, bool, str]] = []
    seen: dict[str, str] = {}
    for row_num, row in enumerate(df.to_dict("records"), start=2):
        if sheet_col and norm_str_series(row.get(sheet_col)) and clean_text(row.get(sheet_col)) != "rol":
            raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} row {row_num}: Sheet must be ROL")

        alias_raw = norm_str_series(row.get(alias_col))
        owner = norm_str_series(row.get(owner_col))
        rule_type = norm_str_series(row.get(rule_type_col))
        active = _truthy(row.get(active_col))
        required_owner_flag = _truthy(row.get(required_owner_col)) if required_owner_col else None

        if not alias_raw and not owner and not rule_type and not active:
            continue
        if not active:
            continue

        alias = clean_text(alias_raw)
        if not alias:
            raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} row {row_num}: Category Alias is required")

        is_ignored = clean_text(rule_type) == LEDGER_OWNER_IGNORED_RULE_TYPE
        if is_ignored or required_owner_flag is False:
            owner_value = ""
        else:
            if not owner:
                raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} row {row_num}: blank Owner is allowed only for Rule Type Ignored")
            owner_value = owner

        existing_owner = seen.get(alias)
        if existing_owner is not None and existing_owner.casefold() != owner_value.casefold():
            raise ValueError(
                f"{LEDGER_OWNER_MAP_FILENAME} has conflicting active owner rules for Category Alias '{alias_raw}'"
            )
        seen[alias] = owner_value
        required_owner = (
            required_owner_flag
            if required_owner_flag is not None
            else bool(owner_value and owner_value not in LEDGER_NON_TEAM_OWNERS)
        )
        rows.append((alias, owner_value, rule_type, active, required_owner, alias_raw))

    if not rows:
        raise ValueError(f"{LEDGER_OWNER_MAP_FILENAME} has no active Ledger owner rows")
    return tuple(rows)


def load_ledger_owner_map_rows(path: Optional[str | os.PathLike[str]] = None) -> tuple[tuple[str, str, str, bool, bool, str], ...]:
    map_path = Path(path) if path is not None else default_ledger_owner_map_path()
    if not map_path.exists():
        return tuple()
    key = _ledger_owner_map_cache_key(map_path)
    return _load_ledger_owner_rows_cached(*key)


def _build_owner_map_rol_from_file(path: Optional[str | os.PathLike[str]] = None) -> Dict[str, str]:
    rows = load_ledger_owner_map_rows(path)
    return {alias: owner for alias, owner, _rule_type, _active, _required, _raw in rows}


def _is_transient_owner_map_access_error(exc: BaseException) -> bool:
    current: Optional[BaseException] = exc
    while current is not None:
        if isinstance(current, (PermissionError, FileNotFoundError)):
            return True
        current = current.__cause__
    return False


def _owners_from_rows(rows: tuple[tuple[str, str, str, bool, bool, str], ...]) -> Set[str]:
    return {
        owner
        for _alias, owner, rule_type, _active, _required, _raw in rows
        if owner and clean_text(rule_type) != LEDGER_OWNER_IGNORED_RULE_TYPE and owner not in LEDGER_NON_TEAM_OWNERS
    }


def current_ledger_owners(
    owner_map_path: Optional[str | os.PathLike[str]] = None,
    as_of: Optional[date] = None,
) -> Set[str]:
    """Owners that are active in the current Ledger owner map."""
    map_path = Path(owner_map_path) if owner_map_path is not None else default_ledger_owner_map_path()
    if map_path.exists():
        try:
            return _owners_from_rows(load_ledger_owner_map_rows(map_path))
        except ValueError as exc:
            if not _is_transient_owner_map_access_error(exc):
                raise
            logger.warning("%s is temporarily unavailable; using the built-in Ledger owner fallback", map_path)
    return {
        owner
        for owner in _default_owner_map_rol(as_of=as_of).values()
        if owner and owner not in LEDGER_NON_TEAM_OWNERS
    }


def historical_ledger_owners(as_of: Optional[date] = None) -> Set[str]:
    """Owners accepted for historical rebuilds and old Ledger tabs."""
    fallback_owners = {
        owner
        for owner in _default_owner_map_rol(as_of=as_of).values()
        if owner and owner not in LEDGER_NON_TEAM_OWNERS
    }
    return fallback_owners | LEDGER_HISTORICAL_OWNER_FALLBACKS | {ROL_LEGACY_OWNER, ROL_TRANSFER_OWNER}


def allowed_ledger_owners(
    owner_map_path: Optional[str | os.PathLike[str]] = None,
    as_of: Optional[date] = None,
) -> Set[str]:
    """Current Ledger owners plus historical owners accepted during rebuilds."""
    return current_ledger_owners(owner_map_path=owner_map_path, as_of=as_of) | historical_ledger_owners(as_of=as_of) | LEDGER_NON_TEAM_OWNERS


def required_rol_owners(as_of: Optional[date] = None) -> Set[str]:
    return current_ledger_owners(as_of=as_of)


def known_rol_owners(as_of: Optional[date] = None) -> Set[str]:
    return allowed_ledger_owners(as_of=as_of)

# -------------------------
# Owner maps (ROL/Query)
# -------------------------
def build_owner_map_rol(as_of: Optional[date] = None, owner_map_path: Optional[str | os.PathLike[str]] = None) -> Dict[str, str]:
    """Unified ROL owner map (merged ROL + Query teams)."""
    map_path = Path(owner_map_path) if owner_map_path is not None else default_ledger_owner_map_path()
    if map_path.exists():
        try:
            return _build_owner_map_rol_from_file(map_path)
        except ValueError as exc:
            if not _is_transient_owner_map_access_error(exc):
                raise
            logger.warning("%s is temporarily unavailable; using the built-in Ledger owner fallback", map_path)
    return _default_owner_map_rol(as_of=as_of)


def lookup_rol_owner_from_category(
    vendor_category: str,
    as_of: Optional[date] = None,
    owner_map: Optional[Dict[str, str]] = None,
) -> tuple[bool, str]:
    cat = clean_text(vendor_category)
    if not cat:
        return False, ""
    active_owner_map = owner_map if owner_map is not None else build_owner_map_rol(as_of=as_of)
    return cat in active_owner_map, active_owner_map.get(cat, "")


def owner_from_category(
    sheet: str,
    vendor_category: str,
    as_of: Optional[date] = None,
    owner_map: Optional[Dict[str, str]] = None,
) -> str:
    cat = clean_text(vendor_category)
    if not cat:
        return ""
    sh = clean_text(sheet)
    if sh == "rol":
        active_owner_map = owner_map if owner_map is not None else build_owner_map_rol(as_of=as_of)
        return active_owner_map.get(cat, "")
    if sh == "key":
        return ""
    return ""


# -------------------------
# Amount cleanup
# -------------------------
def normalize_amount_series(s: pd.Series) -> pd.Series:
    """
    Limpa strings de valor financeiro (remove moedas, converte , e .).
    """
    raw = s.fillna("").astype(str)

    # remove codigos de moeda e espacos
    cleaned = raw.str.replace("\u00a0", " ", regex=False)
    cleaned = cleaned.str.replace(r"\b[A-Za-z]{3}\b", "", regex=True)

    # mantem digitos, sinais e pontuacao
    cleaned = cleaned.str.replace(r"[^\d,\.\-\(\)]", "", regex=True)

    # trata formato (123) como negativo -123
    cleaned = cleaned.str.replace(r"^\((.*)\)$", r"-\1", regex=True)

    def fix(x: str) -> str:
        if not x: return ""
        # decimal com virgula e sem ponto (ex: 100,50)
        if x.count(",") > 0 and x.count(".") == 0:
            return x.replace(".", "").replace(",", ".")
        # milhar com virgula e decimal com ponto (ex: 1,000.50)
        if x.count(",") > 0 and x.count(".") > 0:
            return x.replace(",", "")
        return x

    cleaned = cleaned.map(fix)
    num = pd.to_numeric(cleaned, errors="coerce")
    out = num.map(lambda v: "" if pd.isna(v) else str(v))
    out = out.str.replace(r"\.0$", "", regex=True)
    return out


# -------------------------
# File discovery / Maps Loading
# -------------------------
def find_latest_csv_for_token(base_dir: str, token: str) -> Optional[str]:
    """
    Encontra o CSV mais recente na pasta base_dir que contenha o token no nome.
    """
    pattern = os.path.join(base_dir, f"*{token}*.csv")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return files[0]


def _read_csv_header(path: str, encoding: str = "utf-8-sig") -> list[str]:
    with open(path, "r", encoding=encoding, newline="") as f:
        return next(csv.reader(f))


def load_vendor_category_map(path: str) -> Dict[str, str]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Vendor matrix nao encontrado: {path}")

    header = _read_csv_header(path)
    cols = {clean_text(c): c for c in header}

    ur_col = cols.get("unique ref")
    cat_col = cols.get("category")
    snap_col = cols.get("snapshot date")

    if not ur_col or not cat_col:
        raise ValueError("Vendor matrix precisa ter colunas: 'Unique ref' e 'Category'.")

    usecols = [ur_col, cat_col] + ([snap_col] if snap_col else [])
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", usecols=usecols, low_memory=False)

    df[ur_col] = df[ur_col].fillna("").astype(str).str.strip()
    df[cat_col] = df[cat_col].fillna("").astype(str).str.strip()
    df.loc[df[cat_col] == "", cat_col] = "Uncategorised"
    df = df[df[ur_col] != ""]

    if snap_col:
        df["_snap_"] = parse_date_series(df[snap_col])
        df.sort_values(by=["_snap_"], inplace=True)
        df = df.drop_duplicates(subset=[ur_col], keep="last")
    else:
        df = df.drop_duplicates(subset=[ur_col], keep="last")

    return dict(zip(df[ur_col], df[cat_col]))


def load_vendor_name1_map(path: str) -> Dict[str, str]:
    """
    Load canonical supplier names from the Vendor Master Matrix.

    Key and Ledger reports group rows by Unique Ref, so every row for the same
    supplier identity must carry the same Name 1 value.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Vendor matrix nao encontrado: {path}")

    header = _read_csv_header(path)
    cols = {clean_text(c): c for c in header}

    ur_col = cols.get("unique ref")
    name_col = (
        cols.get("name 1")
        or cols.get("vendor name 1")
        or cols.get("supplier name")
        or cols.get("name")
    )
    snap_col = cols.get("snapshot date")

    if not ur_col:
        raise ValueError("Vendor matrix precisa ter coluna: 'Unique ref'.")

    if not name_col:
        return {}

    usecols = [ur_col, name_col] + ([snap_col] if snap_col else [])
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", usecols=usecols, low_memory=False)

    df[ur_col] = df[ur_col].fillna("").astype(str).str.strip()
    df[name_col] = df[name_col].fillna("").astype(str).str.strip()
    df = df[(df[ur_col] != "") & (df[name_col] != "")]

    if snap_col:
        df["_snap_"] = parse_date_series(df[snap_col])
        df.sort_values(by=["_snap_"], inplace=True)
        df = df.drop_duplicates(subset=[ur_col], keep="last")
    else:
        df = df.drop_duplicates(subset=[ur_col], keep="last")

    return dict(zip(df[ur_col], df[name_col]))


def canonicalize_masterdata_name1_by_unique_ref(
    path: str,
    *,
    vendor_name_map: Optional[Dict[str, str]] = None,
) -> dict[str, int]:
    """
    Ensure each Unique Ref has one Name 1 before Key/Ledger report generation.

    The report generators group by Unique Ref and validate that detail rows in
    the same group have a single Name 1. Prefer the Vendor Master Matrix name
    when available; otherwise use the most frequent non-empty Name 1 in the
    generated MasterData file.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"MasterData nao encontrado: {path}")

    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", low_memory=False)
    if "Unique Ref" not in df.columns or "Name 1" not in df.columns:
        return {"groups_changed": 0, "rows_changed": 0}

    df["Unique Ref"] = df["Unique Ref"].fillna("").astype(str).str.strip()
    df["Name 1"] = df["Name 1"].fillna("").astype(str).str.strip()
    name_map = vendor_name_map or {}

    grouped_names = (
        df.loc[(df["Unique Ref"] != "") & (df["Name 1"] != ""), ["Unique Ref", "Name 1"]]
        .drop_duplicates()
        .groupby("Unique Ref")["Name 1"]
        .nunique()
    )
    divergent_refs = set(grouped_names[grouped_names > 1].index)

    canonical: Dict[str, str] = {
        ur: name.strip()
        for ur, name in name_map.items()
        if str(ur).strip() and str(name).strip()
    }

    missing_refs = set(df.loc[df["Unique Ref"] != "", "Unique Ref"].unique()) - set(canonical)
    if missing_refs:
        counts = (
            df.loc[df["Unique Ref"].isin(missing_refs) & (df["Name 1"] != ""), ["Unique Ref", "Name 1"]]
            .value_counts(sort=True)
            .rename("count")
            .reset_index()
            .sort_values(["Unique Ref", "count", "Name 1"], ascending=[True, False, True])
        )
        fallback = counts.drop_duplicates(subset=["Unique Ref"], keep="first")
        canonical.update(dict(zip(fallback["Unique Ref"], fallback["Name 1"])))

    mapped = df["Unique Ref"].map(canonical).fillna("")
    change_mask = (mapped != "") & (df["Name 1"] != mapped)
    rows_changed = int(change_mask.sum())
    if rows_changed == 0:
        return {"groups_changed": 0, "rows_changed": 0}

    changed_refs = set(df.loc[change_mask, "Unique Ref"].unique())
    df.loc[change_mask, "Name 1"] = mapped.loc[change_mask]

    tmp_path = path + ".name1_tmp"
    df.to_csv(tmp_path, index=False, encoding="utf-8-sig")
    os.replace(tmp_path, path)

    return {
        "groups_changed": len(changed_refs | divergent_refs),
        "rows_changed": rows_changed,
    }


def load_vendor_pm_name_map(path: str) -> Dict[str, str]:
    """
    Le o "PM Name" do Vendor Master Matrix (ou arquivo equivalente) e retorna:
        Unique Ref -> PM Name

    Colunas esperadas (case-insensitive):
      - Unique ref
      - PM Name   (ou variacoes: 'PM', 'PMName')

    Se a coluna nao existir, retorna {} (o MasterData sera gerado com PM Name em branco).
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Vendor matrix nao encontrado: {path}")

    header = _read_csv_header(path)
    cols = {clean_text(c): c for c in header}

    ur_col = cols.get("unique ref")
    pm_col = cols.get("pm name") or cols.get("pmname") or cols.get("pm")
    snap_col = cols.get("snapshot date")

    if not ur_col:
        raise ValueError("Vendor matrix precisa ter coluna: 'Unique ref'.")

    if not pm_col:
        # Sem PM Name no arquivo -> segue com mapa vazio
        return {}

    usecols = [ur_col, pm_col] + ([snap_col] if snap_col else [])
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", usecols=usecols, low_memory=False)

    df[ur_col] = df[ur_col].fillna("").astype(str).str.strip()
    df[pm_col] = df[pm_col].fillna("").astype(str).str.strip()
    df = df[df[ur_col] != ""]

    if snap_col:
        df["_snap_"] = parse_date_series(df[snap_col])
        df.sort_values(by=["_snap_"], inplace=True)
        df = df.drop_duplicates(subset=[ur_col], keep="last")
    else:
        df = df.drop_duplicates(subset=[ur_col], keep="last")

    return dict(zip(df[ur_col], df[pm_col]))


def load_key_owner_map_from_vba_txt(path: str) -> Dict[str, str]:
    """
    Carrega o Owner map (Unique Ref -> Owner) para a aba Key.

    Compatibilidade total:
      - Se existir um CSV no mesmo diretorio (Owner_map.csv / OwnerMap.csv), ele e priorizado.
      - Caso contrario, faz parse do formato antigo (owner_map_vba.txt).
    """
    if not path:
        return {}

    base_dir = os.path.dirname(path) or "."
    csv_candidates = [
        os.path.join(base_dir, "Owner_map.csv"),
        os.path.join(base_dir, "OwnerMap.csv"),
        os.path.join(base_dir, "owner_map.csv"),
        os.path.join(base_dir, "ownerMap.csv"),
    ]

    for csv_path in csv_candidates:
        if not os.path.exists(csv_path):
            continue

        try:
            df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig", low_memory=False)
        except Exception:
            # fallback para arquivos salvos em encoding diferente
            df = pd.read_csv(csv_path, dtype=str, encoding="latin1", low_memory=False)

        # headers case-insensitive
        cols = {clean_text(c): c for c in df.columns}
        ur_col = cols.get("unique ref")
        owner_col = cols.get("owner")
        if not ur_col or not owner_col:
            continue

        ur = df[ur_col].fillna("").astype(str).str.strip()
        ow = df[owner_col].fillna("").astype(str).str.strip()

        out_csv: Dict[str, str] = {}
        for u, o in zip(ur, ow):
            if u:
                out_csv[u] = o
        return out_csv

    # --- fallback: TXT antigo (VBA dict("UR") = "Owner") ---
    if not os.path.exists(path):
        return {}

    pattern = re.compile(r'dict\("([^"]+)"\)\s*=\s*"([^"]*)"', re.IGNORECASE)
    out: Dict[str, str] = {}

    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            m = pattern.search(line)
            if not m:
                continue
            ur = m.group(1).strip()
            owner = m.group(2).strip()
            if ur:
                out[ur] = owner

    return out


def load_key_team_owners(path: str) -> Set[str]:
    """Return Key team owners derived from Owner_map.csv, with legacy fallback."""
    owners = {
        owner
        for owner in load_key_owner_map_from_vba_txt(path).values()
        if owner and owner.casefold() not in KEY_TEAM_EXCLUDED_OWNERS
    }
    return owners or set(KEY_TEAM_FALLBACK_OWNERS)




def load_prev_assignments(path: str) -> Dict[str, str]:
    if not path or not os.path.exists(path):
        return {}

    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", low_memory=False)
    df.columns = [clean_text(c) for c in df.columns]

    if "unique ref" not in df.columns:
        raise ValueError("previous_assignments precisa ter coluna 'Unique Ref'.")

    if "sheet" in df.columns:
        cat_col = "sheet"
    elif "actual category" in df.columns:
        cat_col = "actual category"
    else:
        raise ValueError("previous_assignments precisa ter coluna 'Sheet' ou 'Actual Category'.")

    ur = df["unique ref"].fillna("").astype(str).str.strip()
    cat = df[cat_col].fillna("").astype(str).str.strip()

    out: Dict[str, str] = {}
    for u, c in zip(ur, cat):
        if not u: continue
        cc = clean_text(c)
        if cc == "query":
            out[u] = "ROL"
        elif cc == "key":
            out[u] = "Key"
        elif cc == "rol":
            out[u] = "ROL"
        else:
            out[u] = ""
    return out



def load_prev_owners(path: str) -> Dict[str, str]:
    """
    Le previous_assignments.csv e retorna:
        Unique Ref -> Owner

    Compativel com a versao antiga (que tinha so Unique Ref/Sheet).
    Se a coluna 'Owner' nao existir, retorna {}.
    """
    if not path or not os.path.exists(path):
        return {}

    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", low_memory=False)
    df.columns = [clean_text(c) for c in df.columns]

    if "unique ref" not in df.columns:
        return {}

    # coluna de owner pode variar (owner / previous owner / key owner)
    owner_col = None
    for cand in ("owner", "previous owner", "prev owner", "key owner"):
        if cand in df.columns:
            owner_col = cand
            break

    if not owner_col:
        return {}

    ur = df["unique ref"].fillna("").astype(str).str.strip()
    ow = df[owner_col].fillna("").astype(str).str.strip()

    out: Dict[str, str] = {}
    for u, o in zip(ur, ow):
        if not u:
            continue
        if o:
            out[u] = o
    return out



# -------------------------
# Column pickers / parsing
# -------------------------
def pick(series_by_name: Dict[str, pd.Series], candidates: list[str]) -> pd.Series:
    for c in candidates:
        k = clean_text(c)
        if k in series_by_name:
            return series_by_name[k]
    first = next(iter(series_by_name.values()))
    return pd.Series([""] * len(first), index=first.index)


def parse_date_series(s: pd.Series) -> pd.Series:
    """Parse known ERP date formats without inferring or swapping values.

    ERP exports are mixed: ERP1 commonly uses ISO ``YYYY-MM-DD`` while other
    systems commonly use day-first ``DD/MM/YYYY`` or ``DD.MM.YYYY``. A single
    pandas inference rule can coerce one format to ``NaT`` or reinterpret an
    ISO date. Parse each value against explicit formats instead.
    """
    formats = (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y/%m/%d",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M:%S",
        "%Y%m%d",
    )

    parsed = []
    for value in s.tolist():
        if value is None or value is pd.NaT:
            parsed.append(pd.NaT)
            continue
        if isinstance(value, (pd.Timestamp, datetime)):
            parsed.append(pd.Timestamp(value))
            continue
        if isinstance(value, date):
            parsed.append(pd.Timestamp(value))
            continue

        text = str(value).strip()
        if not text or text.casefold() in {"nan", "nat", "none"}:
            parsed.append(pd.NaT)
            continue

        parsed_value = pd.NaT
        for fmt in formats:
            try:
                parsed_value = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        parsed.append(parsed_value)

    # Keep malformed/out-of-range dates as NaT so one bad tail value cannot
    # abort the report; recover_date_from_tail preserves the original text.
    return pd.Series(
        pd.to_datetime(parsed, errors="coerce"),
        index=s.index,
        name=s.name,
    )


def country_from_code(code_series: pd.Series) -> pd.Series:
    s = code_series.fillna("").astype(str).str.strip().str.upper()
    out = pd.Series(["Unknown"] * len(s), index=s.index)

    out[s.str.startswith("GB")] = "UK"
    out[s.isin(["FR05", "FR02"])] = "FRANCE"
    out[s.str.startswith(("BE", "NL", "LU"))] = "BENELUX"
    out[s.isin(["IT01", "IT02"])] = "ITALY"
    out[s.isin(["SYN-CC-004", "SYN-CC-005", "SYN-CC-006", "SYN-CC-007", "SYN-CC-008", "SYN-CC-009", "SYN-CC-002", "SYN-CC-001", "SYN-CC-003"])] = "GERMANY"
    return out


# -------------------------
# LOG PARSER (Payment issues)
# -------------------------
def _extract_fz_code(line: str) -> Optional[int]:
    m = re.search(r"\bFZ\s+(\d{3})\b", line)
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _clean_erp_msg(line: str) -> str:
    """
    Remove colunas finais (classe FZ etc.) e metadados, mantendo so a mensagem.
    Funciona com separacao por tabs ou espacos.
    """
    s = (line or "").rstrip("\n")
    # remove trailing "FZ 123 S"
    s = re.sub(r"\s+FZ\s+\d+\s+\w\s*$", "", s)
    # remove leading date/time (tab or spaces)
    s = re.sub(r"^\s*\d{2}\.\d{2}\.\d{4}[\t ]+\d{2}:\d{2}:\d{2}[\t ]+", "", s)
    # remove leading ">"
    s = s.strip().lstrip(">").strip()
    return s


def _format_issue(code: Optional[int], msg: str) -> str:
    msg = (msg or "").strip()
    if not msg:
        return ""
    if code is None:
        return msg
    tag = ISSUE_CODE_MAP.get(code)
    if tag:
        return f"{tag} (FZ{code}): {msg}"
    return f"FZ{code}: {msg}"


def _add_issue(target: Dict[str, str], key: str, issue: str) -> None:
    """
    Evita sobrescrever e perder informacao (ex: FZ651 + FZ608 no mesmo documento).
    Armazena multiplos issues separados por ' | ' (sem duplicar).
    """
    issue = (issue or "").strip()
    if not issue:
        return
    prev = target.get(key, "")
    if not prev:
        target[key] = issue
        return
    if issue in prev:
        return
    target[key] = f"{prev} | {issue}"


def load_payment_issues(base_dir: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Le logs TXT e extrai issues de pagamento com granularidade:
      - vendor_issues: key "CC Vendor"      (Supplier sem zeros a esquerda)
      - doc_issues:    key "CC Vendor Doc"  (Doc sem zeros a esquerda)

    IMPORTANTE:
      * doc_issues tem prioridade (mais especifico).
      * Mantem multiplos issues no mesmo key (join com ' | ').

    Retorna: (vendor_issues, doc_issues)
    """
    vendor_issues: Dict[str, str] = {}
    doc_issues: Dict[str, str] = {}

    site_codes_dir = os.path.join(base_dir, "site_codes")
    if not os.path.isdir(site_codes_dir):
        return {}, {}
    all_txt = glob.glob(os.path.join(site_codes_dir, "*.txt"))
    target_files = [f for f in all_txt if _txt_matches_any_company_code(f)]

    if not target_files:
        return {}, {}

    for filepath in target_files:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            curr_v: Optional[str] = None
            curr_cc: Optional[str] = None
            curr_doc: Optional[str] = None

            pending_vendor_info: Optional[Tuple[str, str]] = None  # (vendor, cc)
            issue_buffer: str = ""  # issue "orfao" para anexar quando aparecer o vendor

            for raw in f:
                code = _extract_fz_code(raw)
                msg = _clean_erp_msg(raw)
                if not msg:
                    continue

                # 1) Linha direta de vendor (ex: "Vendor 2569 NL11: ...")
                m = re.search(r"\bVendor\s+(\d+)\s+([A-Z0-9]{4}):\s+(.+)", msg)
                if m:
                    v, cc, txt = m.groups()
                    v_key = v.lstrip("0")
                    k = f"{cc} {v_key}"
                    # Caso especial: "contained in proposal"
                    if "is contained in proposal" in txt.lower():
                        _add_issue(vendor_issues, k, _format_issue(349, txt))
                    else:
                        _add_issue(vendor_issues, k, _format_issue(code, txt))
                    pending_vendor_info = None
                    issue_buffer = ""
                    continue

                # 2) Vendor "contained in proposal" (sem dois-pontos)
                m = re.search(
                    r"\bVendor\s+(\d+)\s+([A-Z0-9]{4})\s+is contained in proposal\b(.+)",
                    msg,
                    flags=re.I,
                )
                if m:
                    v, cc, rest = m.groups()
                    v_key = v.lstrip("0")
                    k = f"{cc} {v_key}"
                    _add_issue(vendor_issues, k, _format_issue(349, f"is contained in proposal{rest}".strip()))
                    pending_vendor_info = None
                    issue_buffer = ""
                    continue

                # 3) Abre bloco "Additional log for vendor ..."
                m = re.search(r"Additional log for vendor\s+(\d+)\s+company code\s+([A-Z0-9]{4})", msg)
                if m:
                    curr_v, curr_cc = m.groups()
                    curr_doc = None
                    pending_vendor_info = None
                    issue_buffer = ""
                    continue

                # 4) Dentro do bloco adicional: pega documento atual + block keys
                if curr_v and curr_cc:
                    m_doc = re.search(r"Document\s+(\d+)\s+line item", msg)
                    if m_doc:
                        curr_doc = m_doc.group(1)

                    m_item_blk = re.search(r"Item is blocked with block key\s+([A-Z0-9])", msg)
                    if m_item_blk:
                        v_key = curr_v.lstrip("0")
                        d_key = (curr_doc or "").lstrip("0")
                        issue = _format_issue(716, f"Item is blocked with block key {m_item_blk.group(1)}")
                        if d_key:
                            _add_issue(doc_issues, f"{curr_cc} {v_key} {d_key}", issue)
                        else:
                            _add_issue(vendor_issues, f"{curr_cc} {v_key}", issue)
                        continue

                    m_acc_blk = re.search(r"Account is blocked with block key\s+([A-Z0-9])", msg)
                    if m_acc_blk:
                        v_key = curr_v.lstrip("0")
                        d_key = (curr_doc or "").lstrip("0")
                        issue = _format_issue(717, f"Account is blocked with block key {m_acc_blk.group(1)}")
                        if d_key:
                            _add_issue(doc_issues, f"{curr_cc} {v_key} {d_key}", issue)
                        else:
                            _add_issue(vendor_issues, f"{curr_cc} {v_key}", issue)
                        continue

                # 5) "Information re. vendor ..." (identifica vendor/cc para amarrar mensagem)
                m = re.search(r"Information re\. vendor\s+(\d+)\s+.*company code\s+([A-Z0-9]{4})", msg)
                if m:
                    v, cc = m.groups()
                    pending_vendor_info = (v, cc)

                    # Se o root-cause veio na(s) linha(s) anterior(es), cola agora
                    if issue_buffer:
                        _add_issue(vendor_issues, f"{cc} {v.lstrip('0')}", issue_buffer)
                        issue_buffer = ""
                    continue

                # 6) Root-cause codes importantes (FZ347/390/608/651 etc)
                if code is not None and code in ISSUE_CODE_MAP:
                    issue = _format_issue(code, msg)

                    if curr_v and curr_cc and curr_doc:
                        _add_issue(doc_issues, f"{curr_cc} {curr_v.lstrip('0')} {curr_doc.lstrip('0')}", issue)
                    elif curr_v and curr_cc:
                        _add_issue(vendor_issues, f"{curr_cc} {curr_v.lstrip('0')}", issue)
                    elif pending_vendor_info:
                        v, cc = pending_vendor_info
                        _add_issue(vendor_issues, f"{cc} {v.lstrip('0')}", issue)
                    else:
                        issue_buffer = issue
                    continue

                # 7) Outros FZ codes nao mapeados: guarda se houver contexto (sem perder)
                ignore = {693, 691, 699, 700, 701, 707, 708, 709, 720, 721, 726, 728, 741, 743, 744, 798, 799}
                if code is not None and code not in ignore:
                    issue = _format_issue(code, msg)
                    if curr_v and curr_cc and curr_doc:
                        _add_issue(doc_issues, f"{curr_cc} {curr_v.lstrip('0')} {curr_doc.lstrip('0')}", issue)
                    elif pending_vendor_info:
                        v, cc = pending_vendor_info
                        _add_issue(vendor_issues, f"{cc} {v.lstrip('0')}", issue)
                    else:
                        issue_buffer = issue

    return vendor_issues, doc_issues



def decide_sheet_grouped(unique_ref: str, key_owner_map: Dict[str, str]) -> str:
    # 1. KEY: Se esta no mapa de owners definidos, e Key
    if unique_ref in key_owner_map:
        return "Key"
    # 2. ROL: Default (Query sheet eliminated - all non-Key go to ROL)
    return "ROL"


# -------------------------
# Writer (monthly/weekly)
# -------------------------
def _detect_source_csv_encoding(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8-sig") as source:
            while source.read(1024 * 1024):
                pass
    except UnicodeDecodeError:
        logger.warning("%s is not UTF-8; reading with latin1 fallback", Path(path).name)
        return "latin1"
    return "utf-8-sig"


def process_file_grouped(
    path: str,
    out_csv: str,
    vendor_cat_map: Dict[str, str],
    key_owner_map: Dict[str, str],
    prev_map: Dict[str, str],
    prev_owner_map: Optional[Dict[str, str]] = None,
    pm_name_map: Optional[Dict[str, str]] = None,
    vendor_pm_map: Optional[Dict[str, str]] = None,
    vendor_name_map: Optional[Dict[str, str]] = None,
    vendor_issues: Optional[Dict[str, str]] = None,
    doc_issues: Optional[Dict[str, str]] = None,
    mode: str = "weekly",  # "monthly" | "weekly"
    as_of: Optional[date] = None,
    write_header: bool = True,
    chunksize: int = 200_000,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> bool:
    encoding = _detect_source_csv_encoding(path)
    reader = pd.read_csv(path, dtype=str, encoding=encoding, chunksize=chunksize, low_memory=False)

    pm_map: Dict[str, str] = (pm_name_map or vendor_pm_map or {})
    name_map: Dict[str, str] = vendor_name_map or {}

    v_issues: Dict[str, str] = vendor_issues or {}
    d_issues: Dict[str, str] = doc_issues or {}
    if as_of is None:
        as_of = date.today()
    rol_owner_map: Optional[Dict[str, str]] = None

    def get_rol_owner_map() -> Dict[str, str]:
        nonlocal rol_owner_map
        if rol_owner_map is None:
            rol_owner_map = build_owner_map_rol(as_of=as_of)
        return rol_owner_map

    for chunk_index, chunk in enumerate(reader, start=1):
        chunk.columns = [clean_text(c) for c in chunk.columns]
        series = {c: chunk[c].fillna("").astype(str).str.strip() for c in chunk.columns}

        # Extracao de colunas com fallback inteligente
        company = pick(series, ["Company Code"])
        supplier = pick(series, ["Supplier", "Vendor", "Vendor Account"])

        currency = pick(series, [
            "Document currency", "Transaction Currency", "Local Currency", "Company Code Currency Key",
        ])

        docnum = pick(series, ["Document Number", "Journal Entry", "Journal Entry Number"])

        # Payment Issues: doc-level tem prioridade; fallback para vendor-level
        issues_list: List[str] = []
        for co, supp, doc in zip(company.tolist(), supplier.tolist(), docnum.tolist()):
            c_norm = str(co).strip().upper()
            s_norm = str(supp).strip().lstrip('0')
            d_norm = str(doc).strip().lstrip('0')
            issue = ''
            if c_norm and s_norm and d_norm:
                issue = d_issues.get(f"{c_norm} {s_norm} {d_norm}", '')
            if not issue and c_norm and s_norm:
                issue = v_issues.get(f"{c_norm} {s_norm}", '')
            issues_list.append(issue)
        ref = pick(series, ["Reference", "Assignment", "Reference Key (header) 1", "Reference Key 2", "Reference key 2"])
        postper = pick(series, ["Posting Period", "Posting period"])

        amount_raw = pick(series, [
            "Amount in doc. curr.", "Amount in document currency", "Amount in local currency",
            "Company Code Currency Value", "Amount (CoCode Crcy)", "Amount (Tran Cur.)",
        ])
        amount = normalize_amount_series(amount_raw)

        doctype = pick(series, ["Document Type", "Document type", "Journal Entry Type"])

        name1 = pick(series, [
            "Vendor Name 1", "Name 1", "Supplier Name",
            "Vendor Account Name 1", "Vendor Account: Name 1", "Name",
        ])

        docdate = pick(series, ["Document Date", "Journal Entry Date"])
        netdue = pick(series, ["Net due date", "Net Due Date"])
        text = pick(series, ["Text", "Item Text", "Item Text ID", "Item Text Id"])
        
        user = pick(series, [
            "User name", "User Name", "Journal Entry Created By",
            "Created By", "Created by", "Posted by", "Posted By",
        ])
        postdt = pick(series, ["Posting Date"])
        payblk = pick(series, ["Payment Block", "Item Payment Block"])

        unique_ref_series = (company + " " + supplier).str.strip()
        vendor_category = unique_ref_series.map(vendor_cat_map).fillna("Uncategorised")
        canonical_name1 = unique_ref_series.map(name_map).fillna("")
        name1 = canonical_name1.where(canonical_name1 != "", name1)

        pm_name_series = unique_ref_series.map(pm_map).fillna("")

        # Sheet decision (Logica Weekly vs Monthly)
        sheets: list[str] = []
        for ur in unique_ref_series.tolist():
            if mode == "weekly":
                # Weekly: Tenta respeitar Key e Historico anterior
                if ur in key_owner_map:
                    sheets.append("Key")
                    continue

                prev = prev_map.get(ur, "")
                if prev == "Query":
                    prev = "ROL"
                if prev in ("Key", "ROL"):
                    sheets.append(prev)
                    continue

                # Se nao tem historico, usa regra normal
                sheets.append(decide_sheet_grouped(ur, key_owner_map))
            else:
                # Monthly (Clean Slate): Recalcula tudo do zero
                sheets.append(decide_sheet_grouped(ur, key_owner_map))

        sheet_series = pd.Series(sheets, index=chunk.index)

        # Owner Mapping
        owners: list[str] = []
        prev_ow: Dict[str, str] = prev_owner_map or {}
        for ur, sh, cat in zip(unique_ref_series.tolist(), sheet_series.tolist(), vendor_category.tolist()):
            if sh == "Key":
                # KEY e sempre absoluto (vem do owner_map_vba.txt / key_owner_map)
                owners.append(key_owner_map.get(ur, ""))
                continue

            active_rol_owner_map = get_rol_owner_map()
            mapped, mapped_owner = lookup_rol_owner_from_category(cat, as_of=as_of, owner_map=active_rol_owner_map)
            # If the current Ledger owner map knows the category, it is the source
            # of truth for the new snapshot. Existing SQLite/archive rows keep
            # their historical owner because daily/latest only replaces the
            # snapshot being loaded.
            if mapped:
                owners.append(mapped_owner)
                continue

            # WEEKLY: preserve previous ROL owner only when the category has no
            # active Ledger owner-map rule.
            if mode == "weekly":
                prev_sheet = prev_map.get(ur, "")
                if prev_sheet.upper() == "QUERY":
                    prev_sheet = "ROL"  # historical backward compat
                if prev_sheet == "ROL":
                    keep = prev_ow.get(ur, "")
                    if keep and "uncategorised" not in keep.lower():
                        owners.append(normalize_rol_owner_for_date(keep, as_of=as_of))
                        continue
                    if keep:
                        owners.append(keep)
                        continue

            # Default: recalculate by category (ROL)
            owners.append(owner_from_category(sh, cat, as_of=as_of, owner_map=active_rol_owner_map))

        owner_series = pd.Series(owners, index=chunk.index)

        # Output DataFrame
        out = pd.DataFrame(
            {
                "Country": country_from_code(company),
                "Vendor category": vendor_category,
                "Sheet": sheet_series,
                "Owner": owner_series,
                "PM Name": pm_name_series,
                "Company Code": company,
                "Document currency": currency,
                "Supplier": supplier,
                "Document Number": docnum,
                "Reference": ref,
                "Posting Period": postper,
                "Amount in doc. curr.": amount,
                "Document Type": doctype,
                "Name 1": name1,
                "Document Date": docdate,
                "Net due date": netdue,
                "Text": text,
                "User name": user,
                "Posting Date": postdt,
                "Payment Block": payblk,
                "Unique Ref": unique_ref_series,
                "Payment Issues": pd.Series(issues_list, index=chunk.index),
            },
            columns=MASTER_COLS,
        )

        # Filtro de linhas vazias e escrita
        out = out[(out["Company Code"].astype(str).str.len() > 0) & (out["Supplier"].astype(str).str.len() > 0)]
        append_df_to_csv(out, out_csv, header=write_header)
        if progress_callback is not None:
            progress_callback(chunk_index, int(len(out)))
        write_header = False

    return write_header
