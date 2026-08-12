# -*- coding: utf-8 -*-
"""
SyntheticReview helper functions for enrichment and file detection.
Reusable by synthetic_review_loader.py and future SyntheticReview API integration.
"""
from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

logger = logging.getLogger(__name__)


def build_unique_ref(division_ref, vendor_no) -> str:
    """Build Unique Ref from Division Ref + Vendor No (e.g. 'SYN-CC-006 100719')."""
    dr = str(division_ref).strip() if division_ref else ''
    vn = str(vendor_no).strip() if vendor_no else ''
    # Remove .0 from float-like vendor numbers
    if vn.endswith('.0'):
        vn = vn[:-2]
    # Strip leading zeros (SyntheticReview exports 0020062904, MasterData uses 20062904)
    if vn:
        vn = vn.lstrip('0') or '0'
    return f"{dr} {vn}" if dr and vn else ''


def assign_owner(
    unique_ref: str,
    category: str,
    key_map: Dict[str, str],
    rol_map: Dict[str, str],
) -> str:
    """
    2-tier owner assignment:
    1. Key owner map (by unique_ref)
    2. ROL category map (by category)  Query team merged into ROL
    3. Fallback: "Unassigned"
    """
    from .masterdata_core import clean_text

    # Tier 1: Key owner
    if unique_ref in key_map:
        owner = key_map[unique_ref]
        if owner:
            return owner

    # Tier 2: Category-based (unified ROL map)
    cat = clean_text(category) if category else ''
    if cat:
        if cat in rol_map and rol_map[cat]:
            return rol_map[cat]

    return 'Unassigned'


def determine_team(
    unique_ref: str,
    category: str,
    key_map: Dict[str, str],
    rol_map: Dict[str, str],
) -> str:
    """
    Determine which tier (team) assigned the owner:
    'Key', 'ROL', or '' if unassigned. Query team eliminated  merged into ROL.
    """
    from .masterdata_core import clean_text

    if unique_ref in key_map and key_map[unique_ref]:
        return 'Key'

    cat = clean_text(category) if category else ''
    if cat:
        if cat in rol_map and rol_map[cat]:
            return 'ROL'

    return ''


def find_latest_masterdata(master_dir: Path) -> Optional[Path]:
    """
    Find the most recent MasterData_WWYY.csv file.
    Naming: MasterData_0926.csv = week 09, year 2026.
    Searches both master_dir and master_dir/archive/.
    Returns path to latest file, or None if none found.
    """
    pat = re.compile(r'^MasterData_(\d{2})(\d{2})\.csv$', re.IGNORECASE)
    candidates = []
    # Search main dir and archive/
    search_dirs = [master_dir]
    archive_dir = master_dir / "archive"
    if archive_dir.exists():
        search_dirs.append(archive_dir)
    for d in search_dirs:
        for f in d.glob('MasterData_*.csv'):
            m = pat.match(f.name)
            if m:
                week, year = int(m.group(1)), int(m.group(2))
                candidates.append((year, week, f))
    if not candidates:
        return None
    candidates.sort(reverse=True)  # highest year+week first
    return candidates[0][2]


def load_masterdata_lookup(masterdata_path: Path) -> Dict[str, dict]:
    """
    Load MasterData CSV and build lookup: UniqueRef -> {Category, Owner, Team}.
    Deduplicates by Unique Ref (first occurrence wins, all rows have same owner).
    """
    import pandas as pd

    required = ['Unique Ref', 'Vendor category', 'Owner', 'Sheet']
    df = pd.read_csv(str(masterdata_path), low_memory=False)
    # Only keep columns that exist (Country may be absent in older CSVs)
    cols = [c for c in required + ['Country'] if c in df.columns]
    df = df[cols]
    df = df.dropna(subset=['Unique Ref'])
    df = df.drop_duplicates(subset='Unique Ref', keep='first')

    # Vectorized: clean Unique Ref column
    df['Unique Ref'] = df['Unique Ref'].astype(str).str.strip()
    # Filter out empty refs
    df = df[df['Unique Ref'] != '']

    # Vectorized: fill NaN and strip all string columns
    for col in ['Vendor category', 'Owner', 'Sheet', 'Country']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()
        else:
            df[col] = ''

    # Build lookup dict using vectorized approach
    lookup = {}
    for ref, cat, owner, team, country in zip(
        df['Unique Ref'],
        df['Vendor category'],
        df['Owner'],
        df['Sheet'],
        df['Country']
    ):
        lookup[ref] = {
            'Category': cat,
            'Owner': owner,
            'Team': team,
            'Country': country,
        }
    return lookup


def load_vendor_matrix_country_map(vendor_matrix_path: Path) -> Dict[str, str]:
    """
    Load Vendor Master Matrix CSV and build lookup: UniqueRef -> Country.
    Used as fallback when MasterData doesn't have the entry.
    """
    import pandas as pd
    from .masterdata_core import clean_text

    path = str(vendor_matrix_path)
    if not Path(path).exists():
        return {}

    import csv
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        header = next(csv.reader(f))
    cols = {clean_text(c): c for c in header}

    ur_col = cols.get("unique ref")
    country_col = cols.get("country")
    if not ur_col or not country_col:
        return {}

    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig",
                     usecols=[ur_col, country_col], low_memory=False)
    df[ur_col] = df[ur_col].fillna("").astype(str).str.strip()
    df[country_col] = df[country_col].fillna("").astype(str).str.strip()
    df = df[df[ur_col] != ""]
    df = df[df[country_col] != ""]
    df = df.drop_duplicates(subset=[ur_col], keep="last")

    return dict(zip(df[ur_col], df[country_col]))


def build_synthetic_review_owner_map(
    vendor_matrix_path: Path,
    owner_map_path: Path,
    master_dir: Optional[Path] = None,
) -> Tuple[Dict[str, dict], Dict[str, str], Dict[str, str], Dict[str, str], Dict[str, str]]:
    """
    Build all lookup maps needed for SyntheticReview enrichment.
    Returns: (md_lookup, category_map, key_owner_map, rol_owner_map, vm_country_map)

    md_lookup: MasterData primary lookup (UniqueRef -> {Category, Owner, Team, Country})
    category_map + key/rol maps: Fallback when UniqueRef not in MasterData.
    vm_country_map: Vendor Matrix country lookup (UniqueRef -> Country) for fallback.
    Note: Query team eliminated  all categories now covered by unified ROL map.
    """
    from .masterdata_core import (
        load_vendor_category_map,
        load_key_owner_map_from_vba_txt,
        build_owner_map_rol,
    )

    # Sources are independent: one unavailable owner file must not discard
    # Category/Country maps that were already loaded successfully.
    md_lookup: Dict[str, dict] = {}
    if master_dir:
        try:
            md_file = find_latest_masterdata(master_dir)
            if md_file:
                md_lookup = load_masterdata_lookup(md_file)
                logger.info("MasterData primary: %s (%d refs)", md_file.name, len(md_lookup))
        except Exception as exc:
            logger.warning("MasterData lookup unavailable; using Vendor Matrix fallback: %s", exc)

    # Fallback: Vendor Matrix + Key/ROL maps (Query team merged into ROL)
    try:
        category_map = load_vendor_category_map(str(vendor_matrix_path))
    except Exception as exc:
        logger.warning("Vendor Matrix category lookup unavailable; using MasterData only: %s", exc)
        category_map = {}

    try:
        key_owner_map = load_key_owner_map_from_vba_txt(str(owner_map_path))
    except Exception as exc:
        logger.warning("Key owner lookup unavailable; continuing without Key owner fallback: %s", exc)
        key_owner_map = {}

    try:
        rol_map = build_owner_map_rol()
    except Exception as exc:
        logger.warning("Ledger owner lookup unavailable; preserving Category lookups: %s", exc)
        rol_map = {}

    # Country fallback from Vendor Matrix
    try:
        vm_country_map = load_vendor_matrix_country_map(vendor_matrix_path)
    except Exception as exc:
        logger.warning("Vendor Matrix country lookup unavailable; using MasterData country only: %s", exc)
        vm_country_map = {}
    if vm_country_map:
        logger.info("Vendor Matrix country map: %d refs", len(vm_country_map))

    return md_lookup, category_map, key_owner_map, rol_map, vm_country_map


def merge_duplicate_pair_owners(df) -> None:
    """
    For duplicate pairs (SourceType='Duplicate Invoice'), if a PairID has
    rows with different owners, merge them: 'Synthetic Owner 022 / Synthetic Owner 019'.
    Modifies df in-place.
    """
    if 'SourceType' not in df.columns or 'PairID' not in df.columns:
        return
    mask = (df['SourceType'] == 'Duplicate Invoice') & df['PairID'].ne('') & df['PairID'].notna()
    dupes = df.loc[mask]
    if dupes.empty:
        return

    # Group by PairID, find pairs with multiple unique owners
    for pair_id, group in dupes.groupby('PairID'):
        owners = [o for o in group['Owner'].unique() if o and o != 'Unassigned']
        if len(owners) > 1:
            merged = ' / '.join(sorted(owners))
            df.loc[group.index, 'Owner'] = merged


def detect_synthetic_review_file_type(file_path: Path) -> Optional[str]:
    """
    Detect SyntheticReview file type by inspecting column headers.
    Returns: 'errors', 'duplicates', or None
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    result = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(max_row=1))]
        if 'Duplicate pair ID' in headers:
            result = 'duplicates'
            break
        if 'Error type' in headers:
            result = 'errors'
            break
    wb.close()
    return result


def normalize_date(val) -> str:
    """Convert various date formats to YYYY-MM-DD."""
    if val is None or val == '':
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s
