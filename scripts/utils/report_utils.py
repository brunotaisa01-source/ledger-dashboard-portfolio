# -*- coding: utf-8 -*-
"""Shared utility functions for build_key_report and build_ledger_report."""
from __future__ import annotations

import io
import os
import re
import shutil
import tempfile
import time
from contextlib import contextmanager
from datetime import datetime

import numpy as np
import pandas as pd

from .masterdata_core import (
    norm_str_series as _norm_str_series,
    parse_date_series as _parse_date_series,
)


# ---------------------------------------------------------------------------
# Report write helpers
# ---------------------------------------------------------------------------

@contextmanager
def report_build_lock(lock_path: str, metadata: str):
    if os.path.exists(lock_path):
        age_min = (time.time() - os.path.getmtime(lock_path)) / 60
        raise RuntimeError(
            f"Build lock exists (age {age_min:.0f}min): {lock_path}\n"
            f"Outro build rodando, ou o ultimo crashou. Apague o .lock se tiver certeza."
        )

    lock_fd = os.open(lock_path, os.O_WRONLY | os.O_CREAT | os.O_EXCL)
    with os.fdopen(lock_fd, "w", encoding="utf-8") as lock_file:
        lock_file.write(metadata)

    try:
        yield
    finally:
        try:
            os.remove(lock_path)
        except FileNotFoundError:
            pass


def report_staging_path(output_path: str) -> str:
    staging_dir = os.environ.get("SYNTHETIC_REPORTING_STAGING_DIR")
    if not staging_dir:
        staging_dir = os.path.join(tempfile.gettempdir(), "synthetic_reporting", "report_staging")
    os.makedirs(staging_dir, exist_ok=True)
    stem, ext = os.path.splitext(os.path.basename(output_path))
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return os.path.join(staging_dir, f"{stem}.{os.getpid()}.{stamp}{ext or '.xlsx'}")


def _wait_for_stable_file(
    path: str,
    *,
    timeout_s: float = 600.0,
    interval_s: float = 5.0,
    stable_checks: int = 3,
) -> int:
    deadline = time.monotonic() + timeout_s
    last_size = -1
    stable_count = 0
    last_error: Exception | None = None

    while time.monotonic() < deadline:
        try:
            if os.path.exists(path):
                size = os.path.getsize(path)
                if size > 0 and size == last_size:
                    stable_count += 1
                else:
                    last_size = size
                    stable_count = 1 if size > 0 else 0
                if stable_count >= stable_checks:
                    return size
        except OSError as err:
            last_error = err
        time.sleep(interval_s)

    suffix = f" Last error: {last_error}" if last_error else ""
    raise TimeoutError(f"Timed out waiting for stable report file: {path}.{suffix}")


def _assert_workbook_opens(path: str) -> None:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as err:
        raise RuntimeError(f"Report workbook could not be opened: {path}") from err
    try:
        workbook.sheetnames
    finally:
        workbook.close()


def _remove_if_exists(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


def _retry_until_deadline(action_name: str, deadline: float, interval_s: float, func) -> None:
    last_error: Exception | None = None
    while time.monotonic() < deadline:
        try:
            func()
            return
        except OSError as err:
            last_error = err
            time.sleep(interval_s)
    raise TimeoutError(f"Timed out during {action_name}. Last error: {last_error}")


def _remaining_timeout(deadline: float) -> float:
    remaining = deadline - time.monotonic()
    if remaining <= 0:
        raise TimeoutError("Timed out during report publish validation.")
    return remaining


def _copy_report_for_publish(source_path: str, publish_tmp: str) -> None:
    shutil.copyfile(source_path, publish_tmp)
    os.utime(publish_tmp, None)


def publish_report_output(
    staging_path: str,
    output_path: str,
    *,
    timeout_s: float = 600.0,
    interval_s: float = 5.0,
    stable_checks: int = 3,
) -> int:
    if not os.path.exists(staging_path):
        raise RuntimeError(f"Report staging file not found: {staging_path}")

    deadline = time.monotonic() + timeout_s
    bounded_interval_s = max(1.0, min(interval_s, 5.0))
    source_size = _wait_for_stable_file(
        staging_path,
        timeout_s=_remaining_timeout(deadline),
        interval_s=bounded_interval_s,
        stable_checks=stable_checks,
    )
    _assert_workbook_opens(staging_path)

    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    stem, ext = os.path.splitext(os.path.basename(output_path))
    publish_tmp = os.path.join(
        output_dir,
        f".{stem}.{os.getpid()}.{datetime.now().strftime('%Y%m%d%H%M%S')}.publish{ext or '.xlsx'}",
    )
    publish_started = time.time()

    try:
        _retry_until_deadline(
            "removing stale publish temp report",
            deadline,
            bounded_interval_s,
            lambda: _remove_if_exists(publish_tmp),
        )
        _retry_until_deadline(
            "copying staged report to operational folder",
            deadline,
            bounded_interval_s,
            lambda: _copy_report_for_publish(staging_path, publish_tmp),
        )
        copied_size = _wait_for_stable_file(
            publish_tmp,
            timeout_s=_remaining_timeout(deadline),
            interval_s=bounded_interval_s,
            stable_checks=stable_checks,
        )
        if copied_size != source_size:
            raise RuntimeError(
                f"Publish temp size mismatch for {publish_tmp}: {copied_size} != {source_size}"
            )
        _assert_workbook_opens(publish_tmp)

        _retry_until_deadline(
            "renaming published temp report to final workbook",
            deadline,
            bounded_interval_s,
            lambda: os.replace(publish_tmp, output_path),
        )
    except PermissionError as err:
        raise PermissionError(
            "Could not publish report output. Close the existing workbook in Excel, "
            f"wait for Local Fixture Store to release it, and rerun. Staging file kept at: {staging_path}; "
            f"destination: {output_path}"
        ) from err
    except OSError as err:
        raise OSError(
            f"Could not publish report output {staging_path} -> {output_path}: {err}"
        ) from err
    finally:
        _remove_if_exists(publish_tmp)

    final_size = _wait_for_stable_file(
        output_path,
        timeout_s=_remaining_timeout(deadline),
        interval_s=bounded_interval_s,
        stable_checks=stable_checks,
    )
    if final_size != source_size:
        raise RuntimeError(
            f"Published report size mismatch for {output_path}: {final_size} != {source_size}"
        )
    final_mtime = os.path.getmtime(output_path)
    if final_mtime < publish_started - 2:
        raise RuntimeError(
            "Published report timestamp did not refresh: "
            f"{output_path} (mtime={final_mtime:.0f}, publish_started={publish_started:.0f})"
        )
    _assert_workbook_opens(output_path)
    return final_size


# ---------------------------------------------------------------------------
# String / value helpers
# ---------------------------------------------------------------------------

def clean_issue_name(val):
    if pd.isna(val): return ""
    return str(val).strip()


def clean_key(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    if s.endswith(".0"): return s[:-2]
    return s


def sanitize_value(val):
    if pd.isna(val): return ""
    if isinstance(val, (pd.Timestamp, datetime)): return val
    if hasattr(val, "item"): return val.item()
    return val


# ---------------------------------------------------------------------------
# Numeric parsing
# ---------------------------------------------------------------------------

def parse_amount_series(s: pd.Series) -> pd.Series:
    s = s.fillna("").astype(str).str.strip()
    s = s.str.replace("\u00A0", "", regex=False)
    s = s.str.replace(" ", "", regex=False)
    def norm_one(x: str):
        if not x or x.lower() in {"nan", "none"}: return np.nan
        if "," in x and "." in x:
            if x.rfind(",") > x.rfind("."): x = x.replace(".", "").replace(",", ".")
            else: x = x.replace(",", "")
        else:
            if x.count(",") == 1 and "." not in x: x = x.replace(",", ".")
            else: x = x.replace(",", "")
        try: return float(x)
        except (ValueError, TypeError): return np.nan
    return s.map(norm_one)


def parse_mixed_date_series(s: pd.Series) -> pd.Series:
    """Parse ERP ISO and day-first date formats without inference."""
    return _parse_date_series(s)


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def recover_date_from_tail(val):
    if pd.isna(val) or str(val).strip() == "": return ""
    parsed = _parse_date_series(pd.Series([val])).iloc[0]
    if pd.isna(parsed):
        return val
    return parsed.to_pydatetime()


def _extract_snapshot_date(path):
    """Extract date from filename like 'Key 09.02.xlsx' or 'Ledger 09.02.xlsx'."""
    base = os.path.basename(path)
    match = re.search(r"(\d{2})\.(\d{2})", base)
    if not match:
        return datetime.now()
    day, month = int(match.group(1)), int(match.group(2))
    today = datetime.now()
    best = None
    for y in (today.year, today.year - 1):
        try:
            d = datetime(y, month, day)
            if best is None or abs((d - today).days) < abs((best - today).days):
                best = d
        except ValueError:
            pass
    return best if best else today


def extract_date_from_filename(f):
    base = os.path.basename(f)
    match = re.search(r"(\d{2})\.(\d{2})", base)
    if not match: return (0, 0, 0)
    day, month = int(match.group(1)), int(match.group(2))
    today = datetime.now()
    best = None
    for y in (today.year, today.year - 1):
        try:
            d = datetime(y, month, day)
            if best is None or abs((d - today).days) < abs((best - today).days):
                best = d
        except ValueError:
            pass
    if best: return (best.year, best.month, best.day)
    return (0, 0, 0)


# ---------------------------------------------------------------------------
# Company code mapping
# ---------------------------------------------------------------------------

def system_from_code(code):
    up = str(code).strip().upper() if code is not None else ""
    if up.startswith("GB"): return "ERP5"
    if up in ("SYN-CC-002", "SYN-CC-001"): return "ERP3"
    if up.startswith(("BE", "NL", "LU", "FR05", "FR02", "SYN-CC-004", "IT01", "IT02")): return "ERP2"
    if up in ("SYN-CC-005", "SYN-CC-006", "SYN-CC-007", "SYN-CC-008", "SYN-CC-009"): return "ERP1"
    if up == "SYN-CC-003": return "ERP4"
    return ""


# ---------------------------------------------------------------------------
# Pandas helpers
# ---------------------------------------------------------------------------

def groupby_apply_no_warning(gb, func):
    try: return gb.apply(func, include_groups=False)
    except TypeError: return gb.apply(func)


def _bucket_mask(g, low, high):
    m = g["_IsDetail"] & (g["_Amt"].fillna(0) != 0)
    if high is None: return m & (g["_Days"] > low)
    return m & (g["_Days"] > low) & (g["_Days"] <= high)


def _supplier_key(g):
    if "Unique Ref" in g.columns: return _norm_str_series(g["Unique Ref"]).replace({"": np.nan}).fillna("")
    return _norm_str_series(g.get("Supplier", ""))


# ---------------------------------------------------------------------------
# Payment Issues aggregation
# ---------------------------------------------------------------------------

def agg_payment_issues(d):
    if d.empty or "Payment Issues" not in d.columns:
        return pd.DataFrame(columns=["Issue", "Owner", "Company Code", "Vendor", "Name", "Qty"])
    mask = d["Payment Issues"].notna() & (d["Payment Issues"].astype(str).str.strip() != "")
    sub = d[mask].copy()
    if sub.empty:
        return pd.DataFrame(columns=["Issue", "Owner", "Company Code", "Vendor", "Name", "Qty"])
    sub["Issue"] = sub["Payment Issues"].apply(clean_issue_name)
    sub["Vendor"] = sub["Supplier"].astype(str).str.strip()
    sub["Name"] = sub["Name 1"].astype(str).str.strip()
    sub["Owner"] = sub["Owner"].astype(str).str.strip()
    sub["Company Code"] = sub["Company Code"].astype(str).str.strip() if "Company Code" in sub.columns else ""
    return sub.groupby(["Issue", "Owner", "Company Code", "Vendor", "Name"], dropna=False).size().reset_index(name="Qty")


# ---------------------------------------------------------------------------
# Report output validation (prevents corrupted xlsx from reaching users)
# ---------------------------------------------------------------------------

_DEFAULT_SKIP_TABS = (
    "Summary", "Weekly_Movement", "Payment Issues",
    "SyntheticReview Duplicates", "SyntheticReview Errors", "ZR Blocks",
    "SyntheticReview Pending Recovery", "SyntheticReview Action Required",
)


def _norm_cell(v):
    if v is None:
        return ""
    return str(v).replace(" ", " ").strip()


def validate_owner_tab_integrity(xlsx_path, skip_tabs=_DEFAULT_SKIP_TABS):
    # Multi-check integrity validator for owner-style report tabs.
    # Fails closed: the first error per tab stops that tab's scan, but every
    # tab is scanned so the caller gets a complete picture of what's wrong.
    #
    # Checks (regression guards for the Key 20.04.xlsx incident):
    #   C1  Unique Ref appears in exactly ONE contiguous range per tab
    #   C2  Each Unique Ref group has exactly ONE summary row
    #       (summary = row with empty Document Number)
    #   C3  Summary TOTAL VOL matches the number of detail rows in its group
    #   C4  Name 1 is consistent within a group (after NBSP/whitespace norm)
    #
    # Tabs without a "Unique Ref" column, and tabs in skip_tabs, are ignored.
    from openpyxl import load_workbook

    skip = set(skip_tabs)
    # Load via BytesIO so non-standard extensions like .tmp work (atomic-write path)
    with open(xlsx_path, "rb") as _fh:
        wb = load_workbook(io.BytesIO(_fh.read()), read_only=True)
    errors = []

    def _close_group(tab, st):
        if st is None:
            return
        # C3: summary VOL vs detail count - only when group has BOTH a summary
        # and details. Summary-only (detail_count=0) might look like a VOL
        # mismatch but the real problem is usually fragmentation, which C1
        # detects when the rest of the group shows up later in the sheet.
        if st["expected_vol"] is None or st["detail_count"] == 0:
            return
        try:
            expected = int(st["expected_vol"])
        except (TypeError, ValueError):
            return
        if expected != st["detail_count"]:
            errors.append(
                f"{tab}: Unique Ref {st['ur']!r} summary TOTAL VOL={expected} "
                f"but counted {st['detail_count']} detail rows"
            )

    try:
        for ws in wb.worksheets:
            if ws.title in skip:
                continue
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row or "Unique Ref" not in header_row:
                continue

            idx_ur = header_row.index("Unique Ref")
            idx_doc = header_row.index("Document Number") if "Document Number" in header_row else None
            idx_name = header_row.index("Name 1") if "Name 1" in header_row else None
            idx_vol = header_row.index("TOTAL VOL") if "TOTAL VOL" in header_row else None

            sentinel = object()
            prev_ur = sentinel
            seen_urs = set()
            state = None  # current group tracking
            tab_failed = False

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if tab_failed:
                    break
                ur = row[idx_ur] if idx_ur < len(row) else None

                # summary = empty Document Number
                is_summary = False
                if idx_doc is not None and idx_doc < len(row):
                    is_summary = _norm_cell(row[idx_doc]) == ""

                if ur != prev_ur:
                    # end of previous group
                    _close_group(ws.title, state)
                    if errors and errors[-1].startswith(f"{ws.title}:"):
                        tab_failed = True
                        break

                    # C1: fragmentation
                    if ur in seen_urs:
                        errors.append(f"{ws.title}: Unique Ref {ur!r} fragmented at row {i}")
                        tab_failed = True
                        break
                    seen_urs.add(ur)

                    name_val = None
                    if idx_name is not None and idx_name < len(row):
                        name_val = _norm_cell(row[idx_name])
                    vol_val = None
                    if is_summary and idx_vol is not None and idx_vol < len(row):
                        vol_val = row[idx_vol]

                    state = {
                        "ur": ur,
                        "name1": name_val,
                        "expected_vol": vol_val,
                        "summary_count": 1 if is_summary else 0,
                        "detail_count": 0 if is_summary else 1,
                    }
                    prev_ur = ur
                else:
                    # continuing the same group
                    if is_summary:
                        state["summary_count"] += 1
                        if state["summary_count"] > 1:
                            # C2: multiple summaries
                            errors.append(
                                f"{ws.title}: Unique Ref {ur!r} has multiple summary rows (row {i})"
                            )
                            tab_failed = True
                            break
                        if state["expected_vol"] is None and idx_vol is not None and idx_vol < len(row):
                            state["expected_vol"] = row[idx_vol]
                    else:
                        state["detail_count"] += 1
                        # C4: Name 1 consistency
                        if idx_name is not None and idx_name < len(row):
                            row_name = _norm_cell(row[idx_name])
                            if not state["name1"]:
                                state["name1"] = row_name
                            elif row_name and row_name != state["name1"]:
                                errors.append(
                                    f"{ws.title}: Unique Ref {ur!r} has divergent Name 1 "
                                    f"({state['name1']!r} vs {row_name!r} at row {i})"
                                )
                                tab_failed = True
                                break

            # close the last group of this tab
            if not tab_failed:
                _close_group(ws.title, state)
    finally:
        wb.close()

    if errors:
        raise RuntimeError(
            "Report validation FAILED:\n  " + "\n  ".join(errors)
        )


# ---------------------------------------------------------------------------
# MasterData CSV integrity validator (shared by weekly + monthly builders)
# ---------------------------------------------------------------------------

_MASTERDATA_REQUIRED_HEADERS = (
    "Sheet", "Owner", "Company Code", "Supplier", "Name 1", "Unique Ref",
)

_MASTERDATA_SHEET_VALUES = {"key", "rol", "query", "uncathegorised", "uncategorised"}

_MASTERDATA_KNOWN_OWNERS = {
    "Synthetic Owner 011", "Synthetic Owner 016", "Synthetic Owner 017", "Synthetic Owner 014", "Synthetic Owner 018", "Synthetic Owner 019", "Synthetic Owner 012",
    "Synthetic Owner 010", "Synthetic Owner 020", "Synthetic Owner 021", "Synthetic Owner 022", "Synthetic Owner 023", "Synthetic Owner 024", "Synthetic Owner 025", "Rent", "Synthetic Owner 013",
    "Tom",
    "Synthetic Owner 015", "Synthetic Owner 026", "Fuel", "Unassigned", "",
}


def validate_masterdata_csv(path, *, min_rows=10_000, max_rows=2_000_000, min_bytes=1024,
                            required_headers=_MASTERDATA_REQUIRED_HEADERS):
    # Structural check for MasterData_*.csv before atomic rename.
    # Streams the file (no full load into memory) and fails closed with
    # a clear message. Called from build_masterdata_{weekly,monthly}.py
    # after writing to {out}.tmp, before os.replace(.tmp, .csv).
    import csv
    import os

    if not os.path.exists(path):
        raise RuntimeError(f"MasterData validation: file not found: {path}")
    size = os.path.getsize(path)
    if size < min_bytes:
        raise RuntimeError(f"MasterData validation: file too small ({size} bytes): {path}")

    errors = []
    unknown_owners = set()
    n_data_rows = 0
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            raise RuntimeError(f"MasterData validation: empty file: {path}")

        header_clean = [h.strip() for h in header]
        missing = [h for h in required_headers if h not in header_clean]
        if missing:
            raise RuntimeError(
                f"MasterData validation: missing required headers {missing} in {path}"
            )

        idx = {h: header_clean.index(h) for h in required_headers}

        for row in reader:
            n_data_rows += 1
            ur = row[idx["Unique Ref"]] if idx["Unique Ref"] < len(row) else ""
            if not ur.strip():
                errors.append(f"row {n_data_rows + 1}: empty Unique Ref")
                if len(errors) >= 5:
                    break
                continue
            sheet_val = row[idx["Sheet"]] if idx["Sheet"] < len(row) else ""
            if sheet_val.strip().lower() not in _MASTERDATA_SHEET_VALUES:
                errors.append(f"row {n_data_rows + 1}: unknown Sheet {sheet_val!r}")
                if len(errors) >= 5:
                    break
                continue
            owner_val = row[idx["Owner"]].strip() if idx["Owner"] < len(row) else ""
            if owner_val not in _MASTERDATA_KNOWN_OWNERS:
                unknown_owners.add(owner_val)

    if errors:
        raise RuntimeError(
            f"MasterData validation FAILED ({path}):\n  " + "\n  ".join(errors)
        )

    if n_data_rows < min_rows:
        raise RuntimeError(
            f"MasterData validation: too few rows ({n_data_rows} < {min_rows}) in {path}"
        )
    if n_data_rows > max_rows:
        raise RuntimeError(
            f"MasterData validation: too many rows ({n_data_rows} > {max_rows}) in {path}"
        )

    if unknown_owners:
        import logging
        logging.getLogger(__name__).warning(
            "MasterData validation: unknown Owner values (non-blocking): %s",
            sorted(o for o in unknown_owners if o)[:10],
        )

    return {"rows": n_data_rows, "unknown_owners": sorted(unknown_owners)}
