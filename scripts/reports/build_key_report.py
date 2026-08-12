# -*- coding: utf-8 -*-
"""Build Key AP Report  generates the weekly Key team Excel report from SQLite data."""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import os
import glob
import sys
import re
import time
import logging
from datetime import datetime, date
from pathlib import Path
import warnings

# Ignora avisos
warnings.filterwarnings('ignore', category=UserWarning)

log = logging.getLogger(__name__)

# --- CONFIGURACAO DE OWNERS "FANTASMA" (MESTRE PYTHON) ---
# Estes Owners terao abas criadas, mas NAO entrarao na soma do Summary/KPIs
OWNERS_HIDDEN_FROM_SUMMARY = ["RENT", "ALUGUEL"]

from ..utils.masterdata_core import load_key_team_owners, norm_str_series as _norm_str_series
from ..utils.report_utils import (
    clean_issue_name, parse_amount_series, parse_mixed_date_series,
    groupby_apply_no_warning as _groupby_apply_no_warning,
    system_from_code, clean_key, sanitize_value, recover_date_from_tail,
    _extract_snapshot_date, extract_date_from_filename, _bucket_mask, _supplier_key,
    agg_payment_issues as _agg_pi,
    validate_owner_tab_integrity,
)

COR_HEADER = "002060"; COR_FONTE = "FFFFFF"
THIN_SIDE = Side(border_style="thin", color="000000")
ALL_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE)
from ..utils.paths import MASTER_DATA, MASTER_ARCHIVE, KEY_DATA, KEY_ARCHIVE, archive_old_files
ROOT_DIR = str(MASTER_DATA)
DATA_DIR = str(KEY_DATA)
ARCHIVE_DIR = str(KEY_ARCHIVE)

def get_latest_master_auto(directory):
    pattern = os.path.join(directory, "MasterData_*.csv")
    files = glob.glob(pattern)
    # Also search archive/ for MasterData CSVs
    archive_pattern = os.path.join(str(MASTER_ARCHIVE), "MasterData_*.csv")
    files += glob.glob(archive_pattern)
    if not files: return None
    def _week_key(f):
        m = re.search(r'MasterData_(?P<week>\d{2})(?P<yy>\d{2})', os.path.basename(f))
        return (int(m.group('yy')), int(m.group('week'))) if m else (0, 0)  # (year, week)
    return max(files, key=_week_key)

data_hoje = datetime.now().strftime("%d.%m")
OUTPUT_FILENAME = f"Key {data_hoje}.xlsx"
OUTPUT_PATH = os.path.join(DATA_DIR, OUTPUT_FILENAME)
KEY_OWNER_MAP_SOURCE = os.path.join(ROOT_DIR, "Owner_map.csv")

TAIL_COLS = ["Query type", "Status", "AP Specialist comment", "Next Step", "TL Comment", "Action Date"]
REVIEW_THIS_WEEK_OPTIONS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
COMPLETE_OPTIONS = ["Yes", "No"]
FINAL_HEADERS = [
    "Country", "Vendor category", "Company Code", "Supplier", "Name 1", "Document Date", "Document Number", "Reference", "Amount in doc. curr.", "Document Type",
    "Net due date", "Document currency", "Posting Date", "Payment Block", "0-30 Days overdue", "31-60 Days overdue", "61-90 Days overdue", "91-120 Days Overdue", "121-180 Days Overdue", "180> Days Overdue",
    "TOTAL VALUE", "TOTAL VOL", "Query type", "Status", "AP Specialist comment", "Next Step", "Action Date", "TL Comment", "Open Payment", "Sheet", "Owner", "Text", "Unique Ref", "System", "User name", "Payment Issues",
    "Review this week", "Complete",
]
OPCOES_STATUS = ["Awaiting PR", "Awaiting DD", "In Progress", "Documents Requested", "Documation - Awaiting Posting", "Documation - Awaiting Approval", "Awaiting Manual Posting", "Resolved", "Disputed", "Blocker"]
OPCOES_QUERY = ["Awaiting Payment - BAU", "Awaiting Payment - Overdues", "Blocked", "Posting Error", "Missing Documents", "Incoming Cash", "Credit Note", "Missing payment"]
COLS_GREEN = ["0-30 Days overdue", "31-60 Days overdue", "61-90 Days overdue", "91-120 Days Overdue"]
COLS_RED = ["121-180 Days Overdue", "180> Days Overdue"]
COLS_TOTAL = ["TOTAL VALUE", "TOTAL VOL"]
cf_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
cf_yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
fill_blue_light = PatternFill("solid", fgColor="DAE9F8"); fill_green_light = PatternFill("solid", fgColor="C1F0C8")
fill_red_alert = PatternFill("solid", fgColor="FF0000"); fill_blue_total = PatternFill("solid", fgColor="83CCEB")
fill_summary_gray = PatternFill("solid", fgColor="D9D9D9")
font_black_bold = Font(bold=True, color="000000"); font_white_bold = Font(bold=True, color="FFFFFF")
font_header_custom = Font(bold=True, color=COR_FONTE); fill_header_custom = PatternFill("solid", fgColor=COR_HEADER)

_re_intercompany = re.compile(r"inter\s*company|intercompany", re.IGNORECASE)

BUCKETS = [("0-30", 0, 30), ("31-60", 30, 60), ("61-90", 60, 90), ("91-120", 90, 120), ("121-180", 120, 180), ("180+", 180, None)]
FOCUS = "180+"

#  pure functions 

def get_bucket_rank(days_overdue):
    if pd.isna(days_overdue) or days_overdue < 0: return 0
    if days_overdue > 180: return 6
    if days_overdue >= 121: return 5
    if days_overdue >= 91: return 4
    if days_overdue >= 61: return 3
    if days_overdue >= 31: return 2
    if days_overdue >= 7: return 1
    return 0

def find_latest_file(directory):
    search_pattern = os.path.join(directory, "Key *.xlsx")
    files = glob.glob(search_pattern)
    # Also search archive/ for previous week files
    archive_pattern = os.path.join(ARCHIVE_DIR, "Key *.xlsx")
    files += glob.glob(archive_pattern)
    valid_files = [f for f in files if os.path.basename(f) != OUTPUT_FILENAME and not os.path.basename(f).startswith("~$")]
    if not valid_files: return None
    return max(valid_files, key=extract_date_from_filename)

def _clean_series_for_nunique(s): return s.astype(str).str.strip().replace({"": np.nan, "nan": np.nan, "NaN": np.nan, "NONE": np.nan, "None": np.nan, "NAN": np.nan})
def distinct_unique_refs(g): return int(_clean_series_for_nunique(g[g["_IsDetail"]&(g["_Amt"]!=0)]["Unique Ref"]).nunique(dropna=True))
def distinct_documents(g): return int(len(g[g["_IsDetail"]&(g["_Amt"]!=0)]))
def distinct_suppliers_over(g, days_limit): return int(_clean_series_for_nunique(g[g["_IsDetail"]&(g["_Days"]>days_limit)&(g["_Amt"]!=0)]["Unique Ref"]).nunique(dropna=True))
def distinct_suppliers_between(g, low, high): return int(_clean_series_for_nunique(g[g["_IsDetail"]&(g["_Days"]>low)&(g["_Days"]<=high)&(g["_Amt"]!=0)]["Unique Ref"]).nunique(dropna=True))
def distinct_documents_over(g, days_limit): return int(len(g[g["_IsDetail"]&(g["_Days"]>days_limit)&(g["_Amt"]!=0)]))
def distinct_documents_between(g, low, high): return int(len(g[g["_IsDetail"]&(g["_Days"]>low)&(g["_Days"]<=high)&(g["_Amt"]!=0)]))

def _merge_tail(old, new):
    """Merge tail data dicts  new values only fill empty slots in old."""
    merged = dict(old)
    for k, v in new.items():
        if k.startswith("_"):
            if k not in merged and v: merged[k]=v
            continue
        if v and not merged.get(k): merged[k]=v
    return merged

def _store_tail(tail_db, key, row_tail):
    """Store or merge row_tail data into tail_db at given key."""
    if key in tail_db:
        tail_db[key] = _merge_tail(tail_db[key], row_tail)
    else:
        tail_db[key] = row_tail

def _process_tail_sheet(df_old, sheet_name):
    """Extract tail data (query type, status, comments) from a single Excel sheet.
    Returns dict mapping keys to tail data dicts.

    NOTE: Uses iterrows() instead of itertuples() for correctness.
    itertuples() creates positional _0, _1, _2 attributes for columns with spaces/numbers,
    breaking column access. Performance impact is minimal (processes only historical sheets).
    """
    tail_data = {}
    if "Unique Ref" not in df_old.columns and "Document Number" not in df_old.columns:
        return tail_data

    # Extract non-empty TAIL_COLS using iterrows (safer for columns with spaces)
    for idx, row in df_old.iterrows():
        row_tail_data = {}
        has_data = False
        for t_col in TAIL_COLS:
            if t_col in df_old.columns:
                val = row.get(t_col)
                if val and str(val).strip() != "" and str(val).lower() != "nan":
                    row_tail_data[t_col] = val
                    has_data = True
        if not has_data:
            continue

        ur_val = clean_key(row.get("Unique Ref", ""))
        doc_val = clean_key(row.get("Document Number", ""))
        cc_val = clean_key(row.get("Company Code", ""))
        sup_val = clean_key(row.get("Supplier", ""))

        if ur_val:
            _store_tail(tail_data, f"{sheet_name}||UR||{ur_val}", row_tail_data)
            _store_tail(tail_data, f"UR||{ur_val}", row_tail_data)
        if doc_val and cc_val and sup_val:
            spec = f"{sheet_name}||DOC||{cc_val}||{sup_val}||{doc_val}"
            glob_k = f"DOC||{cc_val}||{sup_val}||{doc_val}"
            _store_tail(tail_data, spec, row_tail_data)
            _store_tail(tail_data, glob_k, row_tail_data)
            if ur_val:
                _store_tail(tail_data, f"{sheet_name}||DOCUR||{cc_val}||{sup_val}||{doc_val}||{ur_val}", row_tail_data)
                _store_tail(tail_data, f"DOCUR||{cc_val}||{sup_val}||{doc_val}||{ur_val}", row_tail_data)

    return tail_data

def capture_tail_smart(directory):
    """Load historical tail data (query type, status, comments) from previous week's Excel report.
    Scans all owner sheets and builds a lookup dict for matching by Unique Ref or Document keys.
    """
    tail_db = {}
    latest_file = find_latest_file(directory)
    if not latest_file:
        return tail_db

    log.info("    -> Lendo historico de: %s", os.path.basename(latest_file))

    try:
        xls = pd.ExcelFile(latest_file)
    except Exception:
        return tail_db

    for sheet_name in xls.sheet_names:
        if "Snapshot" in sheet_name or "MasterData" in sheet_name or "Summary" in sheet_name:
            continue
        try:
            df_old = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df_old.columns = [str(c).strip() for c in df_old.columns]
            # Process this sheet and merge its tail data into tail_db
            sheet_tail = _process_tail_sheet(df_old, sheet_name)
            for key, data in sheet_tail.items():
                _store_tail(tail_db, key, data)
        except Exception:
            pass

    xls.close()
    return tail_db

#  extracted functions (were interleaved in execution zone) 

def _bucket_suppliers(base, min_days, max_days):
    if base.empty: return pd.DataFrame(columns=["Key","Supplier","Name","Owner","Country","Docs","Value"])
    d = base.copy(); d["_Key"] = _supplier_key(d); mask = _bucket_mask(d, min_days, max_days); dd = d[mask].copy()
    if dd.empty: return pd.DataFrame(columns=["Key","Supplier","Name","Owner","Country","Docs","Value"])
    g = dd.groupby("_Key", dropna=False)
    res = _groupby_apply_no_warning(g, lambda x: pd.Series({
        "Supplier": str(x["Supplier"].iloc[0]), "Name": str(x["Name 1"].iloc[0]),
        "Owner": str(x["Owner"].iloc[0]), "Country": str(x["Country"].iloc[0]),
        "Docs": int(len(x)), "Value": float(x["_Amt"].sum())
    })).reset_index().rename(columns={"_Key": "Key"})
    return res

def _compute_owner_metrics(base, buckets, focus_label):
    """
    MODIFICADO: Agora usa LOGICA HIERARCHICAL para Suppliers
    Cada Unique Ref conta APENAS no bucket MAIS ALTO onde tem documentos
    """
    if base.empty: return pd.DataFrame(columns=["Owner"]).set_index("Owner")
    d = base.copy(); d["_Owner"] = _norm_str_series(d["Owner"] if "Owner" in d.columns else pd.Series("", index=d.index)); d["_Key"] = _supplier_key(d)
    out = {}
    for owner in sorted(d["_Owner"].dropna().unique().tolist()):
        g = d[d["_Owner"] == owner]
        row = {}
        for label, low, high in buckets:
            m = _bucket_mask(g, low, high)

            # HIERARCHICAL: Conta Unique Refs APENAS se NAO estiverem em buckets mais altos
            suppliers_in_bucket = set(g.loc[m, "_Key"].replace({"": np.nan}).dropna().unique())

            # Suppliers em buckets mais altos (days > high)
            if high is not None:
                m_higher = g["_IsDetail"] & (g["_Amt"].fillna(0) != 0) & (g["_Days"] > high)
                suppliers_in_higher = set(g.loc[m_higher, "_Key"].replace({"": np.nan}).dropna().unique())
                # Exclusivos = no bucket atual MAS NAO em buckets mais altos
                exclusive_suppliers = suppliers_in_bucket - suppliers_in_higher
                row[f"Suppliers {label}"] = int(len(exclusive_suppliers))
            else:
                # Bucket mais alto (180+) nao precisa excluir ninguem
                row[f"Suppliers {label}"] = int(len(suppliers_in_bucket))

            row[f"Docs {label}"] = int(m.sum()); row[f"Value {label}"] = float(g.loc[m, "_Amt"].sum())
        row[f"Suppliers {focus_label}"] = row.get(f"Suppliers {focus_label}", 0)
        row[f"Docs {focus_label}"] = row.get(f"Docs {focus_label}", 0)
        row[f"Value {focus_label}"] = row.get(f"Value {focus_label}", 0.0)
        out[owner] = row
    return pd.DataFrame.from_dict(out, orient="index").fillna(0)

def _find_previous_output(data_dir, prefix, exclude_path):
    search_pattern = os.path.join(data_dir, f"{prefix}*.xlsx")
    files = glob.glob(search_pattern)
    # Also search archive/ for previous week files
    archive_pattern = os.path.join(ARCHIVE_DIR, f"{prefix}*.xlsx")
    files += glob.glob(archive_pattern)
    valid_files = [f for f in files if os.path.basename(f) != os.path.basename(exclude_path) and not os.path.basename(f).startswith("~$")]
    if not valid_files: return None
    return max(valid_files, key=extract_date_from_filename)

def _read_previous_summary_cw(path, sheet_name="Summary"):
    """Read CW values from previous report's Owner Weekly Dashboard.
    Returns dict: {owner: {metric: value}} where metric names match _compute_owner_metrics columns.
    """
    from openpyxl import load_workbook
    try:
        wb_prev = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    if sheet_name not in wb_prev.sheetnames:
        wb_prev.close()
        return {}
    ws = wb_prev[sheet_name]
    all_rows = []
    for row in ws.iter_rows():
        all_rows.append([cell.value for cell in row])
    wb_prev.close()

    # Find "Owner Weekly Dashboard" header row
    dash_header_idx = None
    for i, row in enumerate(all_rows):
        cell_val = str(row[0] or "").strip()
        if "Owner Weekly Dashboard" in cell_val:
            dash_header_idx = i + 1
            break
    if dash_header_idx is None or dash_header_idx >= len(all_rows):
        return {}

    headers = [str(v or "").strip() for v in all_rows[dash_header_idx]]

    # Map CW columns: "Total Docs CW" -> col index, metric name "Total Docs"
    cw_col_map = {}
    for j, h in enumerate(headers):
        if " CW" in h:
            metric = h.replace(" CW", "").strip()
            cw_col_map[metric] = j

    # Read owner data rows
    result = {}
    for i in range(dash_header_idx + 1, len(all_rows)):
        row = all_rows[i]
        owner = str(row[0] or "").strip()
        if not owner:
            break
        result[owner] = {}
        for metric, j in cw_col_map.items():
            val = row[j] if j < len(row) else 0
            try: val = float(val or 0)
            except (ValueError, TypeError): val = 0.0
            # Map "Total Value 180+" -> "Value 180+" to match _compute_owner_metrics columns
            if metric.startswith("Total Value "):
                metric = metric.replace("Total Value ", "Value ")
            result[owner][metric] = val
    return result

def _load_previous_detail(path):
    try: xls = pd.ExcelFile(path)
    except Exception: return pd.DataFrame()
    dfs = []
    for sn in xls.sheet_names:
        if sn.lower() in {"summary", "weekly_movement", "payment issues",
                          "duplicate invoices", "invoice errors", "z&r blocks"} or sn.startswith("_"): continue
        try: dfs.append(pd.read_excel(xls, sheet_name=sn, dtype=str))
        except Exception: continue
    xls.close()
    if not dfs: return pd.DataFrame()
    prev = pd.concat(dfs, ignore_index=True)
    prev.columns = [str(c).strip() for c in prev.columns]
    prev = prev.loc[:, ~prev.columns.duplicated()].copy()
    if "Amount_Num" not in prev.columns and "Amount in doc. curr." in prev.columns:
        prev["Amount_Num"] = pd.to_numeric(prev["Amount in doc. curr."], errors="coerce")
    prev["_Amt"] = pd.to_numeric(prev.get("Amount_Num", 0), errors="coerce").fillna(0)
    snapshot_date = _extract_snapshot_date(path)
    if "Net due date" in prev.columns:
        prev["Net due date"] = parse_mixed_date_series(prev["Net due date"])
        prev["_Days"] = (pd.Timestamp(snapshot_date) - prev["Net due date"]).dt.days.fillna(-999)
    else: prev["_Days"] = -999
    # BUG 0 FIX: Override _Days using pre-computed bucket columns from Excel
    # (handles NaT Net due dates that occur when reading Excel with dtype=str)
    for _bcol, _bdays in [
        ("0-30 Days overdue", 0), ("31-60 Days overdue", 31),
        ("61-90 Days overdue", 61), ("91-120 Days Overdue", 91),
        ("121-180 Days Overdue", 121), ("180> Days Overdue", 181),
    ]:
        if _bcol in prev.columns:
            _bv = pd.to_numeric(prev[_bcol], errors='coerce').fillna(0)
            prev.loc[_bv != 0, '_Days'] = float(_bdays)
    prev = prev.copy()
    _doc = prev.get("Document Number")
    if _doc is None: prev["_IsDetail"] = True
    else: prev["_IsDetail"] = _doc.fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True).ne("")
    if "Country" in prev.columns: prev["Country"] = _norm_str_series(prev["Country"]).str.upper()
    if "Owner" in prev.columns:
        prev["Owner"] = _norm_str_series(prev["Owner"])
        prev = prev[~prev["Owner"].str.upper().eq("FUEL")].copy()
        prev = prev[~prev["Owner"].str.upper().eq("UNASSIGNED")].copy()
        prev = prev[prev["Owner"].astype(str).str.strip() != ""].copy()
    ic_mask = pd.Series(False, index=prev.index)
    for col in ["Vendor category", "Name 1", "Text", "Supplier", "Unique Ref"]:
        if col in prev.columns:
            ic_mask |= _norm_str_series(prev[col]).str.contains(_re_intercompany)
    prev = prev[~ic_mask].copy()
    return prev

def _compute_owner_totals(base):
    if base.empty: return pd.DataFrame(columns=["Total Docs", "Total Suppliers"])
    d = base.copy(); d["_Owner"] = _norm_str_series(d["Owner"] if "Owner" in d.columns else pd.Series("", index=d.index)); d["_Key"] = _supplier_key(d)
    m = d["_IsDetail"] & (d["_Amt"].fillna(0) != 0)
    if not m.any(): return pd.DataFrame(columns=["Total Docs", "Total Suppliers"])
    g = d.loc[m].groupby("_Owner", dropna=False)
    return g.apply(lambda x: pd.Series({"Total Docs": int(len(x)), "Total Suppliers": int(x["_Key"].nunique())}), include_groups=False)

def _kpi_focus(m):
    if m.empty: return {"sup": 0.0, "docs": 0.0, "val": 0.0}
    return {
        "sup": float(m[f"Suppliers {FOCUS}"].sum()) if f"Suppliers {FOCUS}" in m.columns else 0.0,
        "docs": float(m[f"Docs {FOCUS}"].sum()) if f"Docs {FOCUS}" in m.columns else 0.0,
        "val": float(m[f"Value {FOCUS}"].sum()) if f"Value {FOCUS}" in m.columns else 0.0,
    }

#  main execution 

def main():
    INPUT_CSV = get_latest_master_auto(ROOT_DIR)

    log.info(">>> INICIANDO PYTHON - REPORT KEY ...")
    if not os.path.exists(DATA_DIR): os.makedirs(DATA_DIR)

    if not INPUT_CSV:
        log.error("ERRO FATAL: CSV nao encontrado.")
        sys.exit()
    log.info("    -> Carregando CSV: %s...", os.path.basename(INPUT_CSV))
    try: df = pd.read_csv(INPUT_CSV, dtype=str, encoding='utf-8', on_bad_lines='skip')
    except (UnicodeDecodeError, pd.errors.ParserError): df = pd.read_csv(INPUT_CSV, dtype=str, encoding='latin1', on_bad_lines='skip')

    df.columns = [str(c).strip() for c in df.columns]
    if "Sheet" in df.columns: df = df[df["Sheet"].astype(str).str.casefold() == "key"].copy()
    df["Amount_Num"] = parse_amount_series(df.get("Amount in doc. curr.", pd.Series([""]*len(df))))
    for c in ["Net due date", "Document Date", "Posting Date", "Action Date"]:
        if c in df.columns: df[c] = parse_mixed_date_series(df[c])
    if "Company Code" in df.columns: df["System"] = df["Company Code"].apply(system_from_code)
    if "Unique Ref" not in df.columns:
        sup_col = "Supplier" if "Supplier" in df.columns else "Vendor Option"
        df["Unique Ref"] = df["Company Code"].fillna("") + " " + df[sup_col].fillna("")
    if "Payment Issues" not in df.columns:
        if "Payment issue" in df.columns: df["Payment Issues"] = df["Payment issue"]
        else: df["Payment Issues"] = ""

    tail_history = capture_tail_smart(DATA_DIR)

    log.info("    -> Calculando dias de atraso e KPIs...")
    # Snapshot date from MasterData filename (MasterData_WWYY.csv)
    # WW = ISO week, YY = year  snapshot = Monday of that ISO week
    _snap_match = re.search(r'MasterData_(\d{2})(\d{2})', os.path.basename(INPUT_CSV))
    if _snap_match:
        _iso_wk = int(_snap_match.group(1))
        _iso_yr = 2000 + int(_snap_match.group(2))
        today_ts = pd.Timestamp(date.fromisocalendar(_iso_yr, _iso_wk, 1))
    else:
        today_ts = pd.Timestamp.now()
    log.info("    -> Snapshot reference date: %s", today_ts.strftime('%Y-%m-%d'))
    df["_Days"] = (today_ts - df["Net due date"]).dt.days.fillna(-999)
    df["_Rank"] = df["_Days"].apply(get_bucket_rank)
    df["_Amt"] = df["Amount_Num"].fillna(0)
    _doc = df.get("Document Number")
    if _doc is None: df["_IsDetail"] = True
    else: df["_IsDetail"] = _doc.fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True).ne("")

    if "Country" in df.columns: df["Country"] = _norm_str_series(df["Country"]).str.upper()
    if "Owner" in df.columns: df["Owner"] = _norm_str_series(df["Owner"])

    # --- LOGICA DE EXCLUSAO (FUEL/UNASSIGNED) ---
    # Aqui garantimos que Fuel e Unassigned NAO aparecem nem no Summary nem nas abas (removidos do df global)
    if "Owner" in df.columns:
        df = df[~df["Owner"].str.upper().eq("FUEL")].copy()
        df = df[~df["Owner"].str.upper().eq("UNASSIGNED")].copy()

    ic_mask = pd.Series(False, index=df.index)
    for col in ["Vendor category", "Name 1", "Text", "Supplier", "Unique Ref"]:
        if col in df.columns:
            ic_mask |= _norm_str_series(df[col]).str.contains(_re_intercompany)
    df = df[~ic_mask].copy()

    df["_121_180_Amt_NET"] = np.where(df["_IsDetail"] & (df["_Days"] > 120) & (df["_Days"] <= 180), df["_Amt"], 0)
    df["_Over180_Amt_NET"] = np.where(df["_IsDetail"] & (df["_Days"] > 180), df["_Amt"], 0)

    # --- MESTRE PYTHON: CRIANDO VIEWS PARA O SUMMARY ---
    # O df original (que contem RENT) continua intacto para a geracao das abas.
    # Aqui criamos o df_kpi apenas para as estatisticas do Summary.
    _mask_excl_cw = df["Owner"].astype(str).str.upper().isin([o.upper() for o in OWNERS_HIDDEN_FROM_SUMMARY])
    df_kpi = df[~_mask_excl_cw].copy()

    log.info("    -> Gerando Excel final...")
    wb = Workbook()
    wb.remove(wb.active)

    log.info("      ...Criando aba Summary")
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.merge_cells("A1:K1")
    ws_sum["A1"] = "Comparative Weekly Dashboard (Key Report)"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A1"].alignment = Alignment(horizontal="center")

    prev_path = _find_previous_output(DATA_DIR, prefix="Key ", exclude_path=OUTPUT_PATH)
    if prev_path:
        log.info("    -> Comparando com arquivo anterior: %s", os.path.basename(prev_path))
    df_lw = _load_previous_detail(prev_path) if prev_path else pd.DataFrame()

    # --- MESTRE PYTHON: FILTRAR LW TAMBEM ---
    # Precisamos remover o RENT do historico para nao gerar Deltas falsos
    if not df_lw.empty and "Owner" in df_lw.columns:
        _mask_excl_lw = df_lw["Owner"].astype(str).str.upper().isin([o.upper() for o in OWNERS_HIDDEN_FROM_SUMMARY])
        df_lw_kpi = df_lw[~_mask_excl_lw].copy()
    else:
        df_lw_kpi = df_lw.copy()

    # --- CALCULOS USANDO AS VERSOES _KPI (SEM RENT) ---
    owner_tw = _compute_owner_metrics(df_kpi, BUCKETS, FOCUS)

    tot_tw = _compute_owner_totals(df_kpi)
    owner_tw = owner_tw.join(tot_tw, how="left").fillna(0)

    # --- LW: Computado a partir dos dados brutos do arquivo anterior ---
    # (BUG 0 fix em _load_previous_detail garante bucket correto mesmo para datas NaT)
    owner_lw = _compute_owner_metrics(df_lw_kpi, BUCKETS, FOCUS) if not df_lw_kpi.empty else owner_tw.copy()*0
    tot_lw = _compute_owner_totals(df_lw_kpi) if not df_lw_kpi.empty else pd.DataFrame(columns=["Total Docs", "Total Suppliers"])
    owner_lw = owner_lw.join(tot_lw, how="left").fillna(0)

    # Filter summary to Key team members from Owner_map.csv.
    _KEY_TEAM = {o.casefold() for o in load_key_team_owners(KEY_OWNER_MAP_SOURCE)}
    _all_owners = set(owner_tw.index).union(set(owner_lw.index))
    owners_all = sorted(o for o in _all_owners if str(o).strip().casefold() in _KEY_TEAM)

    # KPI Table
    k_lw = _kpi_focus(owner_lw); k_tw = _kpi_focus(owner_tw)

    ws_sum["A3"] = "KPI Overview (180+)"; ws_sum["A3"].font = Font(bold=True)
    ws_sum.append(["Metric", "LW", "CW", "Delta"])
    for c in range(1, 5):
        cell=ws_sum.cell(4, c); cell.fill=fill_header_custom; cell.font=font_header_custom; cell.alignment=Alignment(horizontal="center")

    def _write_kpi(r, m, l, t):
        ws_sum.cell(r, 1, m); ws_sum.cell(r, 2, l); ws_sum.cell(r, 3, t); ws_sum.cell(r, 4, t - l)
        for cx in [2,3,4]: ws_sum.cell(r, cx).number_format = "#,##0.00"

    _write_kpi(5, "Suppliers 180+", k_lw["sup"], k_tw["sup"])
    _write_kpi(6, "Docs 180+", k_lw["docs"], k_tw["docs"])
    _write_kpi(7, "Value 180+", k_lw["val"], k_tw["val"])
    ws_sum["F4"]="Resolution Rate (Docs)"; ws_sum["F4"].fill=fill_header_custom; ws_sum["F4"].font=font_header_custom
    ws_sum["F5"]="" if k_lw["docs"]==0 else (k_lw["docs"]-k_tw["docs"])/k_lw["docs"]; ws_sum["F5"].number_format="0.00%"

    # Top Drivers
    td_start = ws_sum.max_row + 2
    ws_sum[f"A{td_start}"] = "Top Drivers (Top 10 Vendors by 180+ Docs CW)"; ws_sum[f"A{td_start}"].font = Font(bold=True)
    td_headers = ["Vendor", "Name", "Owner", "Country", "Docs LW", "Docs CW", "Delta Docs"]
    for j, h in enumerate(td_headers, start=1):
        c=ws_sum.cell(td_start+1, j, h); c.fill=fill_header_custom; c.font=font_header_custom

    # Usando df_kpi / df_lw_kpi para Top Drivers
    tw_over = _bucket_suppliers(df_kpi, 180, None)
    lw_over = _bucket_suppliers(df_lw_kpi, 180, None) if not df_lw_kpi.empty else pd.DataFrame(columns=tw_over.columns)
    td = tw_over.merge(lw_over[["Key","Docs"]], on="Key", how="left", suffixes=("_TW","_LW")).fillna(0)
    td["Delta Docs"] = td["Docs_TW"].astype(float) - td["Docs_LW"].astype(float)
    td = td.sort_values(by="Docs_TW", ascending=False).head(10)
    rr_td = td_start + 2
    for x in td.itertuples(index=False):
        ws_sum.cell(rr_td, 1, x.Supplier); ws_sum.cell(rr_td, 2, x.Name); ws_sum.cell(rr_td, 3, x.Owner)
        ws_sum.cell(rr_td, 4, x.Country); ws_sum.cell(rr_td, 5, int(x.Docs_LW)); ws_sum.cell(rr_td, 6, int(x.Docs_TW)); ws_sum.cell(rr_td, 7, int(x[7]))  # Delta Docs
        rr_td += 1

    # Top 10 Owners
    to_start = rr_td + 1
    ws_sum[f"A{to_start}"] = "Top 10 Owners by Delta Docs Increase (180+)"; ws_sum[f"A{to_start}"].font = Font(bold=True)
    for j, h in enumerate(["Owner", "Docs LW", "Docs CW", "Delta"], start=1):
        c = ws_sum.cell(to_start+1, j, h); c.fill = fill_header_custom; c.font = font_header_custom
    tmp = []
    for owner in owners_all:
        d_tw = float(owner_tw.loc[owner, f"Docs {FOCUS}"]) if owner in owner_tw.index else 0.0
        d_lw = float(owner_lw.loc[owner, f"Docs {FOCUS}"]) if owner in owner_lw.index else 0.0
        tmp.append((owner, d_tw, d_lw, d_tw - d_lw))
    tmp.sort(key=lambda x: x[3], reverse=True)
    rr_o = to_start + 2
    for owner, d_tw, d_lw, dd in tmp[:10]:
        ws_sum.cell(rr_o, 1, owner); ws_sum.cell(rr_o, 2, d_lw); ws_sum.cell(rr_o, 3, d_tw); ws_sum.cell(rr_o, 4, dd)
        rr_o += 1

    # Dashboard
    owner_start = rr_o + 2
    ws_sum[f"A{owner_start}"] = "Owner Weekly Dashboard"; ws_sum[f"A{owner_start}"].font = Font(bold=True)
    headers = ["Owner", "Total Docs LW", "Total Docs CW", "Delta Total Docs", "Total Suppliers LW", "Total Suppliers CW", "Delta Total Suppliers", "Docs 180+ LW", "Docs 180+ CW", "Delta Docs 180+"]
    for label, _, _ in BUCKETS:
        if label == "180+": continue
        headers += [f"Suppliers {label} LW", f"Suppliers {label} CW", f"Delta Suppliers {label}", f"Docs {label} LW", f"Docs {label} CW", f"Delta Docs {label}"]
    headers += ["Total Value 180+ LW", "Total Value 180+ CW"]

    for j, h in enumerate(headers, start=1):
        c = ws_sum.cell(owner_start+1, j, h); c.fill = fill_header_custom; c.font = font_header_custom; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_i = owner_start + 2
    for owner in owners_all:
        # FILTRO DASHBOARD: Remover Unassigned e Fuel.
        # Rent/Aluguel can remain as tabs but are hidden from Summary/KPIs upstream.
        if str(owner).upper() in ["UNASSIGNED", "FUEL", ""]: continue

        ws_sum.cell(row_i, 1, owner)
        td_tw = float(owner_tw.loc[owner, "Total Docs"]) if owner in owner_tw.index else 0.0
        td_lw = float(owner_lw.loc[owner, "Total Docs"]) if owner in owner_lw.index else 0.0
        ts_tw = float(owner_tw.loc[owner, "Total Suppliers"]) if owner in owner_tw.index else 0.0
        ts_lw = float(owner_lw.loc[owner, "Total Suppliers"]) if owner in owner_lw.index else 0.0
        d180_tw = float(owner_tw.loc[owner, f"Docs {FOCUS}"]) if owner in owner_tw.index else 0.0
        d180_lw = float(owner_lw.loc[owner, f"Docs {FOCUS}"]) if owner in owner_lw.index else 0.0

        ws_sum.cell(row_i, 2, int(td_lw)); ws_sum.cell(row_i, 3, int(td_tw)); ws_sum.cell(row_i, 4, int(td_tw-td_lw))
        ws_sum.cell(row_i, 5, int(ts_lw)); ws_sum.cell(row_i, 6, int(ts_tw)); ws_sum.cell(row_i, 7, int(ts_tw-ts_lw))
        ws_sum.cell(row_i, 8, int(d180_lw)); ws_sum.cell(row_i, 9, int(d180_tw)); ws_sum.cell(row_i, 10, int(d180_tw-d180_lw))

        col = 11
        for label, _, _ in BUCKETS:
            if label == "180+": continue
            s_tw = float(owner_tw.loc[owner, f"Suppliers {label}"]) if owner in owner_tw.index else 0.0
            s_lw = float(owner_lw.loc[owner, f"Suppliers {label}"]) if owner in owner_lw.index else 0.0
            d_tw = float(owner_tw.loc[owner, f"Docs {label}"]) if owner in owner_tw.index else 0.0
            d_lw = float(owner_lw.loc[owner, f"Docs {label}"]) if owner in owner_lw.index else 0.0
            ws_sum.cell(row_i, col, int(s_lw)); ws_sum.cell(row_i, col+1, int(s_tw)); ws_sum.cell(row_i, col+2, int(s_tw-s_lw))
            ws_sum.cell(row_i, col+3, int(d_lw)); ws_sum.cell(row_i, col+4, int(d_tw)); ws_sum.cell(row_i, col+5, int(d_tw-d_lw))
            col += 6
        v_tw = float(owner_tw.loc[owner, f"Value {FOCUS}"]) if owner in owner_tw.index else 0.0
        v_lw = float(owner_lw.loc[owner, f"Value {FOCUS}"]) if owner in owner_lw.index else 0.0
        ws_sum.cell(row_i, col, v_lw).number_format = "#,##0.00"; ws_sum.cell(row_i, col+1, v_tw).number_format = "#,##0.00"
        row_i += 1

    last_row = row_i - 1
    for c in range(1, len(headers)+1):
        h = str(ws_sum.cell(owner_start+1, c).value or "")
        if h.startswith("Delta "):
            col_l = get_column_letter(c)
            rng = f"{col_l}{owner_start+2}:{col_l}{last_row}"
            ws_sum.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], fill=cf_red_fill))
            ws_sum.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], fill=fill_green_light))

    mix_start = last_row + 2
    ws_sum[f"A{mix_start}"] = "Aging Mix % (Volume Share - Docs)"; ws_sum[f"A{mix_start}"].font=Font(bold=True)
    mix_h = ["Bucket","Docs LW","Share LW","Docs CW","Share CW","Delta pp"]
    for j, h in enumerate(mix_h, start=1):
        c=ws_sum.cell(mix_start+1, j, h); c.fill=fill_header_custom; c.font=font_header_custom; c.alignment=Alignment(horizontal="center")

    # Aging Mix
    tt = int(df_kpi["_IsDetail"].sum())
    tl = int(df_lw_kpi["_IsDetail"].sum()) if not df_lw_kpi.empty else 0
    mix_row = mix_start + 2
    for label, low, high in BUCKETS:
        mt = _bucket_mask(df_kpi, low, high)
        dt = int(mt.sum())
        ml = _bucket_mask(df_lw_kpi, low, high) if not df_lw_kpi.empty else pd.Series(False, index=[0])
        dl = int(ml.sum())
        st = dt/tt if tt else 0; sl = dl/tl if tl else 0
        ws_sum.cell(mix_row, 1, label); ws_sum.cell(mix_row, 2, dl); ws_sum.cell(mix_row, 3, sl).number_format="0.00%"
        ws_sum.cell(mix_row, 4, dt); ws_sum.cell(mix_row, 5, st).number_format="0.00%"; ws_sum.cell(mix_row, 6, st-sl).number_format="0.00%"
        mix_row += 1

    # === QUERY TYPE BREAKDOWN  formulas COUNTIF dinamicas por aba de owner ===
    _qt_col_letter = get_column_letter(FINAL_HEADERS.index("Query type") + 1)
    _doc_col_letter = get_column_letter(FINAL_HEADERS.index("Document Number") + 1)
    # Use same cleaning as sheet creation (line ~812) to match actual tab names
    _qt_tabs = [str(o)[:30].replace("/", " ").replace("\\", " ").replace("?", "").strip()
                for o in df["Owner"].dropna().unique()
                if str(o)[:30].replace("/", " ").replace("\\", " ").replace("?", "").strip()]

    def _qt_formula_key(qt_val):
        # FIX: COUNTIFS para contar apenas Detail rows (Document Number <> "")
        parts = [f"COUNTIFS('{t}'!{_qt_col_letter}:{_qt_col_letter},{chr(34)}{qt_val}{chr(34)},'{t}'!{_doc_col_letter}:{_doc_col_letter},{chr(34)}<>{chr(34)})"
                 for t in _qt_tabs]
        return "=" + "+".join(parts) if parts else "=0"

    qt_start = mix_row + 1
    ws_sum.cell(qt_start, 1, "Query Type Breakdown (Docs - Row Level)").font = Font(bold=True)
    for j, h in enumerate(["Query Type", "Count"], 1):
        c = ws_sum.cell(qt_start + 1, j, h)
        c.fill = fill_header_custom; c.font = font_header_custom; c.alignment = Alignment(horizontal="center")

    qt_row = qt_start + 2
    _qt_data_start = qt_row
    for qt in OPCOES_QUERY:  # lista fixa, sem (Blank)
        ws_sum.cell(qt_row, 1, qt)
        _c = ws_sum.cell(qt_row, 2, _qt_formula_key(qt))
        _c.number_format = "#,##0"
        qt_row += 1

    _total_range = f"B{_qt_data_start}:B{qt_row - 1}"
    _tc = ws_sum.cell(qt_row, 1, "TOTAL"); _tc.font = Font(bold=True)
    _tc2 = ws_sum.cell(qt_row, 2, f"=SUM({_total_range})"); _tc2.font = Font(bold=True); _tc2.number_format = "#,##0"
    qt_row += 1

    # SUMMARY CHARTS  2 graficos simples lado a lado
    if last_row >= owner_start + 2:
        _cats = Reference(ws_sum, min_col=1, min_row=owner_start + 2, max_row=last_row)

        # Chart 1: Docs 180+ LW vs CW (KPI principal)  vermelho/laranja
        ch1 = BarChart()
        ch1.type = "col"; ch1.grouping = "clustered"; ch1.style = 10
        ch1.title = f"Docs {FOCUS}  LW vs CW"; ch1.y_axis.title = "Docs"
        ch1.x_axis.title = "Owner"; ch1.legend.position = "b"
        ch1.width = 20; ch1.height = 14; ch1.gapWidth = 100
        _d1 = Reference(ws_sum, min_col=8, max_col=9, min_row=owner_start + 1, max_row=last_row)
        ch1.add_data(_d1, titles_from_data=True); ch1.set_categories(_cats)
        for _i, _color in enumerate(["C00000", "ED7D31"]):
            if _i < len(ch1.series):
                ch1.series[_i].graphicalProperties.solidFill = _color
                ch1.series[_i].dLbls = DataLabelList()
                ch1.series[_i].dLbls.showVal = True; ch1.series[_i].dLbls.numFmt = '#,##0'
        ws_sum.add_chart(ch1, "L4")

        # Chart 2: Total Docs LW vs CW  azul escuro/claro
        ch2 = BarChart()
        ch2.type = "col"; ch2.grouping = "clustered"; ch2.style = 10
        ch2.title = "Total Docs  LW vs CW"; ch2.y_axis.title = "Docs"
        ch2.x_axis.title = "Owner"; ch2.legend.position = "b"
        ch2.width = 20; ch2.height = 14; ch2.gapWidth = 100
        _d2 = Reference(ws_sum, min_col=2, max_col=3, min_row=owner_start + 1, max_row=last_row)
        ch2.add_data(_d2, titles_from_data=True); ch2.set_categories(_cats)
        for _i, _color in enumerate(["1F4E79", "5B9BD5"]):
            if _i < len(ch2.series):
                ch2.series[_i].graphicalProperties.solidFill = _color
                ch2.series[_i].dLbls = DataLabelList()
                ch2.series[_i].dLbls.showVal = True; ch2.series[_i].dLbls.numFmt = '#,##0'
        ws_sum.add_chart(ch2, "V4")

    ws_sum.column_dimensions["A"].width = 22
    for col in range(2, 11): ws_sum.column_dimensions[get_column_letter(col)].width = 15
    ws_sum.freeze_panes = "A3"

    # WEEKLY MOVEMENT
    ws_mov = wb.create_sheet(title="Weekly_Movement")
    ws_mov.append(["Key Report - Weekly Movement (180+)"]); ws_mov["A1"].font=Font(bold=True, size=12)

    def _wsec(title, df_sec, start):
        ws_mov.cell(start, 1, title).font=Font(bold=True)
        headers = ["Vendor", "Name", "Owner", "Country", "Docs LW", "Docs CW", "Delta Docs", "Value LW", "Value CW", "Delta Value"]
        for j, h in enumerate(headers, start=1):
            c=ws_mov.cell(start+1, j, h); c.fill=fill_header_custom; c.font=font_header_custom
        r0 = start+2
        if df_sec.empty:
            ws_mov.cell(r0, 1, "(none)")
            return r0 + 2
        for x in df_sec.itertuples(index=False):
            ws_mov.cell(r0, 1, x.Supplier)
            ws_mov.cell(r0, 2, x.Name)
            ws_mov.cell(r0, 3, x.Owner)
            ws_mov.cell(r0, 4, x.Country)
            docs = int(x.Docs); val = float(x.Value)
            if "cleared" in title.lower():
                ws_mov.cell(r0, 5, docs); ws_mov.cell(r0, 6, 0); ws_mov.cell(r0, 7, -docs)
                ws_mov.cell(r0, 8, val).number_format="#,##0.00"; ws_mov.cell(r0, 9, 0).number_format="#,##0.00"; ws_mov.cell(r0, 10, -val).number_format="#,##0.00"
            else:
                ws_mov.cell(r0, 5, 0); ws_mov.cell(r0, 6, docs); ws_mov.cell(r0, 7, docs)
                ws_mov.cell(r0, 8, 0).number_format="#,##0.00"; ws_mov.cell(r0, 9, val).number_format="#,##0.00"; ws_mov.cell(r0, 10, val).number_format="#,##0.00"
            r0 += 1
        return r0 + 2

    cleared = lw_over[lw_over["Key"].isin(set(lw_over["Key"]) - set(tw_over["Key"]))].copy()
    new = tw_over[tw_over["Key"].isin(set(tw_over["Key"]) - set(lw_over["Key"]))].copy()
    row_ptr = 3
    row_ptr = _wsec("Cleared from 180+ (LW -> CW)", cleared, row_ptr)
    row_ptr = _wsec("New 180+ Offenders (LW -> CW)", new, row_ptr)

    # --- PAYMENT ISSUES TAB ---
    ws_pay = wb.create_sheet(title="Payment Issues")
    ws_pay["A1"] = "Payment Issues Analysis (LW vs CW)"; ws_pay["A1"].font = Font(bold=True, size=12)

    # Usando df_kpi / df_lw_kpi para Payment Issues
    pi_tw = _agg_pi(df_kpi); pi_lw = _agg_pi(df_lw_kpi) if not df_lw_kpi.empty else pd.DataFrame(columns=pi_tw.columns)
    merged_pi = pd.merge(pi_tw, pi_lw, on=["Issue","Owner","Company Code","Vendor","Name"], how="outer", suffixes=("_CW","_LW")).fillna(0)
    merged_pi["Qty_CW"] = merged_pi["Qty_CW"].astype(int); merged_pi["Qty_LW"] = merged_pi["Qty_LW"].astype(int)
    merged_pi["Delta"] = merged_pi["Qty_CW"] - merged_pi["Qty_LW"]
    # Vectorized growth calculation
    lw = merged_pi["Qty_LW"]
    cw = merged_pi["Qty_CW"]
    merged_pi["Growth"] = np.where(lw != 0, (cw - lw) / lw, np.where(cw > 0, 1.0, 0.0))
    merged_pi = merged_pi.sort_values(by="Qty_CW", ascending=False)

    h_pi = ["Type of payment issue", "Owner", "Company Code", "Vendor Number", "Vendor Name", "Qty Issue LW", "Qty Issue CW", "Delta", "Growth %"]
    for c, h in enumerate(h_pi, 1):
        cell = ws_pay.cell(3, c, h); cell.fill = fill_header_custom; cell.font = font_header_custom

    curr_r = 4
    for row in merged_pi.itertuples(index=False):
        ws_pay.cell(curr_r, 1, row.Issue)
        ws_pay.cell(curr_r, 2, row.Owner)
        ws_pay.cell(curr_r, 3, row[2])  # "Company Code" (space in name)
        ws_pay.cell(curr_r, 4, row.Vendor)
        ws_pay.cell(curr_r, 5, row.Name)
        ws_pay.cell(curr_r, 6, row.Qty_LW)
        ws_pay.cell(curr_r, 7, row.Qty_CW)
        ws_pay.cell(curr_r, 8, row.Delta)
        ws_pay.cell(curr_r, 9, row.Growth).number_format = '0.0%'
        curr_r += 1

    if curr_r > 4:
        delta_range = f"H4:H{curr_r-1}"
        ws_pay.conditional_formatting.add(delta_range, CellIsRule(operator='greaterThan', formula=['0'], fill=cf_red_fill))
        ws_pay.conditional_formatting.add(delta_range, CellIsRule(operator='lessThan', formula=['0'], fill=fill_green_light))

        acol = 13
        ws_pay.cell(3, acol, "Audit Summary").font = Font(bold=True)
        ws_pay.cell(4, acol, "Issues Increased"); ws_pay.cell(4, acol+1, len(merged_pi[merged_pi["Delta"] > 0]))
        ws_pay.cell(5, acol, "Issues Decreased"); ws_pay.cell(5, acol+1, len(merged_pi[merged_pi["Delta"] < 0]))
        ws_pay.cell(6, acol, "Net Volume Change"); ws_pay.cell(6, acol+1, merged_pi["Delta"].sum())

        owner_agg = merged_pi.groupby("Owner")[["Qty_LW", "Qty_CW"]].sum().reset_index().sort_values(by="Qty_CW", ascending=False)
        owner_agg["Owner"] = owner_agg["Owner"].fillna("").astype(str).str.strip()
        owner_agg = owner_agg[~owner_agg["Owner"].str.upper().isin(["UNASSIGNED","FUEL",""])]
        if not owner_agg.empty:
            ch_row = 10
            ws_pay.cell(ch_row, acol, "Owner"); ws_pay.cell(ch_row, acol+1, "LW"); ws_pay.cell(ch_row, acol+2, "CW")
            for r in owner_agg.itertuples(index=False):
                ch_row += 1
                ws_pay.cell(ch_row, acol, r.Owner); ws_pay.cell(ch_row, acol+1, r.Qty_LW); ws_pay.cell(ch_row, acol+2, r.Qty_CW)

            bc = BarChart()
            bc.title = "Payment Issues by Owner (LW vs CW)"
            bc.style = 10
            bc.y_axis.title = "Volume"
            bc.width = 22
            bc.height = 13
            bc.gapWidth = 120
            bc.overlap = -10
            cats = Reference(ws_pay, min_col=acol, min_row=11, max_row=ch_row)
            data = Reference(ws_pay, min_col=acol+1, max_col=acol+2, min_row=10, max_row=ch_row)
            bc.add_data(data, titles_from_data=True); bc.set_categories(cats)
            bc.legend.position = "b"
            # Cores + data labels: cinza (LW) e azul corporativo (CW)
            _pi_colors = ["A6A6A6", "1F4E79"]
            for i, color in enumerate(_pi_colors):
                if i < len(bc.series):
                    s = bc.series[i]
                    s.graphicalProperties.solidFill = color
                    s.dLbls = DataLabelList()
                    s.dLbls.showVal = True
                    s.dLbls.showCatName = False
                    s.dLbls.showSerName = False
                    s.dLbls.numFmt = '#,##0'
            ws_pay.add_chart(bc, "P3")

    ws_pay.column_dimensions["A"].width = 50
    ws_pay.column_dimensions["C"].width = 12
    ws_pay.column_dimensions["E"].width = 30

    # --- LOOP PARA CRIAR AS ABAS INDIVIDUAIS ---
    # ATENCAO: Aqui usamos "df" original (que CONTEM RENT), e nao "df_kpi"
    for owner in df["Owner"].dropna().unique():
        clean_name = str(owner)[:30].replace("/", " ").replace("\\", " ").replace("?", "").strip()
        if not clean_name:
            continue
        # FUEL e UNASSIGNED ja foram removidos do df global, entao nao aparecerao.
        # RENT esta no df global, entao VAI aparecer.
        if clean_name.upper() == "FUEL":
            continue
        log.info("      Processando Owner: %s", clean_name)
        ws = wb.create_sheet(title=clean_name)
        ws.sheet_view.showOutlineSymbols = True
        ws.sheet_properties.outlinePr.summaryBelow = False

        df_own = df[df["Owner"] == owner].copy()
        grouped = df_own.groupby("Unique Ref")

        ws.append(FINAL_HEADERS)
        for cell in ws[1]:
            col_name = str(cell.value).strip()
            cell.border = ALL_BORDER
            if col_name in COLS_RED: cell.fill = fill_red_alert; cell.font = font_white_bold
            elif col_name in COLS_TOTAL: cell.fill = fill_blue_total; cell.font = font_black_bold
            elif col_name in COLS_GREEN: cell.fill = fill_green_light; cell.font = font_black_bold
            elif col_name in FINAL_HEADERS: cell.fill = fill_blue_light; cell.font = font_black_bold
            else: cell.fill = fill_header_custom; cell.font = font_header_custom

        row_idx = 2
        for unique_ref, group in grouped:
            group = group.sort_values(by=["_Rank", "Net due date"], ascending=[False, True])
            total_vol = int(group["_IsDetail"].sum())
            total_val = group["Amount_Num"].sum()
            b_vals = {k: 0.0 for k in ["0-30", "31-60", "61-90", "91-120", "121-180", "180>"]}
            detail_rows = []
            first_row = group.iloc[0]
            has_open = False

            raw_issues = group["Payment Issues"].dropna().astype(str).str.strip()
            unique_issues = set(i for i in raw_issues if i)
            summary_issue_raw = list(unique_issues)[0] if len(unique_issues) == 1 else ""
            summary_payment_issue = clean_issue_name(summary_issue_raw)

            # Convert to dict records for efficient iteration (faster than iterrows)
            for row in group.to_dict('records'):
                amt = row.get("Amount_Num", 0) if pd.notna(row.get("Amount_Num")) else 0
                days = row.get("_Days", -999)
                d_b = {k: None for k in b_vals.keys()}
                if days >= 0:
                    if days <= 30: b_vals["0-30"]+=amt; d_b["0-30"]=amt
                    elif days <= 60: b_vals["31-60"]+=amt; d_b["31-60"]=amt
                    elif days <= 90: b_vals["61-90"]+=amt; d_b["61-90"]=amt
                    elif days <= 120: b_vals["91-120"]+=amt; d_b["91-120"]=amt
                    elif days <= 180: b_vals["121-180"]+=amt; d_b["121-180"]=amt
                    else: b_vals["180>"]+=amt; d_b["180>"]=amt

                if str(row.get("Document Type", "")).upper().strip() in ["KS", "SA", "DZ", "KZ", "ZP", "ZR", "ZE", "AB", "K1", "K5", "SE"]: has_open = True

                d_map = {h: "" for h in FINAL_HEADERS}
                for col in FINAL_HEADERS:
                    if col in df.columns: d_map[col] = sanitize_value(row[col])

                d_map["0-30 Days overdue"] = d_b["0-30"]
                d_map["31-60 Days overdue"] = d_b["31-60"]
                d_map["61-90 Days overdue"] = d_b["61-90"]
                d_map["91-120 Days Overdue"] = d_b["91-120"]
                d_map["121-180 Days Overdue"] = d_b["121-180"]
                d_map["180> Days Overdue"] = d_b["180>"]
                d_map["Amount in doc. curr."] = amt
                if d_map["Payment Issues"]:
                    d_map["Payment Issues"] = clean_issue_name(d_map["Payment Issues"])

                cc_k = clean_key(row.get("Company Code"))
                sup_k = clean_key(row.get("Supplier"))
                doc_k = clean_key(row.get("Document Number"))
                ur_k = clean_key(row.get("Unique Ref", ""))

                doc_key_specific = f"{clean_name}||DOC||{cc_k}||{sup_k}||{doc_k}"
                doc_key_global = f"DOC||{cc_k}||{sup_k}||{doc_k}"
                dup_specific = f"{clean_name}||DUPDOC||{cc_k}||{sup_k}||{doc_k}"
                dup_global = f"DUPDOC||{cc_k}||{sup_k}||{doc_k}"
                is_dup_doc = dup_specific in tail_history or dup_global in tail_history

                tail_row = None
                if is_dup_doc and ur_k:
                    docur_specific = f"{clean_name}||DOCUR||{cc_k}||{sup_k}||{doc_k}||{ur_k}"
                    docur_global = f"DOCUR||{cc_k}||{sup_k}||{doc_k}||{ur_k}"
                    tail_row = tail_history.get(docur_specific) or tail_history.get(docur_global)
                if tail_row is None:
                    tail_row = tail_history.get(doc_key_specific) or tail_history.get(doc_key_global)
                if tail_row:
                    for t_c in TAIL_COLS:
                        if t_c in tail_row:
                            val = tail_row[t_c]
                            if t_c == "Action Date":
                                val = recover_date_from_tail(val)
                            d_map[t_c] = val

                detail_rows.append([d_map[h] for h in FINAL_HEADERS])

            s_map = {h: "" for h in FINAL_HEADERS}
            for c in ["Country", "Vendor category", "Sheet", "Owner", "Company Code", "Document currency", "Supplier", "Name 1", "Unique Ref", "System"]:
                if c in first_row.index: s_map[c] = sanitize_value(first_row[c])
            s_map["0-30 Days overdue"] = b_vals["0-30"]
            s_map["31-60 Days overdue"] = b_vals["31-60"]
            s_map["61-90 Days overdue"] = b_vals["61-90"]
            s_map["91-120 Days Overdue"] = b_vals["91-120"]
            s_map["121-180 Days Overdue"] = b_vals["121-180"]
            s_map["180> Days Overdue"] = b_vals["180>"]
            s_map["TOTAL VOL"] = total_vol
            s_map["TOTAL VALUE"] = total_val
            s_map["Open Payment"] = "Yes" if has_open else ""
            s_map["Payment Issues"] = summary_payment_issue

            ur_val = clean_key(s_map.get("Unique Ref", ""))
            ur_key_specific = f"{clean_name}||UR||{ur_val}"
            ur_key_global = f"UR||{ur_val}"
            tail_row = tail_history.get(ur_key_specific) or tail_history.get(ur_key_global)
            if tail_row:
                for t_c in TAIL_COLS:
                    if t_c in tail_row:
                        val = tail_row[t_c]
                        if t_c == "Action Date": val = recover_date_from_tail(val)
                        s_map[t_c] = val

            ws.append([s_map[h] for h in FINAL_HEADERS])
            for cell in ws[row_idx]:
                cell.fill = fill_summary_gray; cell.font = font_black_bold
                if isinstance(cell.value, (datetime, pd.Timestamp)): cell.number_format = 'dd/mm/yyyy'

            grp_start = row_idx + 1
            row_idx += 1
            for d_row in detail_rows:
                ws.append(d_row)
                for c_name in ["Net due date", "Document Date", "Posting Date", "Action Date", "Action Data"]:
                    if c_name in FINAL_HEADERS:
                        c_idx = FINAL_HEADERS.index(c_name) + 1
                        cell = ws.cell(row=row_idx, column=c_idx)
                        if isinstance(cell.value, (datetime, pd.Timestamp)):
                            cell.number_format = 'dd/mm/yyyy'
                for c_name in ["Document Number", "Reference", "Supplier", "Company Code"]:
                    if c_name in FINAL_HEADERS:
                        c_idx = FINAL_HEADERS.index(c_name) + 1
                        ws.cell(row=row_idx, column=c_idx).number_format = '@'
                row_idx += 1

            grp_end = row_idx - 1
            if grp_end >= grp_start:
                for r in range(grp_start, grp_end + 1):
                    ws.row_dimensions[r].outlineLevel = 1
                    ws.row_dimensions[r].hidden = True

        if row_idx > 2:
            dv_status = DataValidation(
                type="list",
                formula1=f'"{",".join(OPCOES_STATUS)}"',
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a value from the dropdown list only.",
                showInputMessage=True,
                promptTitle="Dropdown required",
                prompt="Select a value from the list.",
            )
            dv_query = DataValidation(
                type="list",
                formula1=f'"{",".join(OPCOES_QUERY)}"',
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a value from the dropdown list only.",
                showInputMessage=True,
                promptTitle="Dropdown required",
                prompt="Select a value from the list.",
            )
            dv_review_this_week = DataValidation(
                type="list",
                formula1=f'"{",".join(REVIEW_THIS_WEEK_OPTIONS)}"',
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a weekday from the dropdown list only.",
            )
            dv_complete = DataValidation(
                type="list",
                formula1=f'"{",".join(COMPLETE_OPTIONS)}"',
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select Yes or No from the dropdown list only.",
            )
            ws.add_data_validation(dv_status)
            ws.add_data_validation(dv_query)
            ws.add_data_validation(dv_review_this_week)
            ws.add_data_validation(dv_complete)

            if "Status" in FINAL_HEADERS:
                col_l = get_column_letter(FINAL_HEADERS.index("Status") + 1)
                dv_status.add(f"{col_l}2:{col_l}{row_idx-1}")
            if "Query type" in FINAL_HEADERS:
                col_l = get_column_letter(FINAL_HEADERS.index("Query type") + 1)
                dv_query.add(f"{col_l}2:{col_l}{row_idx-1}")
            if "Review this week" in FINAL_HEADERS:
                col_l = get_column_letter(FINAL_HEADERS.index("Review this week") + 1)
                dv_review_this_week.add(f"{col_l}2:{col_l}{row_idx-1}")
            if "Complete" in FINAL_HEADERS:
                col_l = get_column_letter(FINAL_HEADERS.index("Complete") + 1)
                dv_complete.add(f"{col_l}2:{col_l}{row_idx-1}")

        vals = ["0-30 Days overdue", "31-60 Days overdue", "61-90 Days overdue",
                "91-120 Days Overdue", "121-180 Days Overdue", "180> Days Overdue", "TOTAL VALUE", "Amount in doc. curr."]
        for v in vals:
            if v in FINAL_HEADERS:
                col_l = get_column_letter(FINAL_HEADERS.index(v) + 1)
                for row in range(2, row_idx):
                    cell = ws[f"{col_l}{row}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        if "TOTAL VALUE" in FINAL_HEADERS:
            col_l = get_column_letter(FINAL_HEADERS.index("TOTAL VALUE") + 1)
            rng = f"{col_l}2:{col_l}{row_idx}"
            ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['50000'], fill=cf_red_fill))
            ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=['-50000'], fill=cf_red_fill))
            ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['10000'], fill=cf_yellow_fill))
            ws.conditional_formatting.add(rng, CellIsRule(operator='lessThanOrEqual', formula=['-10000'], fill=cf_yellow_fill))

        if "Status" in FINAL_HEADERS and row_idx > 2:
            from openpyxl.formatting.rule import FormulaRule
            status_col = get_column_letter(FINAL_HEADERS.index("Status") + 1)
            last_col = get_column_letter(len(FINAL_HEADERS))
            apply_range = f"A2:{last_col}{row_idx-1}"
            formula = f'=TRIM(${status_col}2)="Resolved"'
            ws.conditional_formatting.add(apply_range, FormulaRule(formula=[formula], fill=fill_green_light, stopIfTrue=True))

        # Final pass: force dd/mm/yyyy on all date columns (openpyxl default is mm-dd-yy)
        for c_name in ["Net due date", "Document Date", "Posting Date", "Action Date"]:
            if c_name in FINAL_HEADERS:
                c_idx = FINAL_HEADERS.index(c_name) + 1
                for rr in range(2, row_idx):
                    cell = ws.cell(row=rr, column=c_idx)
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = 'dd/mm/yyyy'

    log.info(">>> Salvando arquivo final...")

    lock_path = OUTPUT_PATH + ".lock"
    tmp_path = OUTPUT_PATH + ".tmp"

    if os.path.exists(lock_path):
        age_min = (time.time() - os.path.getmtime(lock_path)) / 60
        raise RuntimeError(
            f"Build lock exists (age {age_min:.0f}min): {lock_path}\n"
            f"Outro build rodando, ou o ultimo crashou. Apague o .lock se tiver certeza."
        )
    Path(lock_path).write_text(f"pid={os.getpid()} started={datetime.now().isoformat()}")

    try:
        while True:
            try:
                wb.save(tmp_path)
                break
            except PermissionError:
                log.warning("\n" + "="*60)
                log.warning(" [ERRO] O ARQUIVO '%s.tmp' ESTA BLOQUEADO!", OUTPUT_FILENAME)
                log.warning(" Feche-o e aperte ENTER.")
                log.warning("="*60 + "\n")
                input()
            except Exception as e:
                log.error(">>> ERRO CRITICO AO SALVAR: %s", e)
                try: os.remove(tmp_path)
                except OSError: pass
                raise

        try:
            validate_owner_tab_integrity(tmp_path)
        except Exception as e:
            log.error(">>> OUTPUT INVALIDO, mantendo xlsx anterior intacto: %s", e)
            log.error(">>> Arquivo invalido preservado em: %s (apague depois de investigar)", tmp_path)
            raise

        os.replace(tmp_path, OUTPUT_PATH)
        log.info(">>> SUCESSO! Salvo em: %s", OUTPUT_PATH)
    finally:
        try: os.remove(lock_path)
        except OSError: pass

    # Archive old Key xlsx files (keep only today's output)
    archived = archive_old_files(KEY_DATA, KEY_ARCHIVE, keep_pattern=data_hoje)
    if archived:
        log.info(">>> [ARCHIVE] Moved %d old file(s) to archive/", len(archived))


if __name__ == "__main__":
    main()
