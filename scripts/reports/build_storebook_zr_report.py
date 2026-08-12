"""Build Storebook / Z & R daily workbook and SQLite feed."""
from __future__ import annotations

import argparse
import logging
import os
import re
import shutil
import sqlite3
import time
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..utils.paths import (
    MASTER_DATA,
    STOREBOOK_ZR_ARCHIVE,
    STOREBOOK_ZR_DATA,
    STOREBOOK_ZR_DB,
    STOREBOOK_ZR_SOURCE_DATA,
    sync_db_to_fixture_store,
)
from ..utils.report_utils import (
    parse_amount_series,
    report_build_lock,
    report_staging_path,
)

log = logging.getLogger(__name__)

OWNER = "Synthetic Owner 001"
STOREBOOK_SOURCE = "Storebook"
ZR_SOURCE = "Z & R"
STOREBOOK_SOURCE_PATTERNS = ("*PaymentBatch*.xlsx", "*Storebook*.xlsx")

STOREBOOK_OUTPUT_COLUMNS = [
    "Owner",
    "Status sytem",
    "Created Date",
    "Financial Net Price COGS",
    "Site",
    "Site",
    "Supplier",
    "Supplier",
    "Main Storebook #",
    "Main Vendor Doc.",
    "Unique Ref",
    "Category",
    "Comments",
    "Action Date",
    "Status",
]

ZR_OUTPUT_COLUMNS = [
    "Owner",
    "Unique Ref",
    "Document Number",
    "Vendor",
    "Company Code",
    "Vendor Name 1",
    "Document Type",
    "Reference",
    "Document Date",
    "Amount in local currency",
    "Local Currency",
    "Net due date",
    "Tax code",
    "Posting Date",
    "Text",
    "Payment Block",
    "User name",
    "Category",
    "Status",
    "action date",
    "Comments",
]

STOREBOOK_STATUS_OPTIONS = [
    "Awaiting Invoice",
    "Waiting for site to respond",
    "Resolved",
    "Awaiting SMC/AM to confirm prices.",
    "Waiting for CS to correct.",
    "GR Issue",
]
AUTO_MISSING_FROM_SOURCE_STATUS = "Auto Resolved - Missing From Source"
AUTO_MISSING_FROM_SOURCE_CODE = "auto_missing_from_source"
ZR_STATUS_OPTIONS = [
    "Awaiting Response",
    "Removed",
    "Awaiting C/S Reversal",
    "Awaiting TL",
    AUTO_MISSING_FROM_SOURCE_STATUS,
]

COMMON_FIELDS = [
    "source",
    "snapshot_date",
    "source_key",
    "unique_ref",
    "owner",
    "supplier_id",
    "supplier_name",
    "company_or_entity",
    "company_code",
    "category",
    "status_system",
    "value",
    "opened_date",
    "action_date",
    "resolved_date",
    "status",
    "comments",
]

DB_COLUMNS = COMMON_FIELDS + [
    "resolution_source",
    "site_id",
    "site_name",
    "main_storebook_no",
    "main_vendor_doc",
    "payment_number",
    "document_number",
    "vendor_id",
    "reference",
    "posting_date",
    "payment_block",
]

WORKBOOK_COLUMNS = DB_COLUMNS

TAIL_FIELDS = ("Status", "Comments", "Action Date")
STOREBOOK_TAIL_SHEET_KEYS = {
    "storebook",
    "storebook blocked invocies",
    "storebook blocked invoices",
}
ZR_TAIL_SHEET_KEYS = {"z&r blocks", "zr blocks", "z & r"}

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS storebook_zr_lines (
    source TEXT NOT NULL,
    snapshot_date TEXT NOT NULL,
    source_key TEXT NOT NULL,
    unique_ref TEXT,
    owner TEXT,
    supplier_id TEXT,
    supplier_name TEXT,
    company_or_entity TEXT,
    company_code TEXT,
    category TEXT,
    status_system TEXT,
    value REAL,
    opened_date TEXT,
    action_date TEXT,
    resolved_date TEXT,
    status TEXT,
    comments TEXT,
    resolution_source TEXT,
    site_id TEXT,
    site_name TEXT,
    main_storebook_no TEXT,
    main_vendor_doc TEXT,
    payment_number TEXT,
    document_number TEXT,
    vendor_id TEXT,
    reference TEXT,
    posting_date TEXT,
    payment_block TEXT,
    loaded_at TEXT NOT NULL,
    PRIMARY KEY (snapshot_date, source_key)
);
CREATE INDEX IF NOT EXISTS idx_storebook_zr_snapshot
    ON storebook_zr_lines(snapshot_date, source);
CREATE INDEX IF NOT EXISTS idx_storebook_zr_owner_status
    ON storebook_zr_lines(owner, status);
"""


@dataclass(frozen=True)
class StorebookZRConfig:
    snapshot_date: date
    output_dir: Path = STOREBOOK_ZR_DATA
    archive_dir: Path = STOREBOOK_ZR_ARCHIVE
    source_dir: Path = STOREBOOK_ZR_SOURCE_DATA
    storebook_source: Path | None = None
    zr_source: Path | None = None
    reference_workbook: Path | None = STOREBOOK_ZR_DATA / "Storebook ZR KPI.xlsx"
    vendor_matrix_path: Path = MASTER_DATA / "Synthetic_Vendor_Master_Matrix.csv"
    db_path: Path = STOREBOOK_ZR_DB
    sync_db: bool = False


@dataclass(frozen=True)
class BuildResult:
    output_path: Path
    rows_by_source: dict[str, int]
    archived: list[str]
    duplicate_keys: list[str]
    db_path: Path


class MissingStorebookZRSources(RuntimeError):
    """Raised when the daily Storebook/Z&R input drop has not arrived yet."""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"-?\d+\.0", text):
        return text[:-2]
    return text


def _first_present(row: dict[str, Any], names: Iterable[str]) -> Any:
    for name in names:
        if name in row and _clean_text(row.get(name)):
            return row.get(name)
    return ""


def _lookup_key(value: Any) -> str:
    return _clean_text(value).casefold()


def _matrix_column_key(column: Any) -> str:
    return str(column).replace("\xa0", " ").strip().casefold()


def load_vendor_category_lookup(path: Path) -> dict[str, str]:
    """Load Vendor Master Matrix category by normalized Unique ref and Vendor."""
    if not path.exists():
        log.warning("Vendor Master Matrix not found: %s", path)
        return {}

    wanted = {"unique ref", "vendor", "category"}
    try:
        matrix = pd.read_csv(
            path,
            dtype=str,
            encoding="utf-8-sig",
            usecols=lambda col: _matrix_column_key(col) in wanted,
        )
    except UnicodeDecodeError:
        matrix = pd.read_csv(
            path,
            dtype=str,
            encoding="latin1",
            usecols=lambda col: _matrix_column_key(col) in wanted,
        )
    matrix.columns = [_matrix_column_key(column) for column in matrix.columns]

    lookup: dict[str, str] = {}
    for raw in matrix.fillna("").to_dict("records"):
        category = _clean_text(raw.get("category"))
        if not category:
            continue
        unique_ref = _lookup_key(raw.get("unique ref"))
        vendor = _lookup_key(raw.get("vendor"))
        if unique_ref:
            lookup[f"unique_ref:{unique_ref}"] = category
        if vendor and f"vendor:{vendor}" not in lookup:
            lookup[f"vendor:{vendor}"] = category
    return lookup


def _category_for_vendor(lookup: dict[str, str], *, vendor: Any, unique_ref: Any = "") -> str:
    unique_key = _lookup_key(unique_ref)
    vendor_key = _lookup_key(vendor)
    if unique_key:
        category = lookup.get(f"unique_ref:{unique_key}")
        if category:
            return category
    if vendor_key:
        return lookup.get(f"vendor:{vendor_key}", "")
    return ""


def _parse_date(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _parse_value(value: Any) -> float | None:
    parsed = parse_amount_series(pd.Series([value])).iloc[0]
    if pd.isna(parsed):
        return None
    return float(parsed)


def _normalize_headers(columns: Iterable[Any]) -> list[str]:
    return [str(col).replace("\xa0", " ").strip() for col in columns]


def _find_header_row(path: Path, required: set[str], sheet_name: str | int = 0) -> int:
    probe = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=12, dtype=str)
    for idx, row in enumerate(probe.itertuples(index=False, name=None)):
        values = {str(value).replace("\xa0", " ").strip() for value in row}
        if required.issubset(values):
            return idx
    raise ValueError(f"Could not find header row in {path} for {sorted(required)}")


def _read_excel_with_detected_header(
    path: Path,
    required: set[str],
    sheet_name: str | int = 0,
) -> pd.DataFrame:
    header_row = _find_header_row(path, required, sheet_name)
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row, dtype=object)
    df.columns = _normalize_headers(df.columns)
    df = df.dropna(how="all")
    return df


def make_storebook_key(row: dict[str, Any]) -> str:
    main_storebook_no = _clean_text(_first_present(row, ("Main Storebook #", "main_storebook_no")))
    main_vendor_doc = _clean_text(_first_present(row, ("Main Vendor Doc.", "main_vendor_doc")))
    payment_number = _clean_text(_first_present(row, ("Payment Number", "payment_number")))
    if main_storebook_no and main_vendor_doc and payment_number:
        return f"{STOREBOOK_SOURCE}|{main_storebook_no}|{main_vendor_doc}|{payment_number}"

    supplier_id = _clean_text(_first_present(row, ("Supplier", "supplier_id")))
    opened_date = _parse_date(_first_present(row, ("Created Date", "opened_date")))
    value = _clean_text(_first_present(row, ("Financial Net Price COGS", "value")))
    fallback_parts = [main_storebook_no, main_vendor_doc, payment_number, supplier_id, opened_date, value]
    return f"{STOREBOOK_SOURCE}|fallback|" + "|".join(fallback_parts)


def make_storebook_unique_ref(row: dict[str, Any]) -> str:
    supplier_id = _clean_text(_first_present(row, ("Supplier", "supplier_id")))
    return " ".join(part for part in ("9001", supplier_id) if part)


def make_storebook_tail_key(row: dict[str, Any]) -> str:
    main_storebook_no = _clean_text(_first_present(row, ("Main Storebook #", "main_storebook_no")))
    main_vendor_doc = _clean_text(_first_present(row, ("Main Vendor Doc.", "main_vendor_doc")))
    unique_ref = _clean_text(_first_present(row, ("Unique Ref", "unique_ref")))
    return f"{STOREBOOK_SOURCE}|tail|{main_storebook_no}|{main_vendor_doc}|{unique_ref}"


def make_zr_key(row: dict[str, Any]) -> str:
    company_code = _clean_text(_first_present(row, ("Company Code", "company_code")))
    document_number = _clean_text(_first_present(row, ("Document Number", "document_number")))
    vendor_id = _clean_text(_first_present(row, ("Vendor", "vendor_id", "supplier_id")))
    reference = _clean_text(_first_present(row, ("Reference", "reference")))
    posting_date = _parse_date(_first_present(row, ("Posting Date", "posting_date")))
    payment_block = _clean_text(_first_present(row, ("Payment Block", "payment_block")))
    return f"{ZR_SOURCE}|{company_code}|{document_number}|{vendor_id}|{reference}|{posting_date}|{payment_block}"


def _status_to_resolved_date(status: str, action_date: str) -> str:
    if status.strip().casefold() == "resolved" and action_date:
        return action_date
    return ""


def _zr_status_to_resolved_date(status: str, action_date: str) -> str:
    completed_statuses = {"removed", "resolved", "complete", "completed", "closed"}
    if status.strip().casefold() in completed_statuses and action_date:
        return action_date
    return ""


def normalize_storebook_rows(
    df: pd.DataFrame,
    snapshot_date: date,
    category_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in df.to_dict("records"):
        source_key = make_storebook_key(raw)
        unique_ref = make_storebook_unique_ref(raw)
        action_date = _parse_date(_first_present(raw, ("Action Date", "Action Date ", "action_date")))
        status_system = _clean_text(_first_present(raw, ("Status sytem", "Status system", "Status")))
        status = _clean_text(_first_present(raw, ("Status ", "status")))
        supplier_id = _clean_text(_first_present(raw, ("Supplier", "supplier_id")))
        supplier_name = _clean_text(_first_present(raw, ("Supplier.1", "Supplier Name", "supplier_name")))
        site_id = _clean_text(_first_present(raw, ("Site", "site_id")))
        site_name = _clean_text(_first_present(raw, ("Site.1", "Site Name", "company_or_entity")))
        category = _category_for_vendor(category_lookup, vendor=supplier_id, unique_ref=unique_ref)
        row = {
            "source": STOREBOOK_SOURCE,
            "snapshot_date": snapshot_date.isoformat(),
            "source_key": source_key,
            "unique_ref": unique_ref,
            "owner": OWNER,
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "company_or_entity": "9001",
            "company_code": "9001",
            "category": category,
            "status_system": status_system,
            "value": _parse_value(_first_present(raw, ("Financial Net Price COGS", "value"))),
            "opened_date": _parse_date(_first_present(raw, ("Created Date", "opened_date"))),
            "action_date": action_date,
            "resolved_date": _parse_date(_first_present(raw, ("Resolved Date", "resolved_date")))
            or _status_to_resolved_date(status, action_date),
            "status": status,
            "comments": _clean_text(_first_present(raw, ("Comments", "Comments ", "comments"))),
            "site_id": site_id,
            "site_name": site_name,
            "main_storebook_no": _clean_text(_first_present(raw, ("Main Storebook #", "main_storebook_no"))),
            "main_vendor_doc": _clean_text(_first_present(raw, ("Main Vendor Doc.", "main_vendor_doc"))),
            "payment_number": _clean_text(_first_present(raw, ("Payment Number", "payment_number"))),
            "document_number": "",
            "vendor_id": "",
            "reference": "",
            "posting_date": "",
            "payment_block": "",
            "Status sytem": status_system,
            "Owner": OWNER,
            "Created Date": _parse_date(_first_present(raw, ("Created Date", "opened_date"))),
            "Financial Net Price COGS": _parse_value(_first_present(raw, ("Financial Net Price COGS", "value"))),
            "Site": site_id,
            "Site Name": site_name,
            "Supplier": supplier_id,
            "Supplier Name": supplier_name,
            "Main Storebook #": _clean_text(_first_present(raw, ("Main Storebook #", "main_storebook_no"))),
            "Main Vendor Doc.": _clean_text(_first_present(raw, ("Main Vendor Doc.", "main_vendor_doc"))),
            "Unique Ref": unique_ref,
            "Category": category,
            "Comments": _clean_text(_first_present(raw, ("Comments", "Comments ", "comments"))),
            "Action Date": action_date,
            "Status": status,
            "_tail_keys": [source_key, make_storebook_tail_key({"Main Storebook #": _first_present(raw, ("Main Storebook #", "main_storebook_no")), "Main Vendor Doc.": _first_present(raw, ("Main Vendor Doc.", "main_vendor_doc")), "Unique Ref": unique_ref})],
        }
        if row["source_key"]:
            rows.append(row)
    return rows


def normalize_zr_rows(
    df: pd.DataFrame,
    snapshot_date: date,
    category_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in df.to_dict("records"):
        action_date = _parse_date(_first_present(raw, ("action date", "action date ", "Action Date", "action_date")))
        status = _clean_text(_first_present(raw, ("Status", "Status ", "status")))
        vendor_id = _clean_text(_first_present(raw, ("Vendor", "vendor_id")))
        company_code = _clean_text(_first_present(raw, ("Company Code", "company_code")))
        document_number = _clean_text(_first_present(raw, ("Document Number", "document_number")))
        reference = _clean_text(_first_present(raw, ("Reference", "reference")))
        posting_date = _parse_date(_first_present(raw, ("Posting Date", "posting_date")))
        payment_block = _clean_text(_first_present(raw, ("Payment Block", "payment_block")))
        if not any((document_number, vendor_id, company_code, reference, posting_date, payment_block)):
            continue
        source_key = make_zr_key(raw)
        unique_ref = " ".join(part for part in (company_code, vendor_id) if part)
        category = _category_for_vendor(category_lookup, vendor=vendor_id, unique_ref=unique_ref)
        row = {
            "source": ZR_SOURCE,
            "snapshot_date": snapshot_date.isoformat(),
            "source_key": source_key,
            "unique_ref": unique_ref,
            "owner": OWNER,
            "supplier_id": vendor_id,
            "supplier_name": _clean_text(_first_present(raw, ("Vendor Name 1", "supplier_name"))),
            "company_or_entity": company_code,
            "company_code": company_code,
            "category": category,
            "status_system": "",
            "value": _parse_value(_first_present(raw, ("Amount in local currency", "value"))),
            "opened_date": _parse_date(_first_present(raw, ("Document Date", "opened_date"))),
            "action_date": action_date,
            "resolved_date": _zr_status_to_resolved_date(status, action_date),
            "status": status,
            "comments": _clean_text(_first_present(raw, ("Comments", "Comments ", "comments"))),
            "site_id": "",
            "site_name": "",
            "main_storebook_no": "",
            "main_vendor_doc": "",
            "payment_number": "",
            "document_number": document_number,
            "vendor_id": vendor_id,
            "reference": reference,
            "posting_date": posting_date,
            "payment_block": payment_block,
            "Owner": OWNER,
            "Unique Ref": unique_ref,
            "Document Number": document_number,
            "Vendor": vendor_id,
            "Company Code": company_code,
            "Vendor Name 1": _clean_text(_first_present(raw, ("Vendor Name 1", "supplier_name"))),
            "Document Type": _clean_text(_first_present(raw, ("Document Type",))),
            "Reference": reference,
            "Document Date": _parse_date(_first_present(raw, ("Document Date", "opened_date"))),
            "Amount in local currency": _parse_value(_first_present(raw, ("Amount in local currency", "value"))),
            "Local Currency": _clean_text(_first_present(raw, ("Local Currency",))),
            "Net due date": _parse_date(_first_present(raw, ("Net due date",))),
            "Tax code": _clean_text(_first_present(raw, ("Tax code",))),
            "Posting Date": posting_date,
            "Text": _clean_text(_first_present(raw, ("Text",))),
            "Payment Block": payment_block,
            "User name": _clean_text(_first_present(raw, ("User name",))),
            "Category": category,
            "Status": status,
            "action date": action_date,
            "Comments": _clean_text(_first_present(raw, ("Comments", "Comments ", "comments"))),
            "_tail_keys": [source_key],
        }
        if row["source_key"]:
            rows.append(row)
    return rows


def _read_storebook_rows(
    path: Path,
    snapshot_date: date,
    category_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    df = _read_excel_with_detected_header(path, {"Payment Number", "Main Storebook #", "Main Vendor Doc."})
    return normalize_storebook_rows(df, snapshot_date, category_lookup)


def _read_zr_rows(
    path: Path,
    snapshot_date: date,
    category_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    df = _read_excel_with_detected_header(path, {"Document Number", "Vendor", "Company Code", "Payment Block"})
    return normalize_zr_rows(df, snapshot_date, category_lookup)


def _read_reference_storebook(
    path: Path,
    snapshot_date: date,
    category_lookup: dict[str, str],
) -> list[dict[str, Any]]:
    df = _read_excel_with_detected_header(
        path,
        {"Main Storebook #", "Main Vendor Doc."},
        sheet_name="Storebook ",
    )
    return normalize_storebook_rows(df, snapshot_date, category_lookup)


def _latest_file(directory: Path, pattern: str, exclude: Iterable[str] = ()) -> Path | None:
    excluded = tuple(name.casefold() for name in exclude)
    files = [
        path
        for path in directory.glob(pattern)
        if path.is_file()
        and not path.name.startswith("~$")
        and all(token not in path.name.casefold() for token in excluded)
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def _latest_file_any(directory: Path, patterns: Iterable[str], exclude: Iterable[str] = ()) -> Path | None:
    excluded = tuple(name.casefold() for name in exclude)
    files = [
        path
        for pattern in patterns
        for path in directory.glob(pattern)
        if path.is_file()
        and not path.name.startswith("~$")
        and all(token not in path.name.casefold() for token in excluded)
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def resolve_sources(config: StorebookZRConfig) -> tuple[Path | None, Path | None]:
    storebook_source = config.storebook_source or _latest_file_any(
        config.source_dir,
        STOREBOOK_SOURCE_PATTERNS,
        exclude=("storebook_zr_", "storebook zr kpi"),
    )
    zr_source = config.zr_source or _latest_file(
        config.source_dir,
        "*.xlsx",
        exclude=("paymentbatch", "storebook", "storebook_zr_", "storebook zr kpi"),
    )
    if storebook_source is None and config.reference_workbook and config.reference_workbook.exists():
        storebook_source = config.reference_workbook
    return storebook_source, zr_source


def _tail_from_row(row: dict[str, Any]) -> dict[str, Any]:
    return {field: row.get(field, "") for field in TAIL_FIELDS if _clean_text(row.get(field))}


def _capture_tail_from_workbook(path: Path) -> dict[str, dict[str, Any]]:
    tail: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return tail

    try:
        xls = pd.ExcelFile(path)
    except Exception as err:
        log.warning("Could not read tail workbook %s: %s", path, err)
        return tail

    try:
        for sheet in xls.sheet_names:
            sheet_key = sheet.strip().casefold()
            if sheet_key not in STOREBOOK_TAIL_SHEET_KEYS | ZR_TAIL_SHEET_KEYS:
                continue
            try:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
            except Exception:
                continue
            df.columns = _normalize_headers(df.columns)
            records = df.dropna(how="all").to_dict("records")
            for raw in records:
                source_key = _clean_text(raw.get("source_key"))
                if not source_key:
                    if sheet_key in STOREBOOK_TAIL_SHEET_KEYS:
                        source_key = make_storebook_tail_key(raw)
                    else:
                        source_key = make_zr_key(raw)
                row_tail = {
                    "Status": _clean_text(_first_present(raw, ("Status", "Status "))),
                    "Comments": _clean_text(_first_present(raw, ("Comments", "Comments "))),
                    "Action Date": _parse_date(_first_present(raw, ("Action Date", "Action Date ", "action date "))),
                }
                clean_tail = {k: v for k, v in row_tail.items() if v}
                if source_key and clean_tail:
                    tail.setdefault(source_key, {}).update(clean_tail)
    finally:
        xls.close()
    return tail


def _previous_output_candidates(output_dir: Path, archive_dir: Path, current_output: Path) -> list[Path]:
    files = list(output_dir.glob("Storebook_ZR_*.xlsx"))
    if archive_dir.exists():
        files.extend(archive_dir.glob("Storebook_ZR_*.xlsx"))
    return sorted(
        [path for path in files if not path.name.startswith("~$")],
        key=lambda path: path.name,
        reverse=True,
    )


def capture_tail(config: StorebookZRConfig, current_output: Path) -> dict[str, dict[str, Any]]:
    tail: dict[str, dict[str, Any]] = {}
    for candidate in _previous_output_candidates(config.output_dir, config.archive_dir, current_output):
        tail.update(_capture_tail_from_workbook(candidate))
        if tail:
            break
    if config.reference_workbook and config.reference_workbook.exists():
        reference_tail = _capture_tail_from_workbook(config.reference_workbook)
        reference_tail.update(tail)
        return reference_tail
    return tail


def apply_tail(rows: list[dict[str, Any]], tail: dict[str, dict[str, Any]]) -> None:
    for row in rows:
        row_tail = None
        for key in row.get("_tail_keys", [row["source_key"]]):
            row_tail = tail.get(key)
            if row_tail:
                break
        if not row_tail:
            continue
        for field in TAIL_FIELDS:
            value = row_tail.get(field)
            if value:
                row[field] = value
        action_value = row.get("Action Date") or row.get("action date") or ""
        if row.get("source") == ZR_SOURCE and row.get("Action Date"):
            row["action date"] = row["Action Date"]
        row["status"] = _clean_text(row.get("Status"))
        row["comments"] = _clean_text(row.get("Comments"))
        row["action_date"] = _parse_date(action_value)
        if row.get("source") == ZR_SOURCE:
            row["resolved_date"] = _zr_status_to_resolved_date(row["status"], row["action_date"])
        else:
            row["resolved_date"] = _status_to_resolved_date(row["status"], row["action_date"])


def _is_open_active_zr_row(row: dict[str, Any]) -> bool:
    if row.get("source") != ZR_SOURCE:
        return False
    if _parse_date(row.get("resolved_date")):
        return False
    status = _clean_text(row.get("status")).casefold()
    return status not in {"removed", "resolved", "complete", "completed", "closed"}


def _auto_missing_comment(existing: Any) -> str:
    current = _clean_text(existing)
    if not current:
        return AUTO_MISSING_FROM_SOURCE_STATUS
    if AUTO_MISSING_FROM_SOURCE_STATUS.casefold() in current.casefold():
        return current
    return f"{current} | {AUTO_MISSING_FROM_SOURCE_STATUS}"


def _zr_visible_aliases(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "Owner": row.get("owner", ""),
        "Unique Ref": row.get("unique_ref", ""),
        "Document Number": row.get("document_number", ""),
        "Vendor": row.get("vendor_id") or row.get("supplier_id", ""),
        "Company Code": row.get("company_code", ""),
        "Vendor Name 1": row.get("supplier_name", ""),
        "Reference": row.get("reference", ""),
        "Document Date": row.get("opened_date", ""),
        "Amount in local currency": row.get("value", ""),
        "Posting Date": row.get("posting_date", ""),
        "Payment Block": row.get("payment_block", ""),
        "Category": row.get("category", ""),
        "Status": row.get("status", ""),
        "action date": row.get("action_date", ""),
        "Comments": row.get("comments", ""),
    }


def apply_zr_missing_source_auto_resolutions(
    rows: list[dict[str, Any]],
    previous_rows: list[dict[str, Any]],
    snapshot_date: date,
) -> list[dict[str, Any]]:
    current_zr_keys = {
        _clean_text(row.get("source_key"))
        for row in rows
        if row.get("source") == ZR_SOURCE and _clean_text(row.get("source_key"))
    }
    auto_rows: list[dict[str, Any]] = []
    snapshot_iso = snapshot_date.isoformat()
    for previous in previous_rows:
        source_key = _clean_text(previous.get("source_key"))
        if not source_key or source_key in current_zr_keys or not _is_open_active_zr_row(previous):
            continue
        auto_row = {column: previous.get(column, "") for column in DB_COLUMNS}
        auto_row.update(
            {
                "source": ZR_SOURCE,
                "snapshot_date": snapshot_iso,
                "source_key": source_key,
                "action_date": "",
                "resolved_date": snapshot_iso,
                "status": AUTO_MISSING_FROM_SOURCE_STATUS,
                "comments": _auto_missing_comment(previous.get("comments")),
                "resolution_source": AUTO_MISSING_FROM_SOURCE_CODE,
            }
        )
        auto_row.update(_zr_visible_aliases(auto_row))
        auto_rows.append(auto_row)
        current_zr_keys.add(source_key)
    return auto_rows


def detect_duplicate_keys(rows: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for row in rows:
        key = row["source_key"]
        if key in seen and key not in duplicates:
            duplicates.append(key)
        seen.add(key)
    return duplicates


def dedupe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for row in rows:
        key = row["source_key"]
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _validate_rows(rows: list[dict[str, Any]]) -> None:
    missing = [field for field in COMMON_FIELDS if any(field not in row for row in rows)]
    if missing:
        raise RuntimeError(f"Storebook/Z&R rows missing canonical fields: {sorted(set(missing))}")


def _visible_values(source: str, row: dict[str, Any]) -> list[Any]:
    if source == STOREBOOK_SOURCE:
        return [
            row.get("Owner", OWNER),
            row.get("Status sytem", ""),
            row.get("Created Date", ""),
            row.get("Financial Net Price COGS", ""),
            row.get("Site", ""),
            row.get("Site Name", ""),
            row.get("Supplier", ""),
            row.get("Supplier Name", ""),
            row.get("Main Storebook #", ""),
            row.get("Main Vendor Doc.", ""),
            row.get("Unique Ref", ""),
            row.get("Category", ""),
            row.get("Comments", ""),
            row.get("Action Date", ""),
            row.get("Status", ""),
        ]
    if source == ZR_SOURCE:
        return [row.get(column, "") for column in ZR_OUTPUT_COLUMNS]
    raise ValueError(f"Unsupported Storebook/Z&R source: {source}")


def _assert_storebook_zr_workbook(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != [STOREBOOK_SOURCE, ZR_SOURCE]:
            raise RuntimeError(f"Unexpected Storebook/Z&R sheets: {workbook.sheetnames}")
        return int(path.stat().st_size)
    finally:
        workbook.close()


def _publish_storebook_zr_output(staging_path: Path, output_path: Path) -> int:
    """Copy the validated local workbook into Local Fixture Store and verify it opens.

    Storebook/Z&R workbooks are small and user-facing, so direct local-first
    copy is safer here than a Local Fixture Store-side publish temp that Local Fixture Store can
    resize while syncing.
    """

    if not staging_path.exists():
        raise RuntimeError(f"Storebook/Z&R staging workbook not found: {staging_path}")
    source_size = _assert_storebook_zr_workbook(staging_path)
    if source_size <= 0:
        raise RuntimeError(f"Storebook/Z&R staging workbook is empty: {staging_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    last_error: Exception | None = None
    for attempt in range(1, 6):
        try:
            shutil.copy2(staging_path, output_path)
            final_size = _assert_storebook_zr_workbook(output_path)
            if final_size <= 0:
                raise RuntimeError(f"Published Storebook/Z&R workbook is empty: {output_path}")
            log.info(
                "Published Storebook/Z&R workbook to %s (%d bytes, source %d bytes)",
                output_path,
                final_size,
                source_size,
            )
            return final_size
        except (OSError, RuntimeError) as err:
            last_error = err
            log.warning(
                "Storebook/Z&R publish attempt %d failed: %s",
                attempt,
                err,
            )
            time.sleep(min(2 * attempt, 10))
    raise RuntimeError(
        f"Could not publish Storebook/Z&R workbook to {output_path}: {last_error}"
    )


def write_workbook(rows: list[dict[str, Any]], output_path: Path) -> int:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="002060")
    header_font = Font(bold=True, color="FFFFFF")

    source_specs = (
        (STOREBOOK_SOURCE, STOREBOOK_OUTPUT_COLUMNS, STOREBOOK_STATUS_OPTIONS),
        (ZR_SOURCE, ZR_OUTPUT_COLUMNS, ZR_STATUS_OPTIONS),
    )

    for source, columns, status_options in source_specs:
        ws = wb.create_sheet(source)
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in [record for record in rows if record["source"] == source]:
            ws.append(_visible_values(source, row))
        if ws.max_row > 1:
            status_col = get_column_letter(columns.index("Status") + 1)
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(status_options)}"',
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a value from the dropdown list only.",
            )
            ws.add_data_validation(validation)
            validation.add(f"{status_col}2:{status_col}{ws.max_row}")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"
        for col_idx, column in enumerate(columns, start=1):
            width = max(12, min(36, len(column) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    staging_path = Path(report_staging_path(str(output_path)))
    lock_path = str(output_path) + ".lock"
    with report_build_lock(lock_path, f"pid={os.getpid()} started={datetime.now().isoformat()}"):
        wb.save(staging_path)
        _assert_storebook_zr_workbook(staging_path)
        final_size = _publish_storebook_zr_output(staging_path, output_path)
        try:
            staging_path.unlink()
        except FileNotFoundError:
            pass
    return final_size


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA_SQL)
    existing = {
        str(row[1])
        for row in conn.execute("PRAGMA table_info(storebook_zr_lines)").fetchall()
    }
    column_types = {
        "value": "REAL",
        "loaded_at": "TEXT",
    }
    for column in DB_COLUMNS + ["loaded_at"]:
        if column not in existing:
            conn.execute(
                f'ALTER TABLE storebook_zr_lines ADD COLUMN "{column}" {column_types.get(column, "TEXT")}'
            )


def write_sqlite(rows: list[dict[str, Any]], db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    loaded_at = datetime.now().isoformat(timespec="seconds")
    columns = DB_COLUMNS + ["loaded_at"]
    placeholders = ", ".join("?" for _ in columns)
    column_sql = ", ".join(columns)
    with closing(sqlite3.connect(str(db_path))) as conn:
        init_db(conn)
        conn.execute("BEGIN")
        try:
            if not rows:
                conn.execute("DELETE FROM storebook_zr_lines")
                conn.commit()
                return
            conn.execute(
                """
                DELETE FROM storebook_zr_lines
                WHERE source = ?
                  AND COALESCE(unique_ref, '') = ''
                  AND COALESCE(supplier_id, '') = ''
                  AND COALESCE(company_code, '') = ''
                """,
                (ZR_SOURCE,),
            )
            snapshot_dates = sorted({row["snapshot_date"] for row in rows})
            for snapshot_date in snapshot_dates:
                conn.execute(
                    "DELETE FROM storebook_zr_lines WHERE snapshot_date = ?",
                    (snapshot_date,),
                )
            conn.executemany(
                f"INSERT INTO storebook_zr_lines ({column_sql}) VALUES ({placeholders})",
                [tuple(row.get(column, loaded_at if column == "loaded_at" else "") for column in columns) for row in rows],
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise


def load_previous_zr_snapshot_rows(db_path: Path, snapshot_date: date) -> list[dict[str, Any]]:
    if not db_path.exists():
        return []
    with closing(sqlite3.connect(str(db_path))) as conn:
        init_db(conn)
        previous_snapshot = conn.execute(
            """
            SELECT MAX(snapshot_date)
            FROM storebook_zr_lines
            WHERE source = ?
              AND snapshot_date < ?
            """,
            (ZR_SOURCE, snapshot_date.isoformat()),
        ).fetchone()[0]
        if not previous_snapshot:
            return []
        column_sql = ", ".join(f'"{column}"' for column in DB_COLUMNS)
        rows = conn.execute(
            f"""
            SELECT {column_sql}
            FROM storebook_zr_lines
            WHERE source = ?
              AND snapshot_date = ?
            """,
            (ZR_SOURCE, previous_snapshot),
        ).fetchall()
    return [dict(zip(DB_COLUMNS, row)) for row in rows]


def archive_previous_day_output(output_dir: Path, archive_dir: Path, snapshot_date: date) -> list[str]:
    previous_name = f"Storebook_ZR_{(snapshot_date - timedelta(days=1)).isoformat()}.xlsx"
    today_name = f"Storebook_ZR_{snapshot_date.isoformat()}.xlsx"
    if previous_name == today_name:
        return []
    previous_path = output_dir / previous_name
    if not previous_path.exists():
        return []
    archive_dir.mkdir(parents=True, exist_ok=True)
    dest = archive_dir / previous_name
    if dest.exists():
        dest = archive_dir / f"{dest.stem}.conflict-{int(datetime.now().timestamp())}{dest.suffix}"
    shutil.move(str(previous_path), str(dest))
    return [dest.name]


def _date_or_none(value: Any) -> date | None:
    parsed = _parse_date(value)
    if not parsed:
        return None
    return datetime.strptime(parsed, "%Y-%m-%d").date()


def _business_days(start: date, end: date) -> int:
    if end < start:
        return 0
    days = pd.bdate_range(start=start, end=end)
    return int(len(days))


def compute_window_metrics(rows: list[dict[str, Any]], date_from: date, date_to: date) -> dict[str, Any]:
    open_days: list[int] = []
    resolve_days: list[int] = []
    open_count = 0
    resolved_count = 0
    productivity_count = 0
    detail_active_count = 0

    for row in rows:
        opened = _date_or_none(row.get("opened_date"))
        action = _date_or_none(row.get("action_date"))
        resolved = _date_or_none(row.get("resolved_date"))
        is_open_at_to = opened is not None and opened <= date_to and (resolved is None or resolved > date_to)
        is_resolved_in_window = resolved is not None and date_from <= resolved <= date_to
        is_active = opened is not None and opened <= date_to and (resolved is None or resolved >= date_from)

        if is_open_at_to:
            open_count += 1
            open_days.append(_business_days(opened, date_to))
        if is_resolved_in_window:
            resolved_count += 1
            if opened is not None:
                resolve_days.append(_business_days(opened, resolved))
        if action is not None and date_from <= action <= date_to:
            productivity_count += 1
        if is_active:
            detail_active_count += 1

    return {
        "open_count": open_count,
        "resolved_count": resolved_count,
        "avg_days_open": round(sum(open_days) / len(open_days), 2) if open_days else 0.0,
        "avg_days_to_resolve": round(sum(resolve_days) / len(resolve_days), 2) if resolve_days else 0.0,
        "productivity_count": productivity_count,
        "detail_active_count": detail_active_count,
    }


def load_rows(config: StorebookZRConfig) -> list[dict[str, Any]]:
    storebook_source, zr_source = resolve_sources(config)
    category_lookup = load_vendor_category_lookup(config.vendor_matrix_path)
    storebook_rows: list[dict[str, Any]] = []
    zr_rows: list[dict[str, Any]] = []
    if storebook_source is None and zr_source is None:
        log.warning("No Storebook/Z&R source files found in %s; writing empty feed", config.source_dir)
        return []
    if storebook_source is None:
        log.warning("No Storebook source file found in %s", config.source_dir)
    elif storebook_source == config.reference_workbook:
        storebook_rows = _read_reference_storebook(storebook_source, config.snapshot_date, category_lookup)
    else:
        storebook_rows = _read_storebook_rows(storebook_source, config.snapshot_date, category_lookup)
    if zr_source is None:
        log.warning("No Z & R source file found in %s", config.source_dir)
    else:
        zr_rows = _read_zr_rows(zr_source, config.snapshot_date, category_lookup)
    return storebook_rows + zr_rows


def build_report(config: StorebookZRConfig) -> BuildResult:
    output_path = config.output_dir / f"Storebook_ZR_{config.snapshot_date.isoformat()}.xlsx"
    rows = load_rows(config)
    auto_resolution_rows: list[dict[str, Any]] = []

    if rows:
        tail = capture_tail(config, output_path)
        apply_tail(rows, tail)
        previous_zr_rows = load_previous_zr_snapshot_rows(config.db_path, config.snapshot_date)
        if previous_zr_rows and tail:
            apply_tail(previous_zr_rows, tail)
        auto_resolution_rows = apply_zr_missing_source_auto_resolutions(
            rows,
            previous_zr_rows,
            config.snapshot_date,
        )
    db_rows = rows + auto_resolution_rows
    duplicate_keys = detect_duplicate_keys(db_rows)
    if duplicate_keys:
        log.warning("Detected %d duplicate Storebook/Z&R source key(s)", len(duplicate_keys))
    rows = dedupe_rows(rows)
    db_rows = dedupe_rows(db_rows)
    if db_rows:
        _validate_rows(db_rows)

    config.output_dir.mkdir(parents=True, exist_ok=True)
    write_workbook(rows, output_path)
    write_sqlite(db_rows, config.db_path)
    if config.sync_db:
        sync_db_to_fixture_store()
    archived = archive_previous_day_output(config.output_dir, config.archive_dir, config.snapshot_date)
    rows_by_source = {
        STOREBOOK_SOURCE: sum(1 for row in rows if row["source"] == STOREBOOK_SOURCE),
        ZR_SOURCE: sum(1 for row in rows if row["source"] == ZR_SOURCE),
    }
    return BuildResult(
        output_path=output_path,
        rows_by_source=rows_by_source,
        archived=archived,
        duplicate_keys=duplicate_keys,
        db_path=config.db_path,
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Storebook / Z & R daily workbook and SQLite feed")
    parser.add_argument("--date", dest="snapshot_date", help="Snapshot date YYYY-MM-DD; defaults to today")
    parser.add_argument("--storebook-source", type=Path, help="Explicit Storebook PaymentBatch workbook")
    parser.add_argument("--zr-source", type=Path, help="Explicit Z & R workbook")
    parser.add_argument("--sync-db", action="store_true", help="Copy local SQLite backup to repo db/")
    return parser.parse_args()


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    args = _parse_args()
    snapshot = (
        datetime.strptime(args.snapshot_date, "%Y-%m-%d").date()
        if args.snapshot_date
        else date.today()
    )
    result = build_report(
        StorebookZRConfig(
            snapshot_date=snapshot,
            storebook_source=args.storebook_source,
            zr_source=args.zr_source,
            sync_db=args.sync_db,
        )
    )
    log.info(
        "Storebook/Z&R complete: %s rows=%s archived=%s duplicate_keys=%d db=%s",
        result.output_path,
        result.rows_by_source,
        result.archived,
        len(result.duplicate_keys),
        result.db_path,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
