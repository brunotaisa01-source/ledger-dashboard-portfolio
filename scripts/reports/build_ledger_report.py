# -*- coding: utf-8 -*-
"""
Ledger report builder.
Updated:
- FIXED: AttributeError (handling missing columns in _norm_str_series)
- Payment Issues column logic (Detail vs Summary)
- Summary Columns: LW, CW, Delta (Strictly enforced in all tables)
- New Tab: Payment Issues with Bar Chart (Owner) and Audit Table
- Buckets: 0-30
- Charts generation fixed (Inside team block, Dual Axis)
- Removed Fuel and Unassigned (STRICT FILTER)
- FIXED: File sorting by name date
"""
from __future__ import annotations

import glob
import logging
import os
import sys
import re
import time
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

log = logging.getLogger(__name__)

from ..utils.masterdata_core import known_rol_owners, norm_str_series as _norm_str_series
from ..utils.report_utils import (
    parse_amount_series, parse_mixed_date_series,
    groupby_apply_no_warning as _groupby_apply_no_warning,
    system_from_code, clean_key, sanitize_value, recover_date_from_tail,
    _extract_snapshot_date, extract_date_from_filename, _bucket_mask, _supplier_key,
    agg_payment_issues as _agg_pi,
    validate_owner_tab_integrity,
)

COR_HEADER = "002060"
COR_FONTE = "FFFFFF"
fill_header_custom = PatternFill("solid", fgColor=COR_HEADER)
font_header_custom = Font(bold=True, color=COR_FONTE)
fill_blue_light = PatternFill("solid", fgColor="DAE9F8")
fill_green_light = PatternFill("solid", fgColor="C1F0C8")
fill_red_alert = PatternFill("solid", fgColor="FF0000")
fill_blue_total = PatternFill("solid", fgColor="83CCEB")
fill_summary_gray = PatternFill("solid", fgColor="D9D9D9")
font_black_bold = Font(bold=True, color="000000")
font_white_bold = Font(bold=True, color="FFFFFF")
cf_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
cf_pink_fill = PatternFill(start_color="F4B6C2", end_color="F4B6C2", fill_type="solid")
cf_yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
cf_purple_fill = PatternFill(start_color="C9B2FF", end_color="C9B2FF", fill_type="solid")
cf_green_fill = PatternFill(start_color="C1F0C8", end_color="C1F0C8", fill_type="solid")

from ..utils.paths import MASTER_DATA, MASTER_ARCHIVE, LEDGER_DATA, LEDGER_ARCHIVE, archive_old_files
ROOT_DIR = str(MASTER_DATA)
DATA_DIR = str(LEDGER_DATA)
ARCHIVE_DIR = str(LEDGER_ARCHIVE)

def get_latest_master_auto(directory):
    pattern = os.path.join(directory, "MasterData_*.csv")
    files = glob.glob(pattern)
    # Also search archive/ for MasterData CSVs
    archive_pattern = os.path.join(str(MASTER_ARCHIVE), "MasterData_*.csv")
    files += glob.glob(archive_pattern)
    if not files: return None
    return max(files, key=os.path.getmtime)

data_hoje = datetime.now().strftime("%d.%m")
OUTPUT_FILENAME = f"Ledger {data_hoje}.xlsx"
OUTPUT_PATH = os.path.join(DATA_DIR, OUTPUT_FILENAME)

FINAL_HEADERS = [
    "Country", "System", "Vendor category", "Company Code", "Supplier", "Name 1", "Document Date",
    "Document Number", "Reference", "Amount in doc. curr.", "Document Type", "Net due date",
    "Document currency", "Posting Date", "Payment Block", "0-30 Days overdue", "31-60 Days overdue",
    "61-90 Days overdue", "91-120 Days Overdue", "121-180 Days Overdue", "180> Days Overdue",
    "TOTAL VALUE", "TOTAL VOL", "Total Value Over 90", "Query type", "Status",
    "AP Specialist comment", "Next Step", "Action Date", "TL Comment", "Review Date",
    "Open Payment", "Sheet",
    "Previous Owner", "Owner", "Text", "Unique Ref", "User name",
    "Payment Issues", "Review this week", "Complete",
]
TAIL_COLS = ["Query type", "Status", "AP Specialist comment", "Next Step", "Action Date", "TL Comment", "Review Date"]
REVIEW_THIS_WEEK_OPTIONS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
COMPLETE_OPTIONS = ["Yes", "No"]
OPCOES_QUERY = ["Awaiting Payment - BAU", "Awaiting Payment - Overdues", "Blocked", "Posting Error", "Missing Documents", "Incoming Cash", "Credit Note", "Missing payment"]
OPCOES_STATUS = ["Awaiting PR", "Awaiting DD", "In Progress", "Documents Requested", "Documation - Awaiting Posting", "Documation - Awaiting Approval", "Awaiting Manual Posting", "Resolved", "Disputed", "Blocker"]
COLS_GREEN = []
COLS_RED = []
COLS_TOTAL = ["TOTAL VALUE", "TOTAL VOL", "Total Value Over 90"]

DEBUG_TAIL_FALLBACK = False
TAIL_FALLBACK_PRINT_LIMIT = 30

_re_intercompany = re.compile(r"inter\s*company|intercompany", re.IGNORECASE)

BUCKETS = [("0-30", 0, 30), ("31-60", 30, 60), ("61-90", 60, 90), ("91-120", 90, 120), ("121-180", 120, 180), ("180+", 180, None)]
TEAM_FOCUS = {"ROL": ("90+", 90, None)}

#  pure functions 

def get_bucket_rank(days_overdue):
    if pd.isna(days_overdue) or days_overdue < 0: return 0
    if days_overdue > 180: return 6
    if days_overdue >= 121: return 5
    if days_overdue >= 91: return 4
    if days_overdue >= 61: return 3
    if days_overdue >= 31: return 2
    if days_overdue >= 0: return 1
    return 0

def find_latest_file(directory):
    search_pattern = os.path.join(directory, "Ledger *.xlsx")
    files = glob.glob(search_pattern)
    # Also search archive/ for previous week files
    archive_pattern = os.path.join(ARCHIVE_DIR, "Ledger *.xlsx")
    files += glob.glob(archive_pattern)
    valid_files = [f for f in files if os.path.basename(f) != OUTPUT_FILENAME and not os.path.basename(f).startswith("~$")]
    if not valid_files: return None
    return max(valid_files, key=extract_date_from_filename)

def _merge_tail(old: dict, new: dict) -> dict:
    """Merge two tail dictionaries, preserving existing non-empty values."""
    merged = dict(old)
    for k, v in new.items():
        if k.startswith("_"):
            if k not in merged and v: merged[k] = v
            continue
        if v and not merged.get(k): merged[k] = v
    return merged

def _store_tail(tail_db: dict, key: str, row_tail: dict) -> None:
    """Store or merge a tail entry in the tail database."""
    if key in tail_db: tail_db[key] = _merge_tail(tail_db[key], row_tail)
    else: tail_db[key] = row_tail

def _process_sheet_for_tail(df_old: pd.DataFrame, sheet_name: str, tail_db: dict) -> None:
    """Process a single sheet and extract tail data into tail_db."""
    if "Unique Ref" not in df_old.columns and "Document Number" not in df_old.columns:
        return

    col_names = list(df_old.columns)
    for row in df_old.itertuples(index=False):
        row_tail = {}
        has_data = False
        # dict(zip()) preserves original column names (spaces, special chars)
        row_dict = dict(zip(col_names, row))

        for c in TAIL_COLS:
            if c in row_dict:
                v = row_dict.get(c)
                if v and str(v).strip() != "" and str(v).lower() != "nan":
                    row_tail[c] = v
                    has_data = True

        if not has_data:
            continue

        row_tail["_PrevOwner"] = clean_key(row_dict.get("Owner", "")) or sheet_name
        ur_val = clean_key(row_dict.get("Unique Ref", ""))
        doc_val = clean_key(row_dict.get("Document Number", ""))
        cc_val = clean_key(row_dict.get("Company Code", ""))
        sup_val = clean_key(row_dict.get("Supplier", ""))

        if ur_val:
            _store_tail(tail_db, f"{sheet_name}||UR||{ur_val}", row_tail)
            _store_tail(tail_db, f"UR||{ur_val}", row_tail)

        if doc_val and cc_val and sup_val:
            spec = f"{sheet_name}||DOC||{cc_val}||{sup_val}||{doc_val}"
            glob_k = f"DOC||{cc_val}||{sup_val}||{doc_val}"
            _store_tail(tail_db, spec, row_tail)
            _store_tail(tail_db, glob_k, row_tail)
            if ur_val:
                _store_tail(tail_db, f"{sheet_name}||DOCUR||{cc_val}||{sup_val}||{doc_val}||{ur_val}", row_tail)
                _store_tail(tail_db, f"DOCUR||{cc_val}||{sup_val}||{doc_val}||{ur_val}", row_tail)

def capture_tail_smart(directory):
    """Capture tail data from the latest previous report file."""
    tail_db = {}
    latest_file = find_latest_file(directory)
    if not latest_file:
        return tail_db

    try:
        xls = pd.ExcelFile(latest_file)
    except Exception:
        return tail_db

    log.info(f"    -> Lendo historico de: {os.path.basename(latest_file)}")

    for sheet_name in xls.sheet_names:
        if "Snapshot" in sheet_name or "MasterData" in sheet_name or "Summary" in sheet_name:
            continue
        try:
            df_old = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
            df_old.columns = [str(c).strip() for c in df_old.columns]
            _process_sheet_for_tail(df_old, sheet_name, tail_db)
        except Exception:
            pass

    xls.close()
    return tail_db

#  extracted functions (were interleaved in execution zone) 

def _find_previous_output(data_dir, prefix, exclude_path):
    search_pattern = os.path.join(data_dir, f"{prefix}*.xlsx")
    files = glob.glob(search_pattern)
    # Also search archive/ for previous week files
    archive_pattern = os.path.join(ARCHIVE_DIR, f"{prefix}*.xlsx")
    files += glob.glob(archive_pattern)
    valid_files = [f for f in files if os.path.basename(f) != os.path.basename(exclude_path) and not os.path.basename(f).startswith("~$")]
    if not valid_files: return None
    return max(valid_files, key=extract_date_from_filename)

def _read_previous_summary_cw(path, sheet_name):
    """Read CW values from previous report's Owner Weekly Dashboard.
    Returns dict: {owner: {metric: value}} where metric names match _compute_owner_metrics columns.
    """
    from openpyxl import load_workbook
    try:
        wb_prev = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    if sheet_name not in wb_prev.sheetnames:
        wb_prev.close()
        return {}
    ws = wb_prev[sheet_name]
    all_rows = []
    for row in ws.iter_rows():
        all_rows.append([cell.value for cell in row])
    wb_prev.close()

    # Find "Owner Weekly Dashboard" header row
    dash_header_idx = None
    for i, row in enumerate(all_rows):
        cell_val = str(row[0] or "").strip()
        if "Owner Weekly Dashboard" in cell_val:
            dash_header_idx = i + 1
            break
    if dash_header_idx is None or dash_header_idx >= len(all_rows):
        return {}

    headers = [str(v or "").strip() for v in all_rows[dash_header_idx]]

    # Map CW columns: "Total Docs CW" -> col index, metric name "Total Docs"
    cw_col_map = {}
    for j, h in enumerate(headers):
        if " CW" in h:
            metric = h.replace(" CW", "").strip()
            cw_col_map[metric] = j

    # Read owner data rows
    result = {}
    for i in range(dash_header_idx + 1, len(all_rows)):
        row = all_rows[i]
        owner = str(row[0] or "").strip()
        if not owner:
            break
        result[owner] = {}
        for metric, j in cw_col_map.items():
            val = row[j] if j < len(row) else 0
            try: val = float(val or 0)
            except (ValueError, TypeError): val = 0.0
            # Map "Total Value X" -> "Value X" to match _compute_owner_metrics columns
            if metric.startswith("Total Value "):
                metric = metric.replace("Total Value ", "Value ")
            result[owner][metric] = val
    return result

def _load_previous_detail(path):
    try: xls = pd.ExcelFile(path)
    except Exception: return pd.DataFrame()
    dfs = []
    for sn in xls.sheet_names:
        if sn.lower() in {"summary", "summary_query", "summary_rol", "weekly_movement", "payment issues", "statements", "synthetic_review duplicates", "synthetic_review errors", "zr blocks"} or sn.startswith("_"): continue
        try: dfs.append(pd.read_excel(xls, sheet_name=sn, dtype=str))
        except Exception: continue
    xls.close()
    if not dfs: return pd.DataFrame()
    prev = pd.concat(dfs, ignore_index=True)
    prev.columns = [str(c).strip() for c in prev.columns]
    prev = prev.loc[:, ~prev.columns.duplicated()].copy()
    if "Amount_Num" not in prev.columns and "Amount in doc. curr." in prev.columns:
        prev["Amount_Num"] = pd.to_numeric(prev["Amount in doc. curr."], errors="coerce")
    prev["_Amt"] = pd.to_numeric(prev.get("Amount_Num", 0), errors="coerce").fillna(0)
    snapshot_date = _extract_snapshot_date(path)
    if "Net due date" in prev.columns:
        prev["Net due date"] = parse_mixed_date_series(prev["Net due date"])
        prev["_Days"] = (pd.Timestamp(snapshot_date) - prev["Net due date"]).dt.days.fillna(-999)
    else: prev["_Days"] = -999
    # BUG 0 FIX: Override _Days using pre-computed bucket columns from Excel
    # (handles NaT Net due dates that occur when reading Excel with dtype=str)
    for _bcol, _bdays in [
        ("0-30 Days overdue", 1), ("31-60 Days overdue", 31),
        ("61-90 Days overdue", 61), ("91-120 Days Overdue", 91),
        ("121-180 Days Overdue", 121), ("180> Days Overdue", 181),
    ]:
        if _bcol in prev.columns:
            _bv = pd.to_numeric(prev[_bcol], errors='coerce').fillna(0)
            prev.loc[_bv != 0, '_Days'] = float(_bdays)
    prev = prev.copy()
    _doc = prev.get("Document Number")
    if _doc is None: prev["_IsDetail"] = True
    else: prev["_IsDetail"] = _doc.fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True).ne("")
    if "Country" in prev.columns: prev["Country"] = _norm_str_series(prev["Country"]).str.upper()
    if "Owner" in prev.columns:
        prev["Owner"] = _norm_str_series(prev["Owner"])
        prev = prev[~prev["Owner"].str.upper().eq("FUEL")].copy()
        prev = prev[~prev["Owner"].str.upper().eq("UNASSIGNED")].copy()
        prev = prev[prev["Owner"].astype(str).str.strip() != ""].copy()
    ic_mask = pd.Series(False, index=prev.index)
    for col in ["Vendor category", "Name 1", "Text", "Supplier", "Unique Ref"]:
        if col in prev.columns:
            ic_mask |= _norm_str_series(prev[col]).str.contains(r"inter\s*company|intercompany", case=False, regex=True)
    prev = prev[~ic_mask].copy()
    return prev

def _compute_owner_metrics(base, buckets):
    if base.empty: return pd.DataFrame()
    d = base.copy()
    d["_Owner"] = _norm_str_series(d["Owner"] if "Owner" in d.columns else pd.Series("", index=d.index))
    d["_Key"] = _supplier_key(d)
    out = {}
    for owner in sorted(d["_Owner"].unique().tolist()):
        g = d[d["_Owner"] == owner]
        row = {}
        for label, low, high in buckets:
            m = _bucket_mask(g, low, high)
            row[f"Suppliers {label}"] = int(g.loc[m, "_Key"].replace({"": np.nan}).nunique(dropna=True))
            row[f"Docs {label}"] = int(m.sum())
            row[f"Value {label}"] = float(g.loc[m, "_Amt"].sum())
        out[owner] = row
    return pd.DataFrame.from_dict(out, orient="index").fillna(0)

def _focus_suppliers(base, low, high):
    if base.empty: return pd.DataFrame(columns=["Key","Supplier","Name","Owner","Country","Docs","Value","StatusHint","PayHint"])
    d = base.copy()
    d["_Key"] = _supplier_key(d)
    own_s = d["Owner"] if "Owner" in d.columns else pd.Series("", index=d.index)
    d["Owner"] = _norm_str_series(own_s)
    m = _bucket_mask(d, low, high)
    dd = d[m].copy()
    if dd.empty: return pd.DataFrame(columns=["Key","Supplier","Name","Owner","Country","Docs","Value","StatusHint","PayHint"])
    # CORRECAO AQUI: Garante que e Series ou trata vazio
    s_col = dd["Status"] if "Status" in dd.columns else pd.Series([""]*len(dd), index=dd.index)
    p_col = dd["Open Payment"] if "Open Payment" in dd.columns else pd.Series([""]*len(dd), index=dd.index)
    dd["StatusHint"] = _norm_str_series(s_col)
    dd["PayHint"] = _norm_str_series(p_col)
    g = dd.groupby("_Key", dropna=False)
    res = g.apply(lambda x: pd.Series({
        "Supplier": str(x["Supplier"].iloc[0]), "Name": str(x["Name 1"].iloc[0]), "Owner": str(x["Owner"].iloc[0]),
        "Country": str(x.get("Country", "UNKNOWN").iloc[0]), "Docs": int(len(x)), "Value": float(x["_Amt"].sum()),
        "StatusHint": ",".join(set([v for v in x["StatusHint"] if v]))[:80],
        "PayHint": ",".join(set([v for v in x["PayHint"] if v]))[:80]
    }), include_groups=False).reset_index().rename(columns={"_Key": "Key"})
    return res

def _team_df(base, team):
    if base.empty: return base.copy()
    s = _norm_str_series(base["Sheet"]).str.upper()
    return base[s.eq(team)].copy()


def apply_ledger_owner_tab_highlights(ws, first_row: int, last_row: int) -> None:
    """Apply Ledger-only owner-tab conditional highlights."""
    if last_row < first_row:
        return

    col_total_letter = get_column_letter(FINAL_HEADERS.index("TOTAL VALUE") + 1)
    col_sheet_letter = get_column_letter(FINAL_HEADERS.index("Sheet") + 1)
    col_status_letter = get_column_letter(FINAL_HEADERS.index("Status") + 1)
    not_resolved = f'LOWER(TRIM(${col_status_letter}{first_row}))<>"resolved"'
    rng_total = f"{col_total_letter}{first_row}:{col_total_letter}{last_row}"
    formula_rol_purple = (
        f'=AND(UPPER(TRIM(${col_sheet_letter}{first_row}))="ROL", '
        f'${col_total_letter}{first_row}>0, {not_resolved})'
    )
    ws.conditional_formatting.add(
        rng_total,
        FormulaRule(formula=[formula_rol_purple], fill=cf_purple_fill, stopIfTrue=True),
    )

    col_letter = get_column_letter(FINAL_HEADERS.index("Total Value Over 90") + 1)
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    formula_rol_over_90 = (
        f'=AND(UPPER(TRIM(${col_sheet_letter}{first_row}))="ROL", '
        f'${col_letter}{first_row}<>0, {not_resolved})'
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[formula_rol_over_90], fill=cf_pink_fill),
    )

#  main execution 

def main():
    INPUT_CSV = get_latest_master_auto(ROOT_DIR)

    tail_fallback_total_ur = 0
    tail_fallback_total_doc = 0
    tail_fallback_printed = 0

    if not INPUT_CSV:
        log.error("[ERRO FATAL] Nenhum arquivo 'MasterData_*.csv' encontrado.")
        sys.exit(1)

    log.info(f"    -> Carregando CSV: {os.path.basename(INPUT_CSV)}...")
    try: df = pd.read_csv(INPUT_CSV, dtype=str, encoding="utf-8", on_bad_lines="skip")
    except (UnicodeDecodeError, pd.errors.ParserError): df = pd.read_csv(INPUT_CSV, dtype=str, encoding="latin1", on_bad_lines="skip")

    df.columns = [str(c).strip() for c in df.columns]

    if "Sheet" in df.columns:
        allowed = {"query", "rol", "uncathegorised", "uncategorised"}
        df = df[df["Sheet"].str.casefold().isin(allowed)].copy()
        # Normalize legacy "Query" rows to "ROL" (Query team merged into ROL)
        df.loc[df["Sheet"].str.casefold() == "query", "Sheet"] = "ROL"

    df["Amount_Num"] = parse_amount_series(df.get("Amount in doc. curr.", pd.Series([""]*len(df))))
    for c in ["Net due date", "Document Date", "Posting Date", "Action Date", "Review Date"]:
        if c in df.columns: df[c] = parse_mixed_date_series(df[c])

    if "Company Code" in df.columns: df["System"] = df["Company Code"].apply(system_from_code)
    if "Unique Ref" not in df.columns:
        sup_col = "Supplier" if "Supplier" in df.columns else "Vendor Option"
        df["Unique Ref"] = df.get("Company Code", "").fillna("") + " " + df.get(sup_col, "").fillna("")
    if "Payment Issues" not in df.columns:
        if "Payment issue" in df.columns: df["Payment Issues"] = df["Payment issue"]
        else: df["Payment Issues"] = ""

    tail_history = capture_tail_smart(DATA_DIR)

    log.info("    -> Calculando dias de atraso e KPIs...")
    today_ts = pd.Timestamp.now()
    df["_Days"] = (today_ts - df["Net due date"]).dt.days.fillna(-999) if "Net due date" in df.columns else -999
    df["_Rank"] = df["_Days"].apply(get_bucket_rank)
    df["_Amt"] = df["Amount_Num"].fillna(0)
    _doc = df.get("Document Number")
    if _doc is None: df["_IsDetail"] = True
    else: df["_IsDetail"] = _doc.fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True).ne("")

    if "Country" in df.columns: df["Country"] = _norm_str_series(df["Country"]).str.upper()
    if "Owner" in df.columns: df["Owner"] = _norm_str_series(df["Owner"])

    if "Owner" in df.columns:
        df = df[~df["Owner"].str.upper().eq("FUEL")].copy()
        df = df[~df["Owner"].str.upper().eq("UNASSIGNED")].copy()
        df = df[df["Owner"].astype(str).str.strip() != ""].copy()

    ic_mask = pd.Series(False, index=df.index)
    for col in ["Vendor category", "Name 1", "Text", "Supplier", "Unique Ref"]:
        if col in df.columns:
            ic_mask |= _norm_str_series(df[col]).str.contains(_re_intercompany)
    df = df[~ic_mask].copy()

    df["_121_180_Amt_NET"] = np.where(df["_IsDetail"] & (df["_Days"] > 120) & (df["_Days"] <= 180), df["_Amt"], 0)
    df["_Over180_Amt_NET"] = np.where(df["_IsDetail"] & (df["_Days"] > 180), df["_Amt"], 0)

    log.info("    -> Gerando Excel final...")
    wb = Workbook()
    wb.remove(wb.active)

    log.info("      ...Criando aba Summary")
    summary_df = (
        df.dropna(subset=["Owner"])
        .groupby("Owner", dropna=True)
        .pipe(lambda gb: _groupby_apply_no_warning(gb, lambda g: pd.Series({
            "Total Suppliers": int(g[g["_IsDetail"] & (g["_Amt"]!=0)]["Unique Ref"].nunique()),
            "Total Documents": int(len(g[g["_IsDetail"] & (g["_Amt"]!=0)])),
            "Over120 Amount": float(g["_121_180_Amt_NET"].sum()),
            "Over180 Amount": float(g["_Over180_Amt_NET"].sum()),
        }))).reset_index().sort_values(by="Over180 Amount", ascending=False)
    )

    prev_path = _find_previous_output(DATA_DIR, prefix="Ledger ", exclude_path=OUTPUT_PATH)
    if prev_path:
        log.info(f"    -> Comparando com arquivo anterior: {os.path.basename(prev_path)}")
    df_lw = _load_previous_detail(prev_path) if prev_path else pd.DataFrame()

    def _write_team_block(ws_target, team, start_row):
        focus_label, focus_low, focus_high = TEAM_FOCUS[team]
        d_tw = _team_df(df, team)
        d_lw = _team_df(df_lw, team) if not df_lw.empty else pd.DataFrame()
        owner_tw = _compute_owner_metrics(d_tw, BUCKETS)

        # LW: Computado a partir dos dados brutos (BUG 0 fix garante bucket correto para datas NaT)
        owner_lw = _compute_owner_metrics(d_lw, BUCKETS) if not d_lw.empty else owner_tw.copy()*0

        # Aggregate focus columns (e.g. "90+" = sum of 91-120, 121-180, 180+)
        for om in (owner_tw, owner_lw):
            if om.empty: continue
            focus_buckets = [lbl for lbl, lo, _ in BUCKETS if lo is not None and lo >= focus_low]
            for metric in ("Suppliers", "Docs", "Value"):
                cols = [f"{metric} {lbl}" for lbl in focus_buckets if f"{metric} {lbl}" in om.columns]
                om[f"{metric} {focus_label}"] = om[cols].sum(axis=1) if cols else 0.0

        ws_target.cell(start_row, 1, f"Team {team} - KPI Overview ({focus_label})").font = Font(bold=True)
        for j,h in enumerate(["Metric","LW","CW","Delta"], start=1):
            c=ws_target.cell(start_row+1, j, h); c.fill=fill_header_custom; c.font=font_header_custom; c.alignment=Alignment(horizontal="center")

        def _kpi(m):
            if m.empty: return 0.0, 0.0, 0.0
            return float(m[f"Suppliers {focus_label}"].sum()) if f"Suppliers {focus_label}" in m.columns else 0.0, float(m[f"Docs {focus_label}"].sum()) if f"Docs {focus_label}" in m.columns else 0.0, float(m[f"Value {focus_label}"].sum()) if f"Value {focus_label}" in m.columns else 0.0

        sl, dl, vl = _kpi(owner_lw); st, dt, vt = _kpi(owner_tw)
        def _w(r, m, l, t):
            ws_target.cell(r,1,m); ws_target.cell(r,2,l); ws_target.cell(r,3,t); ws_target.cell(r,4,t-l)
            for cx in [2,3,4]: ws_target.cell(r, cx).number_format = "#,##0.00"

        _w(start_row+2, f"Suppliers {focus_label}", sl, st)
        _w(start_row+3, f"Docs {focus_label}", dl, dt)
        _w(start_row+4, f"Value {focus_label}", vl, vt)

        # Resolved KPIs
        tw_focus = _focus_suppliers(d_tw, focus_low, focus_high)
        lw_focus = _focus_suppliers(d_lw, focus_low, focus_high) if not d_lw.empty else pd.DataFrame(columns=tw_focus.columns)
        tw_keys = set(tw_focus["Key"].astype(str))
        lw_keys = set(lw_focus["Key"].astype(str))
        cleared = lw_focus[lw_focus["Key"].astype(str).isin(lw_keys - tw_keys)].copy()
        resolved_sup = int(len(cleared)); resolved_val = float(cleared["Value"].sum()) if not cleared.empty else 0.0
        ws_target.cell(start_row+2, 6, "Resolved Suppliers").fill = fill_header_custom; ws_target.cell(start_row+2, 6).font = font_header_custom
        ws_target.cell(start_row+2, 7, resolved_sup)
        ws_target.cell(start_row+3, 6, "Resolved Value (LW)").fill = fill_header_custom; ws_target.cell(start_row+3, 6).font = font_header_custom
        ws_target.cell(start_row+3, 7, resolved_val).number_format = "#,##0.00"

        # Owner Weekly Dashboard - Order: LW -> CW -> Delta
        owner_start = start_row + 6
        ws_target.cell(owner_start, 1, f"Owner Weekly Dashboard - Team {team}").font = Font(bold=True)

        def _ot(d):
            if d.empty: return pd.DataFrame(columns=["Total Docs", "Total Suppliers"])
            x = d.copy(); x["_Owner"] = _norm_str_series(x["Owner"] if "Owner" in x.columns else pd.Series("", index=x.index)); x["_Key"] = _supplier_key(x)
            m = x["_IsDetail"] & (x["_Amt"].fillna(0)!=0)
            if not m.any(): return pd.DataFrame(columns=["Total Docs", "Total Suppliers"])
            return x.loc[m].groupby("_Owner").apply(lambda z: pd.Series({"Total Docs": int(len(z)), "Total Suppliers": int(z["_Key"].nunique())}), include_groups=False)

        tt = _ot(d_tw)
        owner_tw = owner_tw.join(tt, how="left").fillna(0)
        tl = _ot(d_lw)
        owner_lw = owner_lw.join(tl, how="left").fillna(0)

        headers = ["Owner", "Total Docs LW", "Total Docs CW", "Delta Total Docs", "Total Suppliers LW", "Total Suppliers CW", "Delta Total Suppliers", f"Docs {focus_label} LW", f"Docs {focus_label} CW", f"Delta Docs {focus_label}"]
        for label,_,_ in BUCKETS: headers += [f"Suppliers {label} LW", f"Suppliers {label} CW", f"Delta Suppliers {label}", f"Docs {label} LW", f"Docs {label} CW", f"Delta Docs {label}"]
        headers += [f"Total Value {focus_label} LW", f"Total Value {focus_label} CW"]

        for j, h in enumerate(headers, start=1):
            c=ws_target.cell(owner_start+1, j, h); c.fill=fill_header_custom; c.font=font_header_custom; c.alignment=Alignment(horizontal="center", wrap_text=True)

        # Filter summary to ROL team members only, including legacy/current ROL transition owners.
        _ROL_TEAM = (known_rol_owners() - {"No Owner"}) | {"Synthetic Owner 023"}
        _all_owners = set(owner_tw.index.tolist()) | set(owner_lw.index.tolist())
        owners_all = sorted(_all_owners & _ROL_TEAM)
        rr = owner_start + 2
        for owner in owners_all:
            ws_target.cell(rr, 1, owner)

            td_tw = float(owner_tw.loc[owner, "Total Docs"]) if owner in owner_tw.index else 0.0
            td_lw = float(owner_lw.loc[owner, "Total Docs"]) if owner in owner_lw.index else 0.0
            ts_tw = float(owner_tw.loc[owner, "Total Suppliers"]) if owner in owner_tw.index else 0.0
            ts_lw = float(owner_lw.loc[owner, "Total Suppliers"]) if owner in owner_lw.index else 0.0
            f_tw = float(owner_tw.loc[owner, f"Docs {focus_label}"]) if owner in owner_tw.index else 0.0
            f_lw = float(owner_lw.loc[owner, f"Docs {focus_label}"]) if owner in owner_lw.index else 0.0

            # LW -> CW -> Delta
            ws_target.cell(rr, 2, int(td_lw))
            ws_target.cell(rr, 3, int(td_tw))
            ws_target.cell(rr, 4, int(td_tw - td_lw))

            ws_target.cell(rr, 5, int(ts_lw))
            ws_target.cell(rr, 6, int(ts_tw))
            ws_target.cell(rr, 7, int(ts_tw - ts_lw))

            ws_target.cell(rr, 8, int(f_lw))
            ws_target.cell(rr, 9, int(f_tw))
            ws_target.cell(rr, 10, int(f_tw - f_lw))

            cc = 11
            for label,_,_ in BUCKETS:
                s_tw = float(owner_tw.loc[owner, f"Suppliers {label}"]) if owner in owner_tw.index else 0.0
                s_lw = float(owner_lw.loc[owner, f"Suppliers {label}"]) if owner in owner_lw.index else 0.0
                d_tw = float(owner_tw.loc[owner, f"Docs {label}"]) if owner in owner_tw.index else 0.0
                d_lw = float(owner_lw.loc[owner, f"Docs {label}"]) if owner in owner_lw.index else 0.0

                ws_target.cell(rr, cc, int(s_lw)); ws_target.cell(rr, cc+1, int(s_tw)); ws_target.cell(rr, cc+2, int(s_tw-s_lw))
                ws_target.cell(rr, cc+3, int(d_lw)); ws_target.cell(rr, cc+4, int(d_tw)); ws_target.cell(rr, cc+5, int(d_tw-d_lw))
                cc += 6

            v_tw = float(owner_tw.loc[owner, f"Value {focus_label}"]) if owner in owner_tw.index else 0.0
            v_lw = float(owner_lw.loc[owner, f"Value {focus_label}"]) if owner in owner_lw.index else 0.0
            ws_target.cell(rr, cc, v_lw).number_format = "#,##0.00"; ws_target.cell(rr, cc+1, v_tw).number_format = "#,##0.00"
            rr += 1

        last_row = rr - 1

        # Delta CF
        for c in range(1, len(headers)+1):
            h = str(ws_target.cell(owner_start+1, c).value or "")
            if h.startswith("Delta "):
                col_l = get_column_letter(c)
                rng = f"{col_l}{owner_start+2}:{col_l}{last_row}"
                ws_target.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], fill=cf_red_fill))
                ws_target.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], fill=fill_green_light))

        # SUMMARY CHARTS  2 graficos simples lado a lado
        if last_row >= owner_start + 2:
            _cats = Reference(ws_target, min_col=1, min_row=owner_start + 2, max_row=last_row)

            # Chart 1: Docs {focus_label} LW vs CW (KPI principal)  vermelho/laranja
            ch1 = BarChart()
            ch1.type = "col"; ch1.grouping = "clustered"; ch1.style = 10
            ch1.title = f"Docs {focus_label}  LW vs CW"; ch1.y_axis.title = "Docs"
            ch1.x_axis.title = "Owner"; ch1.legend.position = "b"
            ch1.width = 20; ch1.height = 14; ch1.gapWidth = 100
            _d1 = Reference(ws_target, min_col=8, max_col=9, min_row=owner_start + 1, max_row=last_row)
            ch1.add_data(_d1, titles_from_data=True); ch1.set_categories(_cats)
            for _i, _clr in enumerate(["C00000", "ED7D31"]):
                if _i < len(ch1.series):
                    ch1.series[_i].graphicalProperties.solidFill = _clr
                    ch1.series[_i].dLbls = DataLabelList()
                    ch1.series[_i].dLbls.showVal = True; ch1.series[_i].dLbls.numFmt = '#,##0'
            ws_target.add_chart(ch1, "L4")

            # Chart 2: Total Docs LW vs CW  azul escuro/claro
            ch2 = BarChart()
            ch2.type = "col"; ch2.grouping = "clustered"; ch2.style = 10
            ch2.title = "Total Docs  LW vs CW"; ch2.y_axis.title = "Docs"
            ch2.x_axis.title = "Owner"; ch2.legend.position = "b"
            ch2.width = 20; ch2.height = 14; ch2.gapWidth = 100
            _d2 = Reference(ws_target, min_col=2, max_col=3, min_row=owner_start + 1, max_row=last_row)
            ch2.add_data(_d2, titles_from_data=True); ch2.set_categories(_cats)
            for _i, _clr in enumerate(["1F4E79", "5B9BD5"]):
                if _i < len(ch2.series):
                    ch2.series[_i].graphicalProperties.solidFill = _clr
                    ch2.series[_i].dLbls = DataLabelList()
                    ch2.series[_i].dLbls.showVal = True; ch2.series[_i].dLbls.numFmt = '#,##0'
            ws_target.add_chart(ch2, "V4")

        # Top 10 Owners - Order: LW -> CW -> Delta
        top_start = last_row + 2
        ws_target.cell(top_start, 1, f"Top 10 Owners by Delta Docs ({focus_label}) - Team ({team})").font = Font(bold=True)
        for j,h in enumerate(["Owner","Docs LW","Docs CW","Delta"], start=1):
            cell = ws_target.cell(top_start+1, j, h)
            cell.fill = fill_header_custom
            cell.font = font_header_custom
        tmp = []
        for owner in owners_all:
            _dtw = float(owner_tw.loc[owner, f"Docs {focus_label}"]) if owner in owner_tw.index else 0.0
            _dlw = float(owner_lw.loc[owner, f"Docs {focus_label}"]) if owner in owner_lw.index else 0.0
            tmp.append((owner, _dtw, _dlw, _dtw - _dlw))
        tmp.sort(key=lambda x: x[3], reverse=True)
        rr2 = top_start + 2
        for owner, _dtw, _dlw, dd in tmp[:10]:
            ws_target.cell(rr2, 1, owner)
            ws_target.cell(rr2, 2, int(_dlw))
            ws_target.cell(rr2, 3, int(_dtw))
            ws_target.cell(rr2, 4, int(dd))
            rr2 += 1

        # === QUERY TYPE BREAKDOWN  formulas COUNTIF dinamicas por aba de owner (team) ===
        _qt_cl = get_column_letter(FINAL_HEADERS.index("Query type") + 1)
        _qt_tabs2 = [str(o) for o in _team_df(df, team)["Owner"].dropna().unique()]

        def _qt_formula_ledger(qt_val):
            parts = [f"COUNTIF('{t}'!{_qt_cl}:{_qt_cl},{chr(34)}{qt_val}{chr(34)})"
                     for t in _qt_tabs2]
            return "=" + "+".join(parts) if parts else "=0"

        qt_start2 = rr2 + 1
        ws_target.cell(qt_start2, 1, "Query Type Breakdown (Docs - Row Level)").font = Font(bold=True)
        for j, h in enumerate(["Query Type", "Count"], 1):
            c = ws_target.cell(qt_start2 + 1, j, h)
            c.fill = fill_header_custom; c.font = font_header_custom; c.alignment = Alignment(horizontal="center")

        qt_row2 = qt_start2 + 2
        _qt_data_start2 = qt_row2
        for qt in OPCOES_QUERY:  # lista fixa, sem (Blank)
            ws_target.cell(qt_row2, 1, qt)
            _c2 = ws_target.cell(qt_row2, 2, _qt_formula_ledger(qt))
            _c2.number_format = "#,##0"
            qt_row2 += 1

        _total_range2 = f"B{_qt_data_start2}:B{qt_row2 - 1}"
        _tc2a = ws_target.cell(qt_row2, 1, "TOTAL"); _tc2a.font = Font(bold=True)
        _tc2b = ws_target.cell(qt_row2, 2, f"=SUM({_total_range2})"); _tc2b.font = Font(bold=True); _tc2b.number_format = "#,##0"
        qt_row2 += 1

        return qt_row2 + 2

    ws_sum = wb.create_sheet(title="Summary_ROL", index=1)
    ws_sum.merge_cells("A1:K1"); ws_sum["A1"]="Comparative Weekly Dashboard (Ledger Report) - ROL"; ws_sum["A1"].font=Font(bold=True,size=14); ws_sum["A1"].alignment=Alignment(horizontal="center"); ws_sum.freeze_panes="A3"
    _write_team_block(ws_sum, "ROL", 3)

    # Payment Issues
    ws_pay = wb.create_sheet(title="Payment Issues")
    ws_pay["A1"] = "Payment Issues Analysis (LW vs CW)"; ws_pay["A1"].font = Font(bold=True, size=12)

    pi_tw = _agg_pi(df); pi_lw = _agg_pi(df_lw) if not df_lw.empty else pd.DataFrame(columns=pi_tw.columns)
    merged_pi = pd.merge(pi_tw, pi_lw, on=["Issue","Owner","Company Code","Vendor","Name"], how="outer", suffixes=("_CW","_LW")).fillna(0)
    merged_pi["Qty_CW"] = merged_pi["Qty_CW"].astype(int)
    merged_pi["Qty_LW"] = merged_pi["Qty_LW"].astype(int)
    merged_pi["Delta"] = merged_pi["Qty_CW"] - merged_pi["Qty_LW"]
    merged_pi["Growth"] = merged_pi.apply(lambda r: (r["Qty_CW"]-r["Qty_LW"])/r["Qty_LW"] if r["Qty_LW"]!=0 else (1.0 if r["Qty_CW"]>0 else 0.0), axis=1)
    merged_pi = merged_pi.sort_values(by="Qty_CW", ascending=False)

    h_pi = ["Type of payment issue", "Owner", "Company Code", "Vendor Number", "Vendor Name", "Qty Issue LW", "Qty Issue CW", "Delta", "Growth %"]
    for c, h in enumerate(h_pi, 1):
        cell = ws_pay.cell(3, c, h); cell.fill = fill_header_custom; cell.font = font_header_custom

    curr_r = 4
    for row in merged_pi.itertuples(index=False):
        ws_pay.cell(curr_r, 1, row.Issue)
        ws_pay.cell(curr_r, 2, row.Owner)
        ws_pay.cell(curr_r, 3, row[2])  # "Company Code" (space in name)
        ws_pay.cell(curr_r, 4, row.Vendor)
        ws_pay.cell(curr_r, 5, row.Name)
        ws_pay.cell(curr_r, 6, row.Qty_LW)
        ws_pay.cell(curr_r, 7, row.Qty_CW)
        ws_pay.cell(curr_r, 8, row.Delta)
        ws_pay.cell(curr_r, 9, row.Growth).number_format = '0.0%'
        curr_r += 1

    if curr_r > 4:
        delta_range = f"H4:H{curr_r-1}"
        ws_pay.conditional_formatting.add(delta_range, CellIsRule(operator='greaterThan', formula=['0'], fill=cf_red_fill))
        ws_pay.conditional_formatting.add(delta_range, CellIsRule(operator='lessThan', formula=['0'], fill=fill_green_light))

        # Audit Table
        acol = 13
        ws_pay.cell(3, acol, "Audit Summary").font = Font(bold=True)
        ws_pay.cell(4, acol, "Issues Increased"); ws_pay.cell(4, acol+1, len(merged_pi[merged_pi["Delta"] > 0]))
        ws_pay.cell(5, acol, "Issues Decreased"); ws_pay.cell(5, acol+1, len(merged_pi[merged_pi["Delta"] < 0]))
        ws_pay.cell(6, acol, "Net Volume Change"); ws_pay.cell(6, acol+1, merged_pi["Delta"].sum())

        # BAR CHART: ISSUES BY OWNER (LW vs CW)
        owner_agg = merged_pi.groupby("Owner")[["Qty_LW", "Qty_CW"]].sum().reset_index().sort_values(by="Qty_CW", ascending=False)
        # Filter Unassigned
        owner_agg["Owner"] = owner_agg["Owner"].fillna("").astype(str).str.strip()
        owner_agg = owner_agg[~owner_agg["Owner"].str.upper().isin(["UNASSIGNED", "FUEL", ""])]
        if not owner_agg.empty:
            ch_row = 10
            ws_pay.cell(ch_row, acol, "Owner"); ws_pay.cell(ch_row, acol+1, "LW"); ws_pay.cell(ch_row, acol+2, "CW")
            for r in owner_agg.itertuples(index=False):
                ch_row += 1
                ws_pay.cell(ch_row, acol, r.Owner); ws_pay.cell(ch_row, acol+1, r.Qty_LW); ws_pay.cell(ch_row, acol+2, r.Qty_CW)

            bc = BarChart()
            bc.title = "Payment Issues by Owner (LW vs CW)"
            bc.style = 10
            bc.y_axis.title = "Volume"
            bc.width = 22
            bc.height = 13
            bc.gapWidth = 120
            bc.overlap = -10
            cats = Reference(ws_pay, min_col=acol, min_row=11, max_row=ch_row)
            data = Reference(ws_pay, min_col=acol+1, max_col=acol+2, min_row=10, max_row=ch_row)
            bc.add_data(data, titles_from_data=True); bc.set_categories(cats)
            bc.legend.position = "b"
            # Cores + data labels: cinza (LW) e azul corporativo (CW)
            _pi_colors = ["A6A6A6", "1F4E79"]
            for i, color in enumerate(_pi_colors):
                if i < len(bc.series):
                    s = bc.series[i]
                    s.graphicalProperties.solidFill = color
                    s.dLbls = DataLabelList()
                    s.dLbls.showVal = True
                    s.dLbls.showCatName = False
                    s.dLbls.showSerName = False
                    s.dLbls.numFmt = '#,##0'
            ws_pay.add_chart(bc, "P3")

    ws_pay.column_dimensions["A"].width = 50 # Wider for full message
    ws_pay.column_dimensions["C"].width = 12
    ws_pay.column_dimensions["E"].width = 30

    if prev_path is None:
        ws_sum["A2"] = "LW not found: comparison skipped"
        ws_sum["A2"].font = Font(italic=True, color="666666")

    owners = df["Owner"].dropna().unique()
    for owner in owners:
        clean_name = str(owner)[:30].replace("/", " ").replace("\\", " ").replace("?", "").strip()
        if not clean_name:
            continue
        if clean_name.upper() == "FUEL":
            continue
        log.info(f"      Processando Owner: {clean_name}")
        ws = wb.create_sheet(title=clean_name)

        ws.sheet_view.showOutlineSymbols = True
        ws.sheet_properties.outlinePr.summaryBelow = False

        df_own = df[df["Owner"] == owner].copy()
        grouped = df_own.groupby("Unique Ref")

        # Header row
        ws.append(FINAL_HEADERS)
        for cell in ws[1]:
            col_name = str(cell.value).strip()
            if col_name in COLS_RED:
                cell.fill = fill_red_alert
                cell.font = font_white_bold
            elif col_name in COLS_TOTAL:
                cell.fill = fill_blue_total
                cell.font = font_black_bold
            elif col_name in COLS_GREEN:
                cell.fill = fill_green_light
                cell.font = font_black_bold
            else:
                cell.fill = fill_blue_light
                cell.font = font_black_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2

        for unique_ref, group in grouped:
            group = group.sort_values(by=["_Rank", "Net due date"], ascending=[False, True])

            total_vol = int(group["_IsDetail"].sum())
            total_val = group["Amount_Num"].sum()

            b_vals = {k: 0.0 for k in ["0-30", "31-60", "61-90", "91-120", "121-180", "180>"]}
            detail_rows = []
            first_row = group.iloc[0]
            has_open = False

            # Prepare Summary Line Issue Logic
            raw_issues = group["Payment Issues"].dropna().astype(str).str.strip()
            unique_issues = set(i for i in raw_issues if i)
            summary_issue_raw = list(unique_issues)[0] if len(unique_issues) == 1 else ""
            # DO NOT CLEAN NAME FOR SUMMARY LINE (FULL MESSAGE)
            summary_payment_issue = summary_issue_raw

            # Map column names to indices for itertuples access
            col_list = list(group.columns)
            amt_idx = col_list.index('Amount_Num')
            days_idx = col_list.index('_Days')
            doc_type_idx = col_list.index('Document Type') if 'Document Type' in col_list else None
            cc_idx = col_list.index("Company Code") if "Company Code" in col_list else None
            sup_idx = col_list.index("Supplier") if "Supplier" in col_list else None
            doc_idx = col_list.index("Document Number") if "Document Number" in col_list else None
            ur_idx = col_list.index("Unique Ref") if "Unique Ref" in col_list else None

            # Pre-compute FINAL_HEADERS indices
            header_indices = {}
            for h in FINAL_HEADERS:
                if h in col_list:
                    header_indices[h] = col_list.index(h)

            # Build detail rows
            for row in group.itertuples(index=False):
                amt = row[amt_idx] if pd.notna(row[amt_idx]) else 0.0
                days = row[days_idx]
                d_b = {k: None for k in b_vals.keys()}

                if days >= 0:
                    if days <= 30:
                        b_vals["0-30"] += amt
                        d_b["0-30"] = amt
                    elif days <= 60:
                        b_vals["31-60"] += amt
                        d_b["31-60"] = amt
                    elif days <= 90:
                        b_vals["61-90"] += amt
                        d_b["61-90"] = amt
                    elif days <= 120:
                        b_vals["91-120"] += amt
                        d_b["91-120"] = amt
                    elif days <= 180:
                        b_vals["121-180"] += amt
                        d_b["121-180"] = amt
                    else:
                        b_vals["180>"] += amt
                        d_b["180>"] = amt

                doc_type_val = str(row[doc_type_idx] if doc_type_idx is not None else "").upper().strip()
                if doc_type_val in ["KS", "SA", "DZ", "KZ", "ZP", "ZR", "ZE", "AB", "K1", "K5", "SE"]:
                    has_open = True

                d_map = {h: "" for h in FINAL_HEADERS}
                for h, idx in header_indices.items():
                    d_map[h] = sanitize_value(row[idx])

                # bucket columns
                d_map["0-30 Days overdue"] = d_b["0-30"]
                d_map["31-60 Days overdue"] = d_b["31-60"]
                d_map["61-90 Days overdue"] = d_b["61-90"]
                d_map["91-120 Days Overdue"] = d_b["91-120"]
                d_map["121-180 Days Overdue"] = d_b["121-180"]
                d_map["180> Days Overdue"] = d_b["180>"]
                d_map["Amount in doc. curr."] = amt
                d_map["Total Value Over 90"] = amt if (days > 90 and days >= 0) else None

                # Tail logic
                cc_k = clean_key(row[cc_idx] if cc_idx is not None else "")
                sup_k = clean_key(row[sup_idx] if sup_idx is not None else "")
                doc_k = clean_key(row[doc_idx] if doc_idx is not None else "")
                ur_k = clean_key(row[ur_idx] if ur_idx is not None else "")

                doc_key_specific = f"{clean_name}||DOC||{cc_k}||{sup_k}||{doc_k}"
                doc_key_global = f"DOC||{cc_k}||{sup_k}||{doc_k}"

                dup_specific = f"{clean_name}||DUPDOC||{cc_k}||{sup_k}||{doc_k}"
                dup_global = f"DUPDOC||{cc_k}||{sup_k}||{doc_k}"
                is_dup_doc = dup_specific in tail_history or dup_global in tail_history

                tail_row = None
                used_global = False

                if is_dup_doc and ur_k:
                    docur_specific = f"{clean_name}||DOCUR||{cc_k}||{sup_k}||{doc_k}||{ur_k}"
                    docur_global = f"DOCUR||{cc_k}||{sup_k}||{doc_k}||{ur_k}"
                    tail_row = tail_history.get(docur_specific)
                    if tail_row is None:
                        tail_row = tail_history.get(docur_global)
                        if tail_row is not None:
                            used_global = True

                if tail_row is None:
                    tail_row = tail_history.get(doc_key_specific)
                    if tail_row is None:
                        tail_row = tail_history.get(doc_key_global)
                        if tail_row is not None:
                            used_global = True

                if tail_row:
                    for t_c in TAIL_COLS:
                        if t_c in tail_row:
                            val = tail_row[t_c]
                            if t_c in {"Action Date", "Review Date"}:
                                val = recover_date_from_tail(val)
                            d_map[t_c] = val

                if used_global and tail_row:
                    prev_owner = str(tail_row.get("_PrevOwner", "")).strip()
                    if prev_owner and prev_owner != clean_name:
                        d_map["Previous Owner"] = prev_owner
                    tail_fallback_total_doc += 1

                detail_rows.append([d_map[h] for h in FINAL_HEADERS])

            # Summary row
            s_map = {h: "" for h in FINAL_HEADERS}
            for c in ["Country", "Vendor category", "Sheet", "Owner", "Company Code", "Document currency", "Supplier", "Name 1", "Unique Ref", "System"]:
                if c in first_row.index:
                    s_map[c] = sanitize_value(first_row[c])

            s_map["0-30 Days overdue"] = b_vals["0-30"]
            s_map["31-60 Days overdue"] = b_vals["31-60"]
            s_map["61-90 Days overdue"] = b_vals["61-90"]
            s_map["91-120 Days Overdue"] = b_vals["91-120"]
            s_map["121-180 Days Overdue"] = b_vals["121-180"]
            s_map["180> Days Overdue"] = b_vals["180>"]
            s_map["TOTAL VOL"] = total_vol
            s_map["TOTAL VALUE"] = total_val
            s_map["Total Value Over 90"] = b_vals["91-120"] + b_vals["121-180"] + b_vals["180>"]
            s_map["Open Payment"] = "Yes" if has_open else ""
            s_map["Payment Issues"] = summary_payment_issue

            # Tail by UR
            ur_val = clean_key(s_map.get("Unique Ref", ""))
            ur_key_specific = f"{clean_name}||UR||{ur_val}"
            ur_key_global = f"UR||{ur_val}"

            tail_row = tail_history.get(ur_key_specific)
            used_global = False
            if tail_row is None:
                tail_row = tail_history.get(ur_key_global)
                if tail_row is not None:
                    used_global = True

            if tail_row:
                for t_c in TAIL_COLS:
                    if t_c in tail_row:
                        val = tail_row[t_c]
                        if t_c in {"Action Date", "Review Date"}:
                            val = recover_date_from_tail(val)
                        s_map[t_c] = val

            if used_global and tail_row:
                prev_owner = str(tail_row.get("_PrevOwner", "")).strip()
                if prev_owner and prev_owner != clean_name:
                    s_map["Previous Owner"] = prev_owner
                tail_fallback_total_ur += 1

            ws.append([s_map[h] for h in FINAL_HEADERS])
            for cell in ws[row_idx]:
                cell.fill = fill_summary_gray
                cell.font = font_black_bold
                if isinstance(cell.value, (datetime, pd.Timestamp)):
                    cell.number_format = 'dd/mm/yyyy'

            grp_start = row_idx + 1
            row_idx += 1

            # Detail rows append
            for d_row in detail_rows:
                ws.append(d_row)

                for c_name in ["Net due date", "Document Date", "Posting Date", "Action Date", "Review Date"]:
                    if c_name in FINAL_HEADERS:
                        c_idx = FINAL_HEADERS.index(c_name) + 1
                        cell = ws.cell(row=row_idx, column=c_idx)
                        if isinstance(cell.value, (datetime, pd.Timestamp)):
                            cell.number_format = 'dd/mm/yyyy'

                for c_name in ["Document Number", "Reference", "Supplier", "Company Code"]:
                    if c_name in FINAL_HEADERS:
                        c_idx = FINAL_HEADERS.index(c_name) + 1
                        ws.cell(row=row_idx, column=c_idx).number_format = "@"

                row_idx += 1

            grp_end = row_idx - 1
            if grp_end >= grp_start:
                for r in range(grp_start, grp_end + 1):
                    ws.row_dimensions[r].outlineLevel = 1
                    ws.row_dimensions[r].hidden = True

        if row_idx > 2:
            dv_status = DataValidation(
                type="list",
                formula1=f"\"{','.join(OPCOES_STATUS)}\"",
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a value from the dropdown list only.",
                showInputMessage=True,
                promptTitle="Dropdown required",
                prompt="Select a value from the list.",
            )
            dv_query = DataValidation(
                type="list",
                formula1=f"\"{','.join(OPCOES_QUERY)}\"",
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a value from the dropdown list only.",
                showInputMessage=True,
                promptTitle="Dropdown required",
                prompt="Select a value from the list.",
            )
            dv_review_this_week = DataValidation(
                type="list",
                formula1=f"\"{','.join(REVIEW_THIS_WEEK_OPTIONS)}\"",
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select a weekday from the dropdown list only.",
            )
            dv_complete = DataValidation(
                type="list",
                formula1=f"\"{','.join(COMPLETE_OPTIONS)}\"",
                allow_blank=True,
                errorStyle="stop",
                showErrorMessage=True,
                errorTitle="Invalid selection",
                error="Select Yes or No from the dropdown list only.",
            )

            ws.add_data_validation(dv_status)
            ws.add_data_validation(dv_query)
            ws.add_data_validation(dv_review_this_week)
            ws.add_data_validation(dv_complete)

            col_status = get_column_letter(FINAL_HEADERS.index("Status") + 1)
            col_query = get_column_letter(FINAL_HEADERS.index("Query type") + 1)
            col_review_this_week = get_column_letter(FINAL_HEADERS.index("Review this week") + 1)
            col_complete = get_column_letter(FINAL_HEADERS.index("Complete") + 1)

            dv_status.add(f"{col_status}2:{col_status}{row_idx-1}")
            dv_query.add(f"{col_query}2:{col_query}{row_idx-1}")
            dv_review_this_week.add(f"{col_review_this_week}2:{col_review_this_week}{row_idx-1}")
            dv_complete.add(f"{col_complete}2:{col_complete}{row_idx-1}")

        num_cols = [
            "0-30 Days overdue",
            "31-60 Days overdue",
            "61-90 Days overdue",
            "91-120 Days Overdue",
            "121-180 Days Overdue",
            "180> Days Overdue",
            "TOTAL VALUE",
            "Total Value Over 90",
            "Amount in doc. curr.",
        ]
        for v in num_cols:
            col_l = get_column_letter(FINAL_HEADERS.index(v) + 1)
            for rr in range(2, row_idx):
                cell = ws[f"{col_l}{rr}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        if row_idx > 2:
            apply_ledger_owner_tab_highlights(ws, 2, row_idx - 1)

        status_col_letter = get_column_letter(FINAL_HEADERS.index("Status") + 1)
        last_col_letter = get_column_letter(len(FINAL_HEADERS))
        apply_range = f"A2:{last_col_letter}{row_idx-1}"
        formula_resolved = f'=LOWER(TRIM(${status_col_letter}2))="resolved"'
        ws.conditional_formatting.add(
            apply_range,
            FormulaRule(formula=[formula_resolved], fill=cf_green_fill, stopIfTrue=True),
        )

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A1:{last_col_letter}{row_idx-1}"

        # Final pass: force dd/mm/yyyy on all date columns (openpyxl default is mm-dd-yy)
        for c_name in ["Net due date", "Document Date", "Posting Date", "Action Date", "Review Date"]:
            if c_name in FINAL_HEADERS:
                c_idx = FINAL_HEADERS.index(c_name) + 1
                for rr in range(2, row_idx):
                    cell = ws.cell(row=rr, column=c_idx)
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = 'dd/mm/yyyy'

    log.info(">>> Salvando arquivo final...")

    lock_path = OUTPUT_PATH + ".lock"
    tmp_path = OUTPUT_PATH + ".tmp"

    if os.path.exists(lock_path):
        age_min = (time.time() - os.path.getmtime(lock_path)) / 60
        raise RuntimeError(
            f"Build lock exists (age {age_min:.0f}min): {lock_path}\n"
            f"Outro build rodando, ou o ultimo crashou. Apague o .lock se tiver certeza."
        )
    Path(lock_path).write_text(f"pid={os.getpid()} started={datetime.now().isoformat()}")

    try:
        while True:
            try:
                wb.save(tmp_path)
                break
            except PermissionError:
                log.error("\n" + "=" * 60)
                log.error(f" [ERRO] O ARQUIVO '{OUTPUT_FILENAME}.tmp' ESTA BLOQUEADO!")
                log.error(" Feche-o e aperte ENTER.")
                log.error("=" * 60 + "\n")
                input()
            except Exception as e:
                log.error(f">>> ERRO CRITICO AO SALVAR: {e}")
                try: os.remove(tmp_path)
                except OSError: pass
                raise

        try:
            validate_owner_tab_integrity(tmp_path)
        except Exception as e:
            log.error(f">>> OUTPUT INVALIDO, mantendo xlsx anterior intacto: {e}")
            log.error(f">>> Arquivo invalido preservado em: {tmp_path} (apague depois de investigar)")
            raise

        os.replace(tmp_path, OUTPUT_PATH)
        log.info(f">>> SUCESSO! Salvo em: {OUTPUT_PATH}")
    finally:
        try: os.remove(lock_path)
        except OSError: pass

    # Archive old Ledger xlsx files (keep only today's output)
    archived = archive_old_files(LEDGER_DATA, LEDGER_ARCHIVE, keep_pattern=data_hoje)
    if archived:
        log.info(f">>> [ARCHIVE] Moved {len(archived)} old file(s) to archive/")


if __name__ == "__main__":
    main()
