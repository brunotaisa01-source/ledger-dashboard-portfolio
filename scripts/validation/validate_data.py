#!/usr/bin/env python3
"""validate_data.py  Multi-stage validation for the AP Ledger pipeline.

Stage validations (read-only  never generates or modifies data):
    --stage masterdata   Validate MasterData CSV after creation
    --stage ledger-load  Validate SQLite after ETL load
    --stage key-build    Validate Key Excel report after generation
    --stage dashboard    Validate dashboard_data.js after generation
    --stage cross        Cross-stage consistency checks
    --all-stages         Run all 5 stage validations sequentially

Legacy validation (compares SQLite vs Excel Summary tabs):
    py validate_data.py               # validate last 2 snapshots
    py validate_data.py --weeks 3     # validate last 3 snapshots
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sqlite3
import sys
import zlib
from contextlib import closing, contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

#  Configuration 
from ..utils.paths import KEY_DATA, LEDGER_DATA, MASTER_DATA, DASHBOARD_DIR, LOGS_DIR
from ..utils.paths import KEY_DB, LEDGER_DB, KEY_ARCHIVE, LEDGER_ARCHIVE
from ..utils.masterdata_core import (
    MASTER_COLS,
    known_rol_owners,
    load_key_team_owners,
    required_rol_owners,
    parse_date_series,
)
from ..utils.synthetic_review_helpers import find_latest_masterdata
from ..dashboard.dashboard_config import JSON_LOCAL_DIR, JSON_MODE

# SQL layer  load query templates from sql/07_validation.sql when available
try:
    from ..utils.sql_loader import load_named_queries
    _VAL_QUERIES = load_named_queries("07_validation.sql")
except (ImportError, FileNotFoundError):
    _VAL_QUERIES = {}

KEY_OWNERS = load_key_team_owners(str(MASTER_DATA / "Owner_map.csv"))
ROL_REQUIRED_OWNERS = required_rol_owners()
ROL_OPTIONAL_OWNERS = {"No Owner", "Synthetic Owner 023"}
ROL_OWNERS = known_rol_owners()
ALL_OWNERS = (KEY_OWNERS | ROL_OWNERS
              | {"ROL Uncategorised"}
              | {"Fuel", "Rent", "Unassigned"}
              | {"Synthetic Owner 023"})


def _rol_required_owners_for_week(week_iso: str) -> set[str]:
    try:
        as_of = date.fromisoformat(str(week_iso)[:10])
    except (TypeError, ValueError):
        as_of = date.today()
    return required_rol_owners(as_of=as_of)

VALID_SHEETS = {"Key", "ROL"}

# Aging bucket columns in SQLite (ordered lowest -> highest)
KEY_BUCKETS = [
    ("07-30", "07-30 Days overdue (Unified)"),
    ("31-60", "31-60 Days overdue"),
    ("61-90", "61-90 Days overdue"),
    ("91-120", "91-120 Days Overdue"),
    ("121-180", "121-180 Days Overdue"),
    ("180+", "180> Days Overdue"),
]
LEDGER_BUCKETS = [
    ("0-30", "0-30 Days overdue"),
    ("31-60", "31-60 Days overdue"),
    ("61-90", "61-90 Days overdue"),
    ("91-120", "91-120 Days Overdue"),
    ("121-180", "121-180 Days Overdue"),
    ("180+", "180> Days Overdue"),
]

# Query type options
KEY_QUERY_TYPES = [
    "Missing Invoices", "Missing Allocation", "Awaiting Allocation",
    "Refund", "Bounce Back", "Credit Note", "Blocked", "ERP Error",
    "Missing payment",
]
LEDGER_QUERY_TYPES = [
    "Awaiting Payment Run", "DD Payment", "Refund", "Misc Payments",
    "Blocked", "Other Misc", "Posting Error", "Duplicate Posting",
    "Missing payment",
]


#  Logging 
class ValidationLogger:
    """Tees output to both terminal and log file."""

    def __init__(self, log_dir: Path):
        log_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        self.log_path = log_dir / f"validation_{ts}.log"
        self._fh = open(self.log_path, "w", encoding="utf-8")

    def print(self, msg: str = ""):
        # Use errors='replace' for Windows consoles that can't handle Unicode
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode("ascii", errors="replace").decode("ascii"))
        self._fh.write(msg + "\n")
        self._fh.flush()

    def close(self):
        self._fh.close()

    @staticmethod
    def cleanup(log_dir: Path, keep: int = 30):
        """Remove old log files, keeping the most recent *keep*."""
        logs = sorted(log_dir.glob("validation_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old in logs[keep:]:
            old.unlink(missing_ok=True)


#  Output helpers 
class Results:
    def __init__(self, logger: Optional[ValidationLogger] = None):
        self.passed = 0
        self.failed = 0
        self.warnings = 0
        self.cw_passed = 0
        self.cw_failed = 0
        self.lw_passed = 0
        self.lw_failed = 0
        self._period = "CW"  # current period being validated
        self._logger = logger
        self.meta: Dict[str, Any] = {}  # metadata for cross-stage checks

    def _print(self, msg: str):
        if self._logger:
            self._logger.print(msg)
        else:
            print(msg)

    def set_period(self, period: str):
        self._period = period

    def _record(self, ok: bool):
        if ok:
            self.passed += 1
            if self._period == "CW":
                self.cw_passed += 1
            else:
                self.lw_passed += 1
        else:
            self.failed += 1
            if self._period == "CW":
                self.cw_failed += 1
            else:
                self.lw_failed += 1

    def check(self, label: str, expected, actual, tolerance: int = 0):
        if expected is None:
            self._print(f"    ? SKIP  {label}: no expected value")
            self.warnings += 1
            return
        if actual is None:
            self._print(f"    X FAIL  {label}: expected={expected}, actual=None")
            self._record(False)
            return
        diff = actual - expected
        ok = abs(diff) <= tolerance
        tag = "  PASS" if ok else "X FAIL"
        self._record(ok)
        extra = f" (diff={diff:+g})" if diff != 0 else ""
        self._print(f"    {tag}  {label}: expected={expected}, db={actual}{extra}")

    def check_value(self, label: str, expected, actual, rel_tol: float = 0.005):
        if expected is None:
            self._print(f"    ? SKIP  {label}: no expected value")
            self.warnings += 1
            return
        if actual is None:
            self._print(f"    X FAIL  {label}: expected={expected:.2f}, actual=None")
            self._record(False)
            return
        diff = actual - expected
        denom = max(abs(expected), 1.0)
        ok = abs(diff) / denom <= rel_tol
        tag = "  PASS" if ok else "X FAIL"
        self._record(ok)
        extra = f" (diff={diff:+.2f})" if abs(diff) > 0.005 else ""
        self._print(f"    {tag}  {label}: expected={expected:.2f}, db={actual:.2f}{extra}")

    def check_bool(self, label: str, condition: bool, fail_msg: str = ""):
        """Boolean pass/fail check."""
        if condition:
            self._print(f"    PASS  {label}")
            self._record(True)
        else:
            self._print(f"    X FAIL  {label}" + (f": {fail_msg}" if fail_msg else ""))
            self._record(False)

    def warn(self, label: str, msg: str):
        """Record a warning (not a failure)."""
        self._print(f"    ! WARN  {label}: {msg}")
        self.warnings += 1

    def summary(self):
        if self.cw_passed + self.cw_failed == 0:
            self.failed += 1
            self.cw_failed += 1
            self._print("    X FAIL  Validation contract: NO CHECKS EXECUTED for current week")
        total = self.passed + self.failed + self.warnings
        self._print(f"\n{'=' * 70}")
        self._print(f"  CW: {self.cw_passed} passed, {self.cw_failed} failed")
        self._print(f"  LW: {self.lw_passed} passed, {self.lw_failed} failed")
        self._print(f"  TOTAL: {self.passed} passed, {self.failed} failed, {self.warnings} skipped  (total {total})")
        if self.cw_failed == 0:
            self._print("  CW VALIDATION: ALL CHECKS PASSED")
        else:
            self._print(f"  CW VALIDATION: *** {self.cw_failed} CHECK(S) FAILED ***")
        if self.lw_failed > 0:
            self._print(f"  LW NOTE: {self.lw_failed} LW check(s) failed  expected if previous")
            self._print(f"           week files were built with older script logic (e.g. no hierarchy)")
        self._print(f"{'=' * 70}")
        return self.cw_failed  # only CW failures count as real errors

    def stage_summary(self, stage_name: str) -> int:
        """Print a compact summary for a single stage. Returns failure count."""
        if self.passed + self.failed == 0:
            self.failed += 1
            self.cw_failed += 1
            self._print(f"    X FAIL  {stage_name}: NO CHECKS EXECUTED")
        total = self.passed + self.failed + self.warnings
        self._print(f"\n{'=' * 70}")
        self._print(f"  {stage_name}")
        self._print(f"  TOTAL: {self.passed} passed, {self.failed} failed, {self.warnings} warnings  (total {total})")
        if self.failed == 0:
            self._print(f"  RESULT: ALL CHECKS PASSED")
        else:
            self._print(f"  RESULT: *** {self.failed} CHECK(S) FAILED ***")
        self._print(f"{'=' * 70}")
        return self.failed


#  Utility 
def safe_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None

def safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def get_snapshots(db_path: Path, table: str) -> List[str]:
    with closing(sqlite3.connect(str(db_path))) as conn:
        rows = conn.execute(
            f'SELECT DISTINCT "SnapshotDateISO" FROM "{table}" ORDER BY "SnapshotDateISO" DESC'
        ).fetchall()
    return [r[0] for r in rows]

def find_excel(base_dir: Path, prefix: str, snapshot_iso: str) -> Optional[Path]:
    d = date.fromisoformat(snapshot_iso)
    candidates = [
        base_dir / f"{prefix} {d.strftime('%d.%m')}.xlsx",
        base_dir / f"{prefix} {d.strftime('%d.%m.%Y')}.xlsx",
        base_dir / f"{prefix} {d.strftime('%d.%m.00')}.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    # Glob fallback
    dd_mm = d.strftime('%d.%m')
    for p in base_dir.glob(f"{prefix}*{dd_mm}*.xlsx"):
        if '~$' not in p.name:
            return p
    return None


#  SQLite queries 
@contextmanager
def _conn(db_path: Path):
    conn = sqlite3.connect(str(db_path))
    try:
        yield conn
    finally:
        conn.close()


def db_total_docs_per_owner(db_path: Path, table: str, snap: str, conn=None) -> Dict[str, int]:
    """Total Docs per owner = COUNT(Detail rows WHERE Amount != 0).

    Args:
        db_path: Path to SQLite database
        table: Table name to query
        snap: Snapshot date (ISO format)
        conn: Optional existing connection. If None, creates new connection.

    Returns:
        Dict mapping owner to total doc count
    """
    def _run(c):
        query_name = f"total_docs_per_owner_{table}"
        if query_name in _VAL_QUERIES:
            sql = _VAL_QUERIES[query_name]
            rows = c.execute(sql, {"snap": snap}).fetchall()
        else:
            rows = c.execute(f'''
                SELECT "Owner", COUNT(*) FROM "{table}"
                WHERE "SnapshotDateISO" = ?
                  AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL
                  AND "Amount in doc. curr." != 0
                GROUP BY "Owner"
            ''', (snap,)).fetchall()
        return {r[0]: r[1] for r in rows}

    if conn:
        return _run(conn)
    with _conn(db_path) as c:
        return _run(c)


def db_total_suppliers_per_owner(db_path: Path, table: str, snap: str, conn=None) -> Dict[str, int]:
    """Total Suppliers per owner = COUNT(DISTINCT Unique Ref WHERE Detail AND Amount != 0).

    Args:
        db_path: Path to SQLite database
        table: Table name to query
        snap: Snapshot date (ISO format)
        conn: Optional existing connection. If None, creates new connection.

    Returns:
        Dict mapping owner to unique supplier count
    """
    def _run(c):
        query_name = f"total_suppliers_per_owner_{table}"
        if query_name in _VAL_QUERIES:
            sql = _VAL_QUERIES[query_name]
            rows = c.execute(sql, {"snap": snap}).fetchall()
        else:
            rows = c.execute(f'''
                SELECT "Owner", COUNT(DISTINCT "Unique Ref") FROM "{table}"
                WHERE "SnapshotDateISO" = ?
                  AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL
                  AND "Amount in doc. curr." != 0
                  AND "Unique Ref" IS NOT NULL
                  AND TRIM("Unique Ref") != ''
                GROUP BY "Owner"
            ''', (snap,)).fetchall()
        return {r[0]: r[1] for r in rows}

    if conn:
        return _run(conn)
    with _conn(db_path) as c:
        return _run(c)


def db_bucket_docs_per_owner(db_path: Path, table: str, snap: str,
                              buckets: list, conn=None) -> Dict[str, Dict[str, int]]:
    """Docs per bucket per owner = COUNT(Detail WHERE Amount!=0 AND bucket_col!=0).

    Args:
        db_path: Path to SQLite database
        table: Table name to query
        snap: Snapshot date (ISO format)
        buckets: List of (label, column_name) tuples for aging buckets
        conn: Optional existing connection. If None, creates new connection.

    Returns:
        Dict mapping owner to dict of {bucket_label: doc_count}
    """
    def _run(c):
        result: Dict[str, Dict[str, int]] = {}
        for label, col in buckets:
            rows = c.execute(f'''
                SELECT "Owner", COUNT(*) FROM "{table}"
                WHERE "SnapshotDateISO" = ?
                  AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                  AND "{col}" IS NOT NULL AND "{col}" != 0
                GROUP BY "Owner"
            ''', (snap,)).fetchall()
            for owner, cnt in rows:
                result.setdefault(owner, {})[label] = cnt
        return result

    if conn:
        return _run(conn)
    with _conn(db_path) as c:
        return _run(c)


def db_bucket_value_per_owner(db_path: Path, table: str, snap: str,
                               focus_col: str) -> Dict[str, float]:
    """Value for focus bucket per owner = SUM(Amount) for Detail rows in that bucket."""
    with _conn(db_path) as conn:
        rows = conn.execute(f'''
            SELECT "Owner", SUM("Amount in doc. curr.") FROM "{table}"
            WHERE "SnapshotDateISO" = ?
              AND "RowLevel" = 'Detail'
              AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
              AND "{focus_col}" IS NOT NULL AND "{focus_col}" != 0
            GROUP BY "Owner"
        ''', (snap,)).fetchall()
    return {r[0]: float(r[1]) if r[1] else 0.0 for r in rows}


def db_suppliers_per_bucket_hierarchical(db_path: Path, table: str, snap: str,
                                          buckets: list) -> Dict[str, Dict[str, int]]:
    """KEY hierarchical: supplier counted only in HIGHEST bucket.

    For each owner, for each supplier (Unique Ref), find the highest bucket
    with non-zero value, then count suppliers per bucket.
    """
    with _conn(db_path) as conn:
        # Get all detail rows with their bucket assignments
        bucket_cols = ", ".join(f'"{col}"' for _, col in buckets)
        rows = conn.execute(f'''
            SELECT "Owner", "Unique Ref", {bucket_cols}
            FROM "{table}"
            WHERE "SnapshotDateISO" = ?
              AND "RowLevel" = 'Detail'
              AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
              AND "Unique Ref" IS NOT NULL AND TRIM("Unique Ref") != ''
        ''', (snap,)).fetchall()

    # For each owner+supplier, find highest bucket
    # owner -> supplier -> highest_bucket_index
    supplier_highest: Dict[str, Dict[str, int]] = {}
    for row in rows:
        owner = row[0]
        uref = row[1]
        for i, (label, col) in enumerate(buckets):
            val = row[2 + i]
            if val is not None and val != 0:
                entry = supplier_highest.setdefault(owner, {})
                if uref not in entry or i > entry[uref]:
                    entry[uref] = i

    # Count suppliers per bucket per owner
    result: Dict[str, Dict[str, int]] = {}
    for owner, suppliers in supplier_highest.items():
        counts: Dict[str, int] = {}
        for uref, bucket_idx in suppliers.items():
            label = buckets[bucket_idx][0]
            counts[label] = counts.get(label, 0) + 1
        result[owner] = counts
    return result


def db_suppliers_per_bucket_flat(db_path: Path, table: str, snap: str,
                                  buckets: list) -> Dict[str, Dict[str, int]]:
    """LEDGER non-hierarchical: supplier counted in every bucket where they have docs."""
    with _conn(db_path) as conn:
        result: Dict[str, Dict[str, int]] = {}
        for label, col in buckets:
            rows = conn.execute(f'''
                SELECT "Owner", COUNT(DISTINCT "Unique Ref") FROM "{table}"
                WHERE "SnapshotDateISO" = ?
                  AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                  AND "{col}" IS NOT NULL AND "{col}" != 0
                  AND "Unique Ref" IS NOT NULL AND TRIM("Unique Ref") != ''
                GROUP BY "Owner"
            ''', (snap,)).fetchall()
            for owner, cnt in rows:
                result.setdefault(owner, {})[label] = cnt
    return result


def db_bucket_from_net_due_date(db_path: Path, table: str, snap: str,
                                buckets: list) -> dict:
    """Calculate bucket assignments from Net due date vs snapshot date.

    Returns dict with keys 'docs', 'suppliers' (hierarchical), and 'values'.
    Used for KEY validation where the report calculates Days Overdue dynamically.
    """
    with _conn(db_path) as conn:
        rows = conn.execute(f'''
            SELECT "Owner", "Unique Ref", "Net due date", "Amount in doc. curr."
            FROM "{table}"
            WHERE "SnapshotDateISO" = ?
              AND "RowLevel" = 'Detail'
              AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
              AND "Unique Ref" IS NOT NULL AND TRIM("Unique Ref") != ''
        ''', (snap,)).fetchall()

    snap_date = date.fromisoformat(snap)
    # Build bucket ranges from the provided bucket list
    bucket_ranges = []
    for label, _ in buckets:
        if label in ("07-30", "0-30"):
            bucket_ranges.append((label, 6, 30))
        elif label == "31-60":
            bucket_ranges.append((label, 30, 60))
        elif label == "61-90":
            bucket_ranges.append((label, 60, 90))
        elif label == "91-120":
            bucket_ranges.append((label, 90, 120))
        elif label == "121-180":
            bucket_ranges.append((label, 120, 180))
        elif label == "180+":
            bucket_ranges.append((label, 180, None))

    # Per-owner: docs per bucket, suppliers per bucket (hierarchical), values per bucket
    from collections import defaultdict
    docs_per_owner: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    values_per_owner: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    supplier_buckets: Dict[str, Dict[str, set]] = defaultdict(lambda: defaultdict(set))

    for owner, uref, ndd_str, amt in rows:
        if not ndd_str:
            continue
        try:
            # Try ISO format first, then dd-mm-yyyy
            try:
                ndd = date.fromisoformat(ndd_str)
            except ValueError:
                from datetime import datetime as _dt
                for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y'):
                    try:
                        ndd = _dt.strptime(ndd_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    continue
            days = (snap_date - ndd).days
            for label, low, high in bucket_ranges:
                if high is None:
                    if days > low:
                        docs_per_owner[owner][label] += 1
                        values_per_owner[owner][label] += float(amt) if amt else 0.0
                        supplier_buckets[owner][label].add(uref)
                        break
                else:
                    if days > low and days <= high:
                        docs_per_owner[owner][label] += 1
                        values_per_owner[owner][label] += float(amt) if amt else 0.0
                        supplier_buckets[owner][label].add(uref)
                        break
        except Exception:
            continue

    # Hierarchical supplier counting (highest bucket only)
    supplier_highest: Dict[str, Dict[str, int]] = {}
    bucket_labels = [label for label, _ in buckets]
    for owner in supplier_buckets:
        # For each supplier, find highest bucket
        sup_max: Dict[str, int] = {}
        for i, label in enumerate(bucket_labels):
            for uref in supplier_buckets[owner].get(label, set()):
                if uref not in sup_max or i > sup_max[uref]:
                    sup_max[uref] = i
        # Count per bucket
        counts: Dict[str, int] = {}
        for uref, idx in sup_max.items():
            lbl = bucket_labels[idx]
            counts[lbl] = counts.get(lbl, 0) + 1
        supplier_highest[owner] = counts

    return {
        'docs': dict(docs_per_owner),
        'suppliers': supplier_highest,
        'values': dict(values_per_owner),
    }


def db_query_types(db_path: Path, table: str, snap: str,
                   sheet_filter: Optional[str] = None,
                   owners_filter: Optional[set] = None,
                   detail_only: bool = False) -> Dict[str, int]:
    """Count rows by Query Type.

    detail_only: if True, count only Detail rows (matching COUNTIFS behavior in KEY).
    owners_filter: if provided, only count rows whose Owner is in this set.
    This excludes Uncategorised tabs that aren't in the Excel COUNTIF range.
    """
    with _conn(db_path) as conn:
        where = '''"SnapshotDateISO" = ? AND "Query type" IS NOT NULL AND TRIM("Query type") != '' '''
        params: list = [snap]
        if detail_only:
            where += ' AND "RowLevel" = \'Detail\''
        if sheet_filter:
            where += ' AND UPPER("Sheet") = ?'
            params.append(sheet_filter.upper())
        if owners_filter:
            placeholders = ", ".join("?" for _ in owners_filter)
            where += f' AND "Owner" IN ({placeholders})'
            params.extend(sorted(owners_filter))
        rows = conn.execute(f'''
            SELECT TRIM("Query type"), COUNT(*) FROM "{table}"
            WHERE {where}
            GROUP BY TRIM("Query type")
        ''', params).fetchall()
    return {r[0]: r[1] for r in rows}


#  Excel Summary readers 
def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def read_key_summary(path: Path) -> dict:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb['Summary']
    data = {}

    # KPI Overview (rows 5-7)
    data['kpi'] = {
        'suppliers_focus': {'lw': safe_int(_cell(ws, 5, 2)), 'cw': safe_int(_cell(ws, 5, 3))},
        'docs_focus':      {'lw': safe_int(_cell(ws, 6, 2)), 'cw': safe_int(_cell(ws, 6, 3))},
        'value_focus':     {'lw': safe_float(_cell(ws, 7, 2)), 'cw': safe_float(_cell(ws, 7, 3))},
    }

    # Owner Weekly Dashboard (position varies with active Key owners)
    # Column map: A(1)=Owner, B(2)/C(3)=Total Docs LW/CW, E(5)/F(6)=Total Suppliers LW/CW
    # H(8)/I(9)=Docs 180+ LW/CW
    # Bucket columns (each bucket = 6 cols: SupLW, SupCW, DeltaSup, DocsLW, DocsCW, DeltaDocs)
    # Starting at col 11: 07-30, 17: 31-60, 23: 61-90, 29: 91-120, 35: 121-180
    # Col 41/42 = Value 180+ LW/CW
    owners = {}
    for r in range(33, 43):  # up to 10 rows for safety
        name = _cell(ws, r, 1)
        if not name:
            break
        name = str(name).strip()
        o = {
            'total_docs':    {'lw': safe_int(_cell(ws, r, 2)), 'cw': safe_int(_cell(ws, r, 3))},
            'total_suppliers': {'lw': safe_int(_cell(ws, r, 5)), 'cw': safe_int(_cell(ws, r, 6))},
            'docs_180+':     {'lw': safe_int(_cell(ws, r, 8)), 'cw': safe_int(_cell(ws, r, 9))},
        }
        # Per-bucket Suppliers and Docs (skip 180+ which is already in cols 8-9 for Docs)
        col = 11
        for label in ["07-30", "31-60", "61-90", "91-120", "121-180"]:
            o[f'suppliers_{label}'] = {'lw': safe_int(_cell(ws, r, col)), 'cw': safe_int(_cell(ws, r, col + 1))}
            o[f'docs_{label}'] = {'lw': safe_int(_cell(ws, r, col + 3)), 'cw': safe_int(_cell(ws, r, col + 4))}
            col += 6
        o['value_180+'] = {'lw': safe_float(_cell(ws, r, 41)), 'cw': safe_float(_cell(ws, r, 42))}
        owners[name] = o
    data['owners'] = owners

    # Aging Mix (dynamic row  find "Aging Mix" header)
    aging = {}
    for r in range(38, 55):
        v = _cell(ws, r, 1)
        if v and "Aging Mix" in str(v):
            for ar in range(r + 2, r + 10):
                bucket = _cell(ws, ar, 1)
                if not bucket:
                    break
                bucket = str(bucket).strip()
                aging[bucket] = {
                    'docs_lw': safe_int(_cell(ws, ar, 2)),
                    'docs_cw': safe_int(_cell(ws, ar, 4)),
                }
            break
    data['aging_mix'] = aging

    # Query Type Breakdown (dynamic row  find "Query Type Breakdown")
    qt = {}
    qt_total = None
    for r in range(45, 70):
        v = _cell(ws, r, 1)
        if v and "Query Type Breakdown" in str(v):
            for qr in range(r + 2, r + 15):
                qtype = _cell(ws, qr, 1)
                qcount = safe_int(_cell(ws, qr, 2))
                if qtype and str(qtype).strip() == "TOTAL":
                    qt_total = qcount
                    break
                if qtype and qcount is not None:
                    qt[str(qtype).strip()] = qcount
            break
    data['query_types'] = qt
    data['query_type_total'] = qt_total

    wb.close()
    return data


def read_ledger_summary(path: Path, sheet_name: str) -> dict:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    data = {}

    # KPI (rows 5-7)
    data['kpi'] = {
        'suppliers_focus': {'lw': safe_int(_cell(ws, 5, 2)), 'cw': safe_int(_cell(ws, 5, 3))},
        'docs_focus':      {'lw': safe_int(_cell(ws, 6, 2)), 'cw': safe_int(_cell(ws, 6, 3))},
        'value_focus':     {'lw': safe_float(_cell(ws, 7, 2)), 'cw': safe_float(_cell(ws, 7, 3))},
    }

    # Owner Dashboard (row 10=header, 11+=data)
    # Col 2-3: Total Docs LW/CW, Col 5-6: Total Suppliers LW/CW
    # Col 8-9: Docs {focus} LW/CW
    # Then 6 buckets x 6 cols (Suppliers LW/CW/Delta, Docs LW/CW/Delta) = cols 11-46
    # Buckets: 0-30, 31-60, 61-90, 91-120, 121-180, 180+
    # Col 47-48: Value {focus} LW/CW
    owners = {}
    for r in range(11, 25):  # up to 14 rows for safety
        name = _cell(ws, r, 1)
        if not name:
            break
        name = str(name).strip()
        o = {
            'total_docs':    {'lw': safe_int(_cell(ws, r, 2)), 'cw': safe_int(_cell(ws, r, 3))},
            'total_suppliers': {'lw': safe_int(_cell(ws, r, 5)), 'cw': safe_int(_cell(ws, r, 6))},
            'docs_focus':    {'lw': safe_int(_cell(ws, r, 8)), 'cw': safe_int(_cell(ws, r, 9))},
        }
        # 6 buckets: 0-30, 31-60, 61-90, 91-120, 121-180, 180+
        col = 11
        bucket_labels = ["0-30", "31-60", "61-90", "91-120", "121-180", "180+"]
        for label in bucket_labels:
            o[f'suppliers_{label}'] = {'lw': safe_int(_cell(ws, r, col)), 'cw': safe_int(_cell(ws, r, col + 1))}
            o[f'docs_{label}'] = {'lw': safe_int(_cell(ws, r, col + 3)), 'cw': safe_int(_cell(ws, r, col + 4))}
            col += 6
        o['value_focus'] = {'lw': safe_float(_cell(ws, r, 47)), 'cw': safe_float(_cell(ws, r, 48))}
        owners[name] = o
    data['owners'] = owners

    # Query Type Breakdown (dynamic row)
    qt = {}
    qt_total = None
    for r in range(25, 50):
        v = _cell(ws, r, 1)
        if v and "Query Type Breakdown" in str(v):
            for qr in range(r + 2, r + 15):
                qtype = _cell(ws, qr, 1)
                qcount = safe_int(_cell(ws, qr, 2))
                if qtype and str(qtype).strip() == "TOTAL":
                    qt_total = qcount
                    break
                if qtype and qcount is not None:
                    qt[str(qtype).strip()] = qcount
            break
    data['query_types'] = qt
    data['query_type_total'] = qt_total

    wb.close()
    return data


#  Validation functions 
def validate_key(snap_cw: str, snap_lw: Optional[str], excel_path: Path, res: Results):
    print(f"\n{'=' * 70}")
    print(f"  KEY  Snapshot {snap_cw} (CW)" + (f" vs {snap_lw} (LW)" if snap_lw else ""))
    print(f"  Excel: {excel_path.name}")
    print(f"{'=' * 70}")

    summary = read_key_summary(excel_path)
    if not summary:
        print("  ERROR: Could not read Summary tab")
        res.failed += 1
        return

    focus_col = "180> Days Overdue"

    for period, snap in [("CW", snap_cw), ("LW", snap_lw)]:
        if not snap:
            continue
        print(f"\n  --- {period} (snapshot {snap}) ---")
        res.set_period(period)

        # Get DB data
        total_docs = db_total_docs_per_owner(KEY_DB, "key_lines", snap)
        total_suppliers = db_total_suppliers_per_owner(KEY_DB, "key_lines", snap)
        # Use Net due date-based bucket calculation (matches Key report and BI)
        ndd_data = db_bucket_from_net_due_date(KEY_DB, "key_lines", snap, KEY_BUCKETS)
        bucket_docs = ndd_data['docs']
        bucket_suppliers = ndd_data['suppliers']
        bucket_value = ndd_data['values']

        # --- Per-owner validation ---
        print(f"\n  [Owner Dashboard  {period}]")
        for owner, excel_data in summary['owners'].items():
            if owner not in KEY_OWNERS:
                continue
            print(f"    [{owner}]")
            p = period.lower()

            # Total Docs
            expected_td = excel_data['total_docs'].get(p)
            actual_td = total_docs.get(owner, 0)
            res.check(f"Total Docs {period}", expected_td, actual_td, tolerance=2)

            # Total Suppliers
            expected_ts = excel_data['total_suppliers'].get(p)
            actual_ts = total_suppliers.get(owner, 0)
            res.check(f"Total Suppliers {period}", expected_ts, actual_ts, tolerance=2)

            # Docs 180+
            expected_d180 = excel_data['docs_180+'].get(p)
            actual_d180 = bucket_docs.get(owner, {}).get("180+", 0)
            res.check(f"Docs 180+ {period}", expected_d180, actual_d180, tolerance=2)

            # Value 180+
            expected_v180 = excel_data['value_180+'].get(p)
            actual_v180 = bucket_value.get(owner, {}).get("180+", 0.0)
            if expected_v180 is not None:
                res.check_value(f"Value 180+ {period}", expected_v180, actual_v180)

            # Per-bucket Docs and Suppliers (skip 180+ for Suppliers since it's focus)
            for label in ["07-30", "31-60", "61-90", "91-120", "121-180"]:
                exp_docs = excel_data.get(f'docs_{label}', {}).get(p)
                act_docs = bucket_docs.get(owner, {}).get(label, 0)
                res.check(f"Docs {label} {period}", exp_docs, act_docs, tolerance=2)

                exp_sups = excel_data.get(f'suppliers_{label}', {}).get(p)
                act_sups = bucket_suppliers.get(owner, {}).get(label, 0)
                res.check(f"Suppliers {label} {period} (hierarchical)", exp_sups, act_sups, tolerance=2)

        # --- KPI totals ---
        print(f"\n  [KPI Overview (180+)  {period}]")
        p = period.lower()
        kpi = summary['kpi']

        # Sum across owners
        total_docs_180 = sum(bucket_docs.get(o, {}).get("180+", 0) for o in KEY_OWNERS)
        total_sup_180 = sum(bucket_suppliers.get(o, {}).get("180+", 0) for o in KEY_OWNERS)
        total_val_180 = sum(bucket_value.get(o, {}).get("180+", 0.0) for o in KEY_OWNERS)

        res.check(f"KPI Docs 180+ {period}", kpi['docs_focus'].get(p), total_docs_180, tolerance=5)
        res.check(f"KPI Suppliers 180+ {period}", kpi['suppliers_focus'].get(p), total_sup_180, tolerance=5)
        if kpi['value_focus'].get(p) is not None:
            res.check_value(f"KPI Value 180+ {period}", kpi['value_focus'][p], total_val_180)

        # --- Aging Mix ---
        if period == "CW" and summary.get('aging_mix'):
            print(f"\n  [Aging Mix  {period}]")
            for bucket_label, aging_data in summary['aging_mix'].items():
                db_label = bucket_label.replace("180+", "180+")
                exp = aging_data.get('docs_cw')
                act = sum(bucket_docs.get(o, {}).get(db_label, 0) for o in KEY_OWNERS)
                res.check(f"Aging Mix Docs {bucket_label} CW", exp, act, tolerance=5)

        # --- Query Type Breakdown ---
        if period == "CW" and summary.get('query_types'):
            print(f"\n  [Query Type Breakdown  CW]")
            # Detail only (matching COUNTIFS in Key report)
            query_types_filtered = db_query_types(KEY_DB, "key_lines", snap,
                                                   owners_filter=KEY_OWNERS, detail_only=True)
            for qtype, expected_count in summary['query_types'].items():
                actual_count = query_types_filtered.get(qtype, 0)
                res.check(f"QT '{qtype}'", expected_count, actual_count, tolerance=5)
            if summary.get('query_type_total'):
                total_qt_db = sum(query_types_filtered.get(qt, 0) for qt in KEY_QUERY_TYPES)
                res.check("QT TOTAL", summary['query_type_total'], total_qt_db, tolerance=10)


def validate_ledger_team(snap_cw: str, snap_lw: Optional[str], excel_path: Path,
                          sheet_name: str, team: str, focus_label: str,
                          focus_col: str, team_owners: set,
                          buckets: list, res: Results):
    print(f"\n  --- LEDGER {team} (focus: {focus_label}) ---")

    summary = read_ledger_summary(excel_path, sheet_name)
    if not summary:
        print(f"  ERROR: Could not read {sheet_name}")
        res.failed += 1
        return

    for period, snap in [("CW", snap_cw), ("LW", snap_lw)]:
        if not snap:
            continue
        print(f"\n  [{team}  {period} (snapshot {snap})]")
        res.set_period(period)

        # Query DB with Sheet filter
        sheet_val = team  # "QUERY" or "ROL"

        # Get DB data (filter by Sheet)
        with _conn(LEDGER_DB) as conn:
            # Total Docs per owner (Detail, Amt != 0, Sheet filter)
            total_docs_rows = conn.execute(f'''
                SELECT "Owner", COUNT(*) FROM "ledger_lines"
                WHERE "SnapshotDateISO" = ? AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                  AND UPPER("Sheet") = ?
                GROUP BY "Owner"
            ''', (snap, sheet_val)).fetchall()
            total_docs = {r[0]: r[1] for r in total_docs_rows}

            # Total Suppliers per owner
            total_sup_rows = conn.execute(f'''
                SELECT "Owner", COUNT(DISTINCT "Unique Ref") FROM "ledger_lines"
                WHERE "SnapshotDateISO" = ? AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                  AND UPPER("Sheet") = ?
                  AND "Unique Ref" IS NOT NULL AND TRIM("Unique Ref") != ''
                GROUP BY "Owner"
            ''', (snap, sheet_val)).fetchall()
            total_sup = {r[0]: r[1] for r in total_sup_rows}

            # Bucket Docs per owner
            bucket_docs: Dict[str, Dict[str, int]] = {}
            for label, col in buckets:
                rows = conn.execute(f'''
                    SELECT "Owner", COUNT(*) FROM "ledger_lines"
                    WHERE "SnapshotDateISO" = ? AND "RowLevel" = 'Detail'
                      AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                      AND UPPER("Sheet") = ?
                      AND "{col}" IS NOT NULL AND "{col}" != 0
                    GROUP BY "Owner"
                ''', (snap, sheet_val)).fetchall()
                for owner, cnt in rows:
                    bucket_docs.setdefault(owner, {})[label] = cnt

            # Bucket Suppliers per owner (non-hierarchical for LEDGER)
            bucket_suppliers: Dict[str, Dict[str, int]] = {}
            for label, col in buckets:
                rows = conn.execute(f'''
                    SELECT "Owner", COUNT(DISTINCT "Unique Ref") FROM "ledger_lines"
                    WHERE "SnapshotDateISO" = ? AND "RowLevel" = 'Detail'
                      AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                      AND UPPER("Sheet") = ?
                      AND "{col}" IS NOT NULL AND "{col}" != 0
                      AND "Unique Ref" IS NOT NULL AND TRIM("Unique Ref") != ''
                    GROUP BY "Owner"
                ''', (snap, sheet_val)).fetchall()
                for owner, cnt in rows:
                    bucket_suppliers.setdefault(owner, {})[label] = cnt

            # Value for focus bucket per owner
            bucket_value_rows = conn.execute(f'''
                SELECT "Owner", SUM("Amount in doc. curr.") FROM "ledger_lines"
                WHERE "SnapshotDateISO" = ? AND "RowLevel" = 'Detail'
                  AND "Amount in doc. curr." IS NOT NULL AND "Amount in doc. curr." != 0
                  AND UPPER("Sheet") = ?
                  AND "{focus_col}" IS NOT NULL AND "{focus_col}" != 0
                GROUP BY "Owner"
            ''', (snap, sheet_val)).fetchall()
            bucket_value = {r[0]: float(r[1]) if r[1] else 0.0 for r in bucket_value_rows}

        # Per-owner validation
        p = period.lower()
        for owner, excel_data in summary['owners'].items():
            if owner not in team_owners:
                continue
            print(f"      [{owner}]")

            # Total Docs
            res.check(f"Total Docs {period}", excel_data['total_docs'].get(p),
                       total_docs.get(owner, 0), tolerance=2)

            # Total Suppliers
            res.check(f"Total Suppliers {period}", excel_data['total_suppliers'].get(p),
                       total_sup.get(owner, 0), tolerance=2)

            # Docs focus
            exp_df = excel_data['docs_focus'].get(p)
            act_df = bucket_docs.get(owner, {}).get(focus_label, 0)
            res.check(f"Docs {focus_label} {period}", exp_df, act_df, tolerance=2)

            # Value focus
            exp_vf = excel_data.get('value_focus', {}).get(p)
            act_vf = bucket_value.get(owner, 0.0)
            if exp_vf is not None:
                res.check_value(f"Value {focus_label} {period}", exp_vf, act_vf)

            # Per-bucket Docs and Suppliers (all 6 buckets)
            bucket_labels = ["0-30", "31-60", "61-90", "91-120", "121-180", "180+"]
            for label in bucket_labels:
                # 0-30 bucket: higher tolerance  boundary effects between
                # date-computed _Days and pre-computed aging column values.
                tol = 100 if label == "0-30" else 5

                exp_docs = excel_data.get(f'docs_{label}', {}).get(p)
                act_docs = bucket_docs.get(owner, {}).get(label, 0)
                if exp_docs is not None:
                    res.check(f"Docs {label} {period}", exp_docs, act_docs, tolerance=tol)

                exp_sups = excel_data.get(f'suppliers_{label}', {}).get(p)
                act_sups = bucket_suppliers.get(owner, {}).get(label, 0)
                if exp_sups is not None:
                    res.check(f"Suppliers {label} {period}", exp_sups, act_sups, tolerance=tol)

        # KPI totals
        print(f"\n      [KPI {focus_label}  {period}]")
        kpi = summary['kpi']
        total_docs_focus = sum(bucket_docs.get(o, {}).get(focus_label, 0) for o in team_owners)
        total_sup_focus = sum(bucket_suppliers.get(o, {}).get(focus_label, 0) for o in team_owners)
        total_val_focus = sum(bucket_value.get(o, 0.0) for o in team_owners)

        # Higher tolerance for LEDGER KPI: Excel includes Uncategorised owners
        # (ROL Uncategorised, Query Uncategorised) that are NOT in SQLite.
        # The difference = exactly the Uncategorised contribution (~5-15 docs).
        res.check(f"KPI Docs {focus_label} {period}", kpi['docs_focus'].get(p), total_docs_focus, tolerance=20)
        res.check(f"KPI Suppliers {focus_label} {period}", kpi['suppliers_focus'].get(p), total_sup_focus, tolerance=10)
        if kpi['value_focus'].get(p) is not None:
            # Higher rel_tol: boundary effects between date-computed _Days and
            # pre-computed aging columns cause docs to shift between buckets,
            # affecting the value sum especially in lower buckets.
            res.check_value(f"KPI Value {focus_label} {period}", kpi['value_focus'][p], total_val_focus, rel_tol=0.10)

        # Query Type Breakdown (CW only)
        if period == "CW" and summary.get('query_types'):
            print(f"\n      [Query Type Breakdown  CW]")
            # Filter by team owners only (Excel COUNTIF covers owner tabs, not Uncategorised)
            qt_db = db_query_types(LEDGER_DB, "ledger_lines", snap,
                                   sheet_filter=sheet_val, owners_filter=team_owners)
            for qtype, expected_count in summary['query_types'].items():
                actual_count = qt_db.get(qtype, 0)
                res.check(f"QT '{qtype}'", expected_count, actual_count, tolerance=5)
            if summary.get('query_type_total'):
                # TOTAL = sum of only the listed types (Excel sums specific COUNTIF, not all)
                total_qt_db = sum(qt_db.get(qt, 0) for qt in LEDGER_QUERY_TYPES)
                res.check("QT TOTAL", summary['query_type_total'], total_qt_db, tolerance=10)


def validate_ledger(snap_cw: str, snap_lw: Optional[str], excel_path: Path, res: Results):
    print(f"\n{'=' * 70}")
    print(f"  LEDGER  Snapshot {snap_cw} (CW)" + (f" vs {snap_lw} (LW)" if snap_lw else ""))
    print(f"  Excel: {excel_path.name}")
    print(f"{'=' * 70}")

    # ROL team (focus: 90+  merged ROL + Query)
    validate_ledger_team(
        snap_cw, snap_lw, excel_path,
        sheet_name="Summary_ROL", team="ROL",
        focus_label="180+", focus_col="180> Days Overdue",
        team_owners=ROL_OWNERS, buckets=LEDGER_BUCKETS, res=res,
    )


# 
# STAGE HELPERS (read-only  never generates or modifies data)
# 

def _get_latest_week(db_path: Path, table: str) -> Optional[str]:
    """Return the most recent WeekStartISO from a SQLite table."""
    if not db_path.exists():
        return None
    with closing(sqlite3.connect(str(db_path))) as conn:
        row = conn.execute(
            f'SELECT MAX("WeekStartISO") FROM "{table}"'
        ).fetchone()
    return row[0] if row and row[0] else None


def _find_latest_key_excel() -> Optional[Path]:
    """Find the latest Key *.xlsx in KEY_DATA by modification time."""
    candidates = [p for p in KEY_DATA.glob("Key*.xlsx") if "~$" not in p.name]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _parse_dashboard_js(js_path: Path) -> Optional[dict]:
    """Parse dashboard_data.js into a Python dict (read-only)."""
    content = js_path.read_text(encoding="utf-8")
    marker = "window.DASHBOARD_DATA = "
    idx = content.find(marker)
    if idx < 0:
        return None
    json_str = content[idx + len(marker):].rstrip().rstrip(";")
    return json.loads(json_str)


def _decompress_blob(b64_str: str) -> Optional[dict]:
    """Decompress a zlib+base64 blob to a Python dict."""
    if not b64_str:
        return None
    raw = base64.b64decode(b64_str)
    return json.loads(zlib.decompress(raw))


def _table_exists(db_path: Path, table: str) -> bool:
    """Check if a table exists in a SQLite database."""
    with closing(sqlite3.connect(str(db_path))) as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (table,)
        ).fetchone()
    return bool(row and row[0] > 0)


def _count_rows(db_path: Path, table: str, where: str = "", params: tuple = ()) -> int:
    """Count rows in a SQLite table with optional WHERE clause."""
    with closing(sqlite3.connect(str(db_path))) as conn:
        sql = f'SELECT COUNT(*) FROM "{table}"'
        if where:
            sql += f" WHERE {where}"
        row = conn.execute(sql, params).fetchone()
    return row[0] if row else 0


def _null_count(db_path: Path, table: str, col: str, week: str) -> int:
    """Count rows where col is NULL or empty for a given week."""
    with closing(sqlite3.connect(str(db_path))) as conn:
        row = conn.execute(f'''
            SELECT COUNT(*) FROM "{table}"
            WHERE "WeekStartISO" = ?
              AND ("{col}" IS NULL OR TRIM("{col}") = '')
        ''', (week,)).fetchone()
    return row[0] if row else 0


def _distinct_values(db_path: Path, table: str, col: str, week: str,
                     extra_where: str = "") -> set:
    """Get distinct non-empty values for a column in a given week."""
    with closing(sqlite3.connect(str(db_path))) as conn:
        sql = f'''
            SELECT DISTINCT "{col}" FROM "{table}"
            WHERE "WeekStartISO" = ?
              AND "{col}" IS NOT NULL AND TRIM("{col}") != ''
        '''
        if extra_where:
            sql += f" AND {extra_where}"
        rows = conn.execute(sql, (week,)).fetchall()
    return {r[0] for r in rows}


# 
# STAGE 1: VALIDATE MASTERDATA CSV
# 

def validate_masterdata(csv_path: Path, logger: Optional[ValidationLogger] = None) -> Results:
    """Validate a MasterData CSV file (read-only). Returns Results."""
    res = Results(logger)
    res.set_period("CW")  # all checks count as real errors

    res._print(f"\n{'=' * 70}")
    res._print(f"  STAGE 1: VALIDATE MASTERDATA")
    res._print(f"  File: {csv_path}")
    res._print(f"{'=' * 70}")

    # 1. File exists
    res.check_bool("File exists", csv_path.exists(), str(csv_path))
    if not csv_path.exists():
        return res

    # 2. File > 0 bytes
    fsize = csv_path.stat().st_size
    res.check_bool("File not empty", fsize > 0, f"size={fsize}")
    if fsize == 0:
        return res

    # Read CSV
    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", low_memory=False)
    except Exception as e:
        try:
            df = pd.read_csv(csv_path, encoding="latin1", low_memory=False)
        except Exception:
            res.check_bool("CSV readable", False, str(e))
            return res

    # 3. All 22 MASTER_COLS present
    csv_cols = set(df.columns)
    missing = set(MASTER_COLS) - csv_cols
    res.check_bool("All 22 columns present", len(missing) == 0,
                    f"missing: {sorted(missing)}")

    # 4. No unexpected extra columns
    extra = csv_cols - set(MASTER_COLS)
    if extra:
        res.warn("Extra columns", f"{sorted(extra)}")

    # 5. Row count sanity
    nrows = len(df)
    res.check_bool("Row count > 0", nrows > 0, f"rows={nrows}")
    if nrows == 0:
        return res
    if nrows < 100:
        res.warn("Row count low", f"{nrows} rows (expected 100-500k)")
    elif nrows > 500_000:
        res.warn("Row count high", f"{nrows} rows (expected 100-500k)")
    else:
        res._print(f"    PASS  Row count: {nrows}")
        res._record(True)

    # 6-10. Critical columns not empty
    #   Owner is only required for Key/ROL rows  Query rows may have blank Owner.
    #   All other critical columns must be filled across the entire file.
    for col_name in ["Country", "Sheet", "Unique Ref", "Company Code"]:
        if col_name not in df.columns:
            continue  # already reported as missing column
        empty = df[col_name].isna().sum() + (df[col_name].astype(str).str.strip() == "").sum()
        pct = empty / nrows * 100
        res.check_bool(f"'{col_name}' no blanks", pct < 5,
                        f"{empty} empty ({pct:.1f}%)")

    # Owner check: Key must have owners (Owner_map.csv); ROL/Query get owners
    # from vendor category mapping  some categories (fuel, intercompany) are
    # intentionally left empty, so we use a softer threshold.
    if "Owner" in df.columns and "Sheet" in df.columns:
        for sheet_name, threshold, severity in [("Key", 5, "fail"), ("ROL", 50, "warn"), ("Query", 50, "warn")]:
            subset = df[df["Sheet"] == sheet_name]
            if len(subset) == 0:
                continue
            empty_o = subset["Owner"].isna().sum() + (subset["Owner"].astype(str).str.strip() == "").sum()
            pct_o = empty_o / len(subset) * 100
            if pct_o >= threshold:
                if severity == "fail":
                    res.check_bool(f"'Owner' no blanks ({sheet_name})", False,
                                    f"{empty_o} empty ({pct_o:.1f}%) out of {len(subset)} rows")
                else:
                    res.warn(f"'Owner' mostly blank ({sheet_name})",
                             f"{empty_o} empty ({pct_o:.1f}%) out of {len(subset)} rows")
            else:
                res._print(f"    PASS  'Owner' filled ({sheet_name}): {pct_o:.1f}% empty")
                res._record(True)

    # 11. Sheet values valid
    if "Sheet" in df.columns:
        actual_sheets = set(df["Sheet"].dropna().astype(str).str.strip().unique())
        invalid = actual_sheets - VALID_SHEETS
        res.check_bool("Sheet values valid", len(invalid) == 0,
                        f"invalid: {sorted(invalid)}")

    # 12. Owner values known
    if "Owner" in df.columns:
        actual_owners = set(df["Owner"].dropna().astype(str).str.strip().unique())
        unknown = actual_owners - ALL_OWNERS
        if unknown:
            res.warn("Unknown owners", f"{sorted(unknown)}")
        else:
            res._print(f"    PASS  All owners known ({len(actual_owners)})")
            res._record(True)

    # 13. All 3 sheets present
    if "Sheet" in df.columns:
        sheets_present = set(df["Sheet"].dropna().astype(str).str.strip().unique())
        for s in VALID_SHEETS:
            res.check_bool(f"Sheet '{s}' present", s in sheets_present,
                            f"sheets found: {sorted(sheets_present)}")

    # 14-15. Dates parseable
    for date_col in ["Net due date", "Document Date"]:
        if date_col not in df.columns:
            continue
        parsed = parse_date_series(df[date_col])
        nat_count = parsed.isna().sum() - df[date_col].isna().sum()  # only count parse failures
        pct = nat_count / nrows * 100 if nrows > 0 else 0
        if pct > 2:
            res.warn(f"'{date_col}' parse issues", f"{nat_count} unparseable ({pct:.1f}%)")
        else:
            res._print(f"    PASS  '{date_col}' parseable ({pct:.1f}% failures)")
            res._record(True)

    # 16. Amount parseable
    if "Amount in doc. curr." in df.columns:
        amt = pd.to_numeric(df["Amount in doc. curr."].astype(str).str.replace(",", ""),
                            errors="coerce")
        bad = amt.isna().sum() - df["Amount in doc. curr."].isna().sum()
        pct = bad / nrows * 100 if nrows > 0 else 0
        if pct > 2:
            res.warn("Amount parse issues", f"{bad} unparseable ({pct:.1f}%)")
        else:
            res._print(f"    PASS  Amount parseable ({pct:.1f}% failures)")
            res._record(True)

    # 17. Sheet distribution
    #   Real data: Query ~85-95%, Key ~3-10%, ROL ~0.5-5%
    #   (most vendors are unclassified Query; Key/ROL are the curated subsets)
    if "Sheet" in df.columns:
        dist = df["Sheet"].value_counts(normalize=True) * 100
        for sheet, lo, hi in [("Key", 1, 20), ("ROL", 0.1, 10), ("Query", 70, 99)]:
            pct = dist.get(sheet, 0)
            if pct < lo or pct > hi:
                res.warn(f"Sheet '{sheet}' distribution", f"{pct:.1f}% (expected {lo}-{hi}%)")
            else:
                res._print(f"    PASS  Sheet '{sheet}' distribution: {pct:.1f}%")
                res._record(True)

    # 18. Owners per Sheet consistency
    #   Allow category-based owners (Fuel, Rent, Unassigned) that appear in Owner_map
    #   or are intentionally mapped in the category functions.
    _special_owners = {"Fuel", "Rent", "Unassigned", "ROL Uncategorised", "Synthetic Owner 023", ""}
    if "Sheet" in df.columns and "Owner" in df.columns:
        for sheet_val, expected_set in [("Key", KEY_OWNERS), ("ROL", ROL_OWNERS)]:
            sheet_owners = set(df.loc[df["Sheet"] == sheet_val, "Owner"].dropna().unique())
            unexpected = sheet_owners - expected_set - _special_owners
            if unexpected:
                res.warn(f"'{sheet_val}' unexpected owners", f"{sorted(unexpected)}")
            else:
                res._print(f"    PASS  '{sheet_val}' owners consistent")
                res._record(True)

    # Store metadata for cross-stage
    res.meta["row_count"] = nrows
    if "Owner" in df.columns:
        res.meta["owners"] = set(df["Owner"].dropna().astype(str).str.strip().unique())
    if "Sheet" in df.columns:
        res.meta["sheets"] = set(df["Sheet"].dropna().astype(str).str.strip().unique())
    # Extract week from filename
    m = re.search(r'MasterData_(\d{2})(\d{2})', csv_path.name)
    if m:
        res.meta["week_num"] = int(m.group(1))
        res.meta["year"] = 2000 + int(m.group(2))

    res.stage_summary("STAGE 1: MASTERDATA VALIDATION")
    return res


# 
# STAGE 2: VALIDATE SQLITE LOAD
# 

def validate_ledger_load(week_iso: Optional[str] = None,
                          logger: Optional[ValidationLogger] = None) -> Results:
    """Validate SQLite databases after ETL load (read-only). Returns Results."""
    res = Results(logger)
    res.set_period("CW")

    res._print(f"\n{'=' * 70}")
    res._print(f"  STAGE 2: VALIDATE SQLITE LOAD")
    res._print(f"  KEY DB:    {KEY_DB}")
    res._print(f"  LEDGER DB: {LEDGER_DB}")
    res._print(f"{'=' * 70}")

    # 1-2. DB files exist
    res.check_bool("KEY DB exists", KEY_DB.exists(), str(KEY_DB))
    res.check_bool("LEDGER DB exists", LEDGER_DB.exists(), str(LEDGER_DB))
    if not KEY_DB.exists() or not LEDGER_DB.exists():
        res.stage_summary("STAGE 2: SQLITE LOAD VALIDATION")
        return res

    # 3-4. Tables exist
    res.check_bool("Table 'key_lines' exists", _table_exists(KEY_DB, "key_lines"))
    res.check_bool("Table 'ledger_lines' exists", _table_exists(LEDGER_DB, "ledger_lines"))
    if not _table_exists(KEY_DB, "key_lines") or not _table_exists(LEDGER_DB, "ledger_lines"):
        res.stage_summary("STAGE 2: SQLITE LOAD VALIDATION")
        return res

    # 5. Auto-detect week
    if not week_iso:
        week_iso = _get_latest_week(KEY_DB, "key_lines")
        if not week_iso:
            res.check_bool("Latest week found", False, "no WeekStartISO in key_lines")
            res.stage_summary("STAGE 2: SQLITE LOAD VALIDATION")
            return res
    res._print(f"    INFO  Validating week: {week_iso}")

    # 6-7. Row counts > 0
    key_rows = _count_rows(KEY_DB, "key_lines", '"WeekStartISO" = ?', (week_iso,))
    ledger_rows = _count_rows(LEDGER_DB, "ledger_lines", '"WeekStartISO" = ?', (week_iso,))
    res.check_bool("Key rows > 0", key_rows > 0, f"rows={key_rows}")
    res.check_bool("Ledger rows > 0", ledger_rows > 0, f"rows={ledger_rows}")

    # 8-10. Owners present
    key_db_owners = _distinct_values(KEY_DB, "key_lines", "Owner", week_iso)
    missing_key = KEY_OWNERS - key_db_owners
    if missing_key:
        res.warn("Key missing owners", f"{sorted(missing_key)}")
    else:
        res._print(f"    PASS  All KEY_OWNERS present ({len(key_db_owners)})")
        res._record(True)

    ledger_rol_owners = _distinct_values(LEDGER_DB, "ledger_lines", "Owner", week_iso,
                                          '''UPPER("Sheet") = 'ROL' ''')
    missing_rol = _rol_required_owners_for_week(week_iso) - ledger_rol_owners
    if missing_rol:
        res.warn("Ledger ROL missing owners", f"{sorted(missing_rol)}")
    else:
        res._print(f"    PASS  All ROL_OWNERS present ({len(ledger_rol_owners)})")
        res._record(True)

    # 11-14. NULL checks for KEY
    for col in ["Owner", "Sheet", "Unique Ref", "WeekStartISO"]:
        nc = _null_count(KEY_DB, "key_lines", col, week_iso)
        res.check_bool(f"Key '{col}' no NULLs", nc == 0, f"{nc} NULL/empty rows")

    # 15. NULL checks for LEDGER (same 4 columns)
    for col in ["Owner", "Sheet", "Unique Ref", "WeekStartISO"]:
        nc = _null_count(LEDGER_DB, "ledger_lines", col, week_iso)
        res.check_bool(f"Ledger '{col}' no NULLs", nc == 0, f"{nc} NULL/empty rows")

    # 16-17. RowLevel distribution (Key)
    key_detail = _count_rows(KEY_DB, "key_lines",
                              '"WeekStartISO" = ? AND "RowLevel" = \'Detail\'', (week_iso,))
    if key_rows > 0:
        detail_pct = key_detail / key_rows * 100
        res.check_bool("Key Detail rows present", key_detail > 0)
        if detail_pct < 70:
            res.warn("Key Detail ratio low", f"{detail_pct:.1f}% (expected >70%)")
        else:
            res._print(f"    PASS  Key Detail ratio: {detail_pct:.1f}%")
            res._record(True)

    # 18. DocClass includes Invoice
    if key_rows > 0:
        doc_classes = _distinct_values(KEY_DB, "key_lines", "DocClass", week_iso)
        res.check_bool("Key DocClass has 'Invoice'", "Invoice" in doc_classes,
                        f"found: {sorted(doc_classes)}")

    # 19. SnapshotDateISO format
    with closing(sqlite3.connect(str(KEY_DB))) as conn:
        snap_rows = conn.execute('''
            SELECT DISTINCT "SnapshotDateISO" FROM "key_lines" WHERE "WeekStartISO" = ?
        ''', (week_iso,)).fetchall()
    bad_format = [r[0] for r in snap_rows if r[0] and not re.match(r'^\d{4}-\d{2}-\d{2}$', r[0])]
    res.check_bool("SnapshotDateISO format YYYY-MM-DD", len(bad_format) == 0,
                    f"bad: {bad_format[:5]}")

    # 20. Cross-check with MasterData CSV
    #   Key: CSV Key rows vs SQLite key_lines (tolerance 25%  CSV has all rows
    #     including headers, SQLite may filter by RowLevel/Owner/Fuel exclusions)
    #   Ledger: CSV ROL+Query rows WITH Owner assigned vs SQLite ledger_lines
    #     (unassigned vendors are not loaded into the ledger)
    md_path = find_latest_masterdata(MASTER_DATA)
    if md_path:
        try:
            md_df = pd.read_csv(md_path, encoding="utf-8-sig",
                                usecols=["Sheet", "Owner"], low_memory=False)
            md_df["Sheet"] = md_df["Sheet"].astype(str).str.strip()
            md_df["Owner"] = md_df["Owner"].fillna("").astype(str).str.strip()
            md_key = (md_df["Sheet"] == "Key").sum()
            # Ledger: only count ROL/Query rows that have an owner assigned
            md_ledger_with_owner = md_df[
                (md_df["Sheet"].isin(["ROL", "Query"])) & (md_df["Owner"] != "")
            ].shape[0]
            if md_key > 0:
                ratio = abs(key_rows - md_key) / md_key
                if ratio > 0.25:
                    res.warn("Key rows vs MasterData",
                             f"SQLite={key_rows}, CSV={md_key} ({ratio:.0%} diff)")
                else:
                    res._print(f"    PASS  Key rows vs MasterData: SQLite={key_rows}, CSV={md_key}")
                    res._record(True)
            if md_ledger_with_owner > 0:
                ratio = abs(ledger_rows - md_ledger_with_owner) / md_ledger_with_owner
                if ratio > 0.25:
                    res.warn("Ledger rows vs MasterData",
                             f"SQLite={ledger_rows}, CSV(with owner)={md_ledger_with_owner} ({ratio:.0%} diff)")
                else:
                    res._print(f"    PASS  Ledger rows vs MasterData: SQLite={ledger_rows}, CSV(with owner)={md_ledger_with_owner}")
                    res._record(True)
        except Exception as e:
            res.warn("MasterData cross-check", f"could not read: {e}")
    else:
        res.warn("MasterData cross-check", "no MasterData CSV found")

    # Store metadata
    res.meta["key_rows"] = key_rows
    res.meta["ledger_rows"] = ledger_rows
    res.meta["week"] = week_iso
    res.meta["key_owners"] = key_db_owners
    res.meta["ledger_owners"] = ledger_rol_owners

    res.stage_summary("STAGE 2: SQLITE LOAD VALIDATION")
    return res


# 
# STAGE 3: VALIDATE KEY BUILD
# 

def validate_key_build(excel_path: Optional[Path] = None,
                        logger: Optional[ValidationLogger] = None) -> Results:
    """Validate Key Excel report (read-only). Returns Results."""
    res = Results(logger)
    res.set_period("CW")

    # Auto-detect if not provided
    if excel_path is None:
        excel_path = _find_latest_key_excel()

    res._print(f"\n{'=' * 70}")
    res._print(f"  STAGE 3: VALIDATE KEY BUILD")
    res._print(f"  File: {excel_path}")
    res._print(f"{'=' * 70}")

    # 1. File exists
    if excel_path is None:
        res.check_bool("Key Excel found", False, "no Key*.xlsx in data/key/")
        res.stage_summary("STAGE 3: KEY BUILD VALIDATION")
        return res
    res.check_bool("File exists", excel_path.exists(), str(excel_path))
    if not excel_path.exists():
        res.stage_summary("STAGE 3: KEY BUILD VALIDATION")
        return res

    # 2. Size > 10KB
    fsize = excel_path.stat().st_size
    res.check_bool("File size > 10KB", fsize > 10_000, f"size={fsize}")

    # Load workbook
    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    except Exception as e:
        res.check_bool("Excel readable", False, str(e))
        res.stage_summary("STAGE 3: KEY BUILD VALIDATION")
        return res

    sheet_names = set(wb.sheetnames)

    # 3. Summary tab exists
    res.check_bool("'Summary' tab exists", "Summary" in sheet_names)

    # 4. All KEY_OWNERS have tabs (WARN if missing  owner may have no data)
    for owner in sorted(KEY_OWNERS):
        if owner in sheet_names:
            res._print(f"    PASS  Tab '{owner}' exists")
            res._record(True)
        else:
            res.warn(f"Tab '{owner}' missing", "owner may have no data this period")

    # 5. Unexpected tabs
    expected_tabs = KEY_OWNERS | {"Summary", "Weekly_Movement", "Payment Issues"}
    unexpected = sheet_names - expected_tabs
    if unexpected:
        res.warn("Unexpected tabs", f"{sorted(unexpected)}")

    # 6-8. Summary KPIs
    if "Summary" in sheet_names:
        try:
            summary = read_key_summary(excel_path)
            kpi = summary.get("kpi", {})

            sup = kpi.get("suppliers_focus", {}).get("cw")
            docs = kpi.get("docs_focus", {}).get("cw")
            val = kpi.get("value_focus", {}).get("cw")

            res.check_bool("KPI 180+ Suppliers > 0", sup is not None and sup > 0,
                            f"value={sup}")
            res.check_bool("KPI 180+ Docs > 0", docs is not None and docs > 0,
                            f"value={docs}")
            res.check_bool("KPI 180+ Value != 0", val is not None and val != 0,
                            f"value={val}")

            res.meta["kpi_suppliers"] = sup
            res.meta["kpi_docs"] = docs
            res.meta["kpi_value"] = val
        except Exception as e:
            res.check_bool("Summary readable", False, str(e))

    # 9. Each owner tab has rows
    for owner in sorted(KEY_OWNERS):
        if owner in sheet_names:
            ws = wb[owner]
            row_count = 0
            for _ in ws.iter_rows(min_row=1, max_row=3):
                row_count += 1
            if row_count < 2:
                res.warn(f"Tab '{owner}' rows", f"only {row_count} rows")
            else:
                res._print(f"    PASS  Tab '{owner}' has data")
                res._record(True)

    wb.close()

    # 10-12. Cross-check with SQLite (180+ focus  matching Excel Summary KPIs)
    #   The Excel Summary shows 180+ suppliers/docs/value.
    #   SQLite queries filter to: Detail rows, Fuel excluded, 180> Days Overdue > 0.
    if KEY_DB.exists() and _table_exists(KEY_DB, "key_lines"):
        latest_week = _get_latest_week(KEY_DB, "key_lines")
        if latest_week:
            with closing(sqlite3.connect(str(KEY_DB))) as conn:
                # 180+ suppliers (distinct Unique Ref with 180+ value)
                row = conn.execute('''
                    SELECT COUNT(DISTINCT "Unique Ref") FROM "key_lines"
                    WHERE "WeekStartISO" = ? AND "RowLevel" = 'Detail'
                      AND COALESCE("Owner", '') != 'Fuel'
                      AND "180> Days Overdue" IS NOT NULL
                      AND CAST("180> Days Overdue" AS REAL) != 0
                ''', (latest_week,)).fetchone()
                db_vendors_180 = row[0] if row else 0

                # 180+ value (sum of 180> Days Overdue column)
                row = conn.execute('''
                    SELECT SUM(CAST("180> Days Overdue" AS REAL)) FROM "key_lines"
                    WHERE "WeekStartISO" = ? AND "RowLevel" = 'Detail'
                      AND COALESCE("Owner", '') != 'Fuel'
                      AND "180> Days Overdue" IS NOT NULL
                      AND CAST("180> Days Overdue" AS REAL) != 0
                ''', (latest_week,)).fetchone()
                db_val_180 = float(row[0]) if row and row[0] else 0.0

                # 180+ docs (count of rows with 180+ value)
                row = conn.execute('''
                    SELECT COUNT(*) FROM "key_lines"
                    WHERE "WeekStartISO" = ? AND "RowLevel" = 'Detail'
                      AND COALESCE("Owner", '') != 'Fuel'
                      AND "180> Days Overdue" IS NOT NULL
                      AND CAST("180> Days Overdue" AS REAL) != 0
                ''', (latest_week,)).fetchone()
                db_docs_180 = row[0] if row else 0

            kpi_sup = res.meta.get("kpi_suppliers")
            if kpi_sup and db_vendors_180 > 0:
                ratio = abs(kpi_sup - db_vendors_180) / db_vendors_180
                if ratio > 0.15:
                    res.warn("Vendor count vs SQLite (180+)",
                             f"Excel={kpi_sup}, SQLite={db_vendors_180} ({ratio:.0%} diff)")
                else:
                    res._print(f"    PASS  Vendor count vs SQLite (180+): Excel={kpi_sup}, DB={db_vendors_180}")
                    res._record(True)

            # Value and doc count: show INFO only (Excel is from MasterData CSV with
            # exclusions; SQLite may have different data vintage and filter logic)
            kpi_val = res.meta.get("kpi_value")
            if kpi_val is not None and db_val_180 != 0:
                res._print(f"    INFO  180+ value: Excel={kpi_val:.0f}, SQLite={db_val_180:.0f}")

            kpi_docs = res.meta.get("kpi_docs")
            if kpi_docs is not None and db_docs_180 > 0:
                res._print(f"    INFO  180+ docs: Excel={kpi_docs}, SQLite={db_docs_180}")

    res.meta["owners_with_tabs"] = KEY_OWNERS & sheet_names

    res.stage_summary("STAGE 3: KEY BUILD VALIDATION")
    return res


# 
# STAGE 4: VALIDATE DASHBOARD
# 

def validate_dashboard(data_js_path: Optional[Path] = None,
                        logger: Optional[ValidationLogger] = None) -> Results:
    """Validate dashboard output files (read-only). Returns Results."""
    res = Results(logger)
    res.set_period("CW")

    if data_js_path is None:
        data_js_path = JSON_LOCAL_DIR / "dashboard_data.js"
    html_path = data_js_path.parent / "dashboard.html"

    res._print(f"\n{'=' * 70}")
    res._print(f"  STAGE 4: VALIDATE DASHBOARD")
    res._print(f"  JS:   {data_js_path}")
    res._print(f"  HTML: {html_path}")
    res._print(f"{'=' * 70}")

    # 1-2. Files exist
    res.check_bool("dashboard_data.js exists", data_js_path.exists())
    res.check_bool("dashboard.html exists", html_path.exists())
    if not data_js_path.exists():
        res.stage_summary("STAGE 4: DASHBOARD VALIDATION")
        return res

    # 3-4. File size
    js_size = data_js_path.stat().st_size
    if js_size < 50_000:
        res.warn("JS file small", f"{js_size / 1024:.1f} KB (expected >50 KB)")
    else:
        res._print(f"    PASS  JS file size: {js_size / 1024:.1f} KB")
        res._record(True)
    if js_size > 100_000_000:
        res.warn("JS file very large", f"{js_size / (1024*1024):.1f} MB (may crash browser)")

    # 5. Pattern check
    content_head = data_js_path.read_text(encoding="utf-8")[:500]
    has_marker = "window.DASHBOARD_DATA =" in content_head
    res.check_bool("JS has DASHBOARD_DATA marker", has_marker)
    if not has_marker:
        res.stage_summary("STAGE 4: DASHBOARD VALIDATION")
        return res

    # 6. JSON parseable
    try:
        payload = _parse_dashboard_js(data_js_path)
        res.check_bool("JSON parseable", payload is not None)
    except Exception as e:
        res.check_bool("JSON parseable", False, str(e)[:200])
        res.stage_summary("STAGE 4: DASHBOARD VALIDATION")
        return res

    if payload is None:
        res.stage_summary("STAGE 4: DASHBOARD VALIDATION")
        return res

    # 7. sorted_weeks
    weeks = payload.get("sorted_weeks", [])
    res.check_bool("sorted_weeks not empty", len(weeks) > 0, f"len={len(weeks)}")

    # 8. filters
    filters = payload.get("filters", {})
    res.check_bool("filters.countries exists", "countries" in filters)
    res.check_bool("filters.owners exists", "owners" in filters)

    # 9. compressed_weeks
    cw = payload.get("compressed_weeks", {})
    res.check_bool("compressed_weeks not empty", len(cw) > 0, f"len={len(cw)}")

    # 10. filters.owners
    f_owners = filters.get("owners", [])
    if not f_owners:
        res.warn("filters.owners empty", "no owners in dashboard")
    else:
        res._print(f"    PASS  filters.owners: {len(f_owners)} owners")
        res._record(True)

    # 11. Decompress 1 week
    if weeks and cw:
        first_week = weeks[0]
        blob = cw.get(first_week, "")
        if blob:
            try:
                week_data = _decompress_blob(blob)
                res.check_bool("First week decompresses", week_data is not None)
            except Exception as e:
                res.check_bool("First week decompresses", False, str(e)[:200])
        else:
            res.warn("First week blob", "empty compressed data")

    # 12. SyntheticReview decompress
    synthetic_review_blob = payload.get("synthetic_review_compressed", "")
    if synthetic_review_blob:
        try:
            xd = _decompress_blob(synthetic_review_blob)
            res.check_bool("SyntheticReview blob decompresses", xd is not None)
        except Exception as e:
            res.warn("SyntheticReview decompress", str(e)[:200])
    else:
        res._print(f"    INFO  No SyntheticReview data in dashboard")

    # 13. Statement decompress
    stmt_blob = payload.get("statement_compressed", "")
    if stmt_blob:
        try:
            sd = _decompress_blob(stmt_blob)
            res.check_bool("Statement blob decompresses", sd is not None)
        except Exception as e:
            res.warn("Statement decompress", str(e)[:200])
    else:
        res._print(f"    INFO  No Statement data in dashboard")

    # 14. Cross-check weeks vs SQLite
    if KEY_DB.exists() and _table_exists(KEY_DB, "key_lines"):
        with closing(sqlite3.connect(str(KEY_DB))) as conn:
            db_weeks = conn.execute(
                'SELECT COUNT(DISTINCT "WeekStartISO") FROM "key_lines"'
            ).fetchone()
        db_week_count = db_weeks[0] if db_weeks else 0
        js_week_count = len(weeks)
        diff = abs(js_week_count - db_week_count)
        if diff > 2:
            res.warn("Week count vs SQLite",
                     f"JS={js_week_count}, SQLite={db_week_count} (diff={diff})")
        else:
            res._print(f"    PASS  Week count: JS={js_week_count}, SQLite={db_week_count}")
            res._record(True)

    # Store metadata
    res.meta["weeks_count"] = len(weeks)
    res.meta["owners"] = set(f_owners)
    res.meta["has_synthetic_review"] = bool(synthetic_review_blob)
    res.meta["has_statement"] = bool(stmt_blob)
    res.meta["latest_week"] = weeks[0] if weeks else None

    res.stage_summary("STAGE 4: DASHBOARD VALIDATION")
    return res


# 
# STAGE 5: CROSS-STAGE VALIDATION
# 

def validate_cross_stage(stage_results: Optional[Dict[str, Results]] = None,
                          logger: Optional[ValidationLogger] = None) -> Results:
    """Cross-stage consistency checks (read-only). Returns Results.

    If stage_results is None, reads directly from each data source.
    """
    res = Results(logger)
    res.set_period("CW")

    res._print(f"\n{'=' * 70}")
    res._print(f"  STAGE 5: CROSS-STAGE CONSISTENCY")
    res._print(f"{'=' * 70}")

    # Gather data from sources (read-only)
    # MasterData
    md_owners = set()
    md_key_rows = 0
    md_ledger_rows = 0
    md_sheets = set()
    md_countries = set()
    md_path = find_latest_masterdata(MASTER_DATA)
    if md_path:
        try:
            md_df = pd.read_csv(md_path, encoding="utf-8-sig",
                                usecols=["Owner", "Sheet", "Country"], low_memory=False)
            md_owners = set(md_df["Owner"].dropna().astype(str).str.strip().unique())
            md_sheets = set(md_df["Sheet"].dropna().astype(str).str.strip().unique())
            md_countries = set(md_df["Country"].dropna().astype(str).str.strip().unique())
            md_key_rows = (md_df["Sheet"].str.strip() == "Key").sum()
            md_ledger_rows = md_df["Sheet"].str.strip().isin(["ROL", "Query"]).sum()
        except Exception:
            pass

    # SQLite
    db_key_owners = set()
    db_ledger_owners = set()
    db_key_rows = 0
    db_ledger_rows = 0
    db_latest_week = None
    db_countries = set()
    if KEY_DB.exists() and _table_exists(KEY_DB, "key_lines"):
        db_latest_week = _get_latest_week(KEY_DB, "key_lines")
        if db_latest_week:
            db_key_owners = _distinct_values(KEY_DB, "key_lines", "Owner", db_latest_week)
            db_key_rows = _count_rows(KEY_DB, "key_lines", '"WeekStartISO" = ?', (db_latest_week,))
            db_countries |= _distinct_values(KEY_DB, "key_lines", "Country", db_latest_week)
    if LEDGER_DB.exists() and _table_exists(LEDGER_DB, "ledger_lines"):
        lw = _get_latest_week(LEDGER_DB, "ledger_lines")
        if lw:
            db_ledger_owners = _distinct_values(LEDGER_DB, "ledger_lines", "Owner", lw)
            db_ledger_rows = _count_rows(LEDGER_DB, "ledger_lines", '"WeekStartISO" = ?', (lw,))
            db_countries |= _distinct_values(LEDGER_DB, "ledger_lines", "Country", lw)

    # Dashboard
    js_owners = set()
    js_latest_week = None
    js_path = JSON_LOCAL_DIR / "dashboard_data.js"
    if js_path.exists():
        try:
            payload = _parse_dashboard_js(js_path)
            if payload:
                js_owners = set(payload.get("filters", {}).get("owners", []))
                sw = payload.get("sorted_weeks", [])
                js_latest_week = sw[0] if sw else None
        except Exception:
            pass

    #  Checks 

    # 1. Owner consistency: MasterData vs SQLite
    if md_owners and db_key_owners:
        md_key_only = md_owners & KEY_OWNERS
        missing = md_key_only - db_key_owners
        res.check_bool("Owners MasterData->SQLite (Key)", len(missing) == 0,
                        f"in CSV but not in SQLite: {sorted(missing)}")
    else:
        res.warn("Owners MasterData->SQLite", "insufficient data to compare")

    # 2. Owner consistency: SQLite vs Dashboard
    all_db_owners = db_key_owners | db_ledger_owners
    if all_db_owners and js_owners:
        missing = all_db_owners - js_owners - {"ROL Uncategorised", "ROL Uncategorised"}
        if missing:
            res.warn("Owners SQLite->Dashboard", f"in DB but not in JS: {sorted(missing)}")
        else:
            res._print(f"    PASS  Owners SQLite->Dashboard consistent")
            res._record(True)
    else:
        res.warn("Owners SQLite->Dashboard", "insufficient data to compare")

    # 3. Week consistency: SQLite vs Dashboard
    if db_latest_week and js_latest_week:
        res.check_bool("Latest week SQLite->Dashboard",
                        db_latest_week == js_latest_week,
                        f"SQLite={db_latest_week}, JS={js_latest_week}")
    else:
        res.warn("Week consistency", "insufficient data to compare")

    # 4. Row count flow: MasterData -> SQLite (Key)
    #   Natural differences: CSV may be from different week, SQLite includes Header
    #   rows, Fuel exclusions, etc. Use generous 75% tolerance.
    if md_key_rows > 0 and db_key_rows > 0:
        ratio = abs(md_key_rows - db_key_rows) / max(md_key_rows, db_key_rows)
        if ratio > 0.75:
            res.warn("Row flow MasterData->SQLite (Key)",
                     f"CSV={md_key_rows}, SQLite={db_key_rows} ({ratio:.0%} diff)")
        else:
            res._print(f"    PASS  Row flow Key: CSV={md_key_rows}, SQLite={db_key_rows} ({ratio:.0%} diff)")
            res._record(True)
    else:
        res.warn("Row flow Key", "insufficient data")

    # 5. Row count flow: MasterData -> SQLite (Ledger)
    #   Large differences expected: CSV has ALL ROL+Query rows, SQLite only loads
    #   rows for specific LEDGER_TABS owners. Compare order of magnitude only.
    if md_ledger_rows > 0 and db_ledger_rows > 0:
        res._print(f"    INFO  Row flow Ledger: CSV(all)={md_ledger_rows}, SQLite(owners)={db_ledger_rows}")
        res._record(True)
    else:
        res.warn("Row flow Ledger", "insufficient data")

    # 6. All 3 sheets present across pipeline
    for sheet in VALID_SHEETS:
        in_csv = sheet in md_sheets if md_sheets else True
        # In SQLite, Sheet is stored as uppercase
        if db_latest_week:
            db_sheets_key = _distinct_values(KEY_DB, "key_lines", "Sheet", db_latest_week) if db_key_rows > 0 else set()
            db_sheets_led = _distinct_values(LEDGER_DB, "ledger_lines", "Sheet", _get_latest_week(LEDGER_DB, "ledger_lines") or "") if db_ledger_rows > 0 else set()
            # Normalize for comparison
            db_all_sheets = {s.strip() for s in db_sheets_key | db_sheets_led}
            in_db = sheet in db_all_sheets or sheet.upper() in {s.upper() for s in db_all_sheets}
        else:
            in_db = True
        res.check_bool(f"Sheet '{sheet}' in pipeline", in_csv and in_db,
                        f"CSV={in_csv}, SQLite={in_db}")

    # 7. Country consistency
    if md_countries and db_countries:
        missing = md_countries - db_countries - {"Unknown", ""}
        if missing and len(missing) > 2:
            res.warn("Country consistency", f"in CSV but not in SQLite: {sorted(missing)[:5]}")
        else:
            res._print(f"    PASS  Country consistency ({len(md_countries)} in CSV, {len(db_countries)} in SQLite)")
            res._record(True)
    else:
        res.warn("Country consistency", "insufficient data")

    # 8. No empty/NULL owner in SQLite
    if db_latest_week:
        nc_key = _null_count(KEY_DB, "key_lines", "Owner", db_latest_week)
        lw_led = _get_latest_week(LEDGER_DB, "ledger_lines") or ""
        nc_led = _null_count(LEDGER_DB, "ledger_lines", "Owner", lw_led) if lw_led else 0
        res.check_bool("No empty Owner in SQLite", nc_key + nc_led == 0,
                        f"key={nc_key}, ledger={nc_led} empty")
    else:
        res.warn("Empty Owner check", "no week found")

    # 9. Dashboard has data for latest week
    if db_latest_week and js_path.exists():
        try:
            payload = _parse_dashboard_js(js_path)
            if payload:
                cw_data = payload.get("compressed_weeks", {})
                res.check_bool("Dashboard has latest week data",
                                db_latest_week in cw_data,
                                f"week {db_latest_week} not in compressed_weeks")
        except Exception:
            res.warn("Dashboard latest week check", "could not parse JS")
    else:
        res.warn("Dashboard latest week check", "insufficient data")

    res.stage_summary("STAGE 5: CROSS-STAGE CONSISTENCY")
    return res


# 
# PRE-DEPLOY CHECKS (T12  6 data quality checks)
# 

def _check_recent_week(res: Results) -> None:
    """Check 1: Latest week has data and is recent (< 14 days old)."""
    res._print("\n  [Check 1] Latest week has recent data")

    for label, db_path, table in [("Key", KEY_DB, "key_lines"),
                                   ("Ledger", LEDGER_DB, "ledger_lines")]:
        if not db_path.exists():
            res.check_bool(f"{label} DB exists", False, str(db_path))
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            res.check_bool(f"{label} has data", False, "no WeekStartISO found")
            continue

        rows = _count_rows(db_path, table, '"WeekStartISO" = ?', (week,))
        res.check_bool(f"{label} latest week ({week}) has rows", rows > 0,
                        f"rows={rows}")

        # Check staleness  warn if latest week > 14 days ago
        try:
            week_date = date.fromisoformat(week)
            age_days = (date.today() - week_date).days
            if age_days > 14:
                res.warn(f"{label} data staleness",
                         f"latest week is {age_days} days old  ETL may not have run")
            else:
                res._print(f"    PASS  {label} data fresh ({age_days} days old)")
                res._record(True)
        except ValueError:
            res.warn(f"{label} week format", f"cannot parse '{week}' as date")


def _check_header_total_value(res: Results) -> None:
    """Check 2: HEADER TOTAL VALUE sanity  detect extreme outliers.

    In AP, negative TV is normal (liability = company owes supplier).
    We check for: (a) any zero-TV headers (data gap), (b) extreme outliers
    where |TV| > 10M (possible data error), (c) expected negative-majority pattern.
    """
    res._print("\n  [Check 2] HEADER Total Value sanity")

    for label, db_path, table in [("Key", KEY_DB, "key_lines"),
                                   ("Ledger", LEDGER_DB, "ledger_lines")]:
        if not db_path.exists():
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            continue

        with closing(sqlite3.connect(str(db_path))) as conn:
            row = conn.execute(f'''
                SELECT COUNT(*),
                       SUM(CASE WHEN COALESCE("TOTAL VALUE", 0) = 0 THEN 1 ELSE 0 END),
                       SUM(CASE WHEN ABS(COALESCE("TOTAL VALUE", 0)) > 10000000 THEN 1 ELSE 0 END),
                       SUM(CASE WHEN COALESCE("TOTAL VALUE", 0) < 0 THEN 1 ELSE 0 END)
                FROM "{table}"
                WHERE "WeekStartISO" = ? AND "RowLevel" = 'Header'
            ''', (week,)).fetchone()

        total_headers = row[0] if row else 0
        zero_headers = row[1] if row else 0
        extreme_headers = row[2] if row else 0
        neg_headers = row[3] if row else 0

        if total_headers == 0:
            res.warn(f"{label} HEADER rows", "no Header rows found")
            continue

        # Zero TV headers  should be rare (supplier with no balance = data gap)
        zero_pct = zero_headers / total_headers * 100
        if zero_pct > 20:
            res.warn(f"{label} HEADER zero TV",
                     f"{zero_headers}/{total_headers} ({zero_pct:.1f}%) have TV=0")
        else:
            res._print(f"    PASS  {label} HEADER TV zeros: {zero_headers}/{total_headers} ({zero_pct:.1f}%)")
            res._record(True)

        # Extreme values  possible data errors
        if extreme_headers > 0:
            res.warn(f"{label} HEADER extreme TV",
                     f"{extreme_headers} suppliers with |TV| > 10M")
        else:
            res._print(f"    PASS  {label} HEADER TV no extremes (>10M)")
            res._record(True)

        # AP pattern: majority should be negative (owed to supplier)
        neg_pct = neg_headers / total_headers * 100
        if neg_pct < 30:
            res.warn(f"{label} HEADER TV sign",
                     f"only {neg_pct:.0f}% negative  unusual for AP data")
        else:
            res._print(f"    PASS  {label} HEADER TV sign: {neg_pct:.0f}% negative (normal AP)")
            res._record(True)


def _check_owners_match(res: Results) -> None:
    """Check 3: Owners in DB match expected KEY_OWNERS."""
    res._print("\n  [Check 3] Owners match expected sets")

    for label, db_path, table, expected in [
        ("Key", KEY_DB, "key_lines", KEY_OWNERS),
        ("Ledger ROL", LEDGER_DB, "ledger_lines", ROL_OWNERS),
    ]:
        if not db_path.exists():
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            continue

        extra_where = ""
        if "QUERY" in label:
            extra_where = '''UPPER("Sheet") = 'QUERY' '''
        elif "ROL" in label:
            extra_where = '''UPPER("Sheet") = 'ROL' '''

        actual = _distinct_values(db_path, table, "Owner", week, extra_where)

        required = _rol_required_owners_for_week(week) if label == "Ledger ROL" else expected
        allowed_optional = ROL_OPTIONAL_OWNERS if label == "Ledger ROL" else set()
        missing = required - actual
        unknown = actual - expected - {"Fuel", "Rent", "Unassigned",
                                        "ROL Uncategorised", "ROL Uncategorised", ""} - allowed_optional

        if missing:
            res.warn(f"{label} missing owners", f"{sorted(missing)}")
        else:
            res._print(f"    PASS  {label} all expected owners present")
            res._record(True)

        if unknown:
            res.warn(f"{label} unknown owners", f"{sorted(unknown)}")


def _check_duplicates(res: Results) -> None:
    """Check 4: Detect duplicate rows in DB.

    Uses (UniqueRef, DocumentNumber, CompanyCode, Week) as key.
    The ERP source allows multiple line items per document, so some duplicates
    are expected  we WARN above a threshold rather than FAIL.
    """
    res._print("\n  [Check 4] Duplicate detection")

    for label, db_path, table in [("Key", KEY_DB, "key_lines"),
                                   ("Ledger", LEDGER_DB, "ledger_lines")]:
        if not db_path.exists():
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            continue

        with closing(sqlite3.connect(str(db_path))) as conn:
            # Count total detail rows
            total_row = conn.execute(f'''
                SELECT COUNT(*) FROM "{table}"
                WHERE "WeekStartISO" = ? AND "RowLevel" = 'Detail'
            ''', (week,)).fetchone()
            total_detail = total_row[0] if total_row else 0

            # Count duplicate groups by (UniqueRef, DocNumber, CompanyCode, Week)
            row = conn.execute(f'''
                SELECT COUNT(*), SUM(cnt - 1) FROM (
                    SELECT COUNT(*) as cnt
                    FROM "{table}"
                    WHERE "WeekStartISO" = ?
                      AND "RowLevel" = 'Detail'
                    GROUP BY "Unique Ref", "Document Number", "Company Code", "WeekStartISO"
                    HAVING COUNT(*) > 1
                )
            ''', (week,)).fetchone()

        dup_groups = row[0] if row and row[0] else 0
        dup_excess = row[1] if row and row[1] else 0  # extra rows beyond 1 per group

        if total_detail == 0:
            res.warn(f"{label} duplicates", "no detail rows")
            continue

        dup_pct = dup_excess / total_detail * 100
        if dup_pct > 5:
            res.check_bool(f"{label} duplicates",
                            False,
                            f"{dup_groups} dup groups, {dup_excess} extra rows ({dup_pct:.1f}% of {total_detail})")
        elif dup_groups > 0:
            res.warn(f"{label} duplicates",
                     f"{dup_groups} dup groups, {dup_excess} extra rows ({dup_pct:.1f}%)  likely ERP line items")
        else:
            res._print(f"    PASS  {label} no duplicate detail rows")
            res._record(True)


def _check_company_codes(res: Results) -> None:
    """Check 5: Company codes in DB exist in MasterData."""
    res._print("\n  [Check 5] Company codes valid")

    md_path = find_latest_masterdata(MASTER_DATA)
    if not md_path:
        res.warn("Company codes", "no MasterData CSV found  cannot cross-check")
        return

    try:
        md_df = pd.read_csv(md_path, encoding="utf-8-sig",
                            usecols=["Company Code"], low_memory=False)
        md_codes = set(md_df["Company Code"].dropna().astype(str).str.strip().unique())
    except Exception as e:
        res.warn("Company codes", f"cannot read MasterData: {e}")
        return

    if not md_codes:
        res.warn("Company codes", "no company codes in MasterData")
        return

    for label, db_path, table in [("Key", KEY_DB, "key_lines"),
                                   ("Ledger", LEDGER_DB, "ledger_lines")]:
        if not db_path.exists():
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            continue

        db_codes = _distinct_values(db_path, table, "Company Code", week)
        # Normalize: DB may store as float (e.g. "1000.0"), CSV as "1000"
        db_codes_norm = set()
        for c in db_codes:
            try:
                db_codes_norm.add(str(int(float(c))))
            except (ValueError, TypeError):
                db_codes_norm.add(str(c).strip())

        md_codes_norm = set()
        for c in md_codes:
            try:
                md_codes_norm.add(str(int(float(c))))
            except (ValueError, TypeError):
                md_codes_norm.add(str(c).strip())

        unknown = db_codes_norm - md_codes_norm - {""}
        if unknown:
            res.warn(f"{label} unknown company codes",
                     f"{sorted(unknown)[:10]} ({len(unknown)} total)")
        else:
            res._print(f"    PASS  {label} all {len(db_codes_norm)} company codes in MasterData")
            res._record(True)


def _check_aging_vs_total(res: Results) -> None:
    """Check 6: Aging buckets sum correctly vs Amount (DETAIL level)."""
    res._print("\n  [Check 6] Aging buckets vs Amount consistency")

    checks = [
        ("Key", KEY_DB, "key_lines",
         [col for _, col in KEY_BUCKETS]),
        ("Ledger", LEDGER_DB, "ledger_lines",
         [col for _, col in LEDGER_BUCKETS]),
    ]

    for label, db_path, table, bucket_cols in checks:
        if not db_path.exists():
            continue

        week = _get_latest_week(db_path, table)
        if not week:
            continue

        # Build SQL: sum all aging buckets per DETAIL row, compare with Amount
        coalesce_parts = " + ".join(
            f'COALESCE("{c}", 0)' for c in bucket_cols
        )

        with closing(sqlite3.connect(str(db_path))) as conn:
            # Count DETAIL rows where aging_sum differs from Amount by > 10%
            # Only check overdue rows (aging_sum != 0)
            rows = conn.execute(f'''
                SELECT COUNT(*) as total_overdue,
                       SUM(CASE
                           WHEN ABS(aging_sum - amt) > ABS(amt) * 0.1 + 1
                           THEN 1 ELSE 0
                       END) as mismatches
                FROM (
                    SELECT
                        COALESCE("Amount in doc. curr.", 0) as amt,
                        ({coalesce_parts}) as aging_sum
                    FROM "{table}"
                    WHERE "WeekStartISO" = ?
                      AND "RowLevel" = 'Detail'
                      AND ({coalesce_parts}) != 0
                )
            ''', (week,)).fetchone()

        total_overdue = rows[0] if rows else 0
        mismatches = rows[1] if rows else 0

        if total_overdue == 0:
            res.warn(f"{label} aging check", "no overdue detail rows found")
            continue

        mismatch_pct = mismatches / total_overdue * 100
        if mismatch_pct > 10:
            res.check_bool(f"{label} aging vs amount",
                            False,
                            f"{mismatches}/{total_overdue} ({mismatch_pct:.1f}%) rows have aging != amount")
        elif mismatch_pct > 2:
            res.warn(f"{label} aging vs amount",
                     f"{mismatches}/{total_overdue} ({mismatch_pct:.1f}%) rows have aging != amount")
        else:
            res._print(f"    PASS  {label} aging vs amount: {mismatches}/{total_overdue} mismatches ({mismatch_pct:.1f}%)")
            res._record(True)


def _check_weeks_vs_files(res: Results) -> None:
    """Check 7: SQLite week count must match spreadsheet file count.

    Counts distinct WeekStartISO in each DB and compares with the number
    of Excel files in data/ + archive/. They must match  every spreadsheet
    loaded produces exactly one week in the DB.
    """
    res._print("\n  [Check 7] SQLite weeks vs spreadsheet files")

    checks = [
        ("Key", KEY_DB, "key_lines", KEY_DATA, KEY_ARCHIVE, "Key *.xlsx"),
        ("Ledger", LEDGER_DB, "ledger_lines", LEDGER_DATA, LEDGER_ARCHIVE, "Ledger *.xlsx"),
    ]

    for label, db_path, table, data_dir, archive_dir, glob_pat in checks:
        if not db_path.exists():
            res.check_bool(f"{label} DB exists", False, str(db_path))
            continue

        # Count distinct weeks in SQLite
        with closing(sqlite3.connect(str(db_path))) as conn:
            rows = conn.execute(
                f'SELECT DISTINCT "WeekStartISO" FROM "{table}" '
                'WHERE "WeekStartISO" IS NOT NULL '
                'ORDER BY "WeekStartISO"'
            ).fetchall()
        db_weeks = [r[0] for r in rows]
        n_db = len(db_weeks)

        # Count spreadsheet files in data/ + archive/
        files_main = [f for f in data_dir.glob(glob_pat)
                      if not f.name.startswith("~$")]
        files_archive = []
        if archive_dir.exists():
            files_archive = [f for f in archive_dir.glob(glob_pat)
                             if not f.name.startswith("~$")]
        n_files = len(files_main) + len(files_archive)

        if n_db == n_files:
            res._print(f"    PASS  {label}: {n_db} weeks in DB = {n_files} spreadsheets")
            res._record(True)
        else:
            detail = f"DB has {n_db} weeks but found {n_files} files"
            if n_db > n_files:
                detail += " (missing spreadsheets? check archive/)"
            else:
                detail += " (files not loaded? run ETL with --rebuild)"
            res.check_bool(f"{label} weeks vs files", False, detail)

        # Show week list for debugging
        if db_weeks:
            res._print(f"           DB weeks: {db_weeks[0]} .. {db_weeks[-1]}")
        file_names = sorted([f.name for f in files_main + files_archive])
        if file_names:
            res._print(f"           Files: {file_names[0]} .. {file_names[-1]} ({n_files} total)")


def run_all_checks(logger: Optional[ValidationLogger] = None) -> dict:
    """Run all 7 pre-deploy data quality checks.

    Returns dict with keys: passed, failed, warnings, total, exit_code.
    Usable by deploy.py and smoke tests.
    """
    res = Results(logger)
    res.set_period("CW")

    res._print(f"\n{'=' * 70}")
    res._print(f"  PRE-DEPLOY DATA QUALITY CHECKS")
    res._print(f"  KEY DB:    {KEY_DB}")
    res._print(f"  LEDGER DB: {LEDGER_DB}")
    res._print(f"{'=' * 70}")

    _check_recent_week(res)
    _check_header_total_value(res)
    _check_owners_match(res)
    _check_duplicates(res)
    _check_company_codes(res)
    _check_aging_vs_total(res)
    _check_weeks_vs_files(res)

    failures = res.stage_summary("PRE-DEPLOY DATA QUALITY CHECKS")

    return {
        "passed": res.passed,
        "failed": res.failed,
        "warnings": res.warnings,
        "total": res.passed + res.failed + res.warnings,
        "exit_code": failures,
    }


# 
# MAIN  Legacy + Stage-based validation
# 

def _run_legacy_validation(weeks: int, logger: Optional[ValidationLogger] = None) -> int:
    """Run the existing KEY+LEDGER vs Excel Summary validation."""
    res = Results(logger)

    res._print(f"Validating last {weeks} snapshot(s)...")
    res._print(f"KEY DB:    {KEY_DB}")
    res._print(f"LEDGER DB: {LEDGER_DB}")

    key_snaps = get_snapshots(KEY_DB, "key_lines")
    ledger_snaps = get_snapshots(LEDGER_DB, "ledger_lines")

    if not key_snaps:
        res._print("ERROR: No KEY snapshots found")
        return 1
    if not ledger_snaps:
        res._print("ERROR: No LEDGER snapshots found")
        return 1

    res._print(f"\nKEY snapshots available: {', '.join(key_snaps[:weeks + 1])}")
    res._print(f"LEDGER snapshots available: {', '.join(ledger_snaps[:weeks + 1])}")

    for idx in range(weeks):
        if idx >= len(key_snaps):
            break

        snap_cw = key_snaps[idx]
        snap_lw = key_snaps[idx + 1] if idx + 1 < len(key_snaps) else None

        key_excel = find_excel(KEY_DATA, "Key", snap_cw)
        if key_excel:
            validate_key(snap_cw, snap_lw, key_excel, res)
        else:
            res._print(f"\n  WARNING: No KEY Excel found for snapshot {snap_cw}")
            res.warnings += 1

        ledger_snap_cw = ledger_snaps[idx] if idx < len(ledger_snaps) else None
        ledger_snap_lw = ledger_snaps[idx + 1] if idx + 1 < len(ledger_snaps) else None
        if ledger_snap_cw:
            ledger_excel = find_excel(LEDGER_DATA, "Ledger", ledger_snap_cw)
            if ledger_excel:
                validate_ledger(ledger_snap_cw, ledger_snap_lw, ledger_excel, res)
            else:
                res._print(f"\n  WARNING: No LEDGER Excel found for snapshot {ledger_snap_cw}")
                res.warnings += 1

    return res.summary()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Synthetic Services AP Ledger  Data Validation (read-only)")
    parser.add_argument("--weeks", type=int, default=2,
                        help="Weeks to validate in legacy mode (default: 2)")
    parser.add_argument("--stage",
                        choices=["masterdata", "ledger-load", "key-build", "dashboard", "cross", "predeploy"],
                        help="Run a single stage validation")
    parser.add_argument("--all-stages", action="store_true",
                        help="Run all 5 stage validations sequentially")
    parser.add_argument("--week", help="Week ISO for ledger-load (YYYY-MM-DD)")
    parser.add_argument("--excel", help="Excel path for key-build validation")
    args = parser.parse_args()

    # Set up logging
    logger = ValidationLogger(LOGS_DIR)
    logger.print(f"Validation started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.print(f"Log file: {logger.log_path}")

    try:
        if args.stage:
            # Single stage
            if args.stage == "masterdata":
                md_path = find_latest_masterdata(MASTER_DATA)
                if not md_path:
                    logger.print("ERROR: No MasterData CSV found")
                    return 1
                r = validate_masterdata(md_path, logger)
                return r.failed

            elif args.stage == "ledger-load":
                r = validate_ledger_load(args.week, logger)
                return r.failed

            elif args.stage == "key-build":
                excel = Path(args.excel) if args.excel else None
                r = validate_key_build(excel, logger)
                return r.failed

            elif args.stage == "dashboard":
                r = validate_dashboard(logger=logger)
                return r.failed

            elif args.stage == "cross":
                r = validate_cross_stage(logger=logger)
                return r.failed

            elif args.stage == "predeploy":
                result = run_all_checks(logger=logger)
                return result["exit_code"]

        elif args.all_stages:
            # Run all 5 stages
            total_failures = 0
            stage_results: Dict[str, Results] = {}

            # Stage 1: MasterData
            md_path = find_latest_masterdata(MASTER_DATA)
            if md_path:
                r1 = validate_masterdata(md_path, logger)
                stage_results["masterdata"] = r1
                total_failures += r1.failed
                if r1.failed > 0:
                    logger.print(f"\n  PIPELINE STOPPED: Stage 1 (MasterData) has {r1.failed} failure(s)")
                    return total_failures
            else:
                logger.print("  WARNING: No MasterData CSV found, skipping stage 1")

            # Stage 2: SQLite Load
            r2 = validate_ledger_load(logger=logger)
            stage_results["ledger-load"] = r2
            total_failures += r2.failed
            if r2.failed > 0:
                logger.print(f"\n  PIPELINE STOPPED: Stage 2 (SQLite Load) has {r2.failed} failure(s)")
                return total_failures

            # Stage 3: Key Build
            r3 = validate_key_build(logger=logger)
            stage_results["key-build"] = r3
            total_failures += r3.failed
            if r3.failed > 0:
                logger.print(f"\n  PIPELINE STOPPED: Stage 3 (Key Build) has {r3.failed} failure(s)")
                return total_failures

            # Stage 4: Dashboard
            r4 = validate_dashboard(logger=logger)
            stage_results["dashboard"] = r4
            total_failures += r4.failed
            if r4.failed > 0:
                logger.print(f"\n  PIPELINE STOPPED: Stage 4 (Dashboard) has {r4.failed} failure(s)")
                return total_failures

            # Stage 5: Cross-stage
            r5 = validate_cross_stage(stage_results, logger)
            total_failures += r5.failed

            # Final summary
            logger.print(f"\n{'=' * 70}")
            logger.print(f"  ALL-STAGES SUMMARY")
            logger.print(f"{'=' * 70}")
            stages = [
                ("MasterData", stage_results.get("masterdata")),
                ("SQLite Load", r2),
                ("Key Build", r3),
                ("Dashboard", r4),
                ("Cross-Stage", r5),
            ]
            for name, sr in stages:
                if sr:
                    status = "PASS" if sr.failed == 0 else "FAIL"
                    logger.print(f"  [{status}] {name}: {sr.passed} passed, {sr.failed} failed, {sr.warnings} warnings")
            logger.print(f"{'=' * 70}")
            if total_failures == 0:
                logger.print(f"  ALL STAGE VALIDATIONS PASSED")
            else:
                logger.print(f"  *** {total_failures} TOTAL FAILURE(S) ***")
            logger.print(f"{'=' * 70}")
            return total_failures

        else:
            # Legacy mode (backward compatible)
            return _run_legacy_validation(args.weeks, logger)

    finally:
        logger.close()
        ValidationLogger.cleanup(LOGS_DIR)


if __name__ == "__main__":
    raise SystemExit(main())
