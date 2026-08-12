"""Executable local contract, sanitization, and real-browser gate.

The browser gate serves only this pack over loopback.  It first tries an
explicit ``SYNTHETIC_BROWSER_EXECUTABLE``, then the installed Microsoft Edge
channel/standard Edge locations, and finally a Playwright-managed Chromium.
Browser binaries are never downloaded into the pack.
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import importlib.metadata
import json
import os
import re
import sqlite3
import sys
import tempfile
import threading
import zipfile
import zlib
from contextlib import contextmanager
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import urlparse


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "scripts" / "synthetic_contract.json"
TEXT_EXTENSIONS = {
    ".bat", ".cmd", ".css", ".csv", ".d.ts", ".html", ".js", ".json",
    ".lock", ".map", ".md", ".ps1", ".py", ".sql", ".svg", ".ts",
    ".txt", ".xml", ".yaml", ".yml",
}
SCREENSHOT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
PROHIBITED_DIRS = {
    "node_modules", "venv", ".venv", "env", "__pycache__", ".pytest_cache",
    ".mypy_cache", ".ruff_cache", ".cache", "cache",
}
MOJIBAKE = re.compile(r"(?:\u00c3[\u0080-\u00bf]|\u00c2[\u0080-\u00bf]|\u00e2[\u0080-\u00bf]{2}|\u00f0\u0178|\ufffd)")
PRIVATE_URL = re.compile(
    r"https?://(?!127\.0\.0\.1(?::\d+)?(?![0-9.])|localhost(?::\d+)?(?![a-z0-9.-])|www\.w3\.org/2000/svg(?:[\"'\s>]|$))",
    re.IGNORECASE,
)
REMOTE_PREFIXES = ("http" + "://", "https" + "://", "//")

# Split across physical lines so this scanner can inspect its own source without
# treating its denylist declarations as leaked content.
def _chars(*codes: int) -> str:
    return "".join(map(chr, codes))


_TOKEN_PARTS = (
    (_chars(101, 103),
     _chars(103, 114, 111, 117, 112)),
    ("eu",
     "rogarages"),
    ("cum",
     "berland"),
    ("bru",
     "no"),
    ("alex",
     "rivera"),
    ("sam",
     "taylor"),
    ("huzaifah",
     "patel"),
    (_chars(105, 109, 114, 97, 110),
     "patel"),
    ("zaid",
     "patel"),
    ("zainab",
     "patel"),
    ("ozair",
     "patel"),
    ("ch",
     "ubb"),
    ("trans",
     "gourmet"),
    ("super",
     "group"),
    ("prox",
     "imy"),
    ("br",
     "inks"),
    ("carrefour",
     "proximite"),
    ("hu",
     "bency"),
    ("en",
     "gie"),
)
FORBIDDEN_TOKENS = tuple("".join(parts).casefold() for parts in _TOKEN_PARTS)
FORBIDDEN_COMPACT_TOKENS = frozenset(
    token
    for token in FORBIDDEN_TOKENS
    if len(token) >= 8
    or token == _chars(101, 103, 103, 114, 111, 117, 112)
)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_js_object(path: Path, symbol: str) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    match = re.search(r"(?:(?:window\.)|(?:const\s+))" + re.escape(symbol) + r"\s*=\s*", text)
    if match is None:
        raise ValueError(f"missing JavaScript assignment for {symbol}: {path}")
    payload = text[match.end() :]
    value, _ = json.JSONDecoder().raw_decode(payload)
    if not isinstance(value, dict):
        raise TypeError(f"{symbol} must be an object")
    return value


def unpack(value: str) -> Any:
    return json.loads(zlib.decompress(base64.b64decode(value)).decode("utf-8"))


def load_rows(payload: dict[str, Any], cfg: dict[str, Any]) -> list[dict[str, Any]]:
    kind = cfg["rows_kind"]
    if kind == "projects":
        rows = payload.get("projects", [])
    elif kind == "compressed_snapshots":
        rows = unpack(payload["compressed_snapshots"][cfg["rows_key"]])
    elif kind == "compressed_weeks_raw":
        rows = unpack(payload["compressed_weeks"][cfg["rows_key"]])["raw"]
    elif kind == "compressed_weeks_list":
        rows = unpack(payload["compressed_weeks"][cfg["rows_key"]])
    elif kind == "compressed_invoices":
        rows = unpack(payload["compressed_invoices"])
    else:
        raise ValueError(f"unknown rows_kind: {kind}")
    if not isinstance(rows, list) or not rows:
        raise ValueError("fixture contains no rows")
    if not all(isinstance(row, dict) for row in rows):
        raise TypeError("fixture rows must be objects")
    return rows


def first_value(row: dict[str, Any], keys: tuple[str, ...], default: Any = "") -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def normalize(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        normalized.append(
            {
                "record_id": str(first_value(row, ("ref", "reference", "document_number", "dn", "s"))),
                "amount": float(first_value(row, ("amount", "amt", "tv", "total"), 0) or 0),
                "owner": str(first_value(row, ("owner", "o", "team_member"), "Unassigned")),
                "status": str(first_value(row, ("status", "st", "ps", "sla"), "Open")),
                "country": str(first_value(row, ("country", "co", "cty"), "Unknown")),
            }
        )
    if any(not row["record_id"] for row in normalized):
        raise ValueError("normalized record without record_id")
    return normalized


def check_required_paths(cfg: dict[str, Any]) -> list[str]:
    return [path for path in cfg["required_paths"] if not (ROOT / path).exists()]


def check_filters(payload: dict[str, Any], cfg: dict[str, Any]) -> dict[str, Any]:
    source = payload.get("filters", {})
    values: dict[str, int] = {}
    missing = []
    for key in cfg["filter_keys"]:
        candidate = source.get(key) if isinstance(source, dict) else None
        if candidate is None:
            candidate = payload.get(key)
        values[key] = len(candidate) if isinstance(candidate, list) else 0
        if not isinstance(candidate, list) or not candidate:
            missing.append(key)
    return {"status": "GREEN" if not missing else "RED", "counts": values, "missing": missing}


def check_ui(cfg: dict[str, Any]) -> dict[str, Any]:
    html_path = ROOT / cfg["entrypoint"]
    text = html_path.read_text(encoding="utf-8")
    refs = re.findall(r"(?:src|href)=[\"']([^\"']+)", text)
    remote_refs = [ref for ref in refs if ref.startswith(REMOTE_PREFIXES)]
    local_refs = [ref.split("?", 1)[0] for ref in refs if not ref.startswith(("#", "data:") + REMOTE_PREFIXES)]
    missing = [ref for ref in local_refs if not (html_path.parent / ref).exists()]
    tokens = {token: token in text for token in cfg["ui_tokens"]}
    errors = [f"remote UI resource: {ref}" for ref in remote_refs]
    errors.extend(missing)
    errors.extend(f"missing UI token: {key}" for key, present in tokens.items() if not present)
    return {
        "status": "GREEN" if not errors else "RED",
        "local_refs": local_refs,
        "remote_refs": remote_refs,
        "missing": errors,
        "tokens": tokens,
    }


def _text_findings(text: str, *, check_urls: bool = True) -> list[str]:
    findings = []
    if "\ufffd" in text:
        findings.append("replacement character U+FFFD")
    controls = sorted({ord(char) for char in text if ord(char) < 0x20 and ord(char) not in (0x09, 0x0A, 0x0D)})
    if controls:
        findings.append("C0 control characters: " + ", ".join(f"U+{value:04X}" for value in controls))
    if "\x7f" in text:
        findings.append("DEL control character U+007F")
    if MOJIBAKE.search(text):
        findings.append("mojibake sequence")
    folded = text.casefold()
    compact_lines = [re.sub(r"[^a-z0-9]+", "", line.casefold()) for line in text.splitlines()]
    for token in FORBIDDEN_TOKENS:
        direct = re.search(rf"(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])", folded)
        derived = token in FORBIDDEN_COMPACT_TOKENS and any(token in line for line in compact_lines)
        if direct or derived:
            findings.append(f"forbidden operational token: {token}")
    if check_urls and PRIVATE_URL.search(text):
        findings.append("non-loopback URL")
    return findings


def scan_text(
    path: Path,
    raw: bytes,
    coverage: dict[str, int],
    findings: list[str],
    parse_json: bool = False,
    check_urls: bool = True,
) -> None:
    coverage["text"] += 1
    try:
        text = raw.decode("utf-8", errors="strict")
    except UnicodeDecodeError as exc:
        findings.append(f"invalid UTF-8 {path.relative_to(ROOT)}: {exc}")
        return
    relative = path.relative_to(ROOT)
    findings.extend(f"{finding}: {relative}" for finding in _text_findings(text, check_urls=check_urls))
    if parse_json:
        try:
            json.loads(text)
            coverage["json"] += 1
        except json.JSONDecodeError as exc:
            findings.append(f"invalid JSON {relative}: {exc}")


def _scan_value(label: str, value: Any, findings: list[str]) -> None:
    if value is None:
        return
    findings.extend(f"{finding}: {label}" for finding in _text_findings(str(value)))


def scan_unclassified(path: Path, raw: bytes, coverage: dict[str, int], findings: list[str]) -> None:
    relative = path.relative_to(ROOT)
    try:
        text = raw.decode("utf-8", errors="strict")
    except UnicodeDecodeError:
        text = ""
    if text and "\x00" not in text:
        scan_text(path, raw, coverage, findings)
        return
    coverage["binary"] += 1
    binary_text = raw.decode("latin1", errors="ignore")
    for finding in _text_findings(binary_text):
        if "control" not in finding and "mojibake" not in finding:
            findings.append(f"binary {finding}: {relative}")


def scan_files() -> dict[str, Any]:
    findings: list[str] = []
    coverage = {"text": 0, "json": 0, "workbook": 0, "sqlite": 0, "binary": 0, "screenshots": 0, "unknown": 0}
    for directory in ROOT.rglob("*"):
        if directory.is_dir() and directory.name.casefold() in PROHIBITED_DIRS:
            findings.append(f"prohibited generated directory: {directory.relative_to(ROOT)}")
    for path in ROOT.rglob("*"):
        if not path.is_file() or path == ROOT / "manifest.json" or any(part.casefold() in PROHIBITED_DIRS for part in path.parts):
            continue
        suffix = path.suffix.lower()
        raw = path.read_bytes()
        if suffix in SCREENSHOT_EXTENSIONS:
            coverage["screenshots"] += 1
            findings.append(f"screenshot requires OCR review: {path.relative_to(ROOT)}")
            continue
        if suffix in TEXT_EXTENSIONS or path.name in {"package-lock.json"}:
            scan_text(
                path,
                raw,
                coverage,
                findings,
                parse_json=suffix in {".json", ".map"} or path.name == "package-lock.json",
                check_urls=path.name != "package-lock.json",
            )
            continue
        if suffix in {".xlsx", ".xlsm"}:
            coverage["workbook"] += 1
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
                for field in ("creator", "lastModifiedBy", "title", "subject", "description", "keywords", "category"):
                    _scan_value(f"workbook metadata {path.relative_to(ROOT)}", getattr(workbook.properties, field, None), findings)
                for sheet in workbook.worksheets:
                    _scan_value(f"workbook sheet {path.relative_to(ROOT)}", sheet.title, findings)
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            _scan_value(f"workbook cell {path.relative_to(ROOT)}", cell, findings)
                workbook.close()
            except Exception as exc:
                findings.append(f"workbook inspection failed {path.relative_to(ROOT)}: {exc}")
            continue
        if suffix in {".sqlite", ".db"}:
            coverage["sqlite"] += 1
            try:
                with sqlite3.connect(f"file:{path}?mode=ro", uri=True) as conn:
                    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
                    if integrity != "ok":
                        findings.append(f"SQLite integrity: {path.relative_to(ROOT)}={integrity}")
                    for name, sql in conn.execute("SELECT name, sql FROM sqlite_master WHERE sql IS NOT NULL"):
                        _scan_value(f"SQLite schema {path.relative_to(ROOT)}:{name}", sql, findings)
                    tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
                    for (table,) in tables:
                        columns = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
                        for column in (col[1] for col in columns if col[2].upper() in {"TEXT", ""}):
                            for (value,) in conn.execute(f'SELECT "{column}" FROM "{table}" WHERE "{column}" IS NOT NULL'):
                                _scan_value(f"SQLite value {path.relative_to(ROOT)}:{table}.{column}", value, findings)
            except Exception as exc:
                findings.append(f"SQLite inspection failed {path.relative_to(ROOT)}: {exc}")
            continue
        if suffix in {".zip", ".pptx", ".docx"}:
            coverage["binary"] += 1
            try:
                with zipfile.ZipFile(path) as archive:
                    for member in archive.namelist():
                        try:
                            member_text = archive.read(member).decode("utf-8", errors="strict")
                        except UnicodeDecodeError:
                            continue
                        for finding in _text_findings(member_text):
                            findings.append(f"archive {finding}: {path.relative_to(ROOT)}:{member}")
            except Exception as exc:
                findings.append(f"archive inspection failed {path.relative_to(ROOT)}: {exc}")
            continue
        if suffix in {".woff", ".woff2", ".ttf", ".otf"}:
            coverage["binary"] += 1
            continue
        scan_unclassified(path, raw, coverage, findings)
    return {"status": "GREEN" if not findings else "RED", "findings": sorted(set(findings)), "coverage": coverage}


class _QuietHandler(SimpleHTTPRequestHandler):
    def log_message(self, _format: str, *args: object) -> None:
        return


@contextmanager
def _local_server() -> Iterator[str]:
    handler = lambda *args, **kwargs: _QuietHandler(*args, directory=str(ROOT), **kwargs)  # noqa: E731
    server = ThreadingHTTPServer(("127.0.0.1", 0), handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield "http" + f"://127.0.0.1:{server.server_port}"
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _launch_browser(playwright: Any) -> tuple[Any, str]:
    attempts: list[str] = []
    explicit = os.environ.get("SYNTHETIC_BROWSER_EXECUTABLE")
    candidates: list[tuple[str, dict[str, Any]]] = []
    if explicit:
        candidates.append(("SYNTHETIC_BROWSER_EXECUTABLE", {"executable_path": explicit}))
    candidates.append(("installed Microsoft Edge channel", {"channel": "msedge"}))
    for label, variable in (
        ("Microsoft Edge Program Files x86", "ProgramFiles(x86)"),
        ("Microsoft Edge Program Files", "ProgramFiles"),
    ):
        configured_root = os.environ.get(variable)
        if not configured_root:
            continue
        path = Path(configured_root) / "Microsoft/Edge/Application/msedge.exe"
        if path.is_file():
            candidates.append((label, {"executable_path": str(path)}))
    candidates.append(("Playwright-managed Chromium", {}))
    for label, kwargs in candidates:
        try:
            return playwright.chromium.launch(headless=True, **kwargs), label
        except Exception as exc:  # pragma: no cover - depends on local browser inventory
            attempts.append(f"{label}: {type(exc).__name__}: {exc}")
    raise RuntimeError("no browser fallback launched; " + " | ".join(attempts))


def _action(page: Any, actions: list[dict[str, Any]], name: str, callback: Any) -> None:
    try:
        detail = callback()
        actions.append({"name": name, "status": "GREEN", "detail": detail})
    except Exception as exc:
        actions.append({"name": name, "status": "RED", "error": f"{type(exc).__name__}: {exc}"})


def _project_actions(page: Any, cfg: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    actions: list[dict[str, Any]] = []
    project_id = cfg["project_id"]
    checks: dict[str, Any] = {}
    if project_id == "project-analytics-01-banking":
        checks = page.evaluate("""() => {
            const trend = window.DASHBOARD_DATA && DASHBOARD_DATA.trend_summary;
            const dates = trend ? trend.dates.slice().sort() : [];
            const aggregate = trend ? _aggregateTrendData(dates, trend.by_date || {}) : null;
            return {dates, aggregate, finite: !!aggregate && aggregate.volumes.every(Number.isFinite) && aggregate.values.every(Number.isFinite)};
        }""")
        _action(page, actions, "filter system ERP1", lambda: (page.select_option("#filterSystem", "ERP1"), page.wait_for_function("() => document.querySelector('#filterSystem').value === 'ERP1' && document.body.innerText.includes('ERP1')"), page.locator("#filterSystem").input_value())[-1])
        _action(page, actions, "open Productivity", lambda: (page.locator('.nav-item[data-page="page-productivity"]').click(), page.wait_for_function("() => document.querySelector('.nav-item[data-page=\"page-productivity\"]').classList.contains('active')"), "active")[-1])
    elif project_id == "project-analytics-02-ledger":
        page.wait_for_function("() => !document.getElementById('loadingOverlay') || document.getElementById('loadingOverlay').classList.contains('hidden')")
        _action(page, actions, "filter Overview to ROL", lambda: (page.select_option("#overviewTeamFilter", "ROL"), page.wait_for_function("() => document.querySelector('#overviewTeamFilter').value === 'ROL' && document.querySelector('#overviewTable').innerText.includes('SYN-SUP-001')"), page.locator("#overviewTable").inner_text())[-1])
        _action(page, actions, "search synthetic supplier", lambda: (page.locator("#supplierSearch").fill("SYN-SUP-001"), page.wait_for_function("() => document.querySelector('#supplierSearch').value === 'SYN-SUP-001' && document.querySelector('#overviewTable').innerText.includes('SYN-SUP-001')"), page.locator("#supplierSearch").input_value())[-1])
        _action(page, actions, "refresh filters", lambda: (page.locator("#refreshFiltersBtn").click(), page.wait_for_function("() => document.querySelector('#overviewTeamFilter').value === ''"), "reset")[-1])
        _action(page, actions, "open ROL Team", lambda: (page.locator('.nav-item[data-page="rol"]').click(), page.wait_for_function("() => document.querySelector('.nav-item[data-page=\"rol\"]').classList.contains('active')"), "active")[-1])
        checks = page.evaluate("() => ({visibleRows:Array.from(document.querySelectorAll('tbody tr')).filter(row => row.getClientRects().length).length})")
    elif project_id == "project-analytics-03-dashboard-project":
        _action(page, actions, "open All Projects", lambda: (page.locator('.nav-item[data-page="detail"]').click(), page.wait_for_function("() => document.querySelector('.nav-item[data-page=\"detail\"]').classList.contains('active')"), "active")[-1])
        _action(page, actions, "search synthetic project", lambda: (page.locator("#searchBox").fill("SYN-PROJ-001"), page.wait_for_function("() => document.body.innerText.includes('SYN-PROJ-001')"), page.locator("#searchBox").input_value())[-1])
        _action(page, actions, "toggle theme", lambda: (page.locator("#themeToggle").click(), page.wait_for_function("() => document.body.classList.contains('dark')"), "dark")[-1])
        checks = page.evaluate("() => ({projectCount:window.PROJECT_DATA && PROJECT_DATA.projects.length, visibleRows:Array.from(document.querySelectorAll('tbody tr')).filter(row => row.getClientRects().length).length})")
    elif project_id == "project-analytics-04-invoice-process":
        page.wait_for_function("() => !document.body.innerText.includes('No data available for this view')")
        _action(page, actions, "filter synthetic owner", lambda: (page.select_option("#filterOwner", {"label": "Synthetic Owner 001"}), page.wait_for_function("() => document.querySelector('#filterOwner').selectedOptions[0].textContent.includes('Synthetic Owner 001')"), page.locator("#filterOwner").input_value())[-1])
        _action(page, actions, "open Detail", lambda: (page.locator('.sb-nav-item[data-page="detail"]').click(), page.wait_for_function("() => document.querySelector('.sb-nav-item[data-page=\"detail\"]').classList.contains('active')"), "active")[-1])
        _action(page, actions, "open Trends", lambda: (page.locator('.sb-nav-item[data-page="trends"]').click(), page.wait_for_function("() => document.querySelector('.sb-nav-item[data-page=\"trends\"]').classList.contains('active')"), "active")[-1])
        checks = page.evaluate("() => ({visibleRows:Array.from(document.querySelectorAll('tbody tr')).filter(row => row.getClientRects().length).length, noData:document.body.innerText.includes('No data available for this view'), compressedKeys:Object.keys(DASHBOARD_DATA.compressed_weeks || {})})")
    elif project_id == "project-analytics-05-payment":
        _action(page, actions, "expand fixture date range", lambda: page.evaluate("""async () => { document.querySelector('#dateFromPicker').value='2026-08-09'; document.querySelector('#dateToPicker').value='2026-08-10'; await onDateRangeChange(); return {from:dateFrom,to:dateTo,rows:paymentRenderCache.filtered.length}; }"""))
        _action(page, actions, "filter transaction type Payments", lambda: (page.locator("#docClassFilter").wait_for(state="visible"), page.select_option("#docClassFilter", {"label": "Payments"}), page.wait_for_function("() => window.docClassFilter === 'payment' || docClassFilter === 'payment'"), page.locator("#docClassFilter").input_value())[-1])
        _action(page, actions, "refresh filters", lambda: (page.locator("#refreshFiltersBtn").click(), page.wait_for_function("() => document.querySelector('#docClassFilter').value === ''"), "reset")[-1])
        _action(page, actions, "open Suppliers", lambda: (page.locator("#tab-paymentSuppliers").click(), page.wait_for_function("() => document.querySelector('#tab-paymentSuppliers').classList.contains('active')"), "active")[-1])
        _action(page, actions, "open Blocks", lambda: (page.locator("#tab-paymentBlocks").click(), page.wait_for_function("() => document.querySelector('#tab-paymentBlocks').classList.contains('active')"), "active")[-1])
        checks = page.evaluate("""async () => { const rows=await getWeekData('2026-08-10'); return {rowCount:rows.length, canonical:rows.filter(_pdIsCanonicalMetricRow).length, paid:rows.filter(_pdIsCanonicalMetricRow).filter(_pdIsPaidRow).length, blocks:rows.map(row => row.pb)}; }""")
    else:
        raise ValueError(f"no browser interaction profile for {project_id}")
    return actions, checks


def run_browser_smoke(cfg: dict[str, Any], evidence_dir: Path | None = None) -> dict[str, Any]:
    evidence_dir = evidence_dir or Path(tempfile.gettempdir()) / "project-analytics-e2e" / cfg["project_id"]
    evidence_dir.mkdir(parents=True, exist_ok=True)
    result: dict[str, Any] = {
        "status": "RED", "method": "Playwright with documented local-browser fallback",
        "console_errors": [], "console_warnings": [], "page_errors": [], "request_failures": [],
        "bad_responses": [], "external_requests": [], "actions": [],
    }
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        result["reason"] = f"Playwright unavailable: {type(exc).__name__}: {exc}"
        return result
    try:
        with _local_server() as base_url, sync_playwright() as playwright:
            browser, fallback = _launch_browser(playwright)
            result["browser"] = {
                "engine": "Chromium", "version": browser.version, "fallback": fallback,
                "playwright": importlib.metadata.version("playwright"),
            }
            context = browser.new_context(viewport={"width": 1440, "height": 900}, locale="en-GB")
            page = context.new_page()
            page.set_default_timeout(10_000)
            page.on("console", lambda msg: result["console_errors" if msg.type == "error" else "console_warnings"].append(msg.text) if msg.type in {"error", "warning"} else None)
            page.on("pageerror", lambda error: result["page_errors"].append(str(error)))
            page.on("requestfailed", lambda request: result["request_failures"].append({"url": request.url, "failure": request.failure}))
            page.on("response", lambda response: result["bad_responses"].append({"url": response.url, "status": response.status}) if response.status >= 400 else None)
            page.on("request", lambda request: result["external_requests"].append(request.url) if urlparse(request.url).hostname not in {"127.0.0.1", "localhost"} and urlparse(request.url).scheme in {"http", "https"} else None)
            url = f"{base_url}/{cfg['entrypoint']}"
            response = page.goto(url, wait_until="domcontentloaded", timeout=20_000)
            result["url"] = url
            result["http_status"] = response.status if response else None
            page.wait_for_function(f"() => typeof window[{json.dumps(cfg['data_symbol'])}] === 'object'")
            actions, checks = _project_actions(page, cfg)
            result["actions"] = actions
            result["checks"] = checks
            body_text = page.locator("body").inner_text()
            result["dom_forbidden"] = _text_findings(body_text)
            screenshot = evidence_dir / f"{cfg['project_id']}.png"
            page.screenshot(path=str(screenshot), full_page=True)
            result["screenshot"] = str(screenshot.resolve())
            result["screenshot_sha256"] = hashlib.sha256(screenshot.read_bytes()).hexdigest()
            context.close()
            browser.close()
    except Exception as exc:
        result["reason"] = f"browser execution failed: {type(exc).__name__}: {exc}"
        return result
    failures = (
        result["console_errors"] + result["page_errors"] + result["request_failures"]
        + result["bad_responses"] + result["external_requests"] + result["dom_forbidden"]
        + [action for action in result["actions"] if action["status"] != "GREEN"]
    )
    if result.get("http_status") != 200:
        failures.append(f"entrypoint HTTP status {result.get('http_status')}")
    checks = result.get("checks", {})
    if cfg["project_id"] == "project-analytics-01-banking" and not checks.get("finite"):
        failures.append("trend aggregate contains non-finite values")
    if cfg["project_id"] == "project-analytics-03-dashboard-project" and checks.get("projectCount", 0) < 3:
        failures.append("project dashboard rendered fewer than three projects")
    if cfg["project_id"] == "project-analytics-04-invoice-process" and (checks.get("noData") or checks.get("visibleRows", 0) < 2):
        failures.append("invoice dashboard did not render two fixture rows")
    if cfg["project_id"] == "project-analytics-05-payment" and (checks.get("paid", 0) < 2 or not all(checks.get("blocks", []))):
        failures.append("payment canonical/paid/block contract failed")
    result["failures"] = failures
    result["status"] = "GREEN" if not failures else "RED"
    if failures:
        result["reason"] = "browser smoke captured one or more failures"
    return result


def run_pipeline(write_manifest: bool = True, run_browser: bool = True, evidence_dir: Path | None = None) -> dict[str, Any]:
    cfg = read_json(CONFIG_PATH)
    errors: list[str] = []
    required_missing = check_required_paths(cfg)
    errors.extend(f"missing required path: {path}" for path in required_missing)
    data_path = ROOT / cfg["data_js"]
    payload: dict[str, Any] = {}
    rows: list[dict[str, Any]] = []
    try:
        payload = load_js_object(data_path, cfg["data_symbol"])
        rows = load_rows(payload, cfg)
    except Exception as exc:
        errors.append(f"load/ETL source failed: {exc}")
    normalized: list[dict[str, Any]] = []
    if rows:
        try:
            normalized = normalize(rows)
        except Exception as exc:
            errors.append(f"normalization failed: {exc}")
    filters = check_filters(payload, cfg) if payload else {"status": "RED", "missing": ["payload"]}
    errors.extend(f"filter contract missing: {key}" for key in filters.get("missing", []))
    ui = check_ui(cfg) if (ROOT / cfg["entrypoint"]).exists() else {"status": "RED", "missing": [cfg["entrypoint"]]}
    errors.extend(ui.get("missing", []))
    query = {"status": "GREEN", "row_count": len(normalized), "amount_total": round(sum(row["amount"] for row in normalized), 2)}
    if not normalized:
        query["status"] = "RED"
        errors.append("query returned no normalized rows")
    sqlite_counts: dict[str, int] = {}
    for db_rel in cfg.get("sqlite_paths", []):
        db_path = ROOT / db_rel
        try:
            with sqlite3.connect(f"file:{db_path}?mode=ro", uri=True) as conn:
                sqlite_counts[db_rel] = sum(row[0] for row in conn.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table'"))
        except Exception as exc:
            errors.append(f"handoff SQLite unavailable: {db_rel}: {exc}")
    scan = scan_files()
    errors.extend(scan["findings"])
    browser_smoke = run_browser_smoke(cfg, evidence_dir) if run_browser else {"status": "SKIPPED", "reason": "explicit --no-browser diagnostic mode"}
    if browser_smoke["status"] != "GREEN":
        errors.extend(str(item) for item in browser_smoke.get("failures", []))
        if not browser_smoke.get("failures"):
            errors.append(browser_smoke.get("reason", "browser smoke failed"))
    status = "GREEN" if not errors else "RED"
    result = {
        "schema_version": 2,
        "project_id": cfg["project_id"],
        "status": status,
        "stages": {
            "load": "GREEN" if rows else "RED",
            "etl_transform": "GREEN" if normalized else "RED",
            "query": query,
            "filters": filters,
            "ui_static_smoke": ui,
            "browser_smoke": browser_smoke,
            "handoff": "GREEN" if not sqlite_counts or all(value > 0 for value in sqlite_counts.values()) else "RED",
            "quality_scan": scan,
        },
        "evidence": {"source_rows": len(rows), "normalized_rows": len(normalized), "sqlite_tables": sqlite_counts},
        "errors": errors,
        "scan_scope": {
            "strict_utf8_c0_replacement_mojibake": True,
            "text_json_workbook_sqlite_binary_screenshots": True,
            "dom_and_derived_variants": True,
            "interactive_browser_required_for_definitive": True,
        },
    }
    if write_manifest:
        serialized = json.dumps(result, indent=2, ensure_ascii=True) + "\n"
        (ROOT / "manifest.json").write_text(serialized, encoding="utf-8")
        runtime_manifest = ROOT / "runtime" / "manifests" / "e2e_latest.json"
        runtime_manifest.parent.mkdir(parents=True, exist_ok=True)
        runtime_manifest.write_text(serialized, encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--no-browser", action="store_true", help="diagnostic only; a skipped browser can never produce definitive GREEN")
    parser.add_argument("--no-write-manifest", action="store_true")
    parser.add_argument("--evidence-dir", type=Path)
    args = parser.parse_args()
    result = run_pipeline(not args.no_write_manifest, not args.no_browser, args.evidence_dir)
    print(json.dumps(result, indent=2, ensure_ascii=True))
    return 0 if result["status"] == "GREEN" else 1


if __name__ == "__main__":
    raise SystemExit(main())
