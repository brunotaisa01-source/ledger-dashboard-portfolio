# -*- coding: utf-8 -*-
"""
Escalations loader.

Reads the Escalations workbook, normalizes source-owned fields, and replaces
the escalation_lines table in escalation_daily.sqlite.
"""
from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import shutil
import sqlite3
import time
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..utils.paths import (
    ESCALATION_ARCHIVE,
    ESCALATION_DB,
    ESCALATION_FILE,
    ESCALATION_SOURCE,
    get_escalation_source,
    sync_db_to_fixture_store,
)


def _escalation_archive_disabled() -> bool:
    return os.getenv("Synthetic_REPORTING_DISABLE_ESCALATION_ARCHIVE", "1").strip().lower() not in {"0", "false", "no", "off"}

logger = logging.getLogger(__name__)

SHEET_NAME = "Escalations "
HEADER_ROW = 5
DATA_START_ROW = 6
_EXCEL_SERIAL_DATE_RE = re.compile(r"^\d+(?:\.\d+)?$")
STALE_OPEN_DAYS = 30

SOURCE_TO_DB = {
    "Category": "Category",
    "Mailbox": "Mailbox",
    "From": "FromEmail",
    "Vendor": "VendorNo",
    "Name": "VendorName",
    "Entity": "Entity",
    "Reference": "Reference",
    "Doc Date": "DocDate",
    "Inv ref": "InvRef",
    "Value": "ValueRaw",
    "Action_Type": "ActionType",
    "Status": "StatusRaw",
    "Priority": "Priority",
    "AP Owner": "APOwner",
    "Received_Date": "ReceivedDate",
    "Escalation_Date": "EscalationDate",
    "Working_Notes": "WorkingNotes",
    "Date Resolved": "DateResolved",
    "Days_To_Resolve": "DaysToResolveSource",
    "Internet_Message_ID": "InternetMsgId",
    "UniqueKey": "UniqueKey",
}

INSERT_COLUMNS = [
    "UniqueKey",
    "LoadedAt",
    "Category",
    "Mailbox",
    "FromEmail",
    "VendorNo",
    "VendorName",
    "Entity",
    "EntityCode",
    "Reference",
    "DocDate",
    "InvRef",
    "Value",
    "ValueRaw",
    "ActionType",
    "Status",
    "StatusRaw",
    "IsOpen",
    "Priority",
    "APOwner",
    "ReceivedDate",
    "EscalationDate",
    "WorkingNotes",
    "DateResolved",
    "DaysToResolveSource",
    "DaysToResolveCalc",
    "DaysOpen",
    "InternetMsgId",
    "MasterCategory",
    "MasterPriority",
    "MasterAPOwner",
    "Flags",
]

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS escalation_lines (
    UniqueKey TEXT PRIMARY KEY,
    LoadedAt TEXT NOT NULL,
    Category TEXT,
    Mailbox TEXT,
    FromEmail TEXT,
    VendorNo TEXT,
    VendorName TEXT,
    Entity TEXT,
    EntityCode TEXT,
    Reference TEXT,
    DocDate TEXT,
    InvRef TEXT,
    Value REAL,
    ValueRaw TEXT,
    ActionType TEXT,
    Status TEXT,
    StatusRaw TEXT,
    IsOpen INTEGER,
    Priority TEXT,
    APOwner TEXT,
    ReceivedDate TEXT,
    EscalationDate TEXT,
    WorkingNotes TEXT,
    DateResolved TEXT,
    DaysToResolveSource INTEGER,
    DaysToResolveCalc INTEGER,
    DaysOpen INTEGER,
    InternetMsgId TEXT,
    MasterCategory TEXT,
    MasterPriority TEXT,
    MasterAPOwner TEXT,
    Flags TEXT
);
CREATE INDEX IF NOT EXISTS idx_esc_status ON escalation_lines(Status);
CREATE INDEX IF NOT EXISTS idx_esc_isopen ON escalation_lines(IsOpen);
CREATE INDEX IF NOT EXISTS idx_esc_vendor ON escalation_lines(VendorNo);
CREATE INDEX IF NOT EXISTS idx_esc_entity ON escalation_lines(EntityCode);
CREATE INDEX IF NOT EXISTS idx_esc_actiontype ON escalation_lines(ActionType);
CREATE INDEX IF NOT EXISTS idx_esc_escdate ON escalation_lines(EscalationDate);
CREATE VIEW IF NOT EXISTS esc_open AS
    SELECT * FROM escalation_lines WHERE IsOpen = 1;
CREATE VIEW IF NOT EXISTS esc_closed AS
    SELECT * FROM escalation_lines WHERE IsOpen = 0;
"""

ENTITY_CODE_RE = re.compile(r"^([^-]+)-")


@dataclass(frozen=True)
class LoadResult:
    rows_loaded: int
    source: str
    db_path: str | None
    dry_run: bool = False

    def as_dict(self) -> dict[str, Any]:
        return {
            "rows_loaded": self.rows_loaded,
            "source": self.source,
            "db_path": self.db_path,
            "dry_run": self.dry_run,
        }


def init_db(conn: sqlite3.Connection) -> None:
    """Create escalation schema, indexes, and views."""
    conn.executescript(SCHEMA_SQL)


def refresh_escalation_file(
    source_path: Path | None = None,
    target_path: Path = ESCALATION_FILE,
    archive_dir: Path = ESCALATION_ARCHIVE,
    *,
    attempts: int = 3,
    sleep_seconds: float = 1.0,
) -> Path:
    """Copy Local Fixture Store source to local data folder, archiving previous target."""
    source_path = _resolve_source_path(source_path)
    target_path = Path(target_path)
    archive_dir = Path(archive_dir)
    if not source_path.exists():
        raise FileNotFoundError(f"Escalation source not found: {source_path}")

    target_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        if source_path.resolve() == target_path.resolve():
            logger.info("Escalation source is already pack-local: %s", target_path)
            return target_path
    except OSError:
        pass
    if not _escalation_archive_disabled():
        archive_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = target_path.with_name(f"{target_path.name}.tmp")

    last_error: PermissionError | None = None
    for attempt in range(1, attempts + 1):
        try:
            shutil.copy2(source_path, tmp_path)
            break
        except PermissionError as err:
            last_error = err
            if attempt == attempts:
                raise PermissionError(
                    f"Could not copy locked escalation source after {attempts} attempts: {source_path}"
                ) from err
            time.sleep(sleep_seconds)
    else:
        if last_error is not None:
            raise last_error

    if target_path.exists() and not _escalation_archive_disabled():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        archive_path = archive_dir / f"{target_path.stem}_{timestamp}{target_path.suffix}"
        shutil.move(str(target_path), str(archive_path))

    tmp_path.replace(target_path)
    return target_path


def read_escalation_rows(
    workbook_path: Path,
    *,
    loaded_at: datetime | None = None,
    today: date | None = None,
) -> list[dict[str, Any]]:
    """Read and normalize rows from the Escalations sheet."""
    loaded_at = loaded_at or datetime.now()
    today = today or date.today()
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Escalation workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Workbook missing required sheet {SHEET_NAME!r}")
        ws = wb[SHEET_NAME]
        headers = _read_headers(ws)
        rows: list[dict[str, Any]] = []
        for row_values in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            source_row = _source_dict(headers, row_values)
            if not _has_data(source_row.values()):
                continue
            rows.append(_normalize_row(source_row, loaded_at=loaded_at, today=today))
        _disambiguate_unique_keys(rows)
        return rows
    finally:
        wb.close()


def load_escalation_file(
    workbook_path: Path,
    db_path: Path,
    *,
    loaded_at: datetime | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    """Parse a workbook and replace escalation_lines in the supplied DB path."""
    rows = read_escalation_rows(workbook_path, loaded_at=loaded_at, today=today)
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with closing(sqlite3.connect(str(db_path))) as conn:
        init_db(conn)
        try:
            conn.execute("BEGIN")
            conn.execute("DELETE FROM escalation_lines")
            conn.executemany(_insert_sql(), [_row_tuple(row) for row in rows])
            conn.commit()
        except Exception:
            conn.rollback()
            raise
    return LoadResult(
        rows_loaded=len(rows),
        source=str(workbook_path),
        db_path=str(db_path),
    ).as_dict()


def load_escalations(
    *,
    source_path: Path | None = None,
    target_path: Path = ESCALATION_FILE,
    archive_dir: Path = ESCALATION_ARCHIVE,
    db_path: Path = ESCALATION_DB,
    dry_run: bool = False,
    loaded_at: datetime | None = None,
    today: date | None = None,
    sync_after: bool = True,
) -> dict[str, Any]:
    """Refresh source file, parse it, and load escalation_lines."""
    resolved_source = _resolve_source_path(source_path)
    if dry_run:
        rows = read_escalation_rows(resolved_source, loaded_at=loaded_at, today=today)
        return LoadResult(
            rows_loaded=len(rows),
            source=str(resolved_source),
            db_path=None,
            dry_run=True,
        ).as_dict()

    local_file = refresh_escalation_file(resolved_source, target_path, archive_dir)
    result = load_escalation_file(local_file, db_path, loaded_at=loaded_at, today=today)
    if sync_after:
        sync_db_to_fixture_store()
    return result


def _read_headers(ws) -> list[str | None]:
    values = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    headers: list[str | None] = []
    for value in values:
        header = _strip_text(value, dash_to_none=False)
        headers.append(header)
    return headers


def _source_dict(headers: list[str | None], values: Iterable[Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for header, value in zip(headers, values):
        if header is None:
            continue
        result[header] = value
    return result


def _has_data(values: Iterable[Any]) -> bool:
    return any(_strip_text(value) is not None for value in values)


def _normalize_row(source: dict[str, Any], *, loaded_at: datetime, today: date) -> dict[str, Any]:
    row: dict[str, Any] = {column: None for column in INSERT_COLUMNS}
    row["LoadedAt"] = loaded_at.strftime("%Y-%m-%d %H:%M:%S")
    # D13: Python loader trusts source columns; master enrichment is deferred.
    row["MasterCategory"] = None
    row["MasterPriority"] = None
    row["MasterAPOwner"] = None

    for source_name, db_name in SOURCE_TO_DB.items():
        raw_value = source.get(source_name)
        if db_name == "StatusRaw":
            row[db_name] = _raw_text(raw_value)
        elif db_name in {"DocDate", "ReceivedDate", "EscalationDate", "DateResolved"}:
            row[db_name] = _normalize_date(raw_value)
        elif db_name == "DaysToResolveSource":
            row[db_name] = _to_int(raw_value)
        elif db_name == "APOwner":
            row[db_name] = _strip_text(raw_value, remove_nbsp=True)
        else:
            row[db_name] = _strip_text(raw_value)

    mailbox = row["Mailbox"] or "Unknown"
    row["Mailbox"] = mailbox
    row["Status"] = _strip_text(row["StatusRaw"]) or ""
    row["IsOpen"] = 0 if row["Status"] == "Closed" else 1
    row["EntityCode"] = _entity_code(row["Entity"])

    value_raw = _strip_text(source.get("Value"))
    row["ValueRaw"] = value_raw
    row["Value"] = _parse_float(value_raw)

    row["DaysToResolveCalc"] = _business_days_minus_one(row["EscalationDate"], row["DateResolved"])
    open_start_date = row["EscalationDate"] or row["ReceivedDate"]
    row["DaysOpen"] = (
        _business_days_minus_one(open_start_date, today.isoformat())
        if row["IsOpen"]
        else None
    )
    row["UniqueKey"] = row["UniqueKey"] or _fallback_unique_key(row)
    row["Flags"] = _flags(row)
    return row


def _raw_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.replace("\xa0", " ")
    return str(value)


def _strip_text(value: Any, *, dash_to_none: bool = True, remove_nbsp: bool = True) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value)
    if remove_nbsp:
        text = text.replace("\xa0", " ")
    stripped = text.strip()
    if not stripped:
        return None
    if dash_to_none and stripped == "-":
        return None
    return stripped


def _normalize_date(value: Any) -> str | None:
    cleaned = _strip_text(value)
    if cleaned is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            excel_date = from_excel(value)
            return excel_date.date().isoformat()
        except (TypeError, ValueError):
            return None
    if _EXCEL_SERIAL_DATE_RE.match(cleaned):
        try:
            excel_date = from_excel(float(cleaned))
            return excel_date.date().isoformat()
        except (TypeError, ValueError):
            return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(cleaned, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(cleaned).date().isoformat()
    except ValueError:
        return None


def _parse_float(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = (
        value.replace(",", "")
        .replace("\u00a3", "")
        .replace("$", "")
        .replace("\u20ac", "")
        .strip()
    )
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    cleaned = _strip_text(value)
    if cleaned is None:
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def _entity_code(entity: Any) -> str | None:
    text = _strip_text(entity)
    if text is None:
        return None
    match = ENTITY_CODE_RE.match(text)
    return match.group(1).strip() if match else text


def _business_days_minus_one(start_iso: str | None, end_iso: str | None) -> int | None:
    if start_iso is None or end_iso is None:
        return None
    try:
        end_exclusive = date.fromisoformat(end_iso) + timedelta(days=1)
        days = np.busday_count(start_iso, end_exclusive.isoformat()) - 1
    except ValueError:
        return None
    return int(days)


def _resolve_source_path(source_path: Path | None) -> Path:
    if source_path is not None:
        return Path(source_path)
    if ESCALATION_SOURCE is not None:
        return ESCALATION_SOURCE
    return get_escalation_source()


def _fallback_unique_key(row: dict[str, Any]) -> str:
    raw = f"{row['Mailbox']}|{row['VendorNo']}|{row['Reference']}|{row['EscalationDate']}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _disambiguate_unique_keys(rows: list[dict[str, Any]]) -> None:
    seen_counts: dict[str, int] = {}
    used_keys: set[str] = set()
    for row in rows:
        base_key = str(row["UniqueKey"])
        seen_counts[base_key] = seen_counts.get(base_key, 0) + 1
        occurrence = seen_counts[base_key]
        if occurrence == 1 and base_key not in used_keys:
            used_keys.add(base_key)
            continue

        suffix = occurrence
        candidate = f"{base_key}__dup{suffix}"
        while candidate in used_keys:
            suffix += 1
            candidate = f"{base_key}__dup{suffix}"
        row["UniqueKey"] = candidate
        _append_flag(row, "duplicate_key")
        used_keys.add(candidate)


def _append_flag(row: dict[str, Any], flag: str) -> None:
    flags = [part for part in str(row.get("Flags") or "").split(",") if part]
    if flag not in flags:
        flags.append(flag)
    row["Flags"] = ",".join(flags)


def _flags(row: dict[str, Any]) -> str:
    flags: list[str] = []
    vendor_no = row.get("VendorNo")
    if isinstance(vendor_no, str) and vendor_no.upper() == "#N/A":
        flags.append("vendor_na")
    if row.get("ValueRaw") is not None and row.get("Value") is None:
        flags.append("multi_value")
    days_open = row.get("DaysOpen")
    if row.get("IsOpen") == 1 and days_open is not None and int(days_open) > STALE_OPEN_DAYS:
        flags.append("stale_open")
    return ",".join(flags)


def _insert_sql() -> str:
    placeholders = ", ".join("?" for _ in INSERT_COLUMNS)
    columns = ", ".join(INSERT_COLUMNS)
    return f"INSERT INTO escalation_lines ({columns}) VALUES ({placeholders})"


def _row_tuple(row: dict[str, Any]) -> tuple[Any, ...]:
    return tuple(row.get(column) for column in INSERT_COLUMNS)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Load Escalations workbook into escalation_daily.sqlite")
    parser.add_argument("--dry-run", action="store_true", help="Parse source workbook without DB writes or sync")
    return parser.parse_args()


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    args = _parse_args()
    try:
        result = load_escalations(dry_run=args.dry_run)
    except (FileNotFoundError, ValueError, PermissionError, RuntimeError, sqlite3.Error) as err:
        logger.error("%s", err)
        return 1

    mode = "DRY RUN" if result["dry_run"] else "LOAD"
    logger.info("%s parsed %d row(s) from %s", mode, result["rows_loaded"], result["source"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
